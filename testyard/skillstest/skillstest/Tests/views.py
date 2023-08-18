from django.conf import settings
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.core.context_processors import csrf
from django.views.generic import View
from django.http import HttpResponseBadRequest, HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.shortcuts import render_to_response
from django.template import RequestContext
#from django.utils import simplejson
from django.db.models import Q
from django.template.response import TemplateResponse
from django.utils.http import base36_to_int, is_safe_url
from django.template import Template, Context
from django.template.loader import get_template
from django.contrib.sites.models import get_current_site
from django.contrib.sessions.backends.db import SessionStore
from django.core.mail import send_mail
from passlib.hash import pbkdf2_sha256 # To create hash of passwords
from django.utils.encoding import smart_text
from django.utils import timezone as utilstimezone
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile


# Standard libraries...
import os, sys, re, time, datetime, stat, gzip
from datetime import timedelta
import pytz
from pytz import timezone
import cPickle
import decimal, math
from Crypto.Cipher import AES, DES3
from Crypto import Random
import base64, urllib, urllib2, httplib, requests
import simplejson as json
from openpyxl import load_workbook
from xlrd import open_workbook
import csv, string
import xml.etree.ElementTree as et
from itertools import chain
import pyaudio, wave
import binascii
from BeautifulSoup import BeautifulSoup
from linkedin_v2 import linkedin
from oauthlib import *
import requests
import speech_recognition as sr
from dateutil import tz
import socket

# Application specific libraries...
from skillstest.Auth.models import User, Session, Privilege, UserPrivilege
from skillstest.Subscription.models import Plan, UserPlan, Transaction, PlanExtensions
from skillstest.Tests.models import Topic, Subtopic, Evaluator, Test, UserTest, Challenge, UserResponse, WouldbeUsers, EmailFailure, Schedule, Interview, InterviewQuestions, InterviewCandidates, PostLinkedin
from skillstest import settings as mysettings
from skillstest.errors import error_msg
import skillstest.utils as skillutils
from skillstest.utils import Logger, connectdb, disconnectdb, connectdb_p
from skillstest.Tests.tasks import send_emails


@skillutils.is_session_valid
@skillutils.session_location_match
def getautocompletelists(request):
    if request.method != 'GET':
        message = error_msg('1004')
        # A logging mechanism may be used to track how many and from where
        # such requests come and that may, sometimes, tell a curious story.
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    asrole = ""
    asrole = request.GET.get("asrole", "")
    if asrole == "":
        message = "Critical parameter missing"
        return HttpResponse(json.dumps([message,]))
    testnameslist = []
    testobjects = []
    uniquenames = {}
    if asrole == "ascreator":
        testobjects = Test.objects.filter(creator=userobj)
        for t in testobjects:
            if uniquenames.has_key(t.testname):
                continue
            uniquenames[t.testname] = 1
            testnameslist.append(str(t.testname))
    elif asrole == "asevaluator":
        evaluator_groups = Evaluator.objects.filter(Q(groupmember1=userobj)|Q(groupmember2=userobj)|Q(groupmember3=userobj)| \
                                                Q(groupmember4=userobj)|Q(groupmember5=userobj)|Q(groupmember6=userobj)| \
                                                Q(groupmember7=userobj)|Q(groupmember8=userobj)|Q(groupmember9=userobj)| \
                                                Q(groupmember10=userobj))
        testobjects = Test.objects.filter(evaluator__in=evaluator_groups)
        for t in testobjects:
            if uniquenames.has_key(t.testname):
                continue
            uniquenames[t.testname] = 1
            testnameslist.append(str(t.testname))
    elif asrole == "ascandidate":
        try:
            usertestqset = UserTest.objects.filter(user=userobj)
            wouldbeqset = WouldbeUsers.objects.filter(emailaddr=userobj.emailid)
        except: # Can't say if we will find any records...
            pass
        for usertestobj in usertestqset:
            testobjects.append(usertestobj.test)
        for wbtestobj in wouldbeqset:
            testobjects.append(wbtestobj.test)
        for t in testobjects:
            if uniquenames.has_key(t.testname):
                continue
            uniquenames[t.testname] = 1
            testnameslist.append(str(t.testname))
    elif asrole == "asinterviewer":
        testobjects = Interview.objects.filter(interviewer=userobj) # these would be Interview objects
        for t in testobjects:
            if uniquenames.has_key(t.title):
                continue
            uniquenames[t.title] = 1
            testnameslist.append(str(t.title)) 
    elif asrole == "asinterviewee":
        testobjects = InterviewCandidates.objects.filter(emailaddr=userobj.emailid) # these would be InterviewCandidate objects
        for t in testobjects:
            if uniquenames.has_key(t.interview.title):
                continue
            uniquenames[t.interview.title] = 1
            testnameslist.append(str(t.interview.title)) 
    return HttpResponse(json.dumps(testnameslist))


def get_user_tests(request):
    # If request method is 'GET', then retrieve Session and User info from the DB
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    pageno_ascreator, pageno_asevaluator, pageno_ascandidate, pageno_asinterviewer, pageno_asinterviewee = 1, 1, 1, 1, 1
    startctr_ascreator, startctr_asevaluator, startctr_ascandidate, startctr_asinterviewer, startctr_asinterviewee = 0, 0, 0, 0, 0
    first_ascreator, last_ascreator, first_asevaluator, last_asevaluator, first_ascandidate, last_ascandidate, first_asinterviewer, last_asinterviewer, first_asinterviewee, last_asinterviewee = 1, 100, 1, 100, 1, 100, 1, 100, 1, 100
    endctr_ascreator, endctr_asevaluator, endctr_ascandidate, endctr_asinterviewer, endctr_asinterviewee = mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE
    # Retrieve pageno for all sections from GET query parameters
    if request.GET.has_key('pageno_creator'):
        try:
            pageno_ascreator = int(request.GET.get('pageno_creator', 1))
        except:
            pass
    if request.GET.has_key('pageno_evaluator'):
        try:
            pageno_asevaluator = int(request.GET.get('pageno_evaluator', 1))
        except:
            pass
    if request.GET.has_key('pageno_candidate'):
        try:
            pageno_ascandidate = int(request.GET.get('pageno_candidate', 1))
        except:
            pass
    if request.GET.has_key('pageno_interviewer'):
        try:
            pageno_asinterviewer = int(request.GET.get('pageno_interviewer', 1))
        except:
            pass
    if request.GET.has_key('pageno_interviewee'):
        try:
            pageno_asinterviewee = int(request.GET.get('pageno_interviewee', 1))
        except:
            pass
    # Compute the start and end ctr for all 5 sections
    startctr_ascreator = pageno_ascreator * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_ascreator = pageno_ascreator * mysettings.PAGE_CHUNK_SIZE
    startctr_asevaluator = pageno_asevaluator * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_asevaluator = pageno_asevaluator * mysettings.PAGE_CHUNK_SIZE
    startctr_ascandidate = pageno_ascandidate * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_ascandidate = pageno_ascandidate * mysettings.PAGE_CHUNK_SIZE
    startctr_asinterviewer = pageno_asinterviewer * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_asinterviewer = pageno_asinterviewer * mysettings.PAGE_CHUNK_SIZE
    startctr_asinterviewee = pageno_asinterviewee * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_asinterviewee = pageno_asinterviewee * mysettings.PAGE_CHUNK_SIZE
    # Compute nextpage and prevpage for all 5 sections.
    nextpage_ascreator = pageno_ascreator + 1
    prevpage_ascreator = pageno_ascreator - 1
    nextpage_asevaluator = pageno_asevaluator + 1
    prevpage_asevaluator = pageno_asevaluator - 1
    nextpage_ascandidate = pageno_ascandidate + 1
    prevpage_ascandidate = pageno_ascandidate - 1
    nextpage_asinterviewer = pageno_asinterviewer + 1
    prevpage_asinterviewer = pageno_asinterviewer - 1
    nextpage_asinterviewee = pageno_asinterviewee + 1
    prevpage_asinterviewee = pageno_asinterviewee - 1
    # Note: the above page numbers are to be added in context at the end of the function.
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    testlist_ascreator = Test.objects.filter(creator=userobj).order_by('-createdate')[startctr_ascreator:endctr_ascreator]
    # Determine if the user should be shown the "Create Test" link
    createlink, testtypes, testrules, testtopics, skilltarget, testscope, answeringlanguage, progenv, existingtestnames, assocevalgrps, evalgroupslitags, createtesturl, addeditchallengeurl, savechangesurl, addmoreurl, clearnegativescoreurl, deletetesturl, showuserviewurl, editchallengeurl, showtestcandidatemode, sendtestinvitationurl, manageinvitationsurl, invitationactivationurl, invitationcancelurl, uploadlink, testbulkuploadurl, testevaluationurl, evaluateresponseurl, getevaluationdetailsurl, settestvisibilityurl, getcanvasurl, savedrawingurl, disqualifycandidateurl, copytesturl, gettestscheduleurl, activatetestbycreator, deactivatetestbycreator, interviewlink, createinterviewurl, chkintnameavailabilityurl, uploadrecordingurl, codepadexecuteurl, postonlinkedinurl, linkedinpostsessionurl, showevaluationscreen, max_interviewers_count, codeexecurl, latexkbdurl, addtogooglecalendarurl, showinterviewschedulescreenurl = "", "", "", "", "", "", "", "", "", "var evalgrpsdict = {};", "", mysettings.CREATE_TEST_URL, mysettings.EDIT_TEST_URL, mysettings.SAVE_CHANGES_URL, mysettings.ADD_MORE_URL, mysettings.CLEAR_NEGATIVE_SCORE_URL, mysettings.DELETE_TEST_URL, mysettings.SHOW_USER_VIEW_URL, mysettings.EDIT_CHALLENGE_URL, mysettings.SHOW_TEST_CANDIDATE_MODE_URL, mysettings.SEND_TEST_INVITATION_URL, mysettings.MANAGE_INVITATIONS_URL, mysettings.INVITATION_ACTIVATION_URL, mysettings.INVITATION_CANCEL_URL, "", mysettings.TEST_BULK_UPLOAD_URL, mysettings.TEST_EVALUATION_URL, mysettings.EVALUATE_RESPONSE_URL, mysettings.GET_CURRENT_EVALUATION_DATA_URL, mysettings.SET_VISIBILITY_URL, mysettings.GET_CANVAS_URL, mysettings.SAVE_DRAWING_URL, mysettings.DISQUALIFY_CANDIDATE_URL, mysettings.COPY_TEST_URL, mysettings.GET_TEST_SCHEDULE_URL, mysettings.ACTIVATE_TEST_BY_CREATOR, mysettings.DEACTIVATE_TEST_BY_CREATOR, "", mysettings.CREATE_INTERVIEW_URL, mysettings.CHECK_INT_NAME_AVAILABILITY_URL, mysettings.UPLOAD_RECORDING_URL, mysettings.CODEPAD_EXECUTE_URL, mysettings.POST_ON_LINKEDIN_URL, mysettings.LINKEDINPOSTSESS_URL, mysettings.SHOW_EVAL_SCREEN, mysettings.MAX_INTERVIEWERS_COUNT, mysettings.CODE_EXEC_URL, mysettings.LATEX_KEYBOARD_URL, mysettings.GOOGLE_CALENDAR_URL, mysettings.SHOW_INTERVIEW_SCHEDULE_SCREEN_URL
    # Check for user's subscription plan; If user is a "Business Plan" user, 
    # then we will need to provide a test and interview data download link.
    # NOTE: If the user is on a Business Plan extension, she/he/other should not
    # be able to download the data.
    busplanobj = None
    busplanflag = mysettings.DEFAULT_BUSPLAN_FLAG
    downloadtestninterviewsdatalink = "" 
    try:
        busplanobj = Plan.objects.get(planname="Business Plan")
    except:
        pass
    if busplanobj is not None:
        userplanqset = UserPlan.objects.filter(plan=busplanobj, user=userobj, planstatus=True).order_by('-subscribedon')
        curdate = datetime.datetime.now()
        for userplanobj in userplanqset:
            if curdate >= userplanobj.planstartdate.replace(tzinfo=None) and curdate <  userplanobj.planenddate.replace(tzinfo=None):
                busplanflag = True
                break
    if busplanflag is True:
        downloadtestninterviewsdatalink = "<h3><img src='static/images/download.png' style='height:25px;width:25px;'>&nbsp;&nbsp; <a href='#/' onclick='javascript:showformatoptions();'>Download My Data</a></h3><span id='waiterspan' style='display:none;'></span>"
    test_created_count = Test.objects.filter(creator=userobj).count()
    if test_created_count <= mysettings.NEW_USER_FREE_TESTS_COUNT: # Also add condition to check user's 'plan' (to be done later)
        createlink = "<a href='#' onClick='javascript:showcreatetestform(&quot;%s&quot;);loaddatepicker();'>Create New Test</a>"%userobj.id
        uploadlink = "<a href='#' onClick='javascript:showuploadtestform(&quot;%s&quot;);loaddatepicker();'>Upload New Test</a>"%userobj.id
        interviewlink = "<a href='#/' onClick='javascript:showcreateinterviewform(&quot;%s&quot;);loaddatepicker();'>Start Video Call</a>"%(userobj.id)
        for ttcode in mysettings.TEST_TYPES.keys():
            ttcodeval = ttcode.replace(" ", "__")
            if ttcode == 'COMP':
                testtypes += "<option value=&quot;%s&quot; selected>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
            else:
                testtypes += "<option value=&quot;%s&quot;>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
        for trule in mysettings.RULES_DICT.keys():
            testrules += "<option value=&quot;%s&quot;>%s</option>"%(trule, mysettings.RULES_DICT[trule])

        unique_topics = {}
        for ttopics in mysettings.TEST_TOPICS:
            ttopicsval = ttopics.replace(" ", "__")
            if not unique_topics.has_key(ttopicsval):
                unique_topics[ttopicsval] = 1
                testtopics += "<option value=&quot;%s&quot;>%s</option>"%(ttopicsval, ttopics)
        # Get topics created in the past by this user
        usertopics = Topic.objects.filter(user=userobj, isactive=True)
        for topic in usertopics:
            topicname = topic.topicname.replace(" ", "__")
            topicname = topicname.replace("+", "__")
            if not unique_topics.has_key(topicname):
                unique_topics[topicname] = 1
                testtopics += "<option value=&quot;%s&quot;>%s</option>"%(topicname, topic.topicname)
        for skillcode in mysettings.SKILL_QUALITY.keys():
            skilltarget += "<option value=&quot;%s&quot;>%s</option>"%(skillcode, mysettings.SKILL_QUALITY[skillcode])
        for tscope in mysettings.TEST_SCOPES:
            if tscope == 'private':
                testscope += "<option value=&quot;%s&quot; selected>%s</option>"%(tscope, tscope)
            else:
                testscope += "<option value=&quot;%s&quot;>%s</option>"%(tscope, tscope)
        for alang in mysettings.ANSWER_LANG_DICT.keys():
            if alang == 'enus':
                answeringlanguage += "<option value=&quot;%s&quot; selected>%s</option>"%(alang, mysettings.ANSWER_LANG_DICT[alang])
            else:
                answeringlanguage += "<option value=&quot;%s&quot;>%s</option>"%(alang, mysettings.ANSWER_LANG_DICT[alang])
        progenv += "<option value=&quot;0&quot; selected>None</option>"
        for proglang in mysettings.COMPILER_LOCATIONS.keys():
            progenv += "<option value=&quot;%s&quot;>Yes - %s</option>"%(proglang, proglang)
    last_ascreator = (Test.objects.filter(creator=userobj).count()/mysettings.PAGE_CHUNK_SIZE) + 1
    evaluator_groups = Evaluator.objects.filter(Q(groupmember1=userobj)|Q(groupmember2=userobj)|Q(groupmember3=userobj)| \
                                                Q(groupmember4=userobj)|Q(groupmember5=userobj)|Q(groupmember6=userobj)| \
                                                Q(groupmember7=userobj)|Q(groupmember8=userobj)|Q(groupmember9=userobj)| \
                                                Q(groupmember10=userobj))
    testlist_asevaluator = Test.objects.filter(evaluator__in=evaluator_groups).order_by('-createdate')[startctr_asevaluator:endctr_asevaluator]
    user_creator_other_evaluators_dict = {}
    tests_creator_ordered_createdate = []
    uniqevalgroups = {}
    for test in testlist_ascreator:
        tests_creator_ordered_createdate.append(test.testname)
        user_creator_other_evaluators_dict[test.testname] = ( test.evaluator.groupmember1, test.evaluator.groupmember2, \
                                                              test.evaluator.groupmember3, test.evaluator.groupmember4, test.evaluator.groupmember5, \
                                                              test.evaluator.groupmember6, test.evaluator.groupmember7, test.evaluator.groupmember8, \
                                                              test.evaluator.groupmember9, test.evaluator.groupmember10 )
        evalgrpemaillist = []
        if test.evaluator.groupmember1 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember1.emailid)
        if test.evaluator.groupmember2 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember2.emailid)
        if test.evaluator.groupmember3 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember3.emailid)
        if test.evaluator.groupmember4 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember4.emailid)
        if test.evaluator.groupmember5 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember5.emailid)
        if test.evaluator.groupmember6 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember6.emailid)
        if test.evaluator.groupmember7 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember7.emailid)
        if test.evaluator.groupmember8 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember8.emailid)
        if test.evaluator.groupmember9 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember9.emailid)
        if test.evaluator.groupmember10 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember10.emailid)
        evalgrpemails = ",".join(evalgrpemaillist)
        test.evaluator.evalgroupname = test.evaluator.evalgroupname.replace("@", "__")
        test.evaluator.evalgroupname = test.evaluator.evalgroupname.replace(".", "__")
        assocevalgrps += "evalgrpsdict.%s = '%s';"%(test.evaluator.evalgroupname, evalgrpemails)
        #evalgroupslitags += "<li id=&quot;%s&quot; title=&quot;%s&quot;>%s</li>"%(test.evaluator.evalgroupname, evalgrpemails, test.evaluator.evalgroupname)
        if not uniqevalgroups.has_key(test.evaluator.evalgroupname):
            shortevalgrpname = test.evaluator.evalgroupname
            if shortevalgrpname.__len__() > 20:
                shortevalgrpname = shortevalgrpname[:20] + "..."
            #evalgroupslitags += "<li id=&quot;%s&quot; title=&quot;%s&quot; draggable=&quot;true&quot;>%s</li>"%(test.evaluator.evalgroupname, evalgrpemails, shortevalgrpname)
            evalgroupslitags += "<option value=&quot;%s&quot; title=&quot;%s&quot; draggable=&quot;true&quot;>%s</option>"%(test.evaluator.evalgroupname, evalgrpemails, shortevalgrpname)
            evalgroupslitags = evalgroupslitags.replace('&quot;', '\\"')
            uniqevalgroups[test.evaluator.evalgroupname] = test.evaluator.evalgroupname

    user_evaluator_creator_other_evaluators_dict = {}
    tests_evaluator_ordered_createdate = []
    test = None
    last_asevaluator = (Test.objects.filter(evaluator__in=evaluator_groups).count()/mysettings.PAGE_CHUNK_SIZE) + 1
    for test in testlist_asevaluator:
        testcreator = test.creator
        testname = test.testname
        creator_evaluators = ( testcreator, test.evaluator.groupmember1, test.evaluator.groupmember2, test.evaluator.groupmember3, test.evaluator.groupmember4, \
                          test.evaluator.groupmember5, test.evaluator.groupmember6, test.evaluator.groupmember7, test.evaluator.groupmember8, \
                          test.evaluator.groupmember9, test.evaluator.groupmember10 ) # Basically we keep the creator as the first element. Rest are evaluators.
        tests_evaluator_ordered_createdate.append(testname)
        user_evaluator_creator_other_evaluators_dict[testname] = creator_evaluators
    try:
        usertestqset = UserTest.objects.filter(user=userobj)[startctr_ascandidate:endctr_ascandidate]
    except: # Can't say if we will find any records...
        usertestqset = WouldbeUsers.objects.filter(emailaddr=userobj.emailid)[startctr_ascandidate:endctr_ascandidate]
    testlist_ascandidate = []
    tests_candidate_ordered_createdate = []
    for usertestobj in usertestqset:
        try:
            testlist_ascandidate.append(usertestobj.test)
        except:
            pass
    user_candidate_other_creator_evaluator_dict = {}
    candidatescount = 0
    try:
        candidatescount += UserTest.objects.filter(user=userobj).count()
    except:
        pass
    try:
        candidatescount += WouldbeUsers.objects.filter(emailaddr=userobj.emailid).count()
    except:
        pass
    last_ascandidate = (candidatescount/mysettings.PAGE_CHUNK_SIZE) + 1
    for test in testlist_ascandidate:
        testcreator = test.creator
        creator_evaluators = ( testcreator, test.evaluator.groupmember1, test.evaluator.groupmember2, test.evaluator.groupmember3, test.evaluator.groupmember4, \
                          test.evaluator.groupmember5, test.evaluator.groupmember6, test.evaluator.groupmember7, test.evaluator.groupmember8, \
                          test.evaluator.groupmember9, test.evaluator.groupmember10 )
        if not user_candidate_other_creator_evaluator_dict.has_key(test.testname):
            user_candidate_other_creator_evaluator_dict[test.testname] = creator_evaluators
        else:
            continue
        tests_candidate_ordered_createdate.append(test.testname)
    interviewsasinterviewer = Interview.objects.filter(interviewer=userobj).order_by('-createdate')[startctr_asinterviewer:endctr_asinterviewer]
    interviews_list = {}
    interviewerslist = []
    interviews_list['asinterviewer'] = {}
    asinterviewer_seq = []
    last_asinterviewer = (Interview.objects.filter(interviewer=userobj).count()/mysettings.PAGE_CHUNK_SIZE) + 1
    for interview in interviewsasinterviewer:
        inttitle = interview.title
        asinterviewer_seq.append(inttitle)
        #inttopic = interview.topic
        inttopic = ""
        inttopicname = interview.topicname
        intmedium = interview.medium
        intlanguage = interview.language
        intcreatedate = interview.createdate
        intpublishdate = interview.publishdate
        intstatus = interview.status
        intmaxscore = interview.maxscore
        interviewerslist = []
        intmaxduration_min = interview.maxduration/60
        intmaxduration_hr = 0
        if intmaxduration_min >= 60:
            intmaxduration_hr = intmaxduration_min/60
        if intmaxduration_hr >= 1:
            if intmaxduration_hr == 1:
                intmaxduration = str(intmaxduration_hr) + " hour"
            else:
                intmaxduration = str(intmaxduration_hr) + " hours"
            intmaxduration_min = (intmaxduration_hr - int(intmaxduration_hr)) * 60
            if intmaxduration_min > 0:
                intmaxduration += " " + str(intmaxduration_min) + " minutes"
        else:
            intmaxduration = str(intmaxduration_min) + " minutes"
        intrealtime = interview.realtime
        intlinkid = interview.interviewlinkid
        intdata = (inttitle, inttopic, inttopicname, intmedium, intlanguage, intcreatedate, intpublishdate, intstatus, intmaxscore, intmaxduration, intrealtime, intlinkid, interviewerslist)
        interviews_list['asinterviewer'][inttitle] = intdata
    interviewsasinterviewees = InterviewCandidates.objects.filter(emailaddr=userobj.emailid).order_by('-scheduledtime')[startctr_asinterviewee:endctr_asinterviewee]
    interviews_list['asinterviewee'] = {}
    asinterviewee_seq = []
    last_asinterviewee = (InterviewCandidates.objects.filter(emailaddr=userobj.emailid).count()/mysettings.PAGE_CHUNK_SIZE) + 1
    for intcandidate in interviewsasinterviewees:
        interviewname = intcandidate.interview.title
        asinterviewee_seq.append(interviewname)
        #interviewtopic = intcandidate.interview.topic
        interviewtopic = ""
        interviewtopicname = intcandidate.interview.topicname
        interviewmedium = intcandidate.interview.medium
        interviewlanguage = intcandidate.interview.language
        interviewcreatedate = intcandidate.interview.createdate
        interviewpublishdate = intcandidate.interview.publishdate
        interviewmaxscore = intcandidate.interview.maxscore
        interviewerid = intcandidate.interview.interviewer_id
        interviewerobj = User.objects.get(id=interviewerid)
        interviewername = interviewerobj.displayname
        interviewstatus = intcandidate.interview.status
        interviewmaxduration = intcandidate.interview.maxduration
        intmaxduration_min = intcandidate.interview.maxduration/60
        if intmaxduration_min >= 60:
            intmaxduration_hr = intmaxduration_min/60
        if intmaxduration_hr >= 1:
            if intmaxduration_hr == 1:
                interviewmaxduration = str(intmaxduration_hr) + " hour"
            else:
                interviewmaxduration = str(intmaxduration_hr) + " hours"
            intmaxduration_min = (intmaxduration_hr - int(intmaxduration_hr)) * 60
            if intmaxduration_min > 0:
                interviewmaxduration += " " + str(intmaxduration_min) + " minutes"
        else:
            interviewmaxduration = str(intmaxduration_min) + " minutes"
        interviewrealtime = intcandidate.interview.realtime
        interviewscheduledtime = intcandidate.scheduledtime
        interviewactualstarttime = intcandidate.actualstarttime
        interviewtotaltimetaken = intcandidate.totaltimetaken
        interviewurl = intcandidate.interviewurl
        intdata = (interviewname, interviewtopic, interviewtopicname, interviewmedium, interviewlanguage, interviewcreatedate, interviewpublishdate, interviewstatus, interviewmaxscore, interviewmaxduration, interviewrealtime, interviewscheduledtime, interviewactualstarttime, interviewtotaltimetaken, interviewurl, interviewername)
        interviews_list['asinterviewee'][interviewname] = intdata
    tests_user_dict = {}
    tests_user_dict['downloadmydata'] = downloadtestninterviewsdatalink
    tests_user_dict['downloadmydataurl'] = skillutils.gethosturl(request) + "/" + mysettings.DOWNLOAD_MY_DATA_URL
    tests_user_dict['user_creator_other_evaluators_dict'] = user_creator_other_evaluators_dict
    tests_user_dict['user_evaluator_creator_other_evaluators_dict'] = user_evaluator_creator_other_evaluators_dict
    tests_user_dict['user_candidate_other_creator_evaluator_dict'] = user_candidate_other_creator_evaluator_dict
    tests_user_dict['testlist_ascreator'] = testlist_ascreator
    tests_user_dict['testlist_asevaluator'] = testlist_asevaluator
    tests_user_dict['testlist_ascandidate'] = testlist_ascandidate
    tests_user_dict['interviewlist_asinterviewer'] = interviews_list['asinterviewer']
    tests_user_dict['asinterviewer_seq'] = asinterviewer_seq
    tests_user_dict['interviewlist_asinterviewee'] = interviews_list['asinterviewee']
    tests_user_dict['asinterviewee_seq'] = asinterviewee_seq
    tests_user_dict['profile_image_tag'] = skillutils.getprofileimgtag(request)
    tests_user_dict['displayname'] = userobj.displayname
    tests_user_dict['createlink'] = createlink
    tests_user_dict['uploadlink'] = uploadlink
    tests_user_dict['interviewlink'] = interviewlink
    tests_user_dict['testtypes'] = testtypes
    tests_user_dict['testrules'] = testrules
    tests_user_dict['testtopics'] = testtopics
    tests_user_dict['interviewtopics'] = testtopics
    tests_user_dict['skilltarget'] = skilltarget
    tests_user_dict['answeringlanguage'] = answeringlanguage
    tests_user_dict['testscope'] = testscope
    tests_user_dict['interviewscope'] = testscope
    tests_user_dict['progenv'] = progenv
    tests_user_dict['creatoremail'] = userobj.emailid
    tests_user_dict['existingtestnames'] = "','".join(user_creator_other_evaluators_dict.keys())
    tests_user_dict['existingtestnames'] = "'" + tests_user_dict['existingtestnames'] + "'"
    tests_user_dict['getautocompletelisturl'] = skillutils.gethosturl(request) + "/" + mysettings.GET_AUTOCOMPLETE_LIST_URL
    tests_user_dict['assocevalgrps'] = assocevalgrps
    tests_user_dict['evalgroupslitags'] = evalgroupslitags
    tests_user_dict['createtesturl'] = createtesturl
    tests_user_dict['createinterviewurl'] = createinterviewurl
    tests_user_dict['getsectionurl'] = skillutils.gethosturl(request) + "/" + mysettings.GET_SECTION_URL
    tests_user_dict['addeditchallengeurl'] = skillutils.gethosturl(request) + "/" + mysettings.EDIT_TEST_URL
    tests_user_dict['edittesturl'] = skillutils.gethosturl(request) + "/" + mysettings.EDIT_EXISTING_TEST_URL
    tests_user_dict['testsummaryurl'] = mysettings.TEST_SUMMARY_URL
    tests_user_dict['deletechallengesurl'] = skillutils.gethosturl(request) + "/" + mysettings.DELETE_CHALLENGE_URL
    tests_user_dict['savechangesurl'] = skillutils.gethosturl(request) + "/" + mysettings.SAVE_CHANGES_URL
    tests_user_dict['addmoreurl'] = skillutils.gethosturl(request) + "/" + addmoreurl
    tests_user_dict['clearnegativescoreurl'] = skillutils.gethosturl(request) + "/" + clearnegativescoreurl
    tests_user_dict['deletetesturl'] = skillutils.gethosturl(request) + "/" + deletetesturl
    tests_user_dict['showuserviewurl'] = skillutils.gethosturl(request) + "/" + showuserviewurl
    tests_user_dict['copytesturl'] = skillutils.gethosturl(request) + "/" + copytesturl
    tests_user_dict['editchallengeurl'] = skillutils.gethosturl(request) + "/" + editchallengeurl
    tests_user_dict['showtestcandidatemode'] = skillutils.gethosturl(request) + "/" + showtestcandidatemode
    tests_user_dict['sendtestinvitationurl'] = skillutils.gethosturl(request) + "/" + sendtestinvitationurl
    tests_user_dict['manageinvitationsurl'] = skillutils.gethosturl(request) + "/" + manageinvitationsurl
    tests_user_dict['invitationactivationurl'] = skillutils.gethosturl(request) + "/" + invitationactivationurl
    tests_user_dict['invitationcancelurl'] = skillutils.gethosturl(request) + "/" + invitationcancelurl
    tests_user_dict['hosturl'] = skillutils.gethosturl(request) 
    tests_user_dict['testlinkid'] = skillutils.generate_random_string()
    tests_user_dict['interviewlinkid'] = skillutils.generate_random_string()
    tests_user_dict['testbulkuploadurl'] = skillutils.gethosturl(request) + "/" + testbulkuploadurl
    tests_user_dict['testevaluationurl'] = skillutils.gethosturl(request) + "/" + testevaluationurl
    tests_user_dict['evaluateresponseurl'] = skillutils.gethosturl(request) + "/" + evaluateresponseurl
    tests_user_dict['showevaluationscreen'] = skillutils.gethosturl(request) + "/" + showevaluationscreen
    tests_user_dict['getevaluationdetailsurl'] = skillutils.gethosturl(request) + "/" + getevaluationdetailsurl
    tests_user_dict['settestvisibilityurl'] = skillutils.gethosturl(request) + "/" + settestvisibilityurl
    tests_user_dict['getcanvasurl'] = skillutils.gethosturl(request) + "/" + getcanvasurl
    tests_user_dict['searchbyroleurl'] = skillutils.gethosturl(request) + "/" + mysettings.ROLE_BASED_SEARCH_URL
    tests_user_dict['activatetestbycreator'] = skillutils.gethosturl(request) + "/" + activatetestbycreator
    tests_user_dict['deactivatetestbycreator'] = skillutils.gethosturl(request) + "/" + deactivatetestbycreator
    tests_user_dict['savedrawingurl'] = skillutils.gethosturl(request) + "/" + savedrawingurl
    tests_user_dict['disqualifycandidateurl'] = skillutils.gethosturl(request) + "/" + disqualifycandidateurl
    tests_user_dict['gettestscheduleurl'] = skillutils.gethosturl(request) + "/" + gettestscheduleurl
    tests_user_dict['tests_creator_ordered_createdate'] = tests_creator_ordered_createdate
    tests_user_dict['tests_evaluator_ordered_createdate'] = tests_evaluator_ordered_createdate
    tests_user_dict['tests_candidate_ordered_createdate'] = tests_candidate_ordered_createdate
    tests_user_dict['secret_key'] = mysettings.DES3_SECRET_KEY
    tests_user_dict['realtime'] = 1
    tests_user_dict['repl_token'] = skillutils.repl_token_generator()
    tests_user_dict['codepadexecuteurl'] = skillutils.gethosturl(request) + "/" + codepadexecuteurl
    tests_user_dict['codeexecurl'] = skillutils.gethosturl(request) + "/" + codeexecurl
    tests_user_dict['chkintnameavailabilityurl'] = skillutils.gethosturl(request) + "/" + chkintnameavailabilityurl
    tests_user_dict['postonlinkedinurl'] = skillutils.gethosturl(request) + "/" + postonlinkedinurl
    tests_user_dict['linkedinpostsessionurl'] = skillutils.gethosturl(request) + "/" + linkedinpostsessionurl
    tests_user_dict['maxinvitationspersession'] = mysettings.MAX_INVITES_PER_SESSION
    tests_user_dict['addtogooglecalendarurl'] = skillutils.gethosturl(request) + "/" + addtogooglecalendarurl
    tests_user_dict['savenewinterviewscheduleurl'] = skillutils.gethosturl(request) + "/" + mysettings.SAVE_NEW_INTERVIEW_SCHEDULE_URL
    tests_user_dict['rescheduleinterviewurl'] = skillutils.gethosturl(request) + "/" + mysettings.RESCHEDULE_INTERVIEW_URL
    for i in range(2, mysettings.MAX_INTERVIEWERS_COUNT + 1):
        interviewerslist.append(i)
    tests_user_dict['interviewerslist'] = interviewerslist
    inttitle = skillutils.generate_random_string()
    intqset = Interview.objects.filter(title=inttitle, interviewer=userobj)
    while intqset.__len__() > 0:
        inttitle = skillutils.generate_random_string()
        intqset = Interview.objects.filter(title=inttitle, interviewer=userobj)
    tests_user_dict['interviewtitle'] = inttitle + "-interview"
    tests_user_dict['showinterviewschedulescreenurl'] = skillutils.gethosturl(request) + "/" + showinterviewschedulescreenurl
    # pagination related vars
    tests_user_dict['nextpage_ascreator'] = nextpage_ascreator
    tests_user_dict['prevpage_ascreator'] = prevpage_ascreator
    tests_user_dict['first_ascreator'] = first_ascreator
    tests_user_dict['last_ascreator'] = last_ascreator
    tests_user_dict['nextpage_asevaluator'] = nextpage_asevaluator
    tests_user_dict['prevpage_asevaluator'] = prevpage_asevaluator
    tests_user_dict['first_asevaluator'] = first_asevaluator
    tests_user_dict['last_asevaluator'] = last_asevaluator
    tests_user_dict['nextpage_ascandidate'] = nextpage_ascandidate
    tests_user_dict['prevpage_ascandidate'] = prevpage_ascandidate
    tests_user_dict['first_ascandidate'] = first_ascandidate
    tests_user_dict['last_ascandidate'] = last_ascandidate
    tests_user_dict['nextpage_asinterviewer'] = nextpage_asinterviewer
    tests_user_dict['prevpage_asinterviewer'] = prevpage_asinterviewer
    tests_user_dict['first_asinterviewer'] = first_asinterviewer
    tests_user_dict['last_asinterviewer'] = last_asinterviewer
    tests_user_dict['nextpage_asinterviewee'] = nextpage_asinterviewee
    tests_user_dict['prevpage_asinterviewee'] = prevpage_asinterviewee
    tests_user_dict['first_asinterviewee'] = first_asinterviewee
    tests_user_dict['last_asinterviewee'] = last_asinterviewee
    tests_user_dict['currpage_ascreator'] = pageno_ascreator
    tests_user_dict['currpage_asevaluator'] = pageno_asevaluator
    tests_user_dict['currpage_ascandidate'] = pageno_ascandidate
    tests_user_dict['currpage_asinterviewer'] = pageno_asinterviewer
    tests_user_dict['currpage_asinterviewee'] = pageno_asinterviewee
    tests_user_dict['testspageurl'] = skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL 
    return  tests_user_dict


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def getsectiondata(request):
    message = ''
    if request.method != "GET": # Illegal bad request... 
        message = error_msg('1004')
        # A logging mechanism may be used to track how many and from where
        # such requests come and that may, sometimes, tell a curious story.
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    pageno_ascreator, pageno_asevaluator, pageno_ascandidate, pageno_asinterviewer, pageno_asinterviewee = 1, 1, 1, 1, 1
    startctr_ascreator, startctr_asevaluator, startctr_ascandidate, startctr_asinterviewer, startctr_asinterviewee = 0, 0, 0, 0, 0
    endctr_ascreator, endctr_asevaluator, endctr_ascandidate, endctr_asinterviewer, endctr_asinterviewee = mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE, mysettings.PAGE_CHUNK_SIZE
    currpage_ascreator, first_ascreator, last_ascreator = pageno_ascreator, 1, 100
    # Retrieve pageno for all sections from GET query parameters
    section = ""
    if request.GET.has_key('section'):
        section = request.GET['section']
    if request.GET.has_key('pageno_creator'):
        try:
            pageno_ascreator = int(request.GET.get('pageno_creator', 1))
        except:
            pass
    if request.GET.has_key('pageno_evaluator'):
        try:
            pageno_asevaluator = int(request.GET.get('pageno_evaluator', 1))
        except:
            pass
    if request.GET.has_key('pageno_candidate'):
        try:
            pageno_ascandidate = int(request.GET.get('pageno_candidate', 1))
        except:
            pass
    if request.GET.has_key('pageno_interviewer'):
        try:
            pageno_asinterviewer = int(request.GET.get('pageno_interviewer', 1))
        except:
            pass
    if request.GET.has_key('pageno_interviewee'):
        try:
            pageno_asinterviewee = int(request.GET.get('pageno_interviewee', 1))
        except:
            pass
    # Compute the start and end ctr for all 5 sections
    startctr_ascreator = pageno_ascreator * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_ascreator = pageno_ascreator * mysettings.PAGE_CHUNK_SIZE
    startctr_asevaluator = pageno_asevaluator * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_asevaluator = pageno_asevaluator * mysettings.PAGE_CHUNK_SIZE
    startctr_ascandidate = pageno_ascandidate * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_ascandidate = pageno_ascandidate * mysettings.PAGE_CHUNK_SIZE
    startctr_asinterviewer = pageno_asinterviewer * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_asinterviewer = pageno_asinterviewer * mysettings.PAGE_CHUNK_SIZE
    startctr_asinterviewee = pageno_asinterviewee * mysettings.PAGE_CHUNK_SIZE - mysettings.PAGE_CHUNK_SIZE
    endctr_asinterviewee = pageno_asinterviewee * mysettings.PAGE_CHUNK_SIZE
    # Compute nextpage and prevpage for all 5 sections.
    currpage_ascreator = pageno_ascreator
    first_ascreator = 1
    nextpage_ascreator = pageno_ascreator + 1
    prevpage_ascreator = pageno_ascreator - 1
    currpage_asevaluator = pageno_asevaluator
    first_asevaluator = 1
    nextpage_asevaluator = pageno_asevaluator + 1
    prevpage_asevaluator = pageno_asevaluator - 1
    currpage_ascandidate = pageno_ascandidate
    first_ascandidate = 1
    nextpage_ascandidate = pageno_ascandidate + 1
    prevpage_ascandidate = pageno_ascandidate - 1
    currpage_asinterviewer = pageno_asinterviewer
    first_asinterviewer = 1
    nextpage_asinterviewer = pageno_asinterviewer + 1
    prevpage_asinterviewer = pageno_asinterviewer - 1
    currpage_asinterviewee = pageno_asinterviewee
    first_asinterviewee = 1
    nextpage_asinterviewee = pageno_asinterviewee + 1
    prevpage_asinterviewee = pageno_asinterviewee - 1
    # Note: the above page numbers are to be added in context at the end of the function.
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    
    tests_user_dict = get_user_tests(request)
    if section == "creator":
        testlist_ascreator = tests_user_dict['testlist_ascreator']
        #testlist_ascreator = Test.objects.filter(creator=userobj).order_by('-createdate')
        inc_context = skillutils.includedtemplatevars("Tests", request) # Since this is the 'Tests' page for the user.
        for inc_key in inc_context.keys():
            tests_user_dict[inc_key] = inc_context[inc_key]
        testnames_created_list = list(tests_user_dict['user_creator_other_evaluators_dict'].keys())
        testnames_created_list.sort()
        tests_user_dict['baseURL'] = skillutils.gethosturl(request)
        ruleexplanataions = {}
        for ruleshort in mysettings.RULES_DICT.keys():
            ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
        tests_user_dict['rulenotes'] = json.dumps(ruleexplanataions)
        testnames_created_dict = {}
        #for test_name in testnames_created_list:
        for tobj in testlist_ascreator:
            test_name = str(tobj.testname)
            #test_name = str(test_name)
            try:
                #tobj = Test.objects.filter(testname=test_name, creator=userobj)[0]
                test_topic = str(tobj.topicname)
                fullmarks = tobj.maxscore
                passscore = tobj.passscore
                challenges = Challenge.objects.filter(test=tobj)
                createdscore = 0
                for chlng in challenges:
                    createdscore += chlng.challengescore
                completeness = format(createdscore/fullmarks, ".2f")
                publishdate = tobj.publishdate.strftime("%d %b, %Y %H:%M:%S")
                tid = tobj.id
                duration = tobj.duration
                ruleset = tobj.ruleset
                ruleset = re.sub(re.compile("\#\|\|\#", re.DOTALL), ", ", ruleset)
                testtype = tobj.testtype
                testquality = tobj.quality
                teststandard = mysettings.SKILL_QUALITY[testquality]
                status = tobj.status
                progenv = tobj.progenv
                negativescoring = tobj.negativescoreallowed
                multipleattempts = tobj.allowmultiattempts
                maxattemptscount = tobj.maxattemptscount
                attemptsinterval = tobj.attemptsinterval
                attemptsintervalunit = tobj.attemptsinterval
                scope = tobj.scope
                activationdate = tobj.activationdate.strftime("%d %b, %Y %H:%M:%S")
                testurl = generatetesturl(tobj, userobj, tests_user_dict)
                usertestqset = UserTest.objects.filter(test=tobj)
                wouldbeusersqset = WouldbeUsers.objects.filter(test=tobj)
                combinedlist = []
                for utobj in usertestqset:
                    combinedlist.append(utobj)
                for wbu in wouldbeusersqset:
                    combinedlist.append(wbu)
                testtakerscount = combinedlist.__len__()
                passcount, failcount, notevaluated = 0, 0, 0
                if passscore and passscore > 0:
                    for utobj in combinedlist:
                        if utobj.evalcommitstate == 1 and utobj.score >= passscore:
                            passcount += 1
                        elif utobj.evalcommitstate == 1 and utobj.score < passscore:
                            failcount += 1
                        else:
                            notevaluated += 1
                else:
                    passcount = "No criteria specified."
                    failcount = "No criteria specified."
                disqualifications = 0
                if combinedlist.__len__() > 0:
                    disqualifications += len(usertestqset.filter(disqualified=True))
                    disqualifications += len(wouldbeusersqset.filter(disqualified=True))
                evaluatoruserobjs = tests_user_dict['user_creator_other_evaluators_dict'][str(test_name)]
                evaluatorlinkslist = []
                for evaluserobj in evaluatoruserobjs:
                    if not evaluserobj:
                        continue
                    evallink = "<a href='%s'>%s</s>"%(evaluserobj.id, evaluserobj.emailid)
                    evaluatorlinkslist.append(evallink)
                evaluatorlinks = ", ".join(evaluatorlinkslist)
                testnames_created_dict[str(test_name)] = [str(tid), str(testurl), str(test_topic), str(fullmarks), str(passscore), str(publishdate), activationdate, str(duration), str(ruleset), str(testtype), str(teststandard), str(status), str(progenv), str(negativescoring), str(multipleattempts), str(maxattemptscount), str(attemptsinterval), str(attemptsintervalunit), str(scope), str(evaluatorlinks), str(createdscore), str(completeness), str(testtakerscount), str(passcount), str(failcount), disqualifications]
            except:
                response = "Error Retrieving Tests Where User As Creator: %s"%sys.exc_info()[1].__str__()
                return HttpResponse(response)
        tests_user_dict['creator_tests_info'] = json.dumps(testnames_created_dict)
        tests_user_dict['nextpage_ascreator'] = nextpage_ascreator
        tests_user_dict['prevpage_ascreator'] = prevpage_ascreator
        tests_user_dict['currpage_ascreator'] = currpage_ascreator
        tests_user_dict['first_ascreator'] = first_ascreator
    elif section == "evaluator":
        testnames_evaluated_list = tests_user_dict['user_evaluator_creator_other_evaluators_dict'].keys()
        testnames_evaluated_list.sort()
        testnames_evaluated_dict = {}
        tests_user_dict['baseURL'] = skillutils.gethosturl(request)
        ruleexplanataions = {}
        for ruleshort in mysettings.RULES_DICT.keys():
            ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
        tests_user_dict['rulenotes'] = json.dumps(ruleexplanataions)
        testlist_asevaluator = tests_user_dict['testlist_asevaluator']
        for tobj in testlist_asevaluator:
            try:
                test_name = tobj.testname
                test_topic = tobj.topicname
                tid = tobj.id
                creatorname = tobj.creator.displayname
                fullmarks = tobj.maxscore
                passscore = tobj.passscore
                challenges = Challenge.objects.filter(test=tobj)
                createdscore = 0
                for chlng in challenges:
                    createdscore += chlng.challengescore
                completeness = format(createdscore/fullmarks, ".2f")
                publishdate = tobj.publishdate.strftime("%d %b, %Y %H:%M:%S")
                duration = tobj.duration
                ruleset = tobj.ruleset
                ruleset = re.sub(re.compile("\#\|\|\#", re.DOTALL), ", ", ruleset)
                testtype = tobj.testtype
                testquality = tobj.quality
                teststandard = mysettings.SKILL_QUALITY[testquality]
                status = tobj.status
                progenv = tobj.progenv
                negativescoring = tobj.negativescoreallowed
                multipleattempts = tobj.allowmultiattempts
                maxattemptscount = tobj.maxattemptscount
                attemptsinterval = tobj.attemptsinterval
                attemptsintervalunit = tobj.attemptsinterval
                scope = tobj.scope
                activationdate = tobj.activationdate.strftime("%d %b, %Y %H:%M:%S")
                testurl = generatetesturl(tobj, userobj, tests_user_dict)
                usertestqset = UserTest.objects.filter(test=tobj)
                wouldbeusersqset = WouldbeUsers.objects.filter(test=tobj)
                combinedlist = []
                for utobj in usertestqset:
                    combinedlist.append(utobj)
                for wbu in wouldbeusersqset:
                    combinedlist.append(wbu)
                testtakerscount = combinedlist.__len__()
                passcount, failcount, notevaluated = 0, 0, 0
                if passscore and passscore > 0:
                    for utobj in combinedlist:
                        if utobj.evalcommitstate == 1 and utobj.score >= passscore:
                            passcount += 1
                        elif utobj.evalcommitstate == 1 and utobj.score < passscore:
                            failcount += 1
                        else:
                            notevaluated += 1
                else:
                    passcount = "No criteria specified."
                    failcount = "No criteria specified."
                disqualifications = 0
                if combinedlist.__len__() > 0:
                    disqualifications += len(usertestqset.filter(disqualified=True))
                    disqualifications += len(wouldbeusersqset.filter(disqualified=True))
                evaluatoruserobjs = tests_user_dict['user_evaluator_creator_other_evaluators_dict'][test_name]
                evaluatorlinkslist = []
                for evaluserobj in evaluatoruserobjs:
                    if not evaluserobj:
                        continue
                    evallink = "<a href='%s'>%s</s>"%(evaluserobj.id, evaluserobj.emailid)
                    evaluatorlinkslist.append(evallink)
                evaluatorlinks = ", ".join(evaluatorlinkslist)
                testnames_evaluated_dict[str(test_name)] = [ str(tid), str(testurl), str(test_topic), str(fullmarks), str(passscore), publishdate, activationdate, str(duration), str(ruleset), str(testtype), str(teststandard), str(status), str(progenv), str(negativescoring), multipleattempts, maxattemptscount, str(attemptsinterval), str(attemptsintervalunit), str(scope), str(evaluatorlinks), str(tobj.creator.emailid), str(tobj.creator.displayname), str(createdscore), completeness, str(testtakerscount), str(passcount), str(failcount), str(disqualifications) ]
            except:
                response = "Error Retrieving Tests Where User As Evaluator: %s"%sys.exc_info()[1].__str__()
                continue
                #return HttpResponse(response)
        tests_user_dict['evaluator_tests_info'] = json.dumps(testnames_evaluated_dict)
        tests_user_dict['currpage_asevaluator'] = currpage_asevaluator
        tests_user_dict['first_asevaluator'] = first_asevaluator
    elif section == "candidate":
        testnames_candidature_list = list(tests_user_dict['user_candidate_other_creator_evaluator_dict'].keys())
        testnames_candidature_list.sort()
        testnames_candidature_dict = {}
        tests_user_dict['baseURL'] = skillutils.gethosturl(request)
        ruleexplanataions = {}
        for ruleshort in mysettings.RULES_DICT.keys():
            ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
        tests_user_dict['rulenotes'] = json.dumps(ruleexplanataions)
        testlist_ascandidate = tests_user_dict['testlist_ascandidate']
        uniquedict = {}
        for tobj in testlist_ascandidate:
            try:
                test_name = tobj.testname
                test_topic = tobj.topicname
                creatorname = tobj.creator.displayname
                fullmarks = tobj.maxscore
                passscore = tobj.passscore
                challenges = Challenge.objects.filter(test=tobj)
                createdscore = 0
                for chlng in challenges:
                    createdscore += chlng.challengescore
                completeness = format(createdscore/fullmarks, ".2f")
                publishdate = tobj.publishdate.strftime("%d %b, %Y %H:%M:%S")
                tid = tobj.id
                duration = tobj.duration
                ruleset = tobj.ruleset
                testtype = tobj.testtype
                testquality = tobj.quality
                teststandard = mysettings.SKILL_QUALITY[testquality]
                status = tobj.status
                progenv = tobj.progenv
                negativescoring = tobj.negativescoreallowed
                multipleattempts = tobj.allowmultiattempts
                maxattemptscount = tobj.maxattemptscount
                attemptsinterval = tobj.attemptsinterval
                attemptsintervalunit = tobj.attemptsintervalunit
                scope = tobj.scope
                activationdate = tobj.activationdate.strftime("%d %b, %Y %H:%M:%S")
                usertestqset = UserTest.objects.filter(test=tobj)
                wouldbeusersqset = WouldbeUsers.objects.filter(test=tobj)
                combinedlist = []
                for utobj in usertestqset:
                    combinedlist.append(utobj)
                for wbu in wouldbeusersqset:
                    combinedlist.append(wbu)
                testtakerscount = combinedlist.__len__()
                userresps = UserResponse.objects.filter(test=tobj, emailaddr=userobj.emailid)
                attainedscore = 0
                if userresps.__len__() == 0: # User has not appeared in the test yet
                    attainedscore = "NA"
                utinfocusqset = UserTest.objects.filter(test=tobj, emailaddr=userobj.emailid)
                # if we get more than one UserTest object here, we take the first one. This is a problem
                # that will cease to exist when we allow candidates to take a test only once.
                utinfocus = None
                if utinfocusqset.__len__() > 0:
                    utinfocus = utinfocusqset[0]
                if utinfocus and not utinfocus.evalcommitstate:
                    attainedscore = "NA"
                if attainedscore != "NA":
                    for uresp in userresps:
                        if uresp.evaluation == "" or uresp.evaluation is None:
                            uresp.evaluation = 0.0
                        attainedscore += uresp.evaluation
                percentilescore = ""
                if attainedscore == "NA":
                    percentilescore = "NA"
                else:
                    percentilescore = getpercentilescore(attainedscore, combinedlist)
                evalcompleteflag = False
                for utobj in combinedlist:
                    if utobj.evalcommitstate == 1:
                        evalcompleteflag = True
                        break
                if evalcompleteflag is False:
                    percentilescore = "NA"
                testurl = ""
                testurl = generatetesturl(tobj, userobj, tests_user_dict)
                # The above line is causing some strange issues - Needs investigation with a fresh mind
                evaluatoruserobjs = tests_user_dict['user_candidate_other_creator_evaluator_dict'][test_name]
                evaluatorlinkslist = []
                for evaluserobj in evaluatoruserobjs:
                    if evaluserobj is None:
                        continue
                    evallink = "<a href='%s'>%s</s>"%(evaluserobj.id, evaluserobj.emailid)
                    evaluatorlinkslist.append(evallink)
                evaluatorlinks = ", ".join(evaluatorlinkslist)
                utqset = UserTest.objects.filter(test=tobj, emailaddr=userobj.emailid)
                if utqset.__len__() == 0:
                    wdqset = WouldbeUsers.objects.filter(test=tobj, emailaddr=userobj.emailid)
                    for wdobj in wdqset:
                        utqset.append(wdobj)
                if utqset.__len__() == 0: # If queryset length is still 0, then there is some discrepency in code. Skip this record.
                    continue
                utobj = None
                for utobj in utqset:
                    if utobj and (utobj.score > 0 or utobj.evalcommitstate): # Find which invitation the user has used to take the test. It will have a positive score value or its evalcommitstate will be 1.
                        break
                if not uniquedict.has_key(test_name):
                    uniquedict[test_name] = 1
                else:
                    continue
                candidate_score = "<font color='#AA0000'>Not evaluated yet.</font>"
                if utobj.evalcommitstate:
                    candidate_score = utobj.score
                test_taken_on = str(utobj.starttime)
                next_test_date = ""
                if not tobj.allowmultiattempts:
                    next_test_date = "Not Applicable"
                else:
                    if test_taken_on:
                        next_test_date = get_next_date(str(test_taken_on), attemptsinterval, attemptsintervalunit)
                    else:
                        next_test_date = "Anytime" # If no usertest object exists then the user would be able to take the test anytime.
                visibility = utobj.visibility
                testnames_candidature_dict[str(test_name)] = [ str(tid), str(testurl), str(test_topic), str(fullmarks), str(passscore), publishdate, activationdate, str(duration), str(ruleset), str(testtype), str(teststandard), str(status), str(progenv), str(negativescoring), multipleattempts, str(maxattemptscount), str(attemptsinterval), str(attemptsintervalunit), str(scope), str(evaluatorlinks), str(createdscore), completeness, str(candidate_score), test_taken_on, next_test_date, visibility, str(testtakerscount), str(percentilescore) ]
            except:
                response = "Error Retrieving Tests Where User As Candidate: %s"%(sys.exc_info()[1].__str__())
                return HttpResponse(response)
        tests_user_dict['candidate_tests_info'] = json.dumps(testnames_candidature_dict)
        tests_user_dict['currpage_ascandidate'] = currpage_ascandidate
        tests_user_dict['first_ascandidate'] = first_ascandidate
    elif section == "interviewer":
        tests_user_dict['baseURL'] = skillutils.gethosturl(request)
        ruleexplanataions = {}
        for ruleshort in mysettings.RULES_DICT.keys():
            ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
        tests_user_dict['rulenotes'] = json.dumps(ruleexplanataions)
        asinterviewerdict = tests_user_dict['interviewlist_asinterviewer']
        newinterviewerdict = {}
        for intname in asinterviewerdict.keys():
            intvals = asinterviewerdict[intname]
            intname = str(intname)
            valueslist = []
            for intval in intvals:
                if intval is None:
                    intval = ""
                if type(intval) == datetime.datetime:
                    intval = intval.strftime("%d %b, %Y %H:%M:%S")
                elif type(intval) == bool:
                    pass
                else:
                    intval = str(intval)
                valueslist.append(intval)
            newinterviewerdict[intname] = valueslist
        tests_user_dict['interviewlist_asinterviewer'] = json.dumps(newinterviewerdict)
        tests_user_dict['currpage_asinterviewer'] = currpage_asinterviewer
        tests_user_dict['first_asinterviewer'] = first_asinterviewer
        #interviewsasinterviewer = Interview.objects.filter(interviewer=userobj).order_by('-createdate')[startctr_asinterviewer:endctr_asinterviewer] # We already have the requisite data from 'get_user_tests' call
    elif section == "interviewee":
        tests_user_dict['baseURL'] = skillutils.gethosturl(request)
        ruleexplanataions = {}
        for ruleshort in mysettings.RULES_DICT.keys():
            ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
        tests_user_dict['rulenotes'] = json.dumps(ruleexplanataions)
        asintervieweedict = tests_user_dict['interviewlist_asinterviewee']
        newintervieweedict = {}
        for intname in asintervieweedict.keys():
            intvals = asintervieweedict[intname]
            intname = str(intname)
            valueslist = []
            for intval in intvals:
                if intval is None:
                    intval = ""
                if type(intval) == datetime.datetime:
                    intval = intval.strftime("%d %b, %Y %H:%M:%S")
                elif type(intval) == bool:
                    pass
                else:
                    intval = str(intval)
                valueslist.append(intval)
            newintervieweedict[intname] = valueslist
        tests_user_dict['interviewlist_asinterviewee'] = json.dumps(newintervieweedict)
        tests_user_dict['currpage_asinterviewee'] = currpage_asinterviewee
        tests_user_dict['first_asinterviewee'] = first_asinterviewee
        #interviewsasinterviewees = InterviewCandidates.objects.filter(emailaddr=userobj.emailid).order_by('-scheduledtime')[startctr_asinterviewee:endctr_asinterviewee] # We already have the requisite data from 'get_user_tests' call
    else:
        pass
    for testkey in tests_user_dict.keys():
        try:
            tests_user_dict[testkey] = str(tests_user_dict[testkey]) # stringify it
        except:
            tests_user_dict.pop(testkey, '') # Remove key if it cannot be stringified
    # Return data as json.
    return HttpResponse(json.dumps(tests_user_dict))
    

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def searchbyrole(request):
    message = ''
    if request.method != "POST": # Illegal bad request... 
        message = error_msg('1004')
        # A logging mechanism may be used to track how many and from where
        # such requests come and that may, sometimes, tell a curious story.
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    # Expected params: q, role, pageno (and csrfmiddlewaretoken) TODO: Implementation pending
    q, role, pageno = "", "", 1
    if request.POST.has_key('q'):
        q = request.POST['q']
    if request.POST.has_key('role'):
        role = request.POST['role']
    if request.POST.has_key('pageno'):
        try:
            pageno = int(request.POST['pageno'])
        except:
            pass
    startctr, endctr = 1, mysettings.PAGE_CHUNK_SIZE
    startctr = mysettings.PAGE_CHUNK_SIZE * pageno - mysettings.PAGE_CHUNK_SIZE
    endctr = mysettings.PAGE_CHUNK_SIZE * pageno
    testslist = []
    dbconn, cursor = None, None
    connutils = connectdb()
    dbconn = connutils[0]
    cursor = connutils[1]
    totalcount = 0
    if role == "creator":
        testsqset = Test.objects.filter(creator=userobj, testname__icontains=q)[startctr:endctr]
        totalcount = Test.objects.filter(creator=userobj, testname__icontains=q).count()
        for test in testsqset:
            d = {}
            evaluator = test.evaluator.evalgroupname
            challenges = Challenge.objects.filter(test=test)
            createdscore = 0
            for chlng in challenges:
                createdscore += chlng.challengescore
            completeness = format(createdscore/test.maxscore, ".2f")
            usertestsqset = UserTest.objects.filter(test=test)
            passcount, failcount, disqualifiedcount = 0, 0, 0
            for utobj in usertestsqset:
                if utobj.outcome:
                    if utobj.score < utobj.test.passscore:
                        failcount += 1
                    else:
                        passcount += 1
                if utobj.disqualified:
                    disqualifiedcount += 1
            d[test.testname] = [test.id, test.testname, test.topic.topicname, test.creatorisevaluator, evaluator, test.testtype, test.createdate.strftime("%d-%m-%Y %H:%M:%S"), test.maxscore, test.passscore, test.ruleset, test.duration, test.allowedlanguages, test.challengecount, test.activationdate.strftime("%d-%m-%Y %H:%M:%S"), test.publishdate.strftime("%d-%m-%Y %H:%M:%S"), test.status, test.progenv, test.scope, test.quality, completeness, createdscore, test.negativescoreallowed, test.allowmultiattempts, usertestsqset.count(), passcount, failcount, disqualifiedcount]
            testslist.append(d)
    elif role == "evaluator":
        evaluator_groups = Evaluator.objects.filter(Q(groupmember1=userobj)|Q(groupmember2=userobj)|Q(groupmember3=userobj)| \
                                                Q(groupmember4=userobj)|Q(groupmember5=userobj)|Q(groupmember6=userobj)| \
                                                Q(groupmember7=userobj)|Q(groupmember8=userobj)|Q(groupmember9=userobj)| \
                                                Q(groupmember10=userobj))
        testsqset = Test.objects.filter(evaluator__in=evaluator_groups, testname__icontains=q)[startctr:endctr]
        totalcount = Test.objects.filter(evaluator__in=evaluator_groups, testname__icontains=q).count()
        for test in testsqset:
            d = {}
            challenges = Challenge.objects.filter(test=test)
            createdscore = 0
            for chlng in challenges:
                createdscore += chlng.challengescore
            completeness = format(createdscore/test.maxscore, ".2f")
            usertestsqset = UserTest.objects.filter(test=test)
            passcount, failcount, disqualifiedcount = 0, 0, 0
            for utobj in usertestsqset:
                if utobj.outcome:
                    if utobj.score < utobj.test.passscore:
                        failcount += 1
                    else:
                        passcount += 1
                if utobj.disqualified:
                    disqualifiedcount += 1
            d[test.testname] = [test.id, test.testname, test.creator.displayname, test.topic.topicname, test.creatorisevaluator, test.testtype, test.createdate.strftime("%d-%m-%Y %H:%M:%S"), test.maxscore, test.passscore, test.ruleset, test.duration, test.allowedlanguages, test.challengecount, test.activationdate.strftime("%d-%m-%Y %H:%M:%S"), test.publishdate.strftime("%d-%m-%Y %H:%M:%S"), test.status, test.progenv, test.scope, test.quality, createdscore, completeness, test.negativescoreallowed, test.allowmultiattempts, usertestsqset.count(), passcount, failcount, disqualifiedcount]
            testslist.append(d)
    elif role == "candidate":
        usertestsqset = UserTest.objects.filter(user=userobj)[startctr:endctr]
        totalcount = UserTest.objects.filter(user=userobj).count()
        testsqset = []
        qregex = re.compile(q, re.IGNORECASE|re.DOTALL)
        myscoreslist = []
        testtakendates = []
        visibilitylist = []
        uniquetests = {}
        for usertest in usertestsqset:
            if usertest.test.testname in uniquetests.keys():
                continue
            uniquetests[usertest.test.testname] = 1
            if re.search(qregex, usertest.test.testname):
                testsqset.append(usertest.test)
                myscoreslist.append(usertest.score)
                testtakendates.append(usertest.starttime.strftime("%d-%m-%Y %H:%M:%S"))
                visibilitylist.append(usertest.visibility)
        ctr = 0
        for test in testsqset:
            d = {}
            utqset = UserTest.objects.filter(test=test)
            wouldbeusersqset = WouldbeUsers.objects.filter(test=test)
            combinedlist = []
            for utobj in utqset:
                combinedlist.append(utobj)
            for wbu in wouldbeusersqset:
                combinedlist.append(wbu)
            percentilescore = ""
            if myscoreslist[ctr] == "NA":
                percentilescore = "NA"
            else:
                percentilescore = getpercentilescore(myscoreslist[ctr], combinedlist)
            negativescoreallowed = "No";
            if test.negativescoreallowed:
                negativescoreallowed = "Yes"
            nexttestdate = "NA"
            if test.attemptsintervalunit is not None:
                nexttestdate = get_next_date(testtakendates[ctr], test.attemptsinterval, test.attemptsintervalunit)
            d[test.testname] = [test.id, test.testname, test.creator.displayname, test.topic.topicname, test.creatorisevaluator, test.testtype, test.createdate.strftime("%d-%m-%Y %H:%M:%S"), test.maxscore, test.passscore, test.ruleset, test.duration, test.allowedlanguages, test.challengecount, test.activationdate.strftime("%d-%m-%Y %H:%M:%S"), test.publishdate.strftime("%d-%m-%Y %H:%M:%S"), test.status, test.progenv, test.scope, test.quality, myscoreslist[ctr], utqset.count(), percentilescore, testtakendates[ctr], negativescoreallowed, nexttestdate, visibilitylist[ctr]]
            testslist.append(d)
            ctr += 1
    elif role == "interviewer":
        testsqset = Interview.objects.filter(interviewer=userobj, title__icontains=q)[startctr:endctr]
        totalcount = Interview.objects.filter(interviewer=userobj, title__icontains=q).count()
        for test in testsqset: # Remember these are actually Interview objects.
            d = {}
            d[test.title] = [test.id, test.title, test.topicname, test.medium, test.language, test.createdate.strftime("%d-%m-%Y %H:%M:%S"), test.publishdate.strftime("%d-%m-%Y %H:%M:%S"), test.status, test.maxscore, test.maxduration, test.realtime, test.interviewlinkid]
            testslist.append(d)
    elif role == "interviewee":
        candidatesqset = InterviewCandidates.objects.filter(emailaddr=userobj.emailid)[startctr:endctr]
        totalcount = InterviewCandidates.objects.filter(emailaddr=userobj.emailid).count()
        testsqset = []
        qregex = re.compile(q, re.IGNORECASE|re.DOTALL)
        actualtimes = []
        scheduledtimes = []
        timetakenlist = []
        interviewurlslist = []
        for candidateinterview in candidatesqset:
            if re.search(qregex, candidateinterview.interview.title):
                testsqset.append(candidateinterview.interview)
                if candidateinterview.actualstarttime is not None:
                    actualtimes.append(candidateinterview.actualstarttime.strftime("%d-%m-%Y %H:%M:%S"))
                else:
                    actualtimes.append("NA")
                if candidateinterview.scheduledtime is not None:
                    scheduledtimes.append(candidateinterview.scheduledtime.strftime("%d-%m-%Y %H:%M:%S"))
                else:
                    scheduledtimes.append("NA")
                timetakenlist.append(candidateinterview.totaltimetaken)
                interviewurlslist.append(candidateinterview.interviewurl)
        ctr = 0
        for test in testsqset:
            d = {}
            d[test.title] = [test.id, test.title, test.topicname, test.medium, test.language, test.createdate.strftime("%d-%m-%Y %H:%M:%S"), test.publishdate.strftime("%d-%m-%Y %H:%M:%S"), test.status, test.maxscore, test.maxduration, test.realtime, test.interviewlinkid, actualtimes[ctr], scheduledtimes[ctr], timetakenlist[ctr], test.interviewer.displayname, interviewurlslist[ctr]]
            testslist.append(d)
            ctr += 1
    else:
        message = "Could not identify the type of object requested."
        respdict = {'err' : message}
        response = HttpResponse(json.dumps(respdict))
        return response
    context = {}
    ruleexplanataions = {}
    for ruleshort in mysettings.RULES_DICT.keys():
        ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
    context['rulenotes'] = ruleexplanataions
    context['tests'] = testslist
    context['currpage'] = pageno
    context['nextpage'] = pageno + 1
    context['prevpage'] = pageno - 1
    context['lastpage'] = int(totalcount/mysettings.PAGE_CHUNK_SIZE) + 1
    context['q'] = q
    context['role'] = role
    jsondata = json.dumps(context)
    return HttpResponse(jsondata)



"""
Function to check if a given user is allowed to access 
the details of a test or a challenge in edit mode. At present,
only the creator of a test is allowed to access the test
in edit mode. Returns True if the user is permitted and
False otherwise.
"""
def ispermittedtoedit(userobj, testobj):
    if testobj.creator.id == userobj.id and testobj.creator.displayname == userobj.displayname:
        return True
    return False

"""
Function to check if a given user is allowed to view 
the details of a test or a challenge. At present, the
creator of a test and its evaluators are allowed to 
view the test. Candidates can also view a test, but they
may do so only during taking the test. This function 
allows only creator and evaluator(s) associated with
the test. Returns True if the user is permitted and False
otherwise.
"""
def ispermittedtoview(userobj, testobj, tests_user_dict):
    if not tests_user_dict['user_evaluator_creator_other_evaluators_dict'].has_key(testobj.testname):
        return False
    evalslist = tests_user_dict['user_evaluator_creator_other_evaluators_dict'][testobj.testname]
    evalemailslist = []
    evaldisplaynames = []
    for evalobj in evalslist:
        if not evalobj:
            continue
        evalemailslist.append(str(evalobj.emailid))
        evaldisplaynames.append(str(evalobj.displayname))
    if testobj.creator.id == userobj.id and testobj.creator.displayname == userobj.displayname:
        return True
    else:
        foundemail = False
        founddispname = False
        for evalemail in evalemailslist:
            if userobj.emailid == evalemail:
                foundemail = True
                break
        for evaldispname in evaldisplaynames:
            if userobj.displayname == evaldispname:
                founddispname = True
                break
        if founddispname  and  foundemail:
           return True  
    return False


"""
This function will compute the next test date (if repeats are allowed) from 
the current test date and attempts interval values. The current test date is
in mysql compatible form : yyyy-mm-dd hh:min:ss
"""
def get_next_date(curtestdate, attemptsinterval, attemptsintervalunit):
    datepart, timepart = curtestdate.split(" ")
    timepartslist = timepart.split("+")
    timepart = timepartslist[0]
    YYYY, MM, DD = datepart.split("-")
    hh, mm, ss = timepart.split(":")
    if attemptsintervalunit == 'h':
        hh = int(hh) + int(attemptsinterval)
        if hh > 24:
            hh = int(hh) - 24
            DD = int(DD) + 1
    elif attemptsintervalunit == 'm':
        mm = int(mm) + int(attemptsinterval)
        if mm > 60:
            mm = int(mm) - 60
            hh = int(hh) + 1
    elif attemptsintervalunit == 'd':
        DD = int(DD) + int(attemptsinterval)
        if DD > 30:
            DD = int(DD) - 30
            MM = int(MM) + 1
    elif attemptsintervalunit == 'M':
        MM = int(MM) + int(attemptsinterval)
        if MM > 12:
            MM = int(MM) - 12
            YYYY = int(YYYY) + 1
    elif attemptsintervalunit == 'Y':
        YYYY = int(YYYY) + int(attemptsinterval)
    else:
        msg = "Incorrect attemptsintervalunit (%s)"%attemptsintervalunit
        return(msg)
    next_test_date = str(YYYY) + "-" + str(MM) + "-" + str(DD) + " " + str(hh) + ":" + str(mm) + ":" + str(ss)
    return next_test_date


"""
This view will provide the following functionalities:
Display a table of tests with latest first... These are all tests the user has created. This page will give the user
# access to all those tests in which she/he is a creator, evaluator or candidate. For tests in which  user is
# creator or evaluator, she/he will be able to access the answer scripts of candidates who took those tests. We get
all this data in dashboard too, but from here the user will be able to go into deeper details like exact questions
attempted by a certain candidate, the exact choices/answers the candidate made and how much points/grades the user
conceded to the candidate for that answer. NOTE: This page will also provide the user with the capability to add and
modify tests that are scheduled in future and in which she/he is creator. This page will display a link that will
enable the user to create tests (if he is a premium user or has conducted less than skills_settings.NEW_USER_FREE_TESTS_COUNT since registering), add/modify/delete candidates to those tests, make assessments for tests in which she/he is the evaluator, etc.
The page consists of following:
    1) A list of all tests (of past) in which she/he is creator, ordered latest first. The fields are "Test name", "Test URL"
    (This is the url at which the creator or evaluators can access just the questions of the test in one place (as if written
    down on a physical paper), the "Test rules" associated with the entire test (or rather a popup/popunder that lists out
    the actual rules), "Full Marks", "How many candidates were invited", "How many candidates took the test", "How many suc-
    ceeded", "Percentage of successful candidates", "Percent failure", "Candidates disqualified, if any, and a popunder dis-
    playing a note of the reason of disqualification", "total marks", "negative marking, if set", "Date of test creation",
    "Date of test", "Date of test assessment (if it took more than one day to assess, then the day assessment  was completed",
    "List of the evaluators of the test (the evaluator group name, and the evaluator names too, each evaluator name will be
    a link which will lead to a page that displays the details of that evaluator pertaining to that test)", "topic and sub-
    topic of the test" and  any comments on the test made by an evaluator (or creator herself/himself) about the test.
    2) List of all tests in which the user is evaluator. Fields will be similar to above and with the addition of "Creator name of the test" and number of candidates the user assessed. This will be a link to the page which lists all candidates of the test that the user had evaluated and the names of the candidates will be links to the score (for each
question) made by the candidate and any comment added by the user against it. (The same will be available to creator if the creator is given permission to view evaluator info)
    3) List of all tests in which user had been a candidate and this will display such fields as appropriate.
    4)There will be a create test link on the top (accessible only to paid users and new users whose quota of complimentary tests is not finished)
    5) Next it will list out all upcoming tests in which the user  can register as a candidate (there would link to a page that registers a user for a test). However candidates can take that test only if the creator allows it. That mechanism will be detailed later. Whenever a user registers for a specific test, the creator of that test is sent  an email mentioning that. 
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def manage(request):
    message = ''
    if request.method != "GET": # Illegal bad request... 
        message = error_msg('1004')
        # A logging mechanism may be used to track how many and from where
        # such requests come and that may, sometimes, tell a curious story.
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    tests_user_dict = get_user_tests(request)
    inc_context = skillutils.includedtemplatevars("Tests", request) # Since this is the 'Tests' page for the user.
    for inc_key in inc_context.keys():
        tests_user_dict[inc_key] = inc_context[inc_key]
    testnames_created_list = tests_user_dict['user_creator_other_evaluators_dict'].keys()
    testnames_created_list.sort()
    tests_user_dict['baseURL'] = skillutils.gethosturl(request)
    ruleexplanataions = {}
    for ruleshort in mysettings.RULES_DICT.keys():
        ruleexplanataions[ruleshort] = mysettings.RULES_DICT[ruleshort]
    tests_user_dict['rulenotes'] = ruleexplanataions
    testnames_created_dict = {}
    testlist_ascreator = tests_user_dict['testlist_ascreator']
    #for test_name in testnames_created_list:
    for tobj in testlist_ascreator:
        try:
            test_name = tobj.testname
            #tobj = Test.objects.filter(testname=test_name, creator=userobj)[0]
            test_topic = tobj.topicname
            fullmarks = tobj.maxscore
            passscore = tobj.passscore
            challenges = Challenge.objects.filter(test=tobj)
            createdscore = 0
            for chlng in challenges:
                createdscore += chlng.challengescore
            completeness = format(createdscore/fullmarks, ".2f")
            publishdate = tobj.publishdate
            tid = tobj.id
            duration = tobj.duration
            ruleset = tobj.ruleset
            ruleset = re.sub(re.compile("\#\|\|\#", re.DOTALL), ", ", ruleset)
            testtype = tobj.testtype
            testquality = tobj.quality
            teststandard = mysettings.SKILL_QUALITY[testquality]
            status = tobj.status
            progenv = tobj.progenv
            negativescoring = tobj.negativescoreallowed
            multipleattempts = tobj.allowmultiattempts
            maxattemptscount = tobj.maxattemptscount
            attemptsinterval = tobj.attemptsinterval
            attemptsintervalunit = tobj.attemptsinterval
            scope = tobj.scope
            activationdate = tobj.activationdate
            testurl = generatetesturl(tobj, userobj, tests_user_dict)
            usertestqset = UserTest.objects.filter(test=tobj)
            wouldbeusersqset = WouldbeUsers.objects.filter(test=tobj)
            combinedlist = []
            for utobj in usertestqset:
                combinedlist.append(utobj)
            for wbu in wouldbeusersqset:
                combinedlist.append(wbu)
            testtakerscount = combinedlist.__len__()
            passcount, failcount, notevaluated = 0, 0, 0
            if passscore and passscore > 0:
                for utobj in combinedlist:
                    if utobj.evalcommitstate == 1 and utobj.score >= passscore:
                        passcount += 1
                    elif utobj.evalcommitstate == 1 and utobj.score < passscore:
                        failcount += 1
                    else:
                        notevaluated += 1
            else:
                passcount = "No criteria specified."
                failcount = "No criteria specified."
            disqualifications = 0
            if combinedlist.__len__() > 0:
                disqualifications += len(usertestqset.filter(disqualified=True))
                disqualifications += len(wouldbeusersqset.filter(disqualified=True))
            evaluatoruserobjs = tests_user_dict['user_creator_other_evaluators_dict'][test_name]
            evaluatorlinkslist = []
            for evaluserobj in evaluatoruserobjs:
                if not evaluserobj:
                    continue
                evallink = "<a href='%s'>%s</s>"%(evaluserobj.id, evaluserobj.emailid)
                evaluatorlinkslist.append(evallink)
            evaluatorlinks = ", ".join(evaluatorlinkslist)
            testnames_created_dict[test_name] = [tid, testurl, test_topic, fullmarks, passscore, publishdate, activationdate, duration, ruleset, testtype, teststandard, status, progenv, negativescoring, multipleattempts, maxattemptscount, attemptsinterval, attemptsintervalunit, scope, evaluatorlinks, createdscore, completeness, testtakerscount, passcount, failcount, disqualifications]
        except:
            response = "Error Retrieving Tests Where User As Creator: %s"%sys.exc_info()[1].__str__()
            return HttpResponse(response)
    tests_user_dict['creator_tests_info'] = testnames_created_dict
    testnames_evaluated_list = tests_user_dict['user_evaluator_creator_other_evaluators_dict'].keys()
    testnames_evaluated_list.sort()
    testnames_evaluated_dict = {}
    testlist_asevaluator = tests_user_dict['testlist_asevaluator']
    for tobj in testlist_asevaluator:
        try:
            test_name = tobj.testname
            test_topic = tobj.topicname
            tid = tobj.id
            creatorname = tobj.creator.displayname
            fullmarks = tobj.maxscore
            passscore = tobj.passscore
            challenges = Challenge.objects.filter(test=tobj)
            createdscore = 0
            for chlng in challenges:
                createdscore += chlng.challengescore
            completeness = format(createdscore/fullmarks, ".2f")
            publishdate = tobj.publishdate
            duration = tobj.duration
            ruleset = tobj.ruleset
            ruleset = re.sub(re.compile("\#\|\|\#", re.DOTALL), ", ", ruleset)
            testtype = tobj.testtype
            testquality = tobj.quality
            teststandard = mysettings.SKILL_QUALITY[testquality]
            status = tobj.status
            progenv = tobj.progenv
            negativescoring = tobj.negativescoreallowed
            multipleattempts = tobj.allowmultiattempts
            maxattemptscount = tobj.maxattemptscount
            attemptsinterval = tobj.attemptsinterval
            attemptsintervalunit = tobj.attemptsinterval
            scope = tobj.scope
            activationdate = tobj.activationdate
            testurl = generatetesturl(tobj, userobj, tests_user_dict)
            usertestqset = UserTest.objects.filter(test=tobj)
            wouldbeusersqset = WouldbeUsers.objects.filter(test=tobj)
            combinedlist = []
            for utobj in usertestqset:
                combinedlist.append(utobj)
            for wbu in wouldbeusersqset:
                combinedlist.append(wbu)
            testtakerscount = combinedlist.__len__()
            passcount, failcount, notevaluated = 0, 0, 0
            if passscore and passscore > 0:
                for utobj in combinedlist:
                    if utobj.evalcommitstate == 1 and utobj.score >= passscore:
                        passcount += 1
                    elif utobj.evalcommitstate == 1 and utobj.score < passscore:
                        failcount += 1
                    else:
                        notevaluated += 1
            else:
                passcount = "No criteria specified."
                failcount = "No criteria specified."
            disqualifications = 0
            if combinedlist.__len__() > 0:
                disqualifications += len(usertestqset.filter(disqualified=True))
                disqualifications += len(wouldbeusersqset.filter(disqualified=True))
            evaluatoruserobjs = tests_user_dict['user_evaluator_creator_other_evaluators_dict'][test_name]
            evaluatorlinkslist = []
            for evaluserobj in evaluatoruserobjs:
                if not evaluserobj:
                    continue
                evallink = "<a href='%s'>%s</s>"%(evaluserobj.id, evaluserobj.emailid)
                evaluatorlinkslist.append(evallink)
            evaluatorlinks = ", ".join(evaluatorlinkslist)
            testnames_evaluated_dict[test_name] = [ tid, testurl, test_topic, fullmarks, passscore, publishdate, activationdate, duration, ruleset, testtype, teststandard, status, progenv, negativescoring, multipleattempts, maxattemptscount, attemptsinterval, attemptsintervalunit, scope, evaluatorlinks, tobj.creator.emailid, tobj.creator.displayname, createdscore, completeness, testtakerscount, passcount, failcount, disqualifications ]
        except:
            response = "Error Retrieving Tests Where User As Evaluator: %s"%sys.exc_info()[1].__str__()
            return HttpResponse(response)
    tests_user_dict['evaluator_tests_info'] = testnames_evaluated_dict
    testnames_candidature_list = tests_user_dict['user_candidate_other_creator_evaluator_dict'].keys()
    testnames_candidature_list.sort()
    testnames_candidature_dict = {}
    testlist_ascandidate = tests_user_dict['testlist_ascandidate']
    uniquedict = {}
    for tobj in testlist_ascandidate:
        try:
            test_name = tobj.testname
            test_topic = tobj.topicname
            creatorname = tobj.creator.displayname
            fullmarks = tobj.maxscore
            passscore = tobj.passscore
            challenges = Challenge.objects.filter(test=tobj)
            createdscore = 0
            for chlng in challenges:
                createdscore += chlng.challengescore
            completeness = format(createdscore/fullmarks, ".2f")
            publishdate = tobj.publishdate
            tid = tobj.id
            duration = tobj.duration
            ruleset = tobj.ruleset
            testtype = tobj.testtype
            testquality = tobj.quality
            teststandard = mysettings.SKILL_QUALITY[testquality]
            status = tobj.status
            progenv = tobj.progenv
            negativescoring = tobj.negativescoreallowed
            multipleattempts = tobj.allowmultiattempts
            maxattemptscount = tobj.maxattemptscount
            attemptsinterval = tobj.attemptsinterval
            attemptsintervalunit = tobj.attemptsintervalunit
            scope = tobj.scope
            activationdate = tobj.activationdate
            usertestqset = UserTest.objects.filter(test=tobj)
            wouldbeusersqset = WouldbeUsers.objects.filter(test=tobj)
            combinedlist = []
            for utobj in usertestqset:
                combinedlist.append(utobj)
            for wbu in wouldbeusersqset:
                combinedlist.append(wbu)
            testtakerscount = combinedlist.__len__()
            userresps = UserResponse.objects.filter(test=tobj, emailaddr=userobj.emailid)
            attainedscore = 0
            if userresps.__len__() == 0: # User has not appeared in the test yet
                attainedscore = "NA"
            utinfocusqset = UserTest.objects.filter(test=tobj, emailaddr=userobj.emailid)
            # if we get more than one UserTest object here, we take the first one. This is a problem
            # that will cease to exist when we allow candidates to take a test only once.
            utinfocus = None
            if utinfocusqset.__len__() > 0:
                utinfocus = utinfocusqset[0]
            if utinfocus and not utinfocus.evalcommitstate:
                attainedscore = "NA"
            if attainedscore != "NA":
                for uresp in userresps:
                    if uresp.evaluation == "" or uresp.evaluation is None:
                        uresp.evaluation = 0.0
                    attainedscore += uresp.evaluation
            percentilescore = ""
            if attainedscore == "NA":
                percentilescore = "NA"
            else:
                percentilescore = getpercentilescore(attainedscore, combinedlist)
            evalcompleteflag = False
            for utobj in combinedlist:
                if utobj.evalcommitstate == 1:
                    evalcompleteflag = True
                    break
            if evalcompleteflag is False:
                percentilescore = "NA"
            #testurl = generatetesturl(tobj, userobj, tests_user_dict)
            # The above line is causing some strange issues - Needs investigation with a fresh mind
            evaluatoruserobjs = tests_user_dict['user_candidate_other_creator_evaluator_dict'][test_name]
            evaluatorlinkslist = []
            for evaluserobj in evaluatoruserobjs:
                if evaluserobj is None:
                    continue
                evallink = "<a href='%s'>%s</s>"%(evaluserobj.id, evaluserobj.emailid)
                evaluatorlinkslist.append(evallink)
            evaluatorlinks = ", ".join(evaluatorlinkslist)
            utqset = UserTest.objects.filter(test=tobj, emailaddr=userobj.emailid)
            if utqset.__len__() == 0:
                wdqset = WouldbeUsers.objects.filter(test=tobj, emailaddr=userobj.emailid)
                for wdobj in wdqset:
                    utqset.append(wdobj)
            if utqset.__len__() == 0: # If queryset length is still 0, then there is some discrepency in code. Skip this record.
                continue
            utobj = None
            for utobj in utqset:
                if utobj and (utobj.score > 0 or utobj.evalcommitstate): # Find which invitation the user has used to take the test. It will have a positive score value or its evalcommitstate will be 1.
                    break
            if not uniquedict.has_key(test_name):
                uniquedict[test_name] = 1
            else:
                continue
            candidate_score = "<font color='#AA0000'>Not evaluated yet.</font>"
            if utobj.evalcommitstate:
                candidate_score = utobj.score
            test_taken_on = utobj.starttime
            next_test_date = ""
            if not tobj.allowmultiattempts:
                next_test_date = "Not Applicable"
            else:
                if test_taken_on:
                    next_test_date = get_next_date(str(test_taken_on), attemptsinterval, attemptsintervalunit)
                else:
                    next_test_date = "Anytime" # If no usertest object exists then the user would be able to take the test anytime.
            visibility = utobj.visibility
            testnames_candidature_dict[test_name] = [tid, testurl, test_topic, fullmarks, passscore, publishdate, activationdate, duration, ruleset, testtype, teststandard, status, progenv, negativescoring, multipleattempts, maxattemptscount, attemptsinterval, attemptsintervalunit, scope, evaluatorlinks, createdscore, completeness, candidate_score, test_taken_on, next_test_date, visibility, testtakerscount, percentilescore ]
        except:
            response = "Error Retrieving Tests Where User As Candidate: %s"%(sys.exc_info()[1].__str__())
            return HttpResponse(response)
    tests_user_dict['candidate_tests_info'] = testnames_candidature_dict
    # Now create and render the template here
    tmpl = get_template("tests/tests.html")
    tests_user_dict.update(csrf(request))
    cxt = Context(tests_user_dict)
    managetestshtml = tmpl.render(cxt)
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        managetestshtml = managetestshtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(managetestshtml)


"""
Computes the percentile score from the given arguments:
argument 1 is the score of the candidate. argument 2 is
a list containing the scores (among several other things)
of all users who have taken the test and have been 
evaluated.
"""
def getpercentilescore(attainedscore, combinedlist):
    takerscount = combinedlist.__len__()
    abovecount, belowcount = 0, 0
    for utobj in combinedlist:
        if not utobj.score:
            utobj.score = 0
        if int(utobj.score) > int(attainedscore):
            abovecount += 1
        else:
            belowcount += 1
        #print abovecount, "####", belowcount, "####", attainedscore, "####", utobj.score
    percentile = (float(belowcount)/float(takerscount)) * 100.00
    percentile = "{0:.2f}".format(percentile)
    return percentile


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def create(request):
    message = ''
    if request.method == "GET": # Illegal bad request... 
        message = error_msg('1004')
        # A logging mechanism may be used to track how many and from where
        # such requests come and that may, sometimes, tell a curious story.
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    testlinkid, testname, testtype, testrules, testtopic, othertopicname, \
    totalscore, numchallenges, evendistribution, negativescoring, testduration,\
    testdurationunit, challengeduration, challengedurationunit, evaluators, \
    grpname, publishdate, activedate, skilltarget, testscope,answerlanguage, \
    progenv, multimediareqd, randomsequencing, multipleattemptsallowed, \
    maxattemptscount, attemptsinterval, attemptsintervalunit, testlinkid, \
    csrfmiddlewaretoken  = "", "", "", "", "", "", "", "", "", "", "", "", "", \
     "", "", "", "", "", "", "", "", "", False, True, False, "", "", "", "", ""
    lastchallengectr, challengenumbersstr, passscore = "", "", "" # \
    # 'lastchallengectr' is the counter of the last challenge entered. \
    # 'challengenumbersstr' is a string consisting of all the created \
    # challenge counters separated by '#||#'.
    challengectr, challengestatement, challengetype, responsekey, imageurl, \
    additionalurl, timeframe, challengequality, mustrespond, challengescore, \
    challengeoptions, exist_test_id = "", "", "", "", "", "", "", "", "", "", [], None
    message, activedatemysql, publishdatemysql = "", "", ""
    # First get the testlinkid. If its value doesn't exist in DB, then create 
    # a new test. Otherwise, simply display the challenges of the existing test
    # and allow the user to edit them.
    if request.POST.has_key('testlinkid'):
        testlinkid = request.POST['testlinkid']
    if request.POST.has_key('testname'):
        testname = request.POST['testname']
    if request.POST.has_key('testtype'):
        testtype = request.POST['testtype']
    if request.POST.has_key('testrules'):
        testrules = request.POST['testrules']
    if request.POST.has_key('testtopic'):
        testtopic = request.POST['testtopic']
    if request.POST.has_key('othertopicname'):
        othertopicname = request.POST['othertopicname']
    if request.POST.has_key('totalscore'):
        totalscore = request.POST['totalscore']
    if request.POST.has_key('passscore'):
        passscore = request.POST['passscore']
    if request.POST.has_key('numchallenges'):
        numchallenges = request.POST['numchallenges']
    if request.POST.has_key('evendistribution'):
        evendistribution = request.POST['evendistribution']
    if request.POST.has_key('negativescoring'):
        negativescoring = request.POST['negativescoring']
    if request.POST.has_key('testduration'):
        testduration = request.POST['testduration']
    if request.POST.has_key('testdurationunit'):
        testdurationunit = request.POST['testdurationunit']
    if request.POST.has_key('challengeduration'):
        challengeduration = request.POST['challengeduration']
    if request.POST.has_key('challengedurationunit'):
        challengedurationunit = request.POST['challengedurationunit']
    if request.POST.has_key('evaluators'):
        evaluators = request.POST['evaluators']
    if request.POST.has_key('grpname'):
        grpname = request.POST['grpname']
    if request.POST.has_key('publishdate'):
        publishdate = request.POST['publishdate']
    if request.POST.has_key('activedate'):
        activedate = request.POST['activedate']
    if request.POST.has_key('skilltarget'):
        skilltarget = request.POST['skilltarget']
    if request.POST.has_key('testscope'):
        testscope = request.POST['testscope']
    if request.POST.has_key('answerlanguage'):
        answerlanguage = request.POST['answerlanguage']
    if request.POST.has_key('progenv'):
        progenv = request.POST['progenv']
    if request.POST.has_key('multimediareqd'):
        multimediareqd = True
    if request.POST.has_key('randomsequencing'):
        randomsequencing = True
    if request.POST.has_key('multipleattemptsallowed'):
        multipleattemptsallowed = True
    if request.POST.has_key('maxattemptscount'):
        maxattemptscount = request.POST['maxattemptscount']
    if request.POST.has_key('attemptsinterval'):
        attemptsinterval = request.POST['attemptsinterval']
    if request.POST.has_key('attemptsintervalunit'):
        attemptsintervalunit = request.POST['attemptsintervalunit']
    if request.POST.has_key('testlinkid'):
        testlinkid = request.POST['testlinkid']
    if request.POST.has_key('csrfmiddlewaretoken'):
        csrfmiddlewaretoken = request.POST['csrfmiddlewaretoken']
    if request.POST.has_key('exist_test_id'):
        exist_test_id = request.POST['exist_test_id']
    testqueryset = Test.objects.filter(testlinkid=testlinkid, creator=userobj) # Returns a Queryset object
    testobj = None
    todaynow = datetime.datetime.now()
    if testqueryset.__len__() == 0: # This is a new test being created.
        testobj = Test()
        testobj.createdate = todaynow
    else:
        testobj = testqueryset[0]
        if exist_test_id and exist_test_id != "":
            testobj = Test.objects.filter(id=exist_test_id)[0]
    # Check the allowed number of tests for this user - only in case of new tests
    if exist_test_id is None or exist_test_id == "":
        dbconn, dbcursor = connectdb()
        #dbconn, dbcursor = connectdb_p()
        userplansql = "select pl.testsninterviews, up.plan_id, pl.planname, up.amountdue, up.planstartdate, up.planenddate from Subscription_userplan up, Subscription_plan pl where up.planstartdate < %s and up.planenddate > %s and up.planstatus=TRUE and up.plan_id=pl.id and up.user_id=%s"
        dbcursor.execute(userplansql, (todaynow, todaynow, userobj.id))
        allrecords = dbcursor.fetchall()
        if allrecords.__len__() > 0:
            testsandinterviewsquota = allrecords[0][0]
            planid = allrecords[0][1]
            planname = allrecords[0][2]
            amtdue = allrecords[0][3]
            planstartdate = allrecords[0][4]
            if planstartdate is None or planstartdate == "":
                message = "Error: Plan start date is missing for user '%s'. Please contact support with this error message."%userobj.displayname
                return HttpResponse(message)
            planenddate = allrecords[0][5]
            if planenddate is None or planenddate == "":
                message = "Error: Plan end date is missing for user '%s'. Please contact support with this error message."%userobj.displayname
                return HttpResponse(message)
            if amtdue > 0.00: # User has not paid the total amount for this plan.
                message = "Error: You have an amount due for the '%s' subscription. Please clear the amount and try again."%planname
                return HttpResponse(message)
            # Check how many tests and interviews this user has created using the subscribed plan.
            testscount, interviewscount, testsninterviewscount = 0, 0, 0
            testsql = "select count(*) from Tests_test where creator_id=%s and createdate > %s and createdate < %s"
            dbcursor.execute(testsql, (userobj.id, planstartdate, planenddate))
            alltestrecs = dbcursor.fetchall()
            if alltestrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                testscount = 0
            else:
                testscount = int(alltestrecs[0][0])
            interviewsql = "select count(*) from Tests_interview where interviewer_id=%s and createdate > %s and createdate < %s"
            dbcursor.execute(interviewsql, (userobj.id, planstartdate, planenddate))
            allinterviewrecs = dbcursor.fetchall()
            if allinterviewrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                interviewscount = 0
            else:
                interviewscount = int(allinterviewrecs[0][0])
            testsninterviewscount = testscount + interviewscount
            #print("%s ### %s =============== %s"%(planstartdate.strftime("%Y-%m-%d %H:%M:%S"), planenddate.strftime("%Y-%m-%d %H:%M:%S"), testsninterviewscount))
            if testsninterviewscount >= testsandinterviewsquota and planname != "Free Plan": # User has already consumed the allocated quota.
                message = "Error: You have already consumed the allocated count of tests and interviews in your subscription plan. Please extend your subscription to create more tests."
                return HttpResponse(message)
            else: # Allow user to go ahead and create test (unless user has consumed the number of tests and interviews allowed by Free Plan.
                if planname == "Free Plan" and int(mysettings.NEW_USER_FREE_TESTS_COUNT) != -1 and mysettings.NEW_USER_FREE_TESTS_COUNT <= testsninterviewscount:
                    message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                    return HttpResponse(message)
                elif planname == "Free Plan" and int(mysettings.NEW_USER_FREE_TESTS_COUNT) == -1 and testsninterviewscount >= testsandinterviewsquota:
                     message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                     return HttpResponse(message)
                else:
                    pass # User's free quota of tests and interviews is not exhausted yet. So let the user continue.
        else: # So this is a free user
            testscount, interviewscount, testsninterviewscount = 0, 0, 0
            testsql = "select count(*) from Tests_test where creator_id=%s"
            dbcursor.execute(testsql, (userobj.id,))
            alltestrecs = dbcursor.fetchall()
            if alltestrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                testscount = 0
            else:
                testscount = int(alltestrecs[0][0])
            interviewsql = "select count(*) from Tests_interview where interviewer_id=%s"
            dbcursor.execute(interviewsql, (userobj.id,))
            allinterviewrecs = dbcursor.fetchall()
            if allinterviewrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                interviewscount = 0
            else:
                interviewscount = int(allinterviewrecs[0][0])
            testsninterviewscount = testscount + interviewscount
            freeplansql = "select testsninterviews from Subscription_plan where planname='Free Plan'"
            dbcursor.execute(freeplansql)
            planrecords = dbcursor.fetchall()
            if planrecords.__len__() >= 0:
                testsninterviewsquota = planrecords[0][0]
            else:
                testsninterviewsquota = 5 # Default free quota is this value.
            if int(mysettings.NEW_USER_FREE_TESTS_COUNT) != -1 and mysettings.NEW_USER_FREE_TESTS_COUNT <= testsninterviewscount:
                message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                return HttpResponse(message)
            elif int(mysettings.NEW_USER_FREE_TESTS_COUNT) == -1 and testsninterviewscount >= testsninterviewsquota:
                message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                return HttpResponse(message)
            else:
                pass # Allow user to continue creating tests and interviews.
        disconnectdb(dbconn, dbcursor)
    testobj.testname = testname
    if testname.__len__() > 100:
        message = "Error: Couldn't create test as the test name is too long. Please keep it within 100 characters."
        return HttpResponse(message)
    if testtopic == "Other": # Topic is custom, so testobj.topicname will be "".
        topic = Topic()
        topic.topicname = othertopicname
        topic.user = userobj
        topic.isactive = True
        topic.save()
        testobj.topic = topic
        testobj.topicname = ""
    else: # Topic is built-in, so testobj.testtopic will be None.
        testobj.topic = Topic.objects.filter(id=-1)[0]
        testobj.topicname = testtopic.replace("__", " ")
    testobj.testtype = testtype
    testobj.creator = userobj
    creatoremailid = userobj.emailid
    disallowedchars = re.compile(r"[^\._\-\w\d@\,\s]", re.DOTALL)
    if re.search(disallowedchars, evaluators):
        msg = "Error: Evaluator email Ids contain inappropriate characters"
        return HttpResponse(msg)
    evaluatorslist = evaluators.split(",")
    testobj.creatorisevaluator = False
    evalemailsdict = {} # This will help in removing duplicates, if any.
    for evalem in evaluatorslist:
        evalem = re.sub(skillutils.multiplewhitespacepattern, '', evalem)
        if not evalemailsdict.has_key(evalem):
            evalemailsdict[evalem] = 1
        if evalem == creatoremailid:
            testobj.creatorisevaluator = True
    if  not re.search(skillutils.numericpattern, totalscore) or (numchallenges != "" and not re.search(skillutils.numericpattern, numchallenges)):
        message = error_msg('1047')
        return HttpResponse(message)
    if numchallenges == "":
        numchallenges = -1 # -1 means the user didn't enter a value
    testobj.maxscore = totalscore
    testobj.challengecount = numchallenges
    if passscore == '' or not passscore:
        passscore = None
    testobj.passscore = passscore
    # Note: Both test duration and challenge duration will converted to seconds\
    # and then saved in DB.
    if testdurationunit == 'h':
        testobj.duration = int(testduration) * 3600 # Conversion to seconds from hr.
    elif testdurationunit == 'm':
        testobj.duration = int(testduration) * 60 # Seconds from minutes.
    else:
        testobj.duration = int(testduration)
    if challengedurationunit == 'h':
        challengeduration = int(challengeduration) * 3600
    elif challengedurationunit == 'm':
        challengeduration = int(challengeduration) * 60
    else:
        challengeduration = int(challengeduration)
    activedateparts = activedate.split('-')
    publishdateparts = publishdate.split('-')
    # Verify whether the date formats received are in 'dd-MON-yyyy' format.
    if activedateparts.__len__() < 3 or activedateparts[1].upper() not in mysettings.MONTHS_DICT.keys() or activedateparts[2].__len__() != 4 or not re.search(skillutils.numericpattern, activedateparts[2]) or activedateparts[0].__len__() != 2 or not re.search(skillutils.numericpattern, activedateparts[0]):
        message = error_msg('1042')
        return HttpResponse(message)
    else:
        activemm = mysettings.MONTHS_DICT[activedateparts[1].upper()]
        if activemm in ('01', '03', '05', '07', '08', '10', '12') and int(activedateparts[0]) > 31:
            message = error_msg('1043')
            return HttpResponse(message)
        elif activemm in ('04', '06', '09', '11') and int(activedateparts[0]) > 30:
            message = error_msg('1043')     
            return HttpResponse(message)
        elif activemm == '02' and int(activedateparts[2]) % 4 == 0 and int(activedateparts[2]) % 100 != 0 and int(activedateparts[0]) > 29: # Leap year
            message = error_msg('1043')
            return HttpResponse(message)
        elif activemm == '02' and int(activedateparts[2]) % 100 == 0 and int(activedateparts[2]) % 400 == 0 and int(activedateparts[0]) > 29: # Leap year
            message = error_msg('1043')
            return HttpResponse(message)
        elif activemm == '02' and int(activedateparts[0]) > 28:
            message = error_msg('1043')
            return HttpResponse(message)
        else:
            activedatemysql = activedateparts[2] + "-" + activemm + "-" + activedateparts[0] + " 00:00:00"
    #######################
    if publishdateparts.__len__() < 3 or publishdateparts[1].upper() not in mysettings.MONTHS_DICT.keys() or publishdateparts[2].__len__() != 4 or not re.search(skillutils.numericpattern, publishdateparts[2]) or publishdateparts[0].__len__() != 2 or not re.search(skillutils.numericpattern, publishdateparts[0]):
        message = error_msg('1044')
        return HttpResponse(message)
    else:
        publishmm = mysettings.MONTHS_DICT[publishdateparts[1].upper()]
        if publishmm in ('01', '03', '05', '07', '08', '10', '12') and int(publishdateparts[0]) > 31:
            message = error_msg('1045')
            return HttpResponse(message)
        elif publishmm in ('04', '06', '09', '11') and int(publishdateparts[0]) > 30:
            message = error_msg('1045')
            return HttpResponse(message)
        elif publishmm == '02' and int(publishdateparts[2]) % 4 == 0 and int(publishdateparts[2]) % 100 != 0 and int(publishdateparts[0]) > 29: # Leap year
            message = error_msg('1045')
            return HttpResponse(message)
        elif publishmm == '02' and int(publishdateparts[2]) % 100 == 0 and int(publishdateparts[2]) % 400 == 0 and int(publishdateparts[0]) > 29: # Leap year
            message = error_msg('1045')
            return HttpResponse(message)
        elif publishmm == '02' and int(publishdateparts[0]) > 28:
            message = error_msg('1045')
            return HttpResponse(message)
        else:
            publishdatemysql = publishdateparts[2] + "-" + publishmm + "-" + publishdateparts[0] + " 00:00:00"
    if datetime.datetime(int(activedateparts[2]), int(activemm), int(activedateparts[0]), int('00'), int('00'), int('00')) < datetime.datetime(int(publishdateparts[2]), int(publishmm), int(publishdateparts[0]), int('00'), int('00'), int('00')):
        message = error_msg('1046')
        return HttpResponse(message)
    testobj.activationdate = activedatemysql
    testobj.publishdate = publishdatemysql
    testobj.status = False # Remember to set this to True when we finally save the test along with its constituent challenges.
    testobj.allowedlanguages = answerlanguage
    # 'answerlanguage' may contain '#||#' at the end. If so, remove it.
    if re.search(re.compile(r"#||#$"), answerlanguage):
        testobj.allowedlanguages = answerlanguage[:-4]
    testobj.ruleset = testrules
    # 'testrules' may contain '#||#' at the end too. Remove it.
    if re.search(re.compile(r"#||#$"), testrules):
        testobj.ruleset = testrules[:-4]
    testobj.testlinkid = testlinkid
    testobj.quality = skilltarget
    # Handle evaluator grouping and 'Evaluator' object creation.
    evalobj = Evaluator()
    testsbycreator = Test.objects.filter(creator = userobj)
    evalnamesbycreator = []
    evalidsbycreator = []
    for tobj in testsbycreator:
        existevalobj = tobj.evaluator
        if existevalobj.evalgroupname not in evalnamesbycreator: # This is to ensure that a single evaluator name doesn't get listed more than once.
            evalnamesbycreator.append(existevalobj.evalgroupname)
            evalidsbycreator.append(existevalobj.id)
    if grpname not in evalnamesbycreator:
        evalobj.evalgroupname = grpname
        memberctr = 1
        for evalemailid in evalemailsdict.keys():
            uobjqset = User.objects.filter(emailid=evalemailid)
            if uobjqset.__len__() > 0: # Valid user exists
                uobj = uobjqset[0]
                if memberctr == 1:
                    evalobj.groupmember1 = uobj
                elif memberctr == 2:
                    evalobj.groupmember2 = uobj
                elif memberctr == 3:
                    evalobj.groupmember3 = uobj
                elif memberctr == 4:
                    evalobj.groupmember4 = uobj
                elif memberctr == 5:
                    evalobj.groupmember5 = uobj
                elif memberctr == 6:
                    evalobj.groupmember6 = uobj
                elif memberctr == 7:
                    evalobj.groupmember7 = uobj
                elif memberctr == 8:
                    evalobj.groupmember8 = uobj
                elif memberctr == 9:
                    evalobj.groupmember9 = uobj
                elif memberctr == 10:
                    evalobj.groupmember10 = uobj
                memberctr += 1
            else: # User with the email id in 'evalemailid' doesn't exist.
                pass
        evalobj.save()
        testobj.evaluator = evalobj
    else: # So a group with this name already exists. We need to see if the existing group has the same set of evaluator email Ids. If so, we simply get the existing evaluator object and assign it to this test.
        ctr = 0
        matchedgrpemails = {}
        evalid = -1
        while ctr < evalnamesbycreator.__len__():
            evalid = evalidsbycreator[ctr]
            if grpname == evalnamesbycreator[ctr]:
                matchedevalobj = Evaluator.objects.filter(id=evalid)[0]
                if matchedevalobj.groupmember1 is not None and matchedevalobj.groupmember1.emailid and matchedevalobj.groupmember1.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember1.emailid] = 1
                if matchedevalobj.groupmember2 is not None and matchedevalobj.groupmember2.emailid and matchedevalobj.groupmember2.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember2.emailid] = 1
                if matchedevalobj.groupmember3 is not None and matchedevalobj.groupmember3.emailid and matchedevalobj.groupmember3.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember3.emailid] = 1
                if matchedevalobj.groupmember4 is not None and matchedevalobj.groupmember4.emailid and matchedevalobj.groupmember4.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember4.emailid] = 1
                if matchedevalobj.groupmember5 is not None and matchedevalobj.groupmember5.emailid and matchedevalobj.groupmember5.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember5.emailid] = 1
                if matchedevalobj.groupmember6 is not None and matchedevalobj.groupmember6.emailid and matchedevalobj.groupmember6.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember6.emailid] = 1
                if matchedevalobj.groupmember7 is not None and matchedevalobj.groupmember7.emailid and matchedevalobj.groupmember7.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember7.emailid] = 1
                if matchedevalobj.groupmember8 is not None and matchedevalobj.groupmember8.emailid and matchedevalobj.groupmember8.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember8.emailid] = 1
                if matchedevalobj.groupmember9 is not None and  matchedevalobj.groupmember9.emailid and matchedevalobj.groupmember9.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember9.emailid] = 1
                if matchedevalobj.groupmember10 is not None and  matchedevalobj.groupmember10.emailid and matchedevalobj.groupmember10.emailid != "":
                    matchedgrpemails[matchedevalobj.groupmember10.emailid] = 1
                break
            ctr += 1
        #print matchedgrpemails.keys().__len__()
        #print evalemailsdict.keys().__len__()
        if matchedgrpemails.keys().__len__() == evalemailsdict.keys().__len__():
            for matchemail in matchedgrpemails.keys():
                if not evalemailsdict.has_key(matchemail):
                    message = error_msg('1048')
                    return HttpResponse(message)
            testobj.evaluator = matchedevalobj
        else:
            message = error_msg('1048')
            return HttpResponse(message)
    testobj.allowmultiattempts = multipleattemptsallowed
    if multipleattemptsallowed:
        testobj.maxattemptscount = maxattemptscount
        testobj.attemptsinterval = attemptsinterval
        testobj.attemptsintervalunit = attemptsintervalunit
    else:
        testobj.maxattemptscount = 1
        testobj.attemptsinterval = None
        testobj.attemptsintervalunit = None
    testobj.randomsequencing = randomsequencing
    testobj.multimediareqd = multimediareqd
    testobj.progenv = None
    if progenv != '0':
        testobj.progenv = progenv
    if not testscope:
        testobj.scope = 'private' # The default value
    else:
        testobj.scope = testscope
    if negativescoring and negativescoring != '0':
        testobj.negativescoreallowed = True
    else:
        testobj.negativescoreallowed = False
    testobj.save()
    if request.POST.has_key('lastchallengectr'):
        lastchallengectr = request.POST['lastchallengectr']
    if request.POST.has_key('challengenumbersstr'):
        challengenumbersstr = request.POST['challengenumbersstr']
    message = error_msg('1050') # Posted test meta data successfully.
    # So we create the challenge insertion form and send it as message.
    message += _challenge_edit_form(request, testobj, lastchallengectr,  evendistribution, challengeduration, int(negativescoring))
    return HttpResponse(message)


# Note: The 'challengedurationseconds' value being passed into this function is in seconds.
def _challenge_edit_form(request, testobj, lastchallengectr, evendistribution, challengedurationseconds, negativescoring=False, multiprogenv = ""):
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    totalchallenges = testobj.challengecount
    testobj.status = False # Set this to false since we are editing a challenge
    testlinkid = testobj.testlinkid
    testtype = testobj.testtype
    multimediareqd = testobj.multimediareqd
    totalscore = testobj.maxscore
    edit_challenge_dict = { 'lastchallengectr' : lastchallengectr, 'testlinkid' : testlinkid, 'multimediareqd' : multimediareqd, 'totalscore' : totalscore, 'challengedurationseconds' : challengedurationseconds, 'testtype' : testtype }
    edit_challenge_dict['answeringoptions'] = ""
    if testtype == 'COMP':
        challengetypeslist = "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-weight:bold;justify:left;'>Select Challenge Type</font>:&nbsp;&nbsp;</div><div class='float-right' style='width:50%;'><select name='challengetype' onchange='javascript:displayoptions();' class='form-control input-ty-tiny'>"
        for ttcode in mysettings.TEST_TYPES.keys():
            ttcodeval = ttcode.replace(" ", "__")
            if ttcode == 'SUBJ':
                challengetypeslist += "<option value=%s selected>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
            elif ttcode == 'COMP':
                continue # A challenge cannot be composite
            else:
                challengetypeslist += "<option value=%s>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
        challengetypeslist += "</select></div></span><br /><div id='ansopts' style=''><span class='slide-container'><div class='slide' style='color:#0000AA;font-weight:bold;justify:left;'>Answer should not exceed </div><div class='float-right' style='width:50%;'><input type='number' name='maxsizewords' value='' size='6' maxlength='6' class='form-control input-ty-tiny' placeholder='no. of characters (leave empty for no limit)'></div></span></div>"
        edit_challenge_dict['challengetypeslist'] = challengetypeslist
    elif testtype == 'MULT' or testtype == 'FILB': # For 'CODN', 'ALGO' and 'SUBJ' type challenges, we need not provide any answering options.
        edit_challenge_dict['answeringoptions'] = "<p>"
        if testtype == 'MULT':
            edit_challenge_dict['answeringoptions'] += "<span class='slide-container row'><div class='slide align-items-start' id='ansopts'><font color='#0000AA' style='font-weight:bold;justify:left;'>Can there be more than one correct option:</font></div><div class='align-items-end' style='width:50%;'>&nbsp;<label class='radio-inline control-label' style='color:#0000AA;'><input type='radio' name='oneormore' value='yes' checked=true onchange='javascript:displayresponsekeycontrols();' class='form-control' style='height:35px;'>Yes</label>&nbsp;&nbsp;&nbsp;&nbsp;<label class='radio-inline control-label' style='color:#0000AA;'><input type='radio' name='oneormore' value='no' onchange='javascript:displayresponsekeycontrols();' class='form-control' style='height:35px;'>No</label></div></span><br />"
            edit_challenge_dict['answeringoptions'] += "<font color='#0000AA' style='font-weight:bold;'>Please enter the options you want to be made available for this challenge/question.(max 8 options) </font><br />"
            edit_challenge_dict['answeringoptions'] += "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #a:</font></div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice1' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #b:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice2' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #c:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice3' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #d:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice4' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #e:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice5' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #f:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice6' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #g:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice7' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #h:</div> <div class='float-right' style='width:50%;color:#0000AA;'> <input type='text' name='choice8' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny'></div></span>"
            edit_challenge_dict['responsekeyscontrolslist'] = "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;justify:left;'>Select the correct response(s)<font size='-2'><a href='#' onmouseover='javascript:showwhyresponsekey(this);' class='helptext'>[why should I do this?]</a></font>:</font></div></span><br />"
            # Since the default value for 'oneormore' is yes, we will initialize the 'responsekeyscontrolslist' with checkbox controls.
            edit_challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #a: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #b: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #c: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #d: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #e: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #f: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #g: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br /><span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;;text-align:right;display:inline-block;color:#0000AA;'>Option #h: </div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='' style='display:inline-block;'></div></span><br />"
        else:
            edit_challenge_dict['answeringoptions'] += "<input type='hidden' name='oneormore' value='no'>"
            edit_challenge_dict['responsekeyscontrolslist'] = "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;font-weight:bold;'>Enter the correct response<font size='-2'><a href='#' onmouseover='javascript:showwhyresponsekey(this);' class='helptext'>[why should I do this?]</a></font>:</font></div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='text' name='responsekey' value='' size='10' maxlength='250' class='form-control input-ty-tiny' style='width:300px;'></div></span><br />"
    elif testtype == 'CODN' or testtype == 'ALGO': # For these testtypes user may want some constraints on the size of the response.
        edit_challenge_dict['answeringoptions'] += "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;'>Max Lines in Answer (leave empty for no limit.) </font></div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='text' name='maxsizelines' value='' size='6' maxlength='6' class='form-control input-ty-tiny'> </div></span><br />"
    elif testtype == 'SUBJ':
        edit_challenge_dict['answeringoptions'] += "<span class='slide-container'><div class='slide' id='ansopts'><font color='#0000AA' style='font-style:italic;justify:left;'>Max Characters in Answer (leave empty for no limit)</font></div> <div class='float-right' style='width:50%;color:#0000AA;'><input type='text' name='maxsizewords' value='' size='6' maxlength='6' class='form-control input-ty-tiny'> </div></span><br />"
    lastchallengectr = int(lastchallengectr) + 1
    edit_challenge_dict['testlinkid'] = testlinkid
    edit_challenge_dict['test_id'] = testobj.id
    edit_challenge_dict['lastchallengectr'] = lastchallengectr.__str__()
    edit_challenge_dict['challengenumbersstr'] = lastchallengectr.__str__()
    edit_challenge_dict['evendistribution'] = evendistribution # This will be '1' or '0'.
    edit_challenge_dict['challengescore'] = -1
    if evendistribution and totalchallenges > 0:
        edit_challenge_dict['challengescore'] = str(float(totalscore)/float(totalchallenges))
    elif totalchallenges == -1:
        edit_challenge_dict['challengescore'] = ""
    edit_challenge_dict['negativescoring'] = negativescoring
    edit_challenge_dict['skillqualitylist'] = ""
    for skillqual in mysettings.SKILL_QUALITY.keys():
        if testobj.quality == skillqual:
            edit_challenge_dict['skillqualitylist'] += "<option value=%s selected>%s</option>"%(skillqual, mysettings.SKILL_QUALITY[skillqual])
        else:
            edit_challenge_dict['skillqualitylist'] += "<option value=%s>%s</option>"%(skillqual, mysettings.SKILL_QUALITY[skillqual])
    # Populate the existing challenges list
    challengesqset = Challenge.objects.filter(test=testobj).order_by('id')
    edit_challenge_dict['testname'] = testobj.testname
    edit_challenge_dict['challenge_links_list'] = []
    edit_challenge_dict.update(csrf(request))
    for challenge in challengesqset:
        challengestmt = challenge.statement[:20] + " ..."
        edit_challenge_dict['challenge_links_list'].append((challenge.id, challengestmt, testobj.id, testlinkid)) # So 'challenge_links_list' is a list of tuples whose elements (in order) are challenge Id (0), abbreviated challenge statement (1) (first 20 characters), test Id (2) and the testlinkid (3).
    # Create and render the template
    edit_challenge_dict['usrid'] = testobj.creator.id
    #edit_challenge_dict['treeview'] = _showtreeview(request, testobj)
    tmpl = get_template("tests/edit_challenge.html")
    cxt = Context(edit_challenge_dict)
    editchallengehtml = tmpl.render(cxt)
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        editchallengehtml = editchallengehtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return editchallengehtml


"""
Function to add or edit a challenge
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def edit(request):
    message = ''
    if request.method == "GET": # Illegal bad request... 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    (lastchallengectr, evendistribution, multimediareqd, totalscore, challengenumbersstr, csrfmiddlewaretoken, negativescoring, mediafile, oneormore, multiprogenv, diagramfile) = ("", False, False, 0, "", "", 0, "", False, '', '')
    # Retrieve challenge data and create challenge object...
    editoperationflag = False
    challengeobj = Challenge()
    if request.POST.has_key('challengeid'):
        challengeobj = Challenge.objects.filter(id=request.POST['challengeid'])[0]
        editoperationflag = True
    testobj = None
    if request.POST.has_key('testlinkid'):
        challengeobj.testlinkid = request.POST['testlinkid']
    if request.POST.has_key('statement'):
        challengeobj.statement = smart_text(request.POST['statement'], encoding='utf-8')
        #challengeobj.statement = request.POST['statement']
    if request.POST.has_key('challengetype'):
        challengeobj.challengetype = request.POST['challengetype']
    else:
        if not request.POST.has_key('testtype'):
            errmsg = error_msg('1053')
            return HttpResponse(errmsg)
        challengeobj.challengetype = request.POST['testtype']
    # Sanity check for challenge type:
    if challengeobj.challengetype not in mysettings.TEST_TYPES.keys():
        message = "Challenge type '%s':"%(challengeobj.challengetype) + error_msg('1054')
        return HttpResponse(message)
    if request.POST.has_key('test_id'):
        testobj = Test.objects.filter(id=request.POST['test_id'])[0]
        challengeobj.test = testobj
        challengeobj.testlinkid = testobj.testlinkid
    elif request.POST.has_key('testlinkid'):
        testobj = Test.objects.filter(testlinkid=request.POST['testlinkid'])[0]
        challengeobj.test = testobj
        challengeobj.testlinkid = request.POST['testlinkid']
    if not request.POST.has_key('test_id') and not request.POST.has_key('testlinkid'): # This should never happen if the user is logged in.
        challengeobj.test = None
        challengeobj.testlinkid = None
        errmsg = error_msg('1051') + error_msg('1111')
        return HttpResponse(errmsg)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    if request.POST.has_key('challengedurationseconds'):
        challengeobj.timeframe = request.POST['challengedurationseconds']
    if request.POST.has_key('mustrespond'):
        challengeobj.mustrespond = request.POST['mustrespond']
    if request.POST.has_key('lastchallengectr'):
        lastchallengectr = request.POST['lastchallengectr']
    if request.POST.has_key('evendistribution'):
        evendistribution = request.POST['evendistribution']
    if request.POST.has_key('multimediareqd'):
        multimediareqd = request.POST['multimediareqd']
    if request.POST.has_key('totalscore'):
        totalscore = request.POST['totalscore']
    if request.POST.has_key('challengenumbersstr'):
        challengenumbersstr = request.POST['challengenumbersstr']
    if request.POST.has_key('multiprogenv'):
        multiprogenv = request.POST['multiprogenv']
        multiprogenv = multiprogenv.replace('"', '')
        challengeobj.proglang = multiprogenv;
    if request.POST.has_key('progenv'):
        progenv = request.POST['progenv']
        progenv = progenv.replace('"', '')
        challengeobj.proglang = progenv;
    if request.POST.has_key('csrfmiddlewaretoken'):
        csrfmiddlewaretoken = request.POST['csrfmiddlewaretoken']
    if request.POST.has_key('negativescore'):
        negativescoring = request.POST['negativescore']
        if negativescoring == "":
            negativescoring = 0
        challengeobj.negativescore = negativescoring
    if request.POST.has_key('challengescore'):
        challengeobj.challengescore = request.POST['challengescore']
    if not challengeobj.challengescore or challengeobj.challengescore == "":
        challengeobj.challengescore = 0
    if request.FILES.has_key('mediafile'):
        mediafilename = request.FILES['mediafile'].name.split(".")[0]
        username = userobj.displayname
        fpath, message, challengemedia = skillutils.handleuploadedfile(request.FILES['mediafile'], mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__(), mediafilename)
        challengeobj.mediafile = request.FILES['mediafile'].name
    #if request.POST.has_key('imagecreatedfile') and request.POST['imagecreatedfile'] != '':
    #    challengeobj.mediafile = request.POST['imagecreatedfile'] # An image drawn on canvas by the user overrides an image uploaded by the user.
    if request.POST.has_key('diagramfile') and request.POST['diagramfile'] != "":
        diagramfile = request.POST['diagramfile']
        username = userobj.displayname
        diagrampath = mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__() + os.path.sep + diagramfile
        # Do we have a previously created mediafile for this challenge? If so, we need to delete it.
        # The reason we need to delete it is because a functionally disruptive attack may be launched
        # using this feature to eat away precious data centre memory drives. One may potentially 
        # keep creating data files this way to fill up all memory spaces.
        existingmediafile = challengeobj.mediafile
        if type(existingmediafile) == str and os.path.exists(mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__() + os.path.sep + existingmediafile):
            os.unlink(mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__() + os.path.sep + existingmediafile)
        challengeobj.mediafile = diagramfile
    challengeobj.maxresponsesizeallowable = ""
    if request.POST.has_key('maxsizewords'):
        challengeobj.maxresponsesizeallowable = request.POST['maxsizewords']
        challengeobj.maxresponsesizeallowable = re.sub(re.compile(r"&#39;"), '', challengeobj.maxresponsesizeallowable)
        if challengeobj.maxresponsesizeallowable == "":
            challengeobj.maxresponsesizeallowable = mysettings.MAXRESPONSECHARCOUNT
    elif request.POST.has_key('maxsizelines'):
        challengeobj.maxresponsesizeallowable = request.POST['maxsizelines']
        challengeobj.maxresponsesizeallowable = re.sub(re.compile(r"&#39;"), '', challengeobj.maxresponsesizeallowable)
        if challengeobj.maxresponsesizeallowable == "":
            challengeobj.maxresponsesizeallowable = mysettings.MAXRESPONSECHARCOUNT
    if request.POST.has_key('oneormore'):
        oneormore = request.POST['oneormore']
    if request.POST.has_key('skillquality'):
        challengeobj.challengequality = request.POST['skillquality']
    if request.POST.has_key('mathenv'):
        challengeobj.mathenv = request.POST['mathenv']
    if request.POST.has_key('extresourceurl'):
        challengeobj.additionalurl = request.POST['extresourceurl']
    # Now let us get the available options for 'MULT' type test
    if challengeobj.challengetype == 'MULT':
        fchoices = lambda(x):re.search(re.compile("^choice(\d+)$", re.IGNORECASE), x)
        allmatches = map(fchoices, request.POST)
        for fld in allmatches:
            if fld is None: # This is a different control, not one of the 'choice' controls
                continue
            controlname = fld.string
            ctrlcounter = fld.groups()[0]
            if not request.POST[controlname] or request.POST[controlname].strip() == "":
                continue
            challengeobj.__dict__['option%s'%ctrlcounter] = smart_text(request.POST[controlname].strip(), encoding='utf-8')
        challengeobj.maxresponsesizeallowable = -1
    # Store the responsekey if challengetype value is 'FILB' or 'MULT'
    # A note on the format used to store responsekey(s): For challengetype
    # value of 'FILB', the responsekey will be a single entry and will be
    # stored as is. For challengetype value of 'MULT' and 'oneormore' value of
    # 'no', we can store the value similarly. However, for challengetype 
    # value of 'MULT' and 'oneormore' value of 'yes', there may be multiple
    # entries, so we will store the entries as a single string joined together
    # using the string '#||#'.
    challengeobj.responsekey = None
    if challengeobj.challengetype == 'FILB' and request.POST.has_key('responsekey'):
        challengeobj.responsekey = smart_text(request.POST['responsekey'], encoding='utf-8')
    elif challengeobj.challengetype == 'MULT' and request.POST.has_key('responsekey') or request.POST.has_key('responsekey[]'):
        if oneormore == "no": # Only a single option will be correct
            challengeobj.responsekey = smart_text(request.POST['responsekey'], encoding='utf-8')
            challengeobj.oneormore = False
        elif oneormore == "yes": # Multiple options may be checked
            responses = request.POST.getlist('responsekey[]')
            challengeobj.oneormore = True
            challengeobj.responsekey = smart_text('#||#'.join(responses), encoding='utf-8')
    # ... and finally save the challenge object.
    if challengeobj.maxresponsesizeallowable == "":
        challengeobj.maxresponsesizeallowable = -1
    challengeobj.save()
    savedchallengesqset = Challenge.objects.filter(test=testobj)
    savedchallengescount = savedchallengesqset.__len__()
    totalchallengescount = testobj.challengecount
    if savedchallengescount > totalchallengescount: # reconcile the difference
        testobj.challengecount = savedchallengescount
        testobj.save()
        totalchallengescount = testobj.challengecount
    savedchallengesscore = 0
    for chlng in savedchallengesqset:
        savedchallengesscore += chlng.challengescore
    statusmessage = "<font color='#0000AA'>Number of Challenges framed:<b>%s</b><br>Total Number of Challenges in the test: <b>%s</b><br>Score accounted for: <b>%s</b><br>Total Score: <b>%s</b></font><br>"%(savedchallengescount.__str__(), totalchallengescount.__str__(), savedchallengesscore.__str__(), testobj.maxscore.__str__())
    if editoperationflag == True:
        statusmessage = "<font color='#0000BB'><b><i>The changes were saved successfully</i></b></font>"
    editchallengehtml = _challenge_edit_form(request, testobj, lastchallengectr,  evendistribution, challengeobj.timeframe, int(testobj.negativescoreallowed), multiprogenv)
    return HttpResponse(statusmessage + editchallengehtml)
    


def _showtreeview(request, testobj):
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    treeview = "<center><SCRIPT language='JavaScript' src='%s/javascript/treeview/ua.js'></SCRIPT><SCRIPT language='JavaScript' src='%s/javascript/treeview/ftiens4.js'></SCRIPT><SCRIPT language='JavaScript'> USETEXTLINKS = 1;STARTALLOPEN = 0;ICONPATH = '%s/images/treeview/';"%(mysettings.STATIC_ROOT, mysettings.STATIC_ROOT, mysettings.STATIC_ROOT)
    treeview += "foldersTree = gFld(\"<i><font size='-1' color='#0000AA'>%s</font></i>\", \"diffFolder.gif\", \"<img src=''>\");"%testobj.testname
    challengesqset = Challenge.objects.filter(test=testobj)
    varcounter = 1
    for chlng in challengesqset:
        treeview += "foldersTree.treeID = 'Frameset';"
        statementshort = chlng.statement[:10]
        treeview += "var varservice_%s = \"<a href='#' id='chkservice_%s' onClick='javascript:editchallange(%s);'>%s</a>\";"%(varcounter.__str__(),varcounter.__str__(), chlng.id, statementshort)
        varcounter += 1
    treeview += "</SCRIPT></center><DIV style='position:absolute; top:0; left:0;'><TABLE border=0><TR><TD><FONT size=-2><A style='font-size:7pt;text-decoration:none;color:silver' href='http://www.treemenu.net/'></A></FONT></TD></TR></TABLE></DIV><SCRIPT language='JavaScript'>initializeDocument();</SCRIPT>"
    return treeview


"""
View to display the test summary screen.  This screen allows the creator
to view challenges and various attributes of the challenges on one screen,
enabling the creator to modify/edit/delete them.
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def testsummary(request):
    message = ''
    if request.method != "GET": # Illegal bad request... 
        message = error_msg('1004')
        # A logging mechanism may be used to track how many and from where
        # such requests come and that may, sometimes, tell a curious story.
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    test_id = None
    if not request.GET.has_key('test_id'):
        message = error_msg('1055')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    test_id = request.GET['test_id']
    testobj = None
    tests_summary_dict = {}
    testqset = Test.objects.filter(id=test_id)
    if testqset.__len__() == 0:
        message = error_msg('1056')
        response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    else:
        testobj = testqset[0]
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    challengesqset = Challenge.objects.filter(test=testobj) # This should yield one or more challenges
    tests_summary_dict['challenge_links_list'] = []
    tests_summary_dict['challenges'] = []
    testlinkid = testobj.testlinkid
    for chlng in challengesqset:
        challenge = {}
        challenge['shortform'] = chlng.statement[:12] + " ..."
        challenge['id'] = chlng.id
        challenge['challengescore'] = chlng.challengescore
        challenge['challengetype'] = mysettings.TEST_TYPES[chlng.challengetype]
        challenge['timeframe'] = chlng.timeframe
        challenge['mediafile'] = chlng.mediafile
        challenge['mathenv'] = chlng.mathenv
        challenge['mediafileshortname'] = challenge['mediafile']
        if not challenge['mediafile']:
            challenge['mediafileshortname'] = ""
        else:
            challenge['mediafileshortname'] = chlng.mediafile[:8] + " ..."
        challenge['statement'] = chlng.statement
        for hexkey in mysettings.INV_HEXCODE_CHAR_MAP.keys():
            if hexkey == ' ':
                continue
            challenge['statement'] = challenge['statement'].replace(hexkey, mysettings.INV_HEXCODE_CHAR_MAP[hexkey]).replace("\r\n", "<br>")
        tests_summary_dict['challenge_links_list'].append((chlng.id, chlng.statement[:20] + " ...", testobj.id, testlinkid))
        tests_summary_dict['challenges'].append(challenge)
    tests_summary_dict['testname'] = testobj.testname
    tests_summary_dict['test_id'] = testobj.id
    tests_summary_dict['usrid'] = userobj.id
    tmpl = get_template("tests/test_summary.html")
    tests_summary_dict.update(csrf(request))
    cxt = Context(tests_summary_dict)
    testsummaryhtml = tmpl.render(cxt)
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        testsummaryhtml = testsummaryhtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(testsummaryhtml)


# Function to determine if a test is editable or not.
def iseditable(testobj):
    curdatetime = datetime.datetime.now()
    publishdate = testobj.publishdate.__str__()
    pubdateparts = publishdate.split(" ")
    pubdate = pubdateparts[0]
    pubtime = pubdateparts[1]
    pubyyyy, pubmon, pubdd = pubdate.split("-")
    pubhhmmss = pubtime.split(":")
    pubhh = int(pubhhmmss[0])
    pubmm = int(pubhhmmss[1])
    #pubss = int(pubhhmmss[2])
    publishdatetime = datetime.datetime(int(pubyyyy), int(pubmon), int(pubdd), pubhh, pubmm, 0)
    if publishdatetime < curdatetime and testobj.status:
        return None
    else:
        return 1


"""
This view is responsible for handling all delete challenge requests.
Delete challenge requests come from one of 2 javascript functions in
test.html - 'deletechallenge()' to delete a single challenge and
'deleteselected()' to delete multiple challenges at a single sweep.
NOTE: To be implemented - Challenges may be deleted only before 
publishdate of the test and only when the test is not active (activedate
is in future).
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def deletechallenges(request):
    message = ''
    if request.method != "POST": # Illegal bad request... 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode) # 'sessionobj' is a QuerySet object...
    userobj = sessionobj[0].user
    challenge_id = None
    if not request.POST.has_key('challengeid'):
        message = error_msg('1057')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    else:
        challenge_id = request.POST['challengeid']
    numchallenges = 1
    if request.POST.has_key('numchallenges'):
        numchallenges = request.POST['numchallenges']
    challengeidlist = challenge_id.split("#||#")
    if numchallenges == 1:
        challengeidlist = [challengeidlist[0], ]
    # Find the testobj that contains this challenge
    testobj = None
    try:
        testobj = Challenge.objects.filter(id=challengeidlist[0])[0].test
    except:
        message = error_msg('1061')
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    # Find if this test (or the challenges of the test) are editable...
    # If the current  date is past the publishdate or the status of the test
    # is active, then the test should not be made editable.
    res = iseditable(testobj) and not testobj.status
    if not res:
        message = "<font color='#FF0000'>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    challengesdeleted = []
    missingchidlist = []
    numpattern = re.compile("^\s*\d+\s*$")
    for chid in challengeidlist:
        if not re.search(numpattern, str(chid)): # This doesn't look like a valid challenge id to me.
            continue
        try:
            chlng = Challenge.objects.filter(id=chid)[0]
            chlng.delete()
            challengesdeleted.append(chid)
        except:
            missingchidlist.append(chid)
    message = "The challenges/questions identified by the following Ids have been deleted: " + ", ".join(challengesdeleted) + ". "
    testobj.challengecount = testobj.challengecount - challengesdeleted.__len__()
    testobj.save()
    if missingchidlist.__len__() > 0:
        message += "The following challenge Ids could not be found: " + ", ".join(missingchidlist) + ". "
    if challengesdeleted.__len__() == numchallenges:
        message = "The selected challenges where successfully deleted."
    response = HttpResponse(message)
    return response


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def savechanges(request):
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    challengeduration = {}
    challengescore = {}
    durationpattern = re.compile(r"^challengeduration_(\d+)$")
    scorepattern = re.compile(r"^challengescore_(\d+)$")
    challengeids = []
    for ekey in request.POST.keys():
        durationmatch = durationpattern.search(ekey)
        scorematch = scorepattern.search(ekey)
        if durationmatch:
            challengeduration[durationmatch.groups()[0]] = request.POST[ekey]
            challengeids.append(durationmatch.groups()[0])
        elif scorematch:
            challengescore[scorematch.groups()[0]] = request.POST[ekey]
        else:
            pass
    nonexistentchallenges = []
    #updatedchallenges = 0
    testobj = None
    try:
        testobj = Challenge.objects.filter(id=challengeids[0])[0].test
    except:
        message = "<font color='#FF0000'>%s</font>"%error_msg('1061')
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    if not iseditable(testobj) or testobj.status:
        message = "<font color='#FF0000'>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    for chid in challengeids:
        try:
            challenge = Challenge.objects.filter(id=chid)[0]
        except:
            nonexistentchallenges.append(chid)
        challenge.challengescore = challengescore[chid]
        if challenge.challengescore.strip() == "":
            challenge.challengescore = 0
        challenge.timeframe = challengeduration[chid]
        if challenge.timeframe.strip() == "" or not challenge.timeframe or challenge.timeframe == 'None':
            challenge.timeframe = testobj.duration
        try:
            challenge.save()
        except:
            print sys.exc_info()[1].__str__()
        #updatedchallenges += 1
    message = "challenges were successfully updated. "
    if nonexistentchallenges.__len__() > 0:
        message += "The following challenges could not be found: "
        message += ", ".join(nonexistentchallenges)
    return HttpResponse(message)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def addmorechallenges(request):
    testobj = None
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    # Create the test object from the test Id passed.
    if not request.POST.has_key('testid'):
        message = "<font color='#FF0000'>No Test Id found in request</font>";
        return HttpResponse(message)
    testid = request.POST['testid']
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "<font color='#FF0000'>Test with the specified Test Id (%s) was not found.</font>"%testid
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    if not iseditable(testobj) or testobj.status:
        message = "<font color='#FF0000' size=-1>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    # Get all challenges from the challenge table
    challengeqset = Challenge.objects.filter(test=testobj)
    lastchallengectr = 1
    evendistribution = False
    challengeduration = testobj.duration
    negativescoring = 1 # We make it True by default  as we do not store what the user specified during test creation. Not storing it doesn't make any difference to the user.
    if challengeqset.__len__() == 0:
        message += _challenge_edit_form(request, testobj, lastchallengectr,  evendistribution, challengeduration, int(negativescoring))
        return HttpResponse(message)
    else:
        lastchallengectr = challengeqset.__len__()
        evendistribution = False
        challengeduration = challengeqset[0].timeframe
        # We will set negativescoring to 1 as we can't be sure if the user wants to use it or not.
        #negativescoring = challengeqset[0].negativescore
        negativescoring = testobj.negativescoreallowed
    message = _challenge_edit_form(request, testobj, lastchallengectr,  evendistribution, challengeduration, int(testobj.negativescoreallowed))
    return HttpResponse(message)

"""
The following view handles the requests for editing existing tests.
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def editexistingtest(request):
    testobj = None
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    # Create the test object from the test Id passed.
    if not request.POST.has_key('testid'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1059');
        return HttpResponse(message)
    testid = request.POST['testid']
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "<font color='#FF0000'>Test with the specified Test Id (%s) was not found.</font>"%testid
        return HttpResponse(message)
    if not testobj: #  If we still have a null object (for whatever reason), inform the user that the operation has failed.
        message = "<font color='#FF0000'>%s</font>"%error_msg('1058')
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    if not iseditable(testobj) or testobj.status:
        message = "<font color='#FF0000'>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    create_test_dict = {}
    create_test_dict['testname'] = testobj.testname
    testtype = testobj.testtype
    """
    create_test_dict['testtypes'] = ""
    for ttcode in mysettings.TEST_TYPES.keys():
        ttcodeval = ttcode.replace(" ", "__")
        if ttcode == testtype:
            create_test_dict['testtypes'] += "<option value=&quot;%s&quot; selected>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
        else:
            create_test_dict['testtypes'] += "<option value=&quot;%s&quot;>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
    """
    create_test_dict['testtypesvalue'] = testtype
    ruleset = testobj.ruleset
    rules = ruleset.split("#||#")
    create_test_dict['testrules'] = ""
    for trule in mysettings.RULES_DICT.keys():
        if trule in rules:
            create_test_dict['testrules'] += "<option value=&quot;%s&quot; selected>%s</option>"%(trule, mysettings.RULES_DICT[trule])
        else:
            create_test_dict['testrules'] += "<option value=&quot;%s&quot;>%s</option>"%(trule, mysettings.RULES_DICT[trule])
    create_test_dict['topic'] = testobj.topic
    testtopicname = testobj.topicname
    create_test_dict['testtopics'] = ""
    unique_topics = {}
    for ttopics in mysettings.TEST_TOPICS:
        ttopicsval = ttopics.replace(" ", "__")
        ttopicsval = ttopics.replace('""', '"')
        if testtopicname == ttopics and not unique_topics.has_key(ttopicsval):
            create_test_dict['testtopics'] += "<option value='%s' selected>%s</option>"%(ttopicsval, ttopics)
            unique_topics[ttopicsval] = 1
        elif not unique_topics.has_key(ttopicsval):
            create_test_dict['testtopics'] += "<option value='%s'>%s</option>"%(ttopicsval, ttopics)
            unique_topics[ttopicsval] = 1
        # Get topics created in the past by this user
    usertopics = Topic.objects.filter(user=userobj, isactive=True)
    for topic in usertopics:
        topicname = topic.topicname.replace(" ", "__")
        topicname = topic.topicname.replace('""', '"')
        if testtopicname == topic.topicname and not unique_topics.has_key(topicname):
            create_test_dict['testtopics'] += "<option value='%s' selected>%s</option>"%(topicname, topic.topicname)
            unique_topics[topicname] = 1
        elif not unique_topics.has_key(topicname):
            create_test_dict['testtopics'] += "<option value='%s'>%s</option>"%(topicname, topic.topicname)
            unique_topics[topicname] = 1
    create_test_dict['totalscore'] = testobj.maxscore
    create_test_dict['evendistribution'] = 1 # Need to find from the existing challenges as to what its value should be.
    challengeqset = Challenge.objects.filter(test=testobj)
    prevchalscore = -1
    for challenge in challengeqset:
        if challenge.challengescore != prevchalscore and  prevchalscore != -1:
            create_test_dict['evendistribution'] = 0
            break
        prevchalscore = challenge.challengescore
    create_test_dict['negativescoring'] = testobj.negativescoreallowed
    create_test_dict['passscore'] = testobj.passscore
    if not create_test_dict['passscore']:
        create_test_dict['passscore'] = ""
    create_test_dict['testduration'] = testobj.duration
    create_test_dict['testdurationunit'] = "s"
    testduration_minute = int(testobj.duration)/60
    if testduration_minute > 0:
        create_test_dict['testduration'] = testduration_minute
        create_test_dict['testdurationunit'] = "m"
    # Randomly pick up a Challenge object for this test
    challengeobject = None
    challengeobjectqset = Challenge.objects.filter(test=testobj)
    if challengeobjectqset.__len__() == 0: #there are no challenge objects for this test
        create_test_dict['challengeduration'] = create_test_dict['testduration']  # so set the time duration of the entire test for this challenge - a bit of arbitrary decision.
        create_test_dict['challengedurationunit'] = create_test_dict['testdurationunit']
    else:
        challengeobject = challengeobjectqset[0]
    if challengeobject:
        create_test_dict['challengeduration'] = challengeobject.timeframe
        create_test_dict['challengedurationunit'] = 's'
    # Done with duration calculations...
    create_test_dict['grpname'] = testobj.evaluator
    evalobj = Evaluator.objects.filter(evalgroupname=testobj.evaluator)[0]
    create_test_dict['evaluators'] = ""
    uniqueevaluators = {}
    if evalobj.groupmember1 and evalobj.groupmember1.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember1.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember1.emailid + ","
            uniqueevaluators[evalobj.groupmember1.emailid] = 1
    if evalobj.groupmember2 and evalobj.groupmember2.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember2.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember2.emailid + ","
            uniqueevaluators[evalobj.groupmember2.emailid] = 1
    if evalobj.groupmember3 and evalobj.groupmember3.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember3.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember3.emailid + ","
            uniqueevaluators[evalobj.groupmember3.emailid] = 1
    if evalobj.groupmember4 and evalobj.groupmember4.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember4.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember4.emailid + ","
            uniqueevaluators[evalobj.groupmember4.emailid] = 1
    if evalobj.groupmember5 and evalobj.groupmember5.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember5.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember5.emailid + ","
            uniqueevaluators[evalobj.groupmember5.emailid] = 1
    if evalobj.groupmember6 and evalobj.groupmember6.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember6.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember6.emailid + ","
            uniqueevaluators[evalobj.groupmember6.emailid] = 1
    if evalobj.groupmember7 and evalobj.groupmember7.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember7.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember7.emailid + ","
            uniqueevaluators[evalobj.groupmember7.emailid] = 1
    if evalobj.groupmember8 and evalobj.groupmember8.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember8.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember8.emailid + ","
            uniqueevaluators[evalobj.groupmember8.emailid] = 1
    if evalobj.groupmember9 and evalobj.groupmember9.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember9.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember9.emailid + ","
            uniqueevaluators[evalobj.groupmember9.emailid] = 1
    if evalobj.groupmember10 and evalobj.groupmember10.emailid != "":
        if not uniqueevaluators.has_key(evalobj.groupmember10.emailid):
            create_test_dict['evaluators'] += evalobj.groupmember10.emailid + ","
            uniqueevaluators[evalobj.groupmember10.emailid] = 1
    if not uniqueevaluators.has_key(User.objects.filter(id=testobj.creator_id)[0].emailid):
        create_test_dict['creatoremail'] = User.objects.filter(id=testobj.creator_id)[0].emailid + "," + create_test_dict['evaluators']
    else:
        create_test_dict['creatoremail'] = create_test_dict['evaluators']
    # The variable name 'creatoremail' (above) is very misleading, but since I 
    # had been using it since earlier and don't want to disturb the stable
    # code, I am continuing with its usage. Note that the 'creatoremail'
    # actually holds the comma separated email ids of all evaluators as 
    # well as the email id of the creator.
    create_test_dict['creatoremail'] = re.sub(re.compile(r",\s*$"), "", create_test_dict['creatoremail']) # Removes all trailing comma's
    testlist_ascreator = Test.objects.filter(creator=userobj).order_by('-createdate')
    uniqevalgroups = {}
    evalgroupslitags, assocevalgrps = "", ""
    for test in testlist_ascreator:
        evalgrpemaillist = []
        if test.evaluator.groupmember1 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember1.emailid)
        if test.evaluator.groupmember2 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember2.emailid)
        if test.evaluator.groupmember3 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember3.emailid)
        if test.evaluator.groupmember4 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember4.emailid)
        if test.evaluator.groupmember5 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember5.emailid)
        if test.evaluator.groupmember6 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember6.emailid)
        if test.evaluator.groupmember7 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember7.emailid)
        if test.evaluator.groupmember8 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember8.emailid)
        if test.evaluator.groupmember9 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember9.emailid)
        if test.evaluator.groupmember10 is not None:
            evalgrpemaillist.append(test.evaluator.groupmember10.emailid)
        evalgrpemails = ",".join(evalgrpemaillist)
        test.evaluator.evalgroupname = test.evaluator.evalgroupname.replace("@", "__")
        test.evaluator.evalgroupname = test.evaluator.evalgroupname.replace(".", "__")
        assocevalgrps += "evalgrpsdict.%s = '%s';"%(test.evaluator.evalgroupname, evalgrpemails)
        #evalgroupslitags += "<li id=&quot;%s&quot; title=&quot;%s&quot;>%s</li>"%(test.evaluator.evalgroupname, evalgrpemails, test.evaluator.evalgroupname)
        if not uniqevalgroups.has_key(test.evaluator.evalgroupname):
            shortevalgrpname = test.evaluator.evalgroupname
            if shortevalgrpname.__len__() > 20:
                shortevalgrpname = shortevalgrpname[:20] + "..."
            #evalgroupslitags += "<li id=&quot;%s&quot; title=&quot;%s&quot; draggable=&quot;true&quot;>%s</li>"%(test.evaluator.evalgroupname, evalgrpemails, shortevalgrpname)
            evalgroupslitags += "<option value=&quot;%s&quot; title=&quot;%s&quot; draggable=&quot;true&quot;>%s</option>"%(test.evaluator.evalgroupname, evalgrpemails, shortevalgrpname)
            evalgroupslitags = evalgroupslitags.replace('&quot;', '\\"')
            uniqevalgroups[test.evaluator.evalgroupname] = test.evaluator.evalgroupname

    publishdatetime = testobj.publishdate.__str__()
    publishdateparts = publishdatetime.split(" ")
    publishdate, publishtime = publishdateparts[0], publishdateparts[1]
    publishyyyy, publishmon, publishday = publishdate.split("-")
    publishmon = mysettings.REV_MONTHS_DICT[publishmon]
    create_test_dict['publishdate'] = publishday + "-" + publishmon + "-" + publishyyyy
    activationdate = testobj.activationdate.__str__()
    activationdatetime = activationdate.split(" ")
    activationdatepart, activationtimepart = activationdatetime[0], activationdatetime[1]
    activationdateyyyy, activationdatemon, activationdateday = activationdatepart.split("-")
    activationdatemon = mysettings.REV_MONTHS_DICT[activationdatemon]
    create_test_dict['activedate'] = activationdateday + "-" + activationdatemon + "-" + activationdateyyyy
    skilltarget = testobj.quality
    create_test_dict['skilltarget'] = ""
    for skillcode in mysettings.SKILL_QUALITY.keys():
        if skillcode == skilltarget:
            create_test_dict['skilltarget'] += "<option value=&quot;%s&quot; selected>%s</option>"%(skillcode, mysettings.SKILL_QUALITY[skillcode])
        else:
            create_test_dict['skilltarget'] += "<option value=&quot;%s&quot;>%s</option>"%(skillcode, mysettings.SKILL_QUALITY[skillcode])
    testscope = testobj.scope
    create_test_dict['testscope'] = ""
    for tscope in mysettings.TEST_SCOPES:
        if tscope == testscope:
            create_test_dict['testscope'] += "<option value=&quot;%s&quot; selected>%s</option>"%(tscope, tscope)
        else:
            create_test_dict['testscope'] += "<option value=&quot;%s&quot;>%s</option>"%(tscope, tscope)
    answeringlanguages = testobj.allowedlanguages
    chosenlanguages = answeringlanguages.split('#||#')
    create_test_dict['answeringlanguage'] = ""
    for alang in mysettings.ANSWER_LANG_DICT.keys():
        if alang in chosenlanguages:
            create_test_dict['answeringlanguage'] += "<option value=&quot;%s&quot; selected>%s</option>"%(alang, mysettings.ANSWER_LANG_DICT[alang])
        else:
            create_test_dict['answeringlanguage'] += "<option value=&quot;%s&quot;>%s</option>"%(alang, mysettings.ANSWER_LANG_DICT[alang])
    testprogenv = testobj.progenv
    create_test_dict['progenv'] = ""
    if not testprogenv:
        create_test_dict['progenv'] += "<option value=&quot;0&quot; selected>None</option>"
        for proglang in mysettings.COMPILER_LOCATIONS.keys():
            create_test_dict['progenv'] += "<option value=&quot;%s&quot;>Yes - %s</option>"%(proglang, proglang)
    else:
        create_test_dict['progenv'] += "<option value=&quot;0&quot;>None</option>"
        for proglang in mysettings.COMPILER_LOCATIONS.keys():
            if testprogenv == proglang:
                create_test_dict['progenv'] += "<option value=&quot;%s&quot; selected>Yes - %s</option>"%(proglang, proglang)
            else:
                create_test_dict['progenv'] += "<option value=&quot;%s&quot;>Yes - %s</option>"%(proglang, proglang)
    create_test_dict['multimediareqd'] = testobj.multimediareqd
    create_test_dict['randomsequencing'] = testobj.randomsequencing
    create_test_dict['multipleattemptsallowed'] = testobj.allowmultiattempts
    create_test_dict['testlinkid'] = testobj.testlinkid
    create_test_dict['maxattemptscount'] = testobj.maxattemptscount
    create_test_dict['attemptsinterval'] = testobj.attemptsinterval
    create_test_dict['attemptsintervalunit'] = testobj.attemptsintervalunit
    create_test_dict['exist_test_id'] = testid
    create_test_dict['numchallenges'] = testobj.challengecount
    create_test_dict['createtesturl'] = skillutils.gethosturl(request) + '/' + mysettings.CREATE_TEST_URL
    create_test_dict['evalgroupslitags'] = evalgroupslitags
    #create_test_dict['datepickercode'] = "<link rel='stylesheet' type='text/css' media='all' href='static/datepick/jsDatePick_ltr.min.css' /><script type='text/javascript' src='static/datepick/jsDatePick.min.1.3.js'></script><script>new JsDatePick({ useMode:2, target:'publishdate', dateFormat:'%d-%M-%Y', limitToToday : false });  new JsDatePick({ useMode:2, target:'activedate', dateFormat:'%d-%M-%Y', limitToToday : false });</script>"
    tmpl = get_template("tests/create_test_form.html")
    create_test_dict.update(csrf(request))
    cxt = Context(create_test_dict)
    createtesthtml = tmpl.render(cxt)
    createtesthtml = re.sub(re.compile(r'\\"'), "&quot;", createtesthtml)
    #createtesthtml = re.sub(re.compile(r'"\s+\+\s+uid\s+\+\s+"'), userobj.id.__str__(), createtesthtml)
    #createtesthtml = "var uid='" + userobj.id + "';" + createtesthtml
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        createtesthtml = createtesthtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(createtesthtml)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def clearnegativescoreurl(request):
    testobj = None
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    # Create the test object from the test Id passed.
    if not request.POST.has_key('exist_test_id'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1059');
        return HttpResponse(message)
    testid = request.POST['exist_test_id']
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "<font color='#FF0000'>Test with the specified Test Id (%s) was not found.</font>"%testid
        return HttpResponse(message)
    if not testobj: #  If we still have a null object (for whatever reason), inform the user that the operation has failed.
        message = "<font color='#FF0000'>%s</font>"%error_msg('1058')
        return HttpResponse(message)
    if not iseditable(testobj) or testobj.status:
        message = "<font color='#FF0000'>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    # Now we are sure we have a valid Test object. Find all Challenges. But first, set 'negativescoreallowed' for the test object to False.
    testobj.negativescoreallowed = False
    allchallenges = Challenge.objects.filter(test=testobj)
    if allchallenges.__len__() == 0: # There are no challenges in this test.
        resp = "<font color='#0000BB' size=-1>Negative scoring has been disabled for this test.</font>"
        return HttpResponse(resp)
    for challenge in allchallenges:
        challenge.negativescore = 0
        challenge.save()
    testobj.save()
    resp = "<font color='#0000BB' size=-1>Negative scoring has been disabled for this test.</font>"
    return HttpResponse(resp)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def deletetest(request):
    testobj = None
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    # Create the test object from the test Id passed.
    if not request.POST.has_key('test_id'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1059');
        return HttpResponse(message)
    testid = request.POST['test_id']
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "<font color='#FF0000'>Test with the specified Test Id (%s) was not found.</font>"%testid
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    if testobj and (not iseditable(testobj) or testobj.status):
        message = "<font color='#FF0000' size=-1>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    testname = testobj.testname
    # Now we have the test_id of the test that we need to delete. So we fetch 
    # all the challenges belonging to that test and delete them first. The we 
    # delete the test and return a HttpResponse.
    allchallenges = Challenge.objects.filter(test=testobj)
    chcount = 0
    for challenge in allchallenges:
        challenge.delete() # Deleted challenge...
        chcount += 1
    # Now delete the test object
    testobj.delete()
    message = "<font color='#0000BB' size=-1>The test named '%s' containing %s challenges was successfully deleted.</font>"%(testname, chcount.__str__())
    return HttpResponse(message)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def showuserview(request):
    testobj = None
    challengeobj = None
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    # Create the test object from the test Id passed.
    if not request.POST.has_key('testid'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1059');
        return HttpResponse(message)
    testid = request.POST['testid']
    if not request.POST.has_key('challengeid'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1061');
        return HttpResponse(message)
    challengeid = request.POST['challengeid']
    """
    if not request.POST.has_key('testlinkid'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1062');
        return HttpResponse(message)
    testlinkid = request.POST['testlinkid']
    """
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "<font color='#FF0000'>Test with the specified Test Id (%s) was not found.</font>"%testid
        return HttpResponse(message)
    try:
        challengeobj = Challenge.objects.filter(id=challengeid)[0]
    except:
        message = "<font color='#FF0000'>Challenge with the specified Challenge Id (%s) was not found.</font>"%challengeid
        return HttpResponse(message)
    challenge_dict = {}
    # Is the test editable?
    challenge_dict['editable'] = 1
    if not iseditable(testobj) or testobj.status:
        challenge_dict['editable'] = 0;
    # We need to find whether we can have both 'previous' and 'next' buttons or
    # only one of them
    allchallenges = Challenge.objects.filter(test=testobj).order_by('id')
    # But before doing that we need to see if we got a 'trav' argument. If so
    # we need to alter the challenge Id accordingly.
    if request.POST.has_key('trav'):
        direction = request.POST['trav']
        if direction.lower() == 'prev': # Find the challenge which was added
            #just prior to the challenge identified by the current challenge Id.
            previd = -1
            for chlng in allchallenges:
                if str(chlng.id) == challengeid:
                    if previd == -1: # The challengeid we have is the first one. So do nothing.
                        break
                    else:
                        challengeid = previd
                        challengeobj = Challenge.objects.filter(id=challengeid)[0]
                        break
                previd = chlng.id
        elif direction.lower() == 'next': # Find the next challenge
            nextctr = 0
            for chlng in allchallenges:
                if str(chlng.id) == challengeid:
                    if nextctr == allchallenges.__len__() - 1: # This is the last challenge, no next one. So just break out.
                        break
                    else:
                        challengeid = allchallenges[nextctr + 1].id
                        challengeobj = Challenge.objects.filter(id=challengeid)[0]
                        break
                nextctr += 1
        else: # Unrecognized 'trav' value, so no action need be taken.
            pass
    # By default we have both of them.
    challenge_dict['previous'] = 1
    challenge_dict['next'] = 1
    chctr = 0
    for chlng in allchallenges:
        if chlng.id == challengeobj.id:
            break
        chctr += 1
    if chctr == 0: # This is the first challenge, so no 'previous' button.
        challenge_dict['previous'] = 0
    if chctr == allchallenges.__len__() - 1: # Last challenge, so no 'next' button.
        challenge_dict['next'] = 0
    challenge_dict['testid'] = testid
    challenge_dict['challengeid'] = challengeid
    challengetype = challengeobj.challengetype
    challengemedia = challengeobj.mediafile
    challenge_dict['challengestatement'] = challengeobj.statement.encode('utf-8')
    challenge_dict['challengetypedesc'] = ''
    if challengetype == 'MULT':
        challenge_dict['challengetypedesc'] = 'Multiple Choice'
    elif challengetype == 'FILB':
        challenge_dict['challengetypedesc'] = 'Fill up the Blanks'
    elif challengetype == 'SUBJ':
        challenge_dict['challengetypedesc'] = 'Subjective Type'
    elif challengetype == 'CODN':
        challenge_dict['challengetypedesc'] = 'Code writing/Programming'
    elif challengetype == 'ALGO':
        challenge_dict['challengetypedesc'] = 'Algorithms'
    else:
        challenge_dict['challengetypedesc'] = 'Unsupported Type'
    challenge_dict['challengetype'] = challengetype
    challenge_dict['responsekey'] = challengeobj.responsekey
    challenge_dict['challengemedia'] = challengemedia
    challenge_dict['testlinkid'] = testobj.testlinkid
    if challengemedia:
        username = userobj.displayname
        challenge_dict['challengemedia'] = "media" + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__() + os.path.sep + challengemedia
        #challenge_dict['challengemedia'] = "userdata" + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__() + os.path.sep + challengemedia
    challenge_dict['oneormore'] = challengeobj.oneormore
    challengeoptions = []
    if challengetype == 'MULT': # multiple choice type test
        for ctr in range(8):
            option = "option" + ctr.__str__()
            if challengeobj.__dict__.has_key(option) and challengeobj.__dict__[option] is not None:
                if challengeobj.__dict__[option] != "":
                    challengeoptions.append(challengeobj.__dict__[option])
        if challengeobj.oneormore:
            if challengeobj.responsekey is not None:
                responsekeyslist = challengeobj.responsekey.split('#||#')
                challenge_dict['responsekey'] = "', '".join(responsekeyslist)
                challenge_dict['responsekey'] = "'" + challenge_dict['responsekey'] + "'"
            else:
                challenge_dict['responsekey'] = 'Not available'
    else:
        challengeoptions = None # Except for 'MULT' type tests,
                                # challengeoptions do not have any significance.
    challenge_dict['proglanguage'] = ""
    if challengetype == 'CODN':
        proglanguage = Challenge.objects.filter(id=challengeid)[0].proglang
        challenge_dict['proglanguage'] = proglanguage
    challenge_dict['challengeoptions'] = challengeoptions
    challengescore = challengeobj.challengescore
    challengenegativescore = None
    if testobj.negativescoreallowed:
        challengenegativescore = challengeobj.negativescore
    maxtimeallowed = challengeobj.timeframe
    usermustrespond = challengeobj.mustrespond
    challenge_dict['challengescore'] = challengescore
    challenge_dict['challengenegativescore'] = challengenegativescore
    challenge_dict['maxtimeallowed'] = maxtimeallowed
    challenge_dict['usermustrespondresp'] = 'No'
    if challengeobj.mathenv == True:
        challenge_dict['mathenv'] = 1
    else:
        challenge_dict['mathenv'] = 0
    if usermustrespond:
        challenge_dict['usermustrespondresp'] = 'Yes'
    challenge_dict['additionalurl'] = challengeobj.additionalurl
    challenge_dict['challengequality'] = challengeobj.challengequality
    challenge_dict['oneormoreresp'] = 'Yes'
    if not challengeobj.oneormore:
        challenge_dict['oneormoreresp'] = 'No'
    tmpl = get_template("tests/challenge_user_view.html")
    challenge_dict.update(csrf(request))
    cxt = Context(challenge_dict)
    challengehtml = tmpl.render(cxt)
    return HttpResponse(challengehtml)
    

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def editchallenge(request):
    testobj = None
    challengeobj = None
    message = ''
    if request.method != "POST": 
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionobj = Session.objects.filter(sessioncode=sesscode)
    userobj = sessionobj[0].user
    # Need the test id to create the test object so we that can find the number of challenges it has.
    if not request.POST.has_key('testid'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1059');
        return HttpResponse(message)
    testid = request.POST['testid']
    if not request.POST.has_key('challengeid'):
        message = "<font color='#FF0000'>%s</font>"%error_msg('1061');
        return HttpResponse(message)
    challengeid = request.POST['challengeid']
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "<font color='#FF0000'>Test with the specified Test Id (%s) was not found.</font>"%testid
        return HttpResponse(message)
    if not iseditable(testobj) or testobj.status:
        message = "<font color='#FF0000' size=-1>%s</font>"%error_msg('1060')
        return HttpResponse(message)
    if not ispermittedtoedit(userobj, testobj):
        message = error_msg('1063')
        return HttpResponse(message)
    try:
        challengeobj = Challenge.objects.filter(id=challengeid)[0]
    except:
        message = "<font color='#FF0000'>Challenge with the specified Challenge Id (%s) was not found.</font>"%challengeid
        return HttpResponse(message)
    # So now we have the entire challenge object, so we create the dict for the template.
    challenge_dict = {}
    testchallenges = Challenge.objects.filter(test=testobj).order_by('id')
    challengenumbersstr = 1
    totalscore = 0
    evendistrib = True
    prevscore = 0
    for chlng in testchallenges:
        if chlng.id == challengeobj.id:
            challenge_dict['challengenumbersstr'] = challengenumbersstr
        challengenumbersstr += 1
        totalscore += chlng.challengescore
        if chlng.challengescore != prevscore and prevscore > 0:
            evendistrib = False
        prevscore = chlng.challengescore
    challenge_dict['challengeid'] = challengeid
    challenge_dict['statement'] = challengeobj.statement
    if challengeobj.mathenv == True:
        challenge_dict['mathenv'] = "checked"
    else:
        challenge_dict['mathenv'] = ""
    challenge_dict['multimediareqd'] = testobj.multimediareqd
    challenge_dict['negativescoring'] = testobj.negativescoreallowed
    challenge_dict['testtype'] = testobj.testtype
    challenge_dict['testname'] = testobj.testname
    challenge_dict['mediafile'] = challengeobj.mediafile
    challengemedia = None
    if challengeobj.mediafile:
        challengemedia = challengeobj.mediafile
    challenge_dict['extresourceurl'] = challengeobj.additionalurl
    if challenge_dict['extresourceurl'] is None:
        challenge_dict['extresourceurl'] = ""
    challenge_dict['challengescore'] = challengeobj.challengescore
    challenge_dict['existingchallengescore'] = challengeobj.challengescore
    challenge_dict['negativescore'] = challengeobj.negativescore
    challenge_dict['mustrespond'] = challengeobj.mustrespond
    challenge_dict['challengedurationseconds'] = challengeobj.timeframe
    challenge_dict['maxsizeallowable'] = challengeobj.maxresponsesizeallowable
    if challengeobj.maxresponsesizeallowable == -1: # app limit on response size
        challenge_dict['maxsizeallowable'] = mysettings.MAXRESPONSECHARCOUNT
        challengeobj.maxresponsesizeallowable = mysettings.MAXRESPONSECHARCOUNT
    respkeys = challengeobj.responsekey
    challenge_dict['responsekey'] = []
    challenge_dict['options'] = [] # This will be populated in case of 'MULT' type
    if challengeobj.challengetype == 'MULT':
        for ctr in range(8):
            option = "option" + ctr.__str__()
            if challengeobj.__dict__.has_key(option) and challengeobj.__dict__[option] is not None:
                if challengeobj.__dict__[option] != "":
                    challenge_dict['options'].append(challengeobj.__dict__[option])
    if challengeobj.responsekey and re.search(mysettings.SEPARATOR_PATTERN, challengeobj.responsekey):
        challenge_dict['responsekey'] = challengeobj.responsekey.split('#||#')
    else:
        challenge_dict['responsekey'] = [respkeys, ]
    challenge_dict['challengequality'] = challengeobj.challengequality
    challenge_dict['oneormore'] = challengeobj.oneormore
    challenge_dict['testlinkid'] = challengeobj.testlinkid
    if challengemedia is not None:
        username = userobj.displayname
        challenge_dict['challengemedia'] = "media" + os.path.sep + username + os.path.sep + "tests" + os.path.sep + testobj.id.__str__() + os.path.sep + challengemedia
    else:
        challenge_dict['challengemedia'] = ""
    challenge_dict['skillqualitylist'] = ''
    for skillqual in mysettings.SKILL_QUALITY.keys():
        if challenge_dict['challengequality'] == skillqual:
            challenge_dict['skillqualitylist'] += "<option value='%s' selected>%s</option>"%(skillqual, mysettings.SKILL_QUALITY[skillqual])
        else:
            challenge_dict['skillqualitylist'] += "<option value='%s'>%s</option>"%(skillqual, mysettings.SKILL_QUALITY[skillqual])
    challengetypeslist = ''
    if challenge_dict['testtype'] == 'COMP':
        challengetypeslist += "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-weight:bold;justify:left;'>Select Challenge Type</font></div><div class='float-right' style='width:50%;'><select name='challengetype' onchange='javascript:displayoptions();' class='form-control input-ty-tiny'>"
        for ttcode in mysettings.TEST_TYPES.keys():
            ttcodeval = ttcode.replace(" ", "__")
            if ttcode == challengeobj.challengetype:
                challengetypeslist += "<option value=%s selected>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
            elif ttcode == 'COMP':
                continue # A challenge cannot be composite
            else:
                challengetypeslist += "<option value=%s>%s</option>"%(ttcodeval, mysettings.TEST_TYPES[ttcode])
        challengetypeslist += "</select></div></span><br />"
        if challengeobj.challengetype == 'SUBJ':
            challengetypeslist += "<span class='slide-container'><div class='slide' id='ansopts'><font color='#0000AA' style='font-weight:bold;justify:left;'>Max Answer Length (characters, leave empty for no limit)</font></div><div class='float-right' style='width:50%%;color:#0000AA;'><input type='text' name='maxsizewords' value='%s' size='6' maxlength='6' class='form-control input-ty-tiny' style='width:300px;'> </div></span>"%(challenge_dict['maxsizeallowable'].__str__())
        elif challengeobj.challengetype == 'CODN' or challengeobj.challengetype == 'ALGO':
            challengetypeslist += "<span class='slide-container'><div class='slide' id='ansopts'><font color='#0000AA' style='font-weight:bold;justify:left;'>Max Answer Length (lines, leave empty for no limit)</font></div><div class='float-right' style='width:50%%;color:#0000AA;'> <input type='text' name='maxsizelines' value='%s' size='6' maxlength='6' class='form-control input-ty-tiny'> </div></span>"%(challenge_dict['maxsizeallowable'].__str__())
        else:
            pass
    challenge_dict['challengetypeslist'] = challengetypeslist
    challenge_dict['responsekeyscontrolslist'] = ''
    challenge_dict['answeringoptions'] = ''
    if challengeobj.challengetype == 'MULT':
        challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container' style='width:100%'><div class='slide' style='width:100%' id='ansopts'><font color='#0000AA' style='font-weight:bold;justify:left;'>Select the correct response(s)<font size='-2'><a href='#' onmouseover='javascript:showwhyresponsekey(this);' class='helptext'>[why should I do this?]</a></font>:</font></div></span><br>"
        if challenge_dict['oneormore']:
            challenge_dict['answeringoptions'] += "<span class='slide-container row'><div class='slide align-items-start'><font color='#0000AA' style='font-weight:bold;justify:left;'>Can there be more than one correct option:&nbsp;</font></div><div class='align-items-end' style='width:50%;color:#0000AA;'><label class='radio-inline control-label' style='color:#0000AA;'><input type='radio' name='oneormore' value='yes' checked=true onchange='javascript:displayresponsekeycontrols();' class='form-control' style='height:35px;'>Yes</label>&nbsp;&nbsp;&nbsp;&nbsp;<label class='radio-inline control-label' style='color:#0000AA;'><input type='radio' name='oneormore' value='no' onchange='javascript:displayresponsekeycontrols();' class='form-control' style='height:35px;'>No</label></div></span><br />"
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container row'><div class='slide align-items-start'><font color='#0000AA' style='font-weight:bold;justify:left;'>Can there be more than one correct option:&nbsp;</font></div><div class='align-items-end' style='width:50%;color:#0000AA;'><label class='radio-inline control-label' style='color:#0000AA;'><input type='radio' name='oneormore' value='yes' onchange='javascript:displayresponsekeycontrols();' class='form-control' style='height:35px;'>Yes</label>&nbsp;&nbsp;|&nbsp;&nbsp;<label class='radio-inline control-label' style='color:#0000AA;'><input type='radio' name='oneormore' value='no' checked=true onchange='javascript:displayresponsekeycontrols();' class='form-control' style='height:35px;'>No</label></div></span><br />"
        challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-weight:bold;justify:left;'>Please enter the options you want to be made available for this challenge/question.(max 8 options) </font></div></span><br />"
        #print challenge_dict['options'][4]
        if challenge_dict['options'].__len__() > 0 and challenge_dict['options'][0]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #a:&nbsp;<input type='text' name='choice1' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][0])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #a:<input type='text' name='choice1' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 1 and challenge_dict['options'][1]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #b:<input type='text' name='choice2' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][1])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #b: <input type='text' name='choice2' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 2 and challenge_dict['options'][2]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #c:<input type='text' name='choice3' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][2])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #c: <input type='text' name='choice3' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 3 and challenge_dict['options'][3]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #d:<input type='text' name='choice4' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][3])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #d:<input type='text' name='choice4' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 4 and challenge_dict['options'][4]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #e: <input type='text' name='choice5' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][4])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #e: <input type='text' name='choice5' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 5 and challenge_dict['options'][5]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;justify:left;'>Option #f:<input type='text' name='choice6' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][5])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #f:<input type='text' name='choice6' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 6 and challenge_dict['options'][6]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #g:<input type='text' name='choice7' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"%(challenge_dict['options'][6])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #g: <input type='text' name='choice7' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['options'].__len__() > 7 and challenge_dict['options'][7]:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%%;'><div class='slide' style='width:100%%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #h:<input type='text' name='choice8' value='%s' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span></br>"%(challenge_dict['options'][7])
        else:
            challenge_dict['answeringoptions'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><font color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;'>Option #h:<input type='text' name='choice8' value='' onblur='javascript:displayresponsekeycontrols();' class='form-control input-ty-tiny' style='display:inline-block;'></font></div></span><br />"
        if challenge_dict['oneormore']:
            for option in challenge_dict['options']:
                if option  in challenge_dict['responsekey']:
                    challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><label color='#0000AA' style='font-style:italic;justify:right;text-align:right;display:inline-block;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='" + option  + "' checked class='form-control pull-right input-ty-tiny' style='display:inline-block;'>&nbsp;" + option + "</label></div></span><br>"
                else:
                    challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><label color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;color:#0000AA;'><input type='checkbox' name='responsekey[]' value='" + option + "' class='form-control pull-right input-ty-tiny' style='display:inline-block;'>&nbsp;" + option + "</label></div></span><br>"
        else: # Only one option is correct
            for option in challenge_dict['options']:
                if option  in challenge_dict['responsekey']:
                    challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><label color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;color:#0000AA;'><input type='radio' name='responsekey' value='" + option  + "' checked=true class='form-control pull-right input-ty-tiny' style='display:inline-block;'>&nbsp;" + option + "</label></div></span><br>"
                else:
                    challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container' style='width:100%;'><div class='slide' style='width:100%;'><label color='#0000AA' style='font-style:italic;text-align:right;display:inline-block;color:#0000AA;'><input type='radio' name='responsekey' value='" + option + "' class='form-control pull-right input-ty-tiny' style='display:inline-block;'>&nbsp;" + option + "</label></div></span><br>"
    elif challengeobj.challengetype == 'FILB':
         challenge_dict['responsekeyscontrolslist'] += "<span class='slide-container'><div class='slide'><font color='#0000AA' style='font-style:italic;justify:left;font-weight:bold;'>Enter the correct response<font size='-2'><a href='#' onmouseover='javascript:showwhyresponsekey(this);' class='helptext'>[why should I do this?]</a></font>:</font></div> <div class='float-right' style='width:50%%;color:#0000AA;'><input type='text' name='responsekey' value='%s' size='10' maxlength='250' class='form-control input-ty-tiny' style='width:300px;'></div></span><br>"%(challenge_dict['responsekey'][0])
         challenge_dict['answeringoptions'] += "<input type='hidden' name='oneormore' value='no'>"
    elif challengeobj.challengetype == 'CODN' or challengeobj.challengetype == 'ALGO':
        pass
        #challenge_dict['answeringoptions'] += "<b>Answer should not exceed <input type='text' name='maxsizelines' value='%s' size='6' maxlength='6'> lines. </b>(leave empty for no limit.)</p>"%(challenge_dict['maxsizeallowable'].__str__())
    elif challengeobj.challengetype == 'SUBJ':
        pass
        #challenge_dict['answeringoptions'] += "<b>Answer should not exceed <input type='text' name='maxsizewords' value='%s' size='6' maxlength='6'> characters</b>(leave empty for no limit)</p>"%(challenge_dict['maxsizeallowable'].__str__())
    else:
        pass
    challenge_dict['evendistribution'] = evendistrib
    challenge_dict['test_id'] = testid # or challengeobj.test.id
    challenge_dict['lastchallengectr'] = testchallenges.__len__()
    challenge_dict['totalscore'] = totalscore
    challenge_dict['usrid'] = userobj.id
    challenge_dict['challenge_links_list'] = []
    for challenge in testchallenges:
        challengestmt = challenge.statement[:20] + " ..."
        challenge_dict['challenge_links_list'].append((challenge.id, challengestmt, testobj.id, testobj.testlinkid))
    tmpl = get_template("tests/edit_challenge.html")
    challenge_dict.update(csrf(request))
    cxt = Context(challenge_dict)
    challengehtml = tmpl.render(cxt)
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        challengehtml = challengehtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(challengehtml)


"""
The function below gets the shortened form of the test URL from the DB
if it has already been created  or generates the shortened URL from bit.ly
and stores it as an attribute of the Test model (i.e., as a field in the
Tests_test table in the 'testyard' database). To access the test using
this URL, we use the bit.ly API every time to expand the short URL, get
the parameters from this URL (in order to identify the test to display),
and then display the test to the user.
"""
def gettesturlforuser(targetuseremail, testid, baseurl):
    testobj = None
    targetuserid = targetuseremail # If we can't find the targetuserid, then we will use the targetuseremail.
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "Error: " + error_msg('1056')
        response = HttpResponse(message)
        return response
    try:
        targetuserobj = User.objects.filter(emailid=targetuseremail)[0]
        targetuserid = targetuserobj.id
    except:
        pass
    randstring = skillutils.randomstringgen()
    query_string = "targetuser=" + str(targetuserid) + "&testid=" + str(testobj.id) + "&mode=test&targetemail=" + targetuseremail + "&rand=" + randstring
    encoded_query_string = base64.b64encode(query_string)
    testurl = baseurl + "/" + mysettings.SHOW_TEST_CANDIDATE_MODE_URL + "?" + encoded_query_string
    jsonstringcontent = skillutils.getshorturl(testurl)
    jsonobj = json.loads(jsonstringcontent)
    try:
        shorttesturl = jsonobj['url']
    except:
        print("Error generating short URL from URL Shortener Service: %s"%json.dumps(jsonobj))
        shorttesturl = ""
    return (shorttesturl, randstring)




def gettesturlforuser_bitly_latest(targetuseremail, testid, baseurl):
    testobj = None
    targetuserid = targetuseremail # If we can't find the targetuserid, then we will use the targetuseremail.
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "Error: " + error_msg('1056')
        response = HttpResponse(message)
        return response
    try:
        targetuserobj = User.objects.filter(emailid=targetuseremail)[0]
        targetuserid = targetuserobj.id
    except:
        pass
    randstring = skillutils.randomstringgen()
    query_string = "targetuser=" + str(targetuserid) + "&testid=" + str(testobj.id) + "&mode=test&targetemail=" + targetuseremail + "&rand=" + randstring
    encoded_query_string = base64.b64encode(query_string)
    testurl = baseurl + "/" + mysettings.SHOW_TEST_CANDIDATE_MODE_URL + "?" + encoded_query_string
    # Now bitlyfy the testurl
    bitlyapiurl = mysettings.BITLY_LINK_API_ADDRESS
    #postdatadict = {"group_guid" : mysettings.BITLY_APP_GROUP, "domain" : "bit.ly", "long_url" : testurl}
    # Note: Despite mentioned in the documentation, the group_guid should actually be omitted from the request. Wasted a lot of time.
    postdatadict = {"domain" : "bit.ly", "long_url" : testurl}
    postdata = json.dumps(postdatadict)
    httpheaders = {'Authorization' : 'Bearer %s'%mysettings.BITLY_OAUTH_ACCESS_TOKEN, 'Content-type' : 'application/json'}
    try:
        httpresponse = requests.post(bitlyapiurl, headers=httpheaders, data=postdata)
        jsonstringcontent = httpresponse.content
    except:
        message = "Error: %s"%(sys.exc_info()[1].__str__())
        response = HttpResponse(message)
        return response
    jsonobj = json.loads(jsonstringcontent)
    #print(jsonobj)
    try:
        shorttesturl = jsonobj['link']
    except:
        print("Error generating short URL from Bitly: %s"%json.dumps(jsonobj))
        shorttesturl = ""
    return (shorttesturl, randstring)



def gettesturlforuser_old_bitly_v3(targetuseremail, testid, baseurl):
    testobj = None
    targetuserid = targetuseremail # If we can't find the targetuserid, then we will use the targetuseremail.
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "Error: " + error_msg('1056')
        response = HttpResponse(message)
        return response
    try:
        targetuserobj = User.objects.filter(emailid=targetuseremail)[0]
        targetuserid = targetuserobj.id
    except:
        pass
    randstring = skillutils.randomstringgen()
    query_string = "targetuser=" + str(targetuserid) + "&testid=" + str(testobj.id) + "&mode=test&targetemail=" + targetuseremail + "&rand=" + randstring
    encoded_query_string = base64.b64encode(query_string)
    testurl = baseurl + "/" + mysettings.SHOW_TEST_CANDIDATE_MODE_URL + "?" + encoded_query_string
    # Now bitlyfy the testurl
    bitlyapiurl = mysettings.BITLY_LINK_API_ADDRESS + "/v3/shorten?access_token=" + mysettings.BITLY_OAUTH_ACCESS_TOKEN + "&longUrl=" + skillutils.urlencodestring(testurl)
    try:
        httpresponsejson = urllib2.urlopen(bitlyapiurl)
        jsonstringcontent = httpresponsejson.read()
    except:
        message = "Error: %s"%(sys.exc_info()[1].__str__())
        response = HttpResponse(message)
        return response
    jsonobj = json.loads(jsonstringcontent)
    print(jsonobj)
    shorttesturl = jsonobj['data']['url']
    return (shorttesturl, randstring)


"""
The function below generates a test URL at realtime. For a given test,
it creates the URL every time. But before creating it checks the 
privileges of the user it is servicing, and handles the generation of the
test URL accordingly. The view pointed to by this URL is similar to what
the candidate sees while taking the test. The only difference will be
the controls for entering the response would be readonly and/or disabled.
The user would also be able to view the correct answers (if any) prov-
ided by the creator.
"""
def generatetesturl(testobj, userobj, tests_user_dict):
    if not ispermittedtoview(userobj, testobj, tests_user_dict):
        return ""
    else:
        testurl = tests_user_dict['baseURL'] + "/" + mysettings.VIEW_TEST_URL + "?tlinkid=" + testobj.testlinkid + "&testid=" + str(testobj.id)
    return testurl


"""
This is the main function that is triggered when a user takes a test
or the test is displayed to the user as it would appear to a user who
is taking the test. Now, there are 2 ways to handle this, and both 
have some advantages and disadvantages over the other. 1. Send the 
entire test (the challenges, the rules, etc.) as a json string to the
user, and 2. Send each challenge at a time, as the user presses the
'Next' button. We will be handling the test using the option #1 above
and in order to tackle cases of mischief, we will encrypt the json 
string that we send. On the client side, we will place javascript code
that will decrypt the json, and do the needful to allow the user to 
take the test, and for every challenge/question answered, it will send
the user's response to the server along with the challenge Id and the
test Id (among other variables like the userId. A view on the server 
will accept the response and do the needful (that is, it will enter 
the response in the name of the user to the database).
"""
def showtestcandidatemode(request):
    if request.method != "POST" and request.method != 'GET': # If it is not a
        # POST or GET request, shoot it down. POST request comes when an
        # evaluator or creator tries to view the test as shown to the
        # candidates. GET comes when the user clicks on the bitly/urlshortener test link.
        message = error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    query_string = ""
    try:
        query_string = base64.b64decode(request.META['QUERY_STRING'])
        #query_string = base64.b64decode(request.POST)
    except:
        message = error_msg('1078')
        response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    query_string_parts = query_string.split("&")
    testid, targetemail, mode, rand = -1, "", "", ""
    for qpart in query_string_parts:
        qkey, qval = qpart.split("=")
        if qkey == 'testid':
            testid = qval
        elif qkey == 'targetemail':
            targetemail = qval
        elif qkey == 'mode':
            mode = qval
        elif qkey == 'rand':
            rand = qval
    if not testid or testid == -1:
        message = error_msg('1059')
        response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    testdict = {} # This will be our json object...
    testobj = None
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = error_msg('1056')
        response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    testtakeruserqset = UserTest.objects.filter(emailaddr=targetemail).filter(test=testobj).filter(active=True).filter(cancelled=False).filter(stringid=rand)
    tabtype = 'usertest'
    testtakeruserobj = None
    #print targetemail
    if testtakeruserqset.__len__() == 0:
        testtakeruserqset = WouldbeUsers.objects.filter(emailaddr=targetemail).filter(test=testobj).filter(active=True).filter(cancelled=False).filter(stringid=rand)
        tabtype = 'wouldbeusers'
        if testtakeruserqset.__len__() == 0:
            message = error_msg('1077')
            response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
            return response
    testtakeruserobj = testtakeruserqset[0]
    if testtakeruserobj.starttime and testtakeruserobj.schedule is None: # User had already started taking this test before, so disqualify user since she/he is trying to restart it.
        message = "Error: %s\n"%error_msg('1049')
        response = HttpResponse(message)
        return response
    tabid = testtakeruserobj.id
    userobj = None
    try:
        userobj = User.objects.filter(emailid=testtakeruserobj.emailaddr)[0]
    except:
        pass
    # Check the status and publish date and activation date
    if not testobj.status: # Test is being edited or modified
        message = error_msg('1064')
        response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    currentdatetime = datetime.datetime.now()
    activationdate = skillutils.mysqltopythondatetime(testobj.activationdate.__str__()) # It is the activation date that matters to the user taking the test. A test that has been activated has definitely been published.
    if currentdatetime < activationdate: # This test is not active yet.
        message = error_msg('1065')
        response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    # If this is a scheduled test check if current time is in the interval bounded by validfrom and validtill
    if testtakeruserobj.schedule is not None:
        curdatetime = datetime.datetime.now() 
        curdatetime_year, curdatetime_month, curdatetime_day, curdatetime_hour, curdatetime_min, curdatetime_sec = curdatetime.year, curdatetime.month, curdatetime.day, curdatetime.hour, curdatetime.minute, curdatetime.second
        curdatetime_tzaware = datetime.datetime(curdatetime_year, curdatetime_month, curdatetime_day, curdatetime_hour, curdatetime_min, curdatetime_sec, 0, pytz.UTC)
        validfrom = testtakeruserobj.validfrom
        validtill = testtakeruserobj.validtill
        if curdatetime_tzaware < validfrom: # test is scheduled in the future
            message = error_msg('1164')%(testtakeruserobj.validfrom, testtakeruserobj.validtill)
            response = HttpResponse(message)
            return response
        elif curdatetime_tzaware > validtill: # Test is over
            message = error_msg('1165')%(testtakeruserobj.validfrom, testtakeruserobj.validtill)
            response = HttpResponse(message)
            return response
        else: # Give the test to the candidate.
            pass
    # Now, check if this user is the creator or evaluator of this test.
    # If so, set a flag in the json while sending the test. Tests with
    # this flag set will not be sending back responses (to be implemented
    # in the javascript). By the way, this implies that a creator or an
    # evaluator would never be allowed to take the test, which is not
    # an insane assumption.
    testcreator = testobj.creator
    testevaluator = testobj.evaluator
    testevalemailidlist = []
    if testevaluator.groupmember1:
        testevalemailidlist.append(testevaluator.groupmember1.emailid)
    if testevaluator.groupmember2:
        testevalemailidlist.append(testevaluator.groupmember2.emailid)
    if testevaluator.groupmember3:
        testevalemailidlist.append(testevaluator.groupmember3.emailid)
    if testevaluator.groupmember4:
        testevalemailidlist.append(testevaluator.groupmember4.emailid)
    if testevaluator.groupmember5:
        testevalemailidlist.append(testevaluator.groupmember5.emailid)
    if testevaluator.groupmember6:
        testevalemailidlist.append(testevaluator.groupmember6.emailid)
    if testevaluator.groupmember7:
        testevalemailidlist.append(testevaluator.groupmember7.emailid)
    if testevaluator.groupmember8:
        testevalemailidlist.append(testevaluator.groupmember8.emailid)
    if testevaluator.groupmember9:
        testevalemailidlist.append(testevaluator.groupmember9.emailid)
    if testevaluator.groupmember10:
        testevalemailidlist.append(testevaluator.groupmember10.emailid)
    testdict['usercreatorevaluatorflag'] = 0
    if testtakeruserobj.emailaddr == testcreator.emailid:
        testdict['usercreatorevaluatorflag'] = 1
    elif testtakeruserobj.emailaddr in testevalemailidlist:
        testdict['usercreatorevaluatorflag'] = 1
    else:
        pass
    testdict['testname'] = testobj.testname
    testdict['testid'] = testobj.id
    testdict['topicname'] = testobj.topicname # for built-in topic
    if testobj.topic != "": # topic is not one of the built-in topics
        testdict['topicname'] = testobj.topic.topicname
    testdict['creator'] = testobj.creator.displayname # Not sure if this should be shown to the test taker.
    testdict['maxscore'] = testobj.maxscore
    testdict['passscore'] = testobj.passscore
    testdict['rules'] = testobj.ruleset.split("#||#")
    testdict['duration'] = testobj.duration
    testdict['allowedlanguages'] = testobj.allowedlanguages.split("#||#")
    testdict['randomsequencing'] = testobj.randomsequencing
    testdict['multimediareqd'] = testobj.multimediareqd
    testdict['progenv'] = testobj.progenv
    testdict['quality'] = testobj.quality
    testdict['negativescoreallowed'] = testobj.negativescoreallowed
    testdict['scope'] = testobj.scope
    testdict['sendtestdataurl'] = mysettings.SEND_TEST_DATA_URL
    testdict['targetemail'] = targetemail
    testdict['tabtype'] = tabtype
    testdict['tabid'] = tabid
    testdict['codepadexecuteurl'] = skillutils.gethosturl(request) + "/" + mysettings.CODEPAD_EXECUTE_URL
    testdict['codeexecurl'] = skillutils.gethosturl(request) + "/" + mysettings.CODE_EXEC_URL
    testdict['reportwindowchangeurl'] = skillutils.gethosturl(request) + "/" + mysettings.REPORT_WINDOW_CHANGE_URL
    testdict['gettimeremainingurl'] = skillutils.gethosturl(request) + "/" + mysettings.GET_TIME_REMAINING_URL
    testdict['checkconnectionurl'] = skillutils.gethosturl(request) + "/" + mysettings.CHK_INTERNET_CONN_URL
    #testdict['testlink'] = request.META['HTTP_REFERER']
    # If the test taker is a candidate, we need to check for multiple attempts...
    if not testdict['usercreatorevaluatorflag']: 
        allowmultiattempts = testobj.allowmultiattempts
        usertest = testtakeruserobj
        usertestqset = testtakeruserqset.filter(status=2)
        if allowmultiattempts and not usertest.schedule: # Multiple attempts have meaning only in the context of tests that are not scheduled.
            maxattemptscount = testobj.maxattemptscount
            if usertestqset.__len__() >= int(maxattemptscount):
                message = error_msg('1067')
                response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
                return response
            attemptsinterval = testobj.attemptsinterval
            attemptsintervalunit = testobj.attemptsintervalunit
            if attemptsintervalunit == 'h':
                attemptsintervalseconds = int(attemptsinterval) * 3600
            elif attemptsintervalunit == 'm':
                attemptsintervalseconds = int(attemptsinterval) * 60
            elif attemptsintervalunit == 'd':
                attemptsintervalseconds = int(attemptsinterval) * 86400
            elif attemptsintervalunit == 'M':
                attemptsintervalseconds = int(attemptsinterval) * 86400 * 30 # Taking 30 days as month.
            elif attemptsintervalunit == 'Y':
                attemptsintervalseconds = int(attemptsinterval) * 86400 * 30*12
            else:
                attemptsintervalseconds = int(attemptsinterval) # take the interval at face value.
            if usertest.status > 0 and usertest.starttime is not None: # User has taken the test previously
                lasttimetesttaken = skillutils.mysqltopythondatetime(str(usertest.starttime))
                testtakentimedelta = currentdatetime - lasttimetesttaken
                if testtakentimedelta.total_seconds() < attemptsintervalseconds:
                    secondsremaining = attemptsintervalseconds - testtakentimedelta.total_seconds()
                    message = error_msg('1068') + " Please wait for %s before attempting this test."%skillutils.converttimeunit(secondsremaining)
                    response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
                    return response
            else: # User has never taken this test before.
                pass
        else: # multiple attempts not allowed ...
            if usertestqset and usertestqset.__len__() > 0: # ... and user has taken the test already
                message = error_msg('1066')
                response = HttpResponse(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
                return response
        # Now check if the test is valid now.
        if skillutils.mysqltopythondatetime(str(usertest.validtill)) < currentdatetime and not usertest.schedule:
            message = "Error: %s\n"%error_msg('1074')
            response = HttpResponse(message)
            return response
        # Check if the user has already taken the test or if the test is being taken now.
        if usertest.status == 2 and not usertest.schedule:
            message = "Error: %s\n"%error_msg('1075')
            response = HttpResponse(message)
            return response
    else: # User is creator or evaluator, so can't take the test.
        message = "Error: %s\n"%error_msg('1076')
        response = HttpResponse(message)
        return response
    # If the control comes here, the user can take this test.
    challengesqset = Challenge.objects.filter(test=testobj)
    challengesdict = {}
    testdict['challenges'] = challengesdict
    for challenge in challengesqset:
        statement = challenge.statement
        challengetype = challenge.challengetype
        maxresponsesizeallowable = challenge.maxresponsesizeallowable
        oftheabovePattern = re.compile("\s+of\s+the\s+above", re.IGNORECASE|re.DOTALL)
        #noneabovePattern = re.compile(/None\s+of\s+the\s+above/g, re.DOTALL|re.IGNORECASE)
        challengesdict[statement] = {'challengetype' : challengetype, 'maxresponsesizeallowable' : maxresponsesizeallowable}
        validoptionscount = 0
        if challengetype == 'MULT' or challengetype == 'FILB':
            if challenge.option1 and challenge.option1 != "":
                challengesdict[statement]['option1'] = challenge.option1
                validoptionscount += 1
            if challenge.option2 and challenge.option2 != "":
                challengesdict[statement]['option2'] = challenge.option2
                validoptionscount += 1
            if challenge.option3 and challenge.option3 != "":
                challengesdict[statement]['option3'] = challenge.option3
                validoptionscount += 1
            if challenge.option4 and challenge.option4 != "":
                challengesdict[statement]['option4'] = challenge.option4
                validoptionscount += 1
            if challenge.option5 and challenge.option5 != "":
                challengesdict[statement]['option5'] = challenge.option5
                validoptionscount += 1
            if challenge.option6 and challenge.option6 != "":
                challengesdict[statement]['option6'] = challenge.option6
                validoptionscount += 1
            if challenge.option7 and challenge.option7 != "":
                challengesdict[statement]['option7'] = challenge.option7
                validoptionscount += 1
            if challenge.option8 and challenge.option8 != "":
                challengesdict[statement]['option8'] = challenge.option8
                validoptionscount += 1
            log = Logger(mysettings.LOG_PATH + "/testyard.log")
            opt1taken = 0
            for opt in challengesdict[statement].keys():
                optval = challengesdict[statement][opt]
                if type(optval) != str and type(optval) != unicode:
                    optval = str(optval)
                log.logmessage(optval)
                optval = optval.replace("__LT__", "&lt;")
                optval = optval.replace("__GT__", "&gt;")
                challengesdict[statement][opt] = optval
                if oftheabovePattern.search(optval.encode('utf-8')):
                    tempval = challengesdict[statement][opt]
                    if not opt1taken:
                        #lastoption = 'option' + str(validoptionscount)
                        lastoption = 'option1'
                        opt1taken = 1
                    elif opt1taken == 1:
                        #lastoption = 'option' + str(validoptionscount - 1)
                        lastoption = 'option2'
                        opt1taken += 1
                    else:
                        #lastoption = 'option' + str(validoptionscount - 2)
                        lastoption = 'option3'
                        opt1taken += 1
                    challengesdict[statement][opt] = challengesdict[statement][lastoption]
                    challengesdict[statement][lastoption] = tempval
                else:
                    pass
        challengesdict[statement]['challengescore'] = challenge.challengescore
        challengesdict[statement]['negativescore'] = challenge.negativescore
        challengesdict[statement]['mustrespond'] = challenge.mustrespond
        challengesdict[statement]['mediafile'] = challenge.mediafile
        challengesdict[statement]['additionalurl'] = challenge.additionalurl
        challengesdict[statement]['timeframe'] = challenge.timeframe
        challengesdict[statement]['challengequality'] = challenge.challengequality
        challengesdict[statement]['oneormore'] = challenge.oneormore
        challengesdict[statement]['chid'] = challenge.id
        challengesdict[statement]['progenv'] = challenge.test.progenv
        if challenge.mathenv == True:
            challengesdict[statement]['mathenv'] = 1
        else:
            challengesdict[statement]['mathenv'] = 0
        challengesdict[statement]['maxresponsesizeallowable'] = challenge.maxresponsesizeallowable
        if challenge.maxresponsesizeallowable == "" or challenge.maxresponsesizeallowable == -1:
            challengesdict[statement]['maxresponsesizeallowable'] = mysettings.MAXRESPONSECHARCOUNT
            challenge.maxresponsesizeallowable = mysettings.MAXRESPONSECHARCOUNT
        if challenge.test.progenv == "multi":
            challengesdict[statement]['progenv'] = challenge.proglang
        alloptionslist = challengesdict[statement].keys()
        alloptionslist.sort(reverse = True)
        challengesdict[statement]['sortedoptions'] = alloptionslist
    testdict['challenges'] = challengesdict
    testdict['rulesdict'] = mysettings.RULES_DICT
    testdict['qualitydict'] = mysettings.SKILL_QUALITY
    testdict['answeringlanguagedict'] = mysettings.ANSWER_LANG_DICT
    jsonstr = json.dumps(testdict)
    # Now, encrypt jsonstr...
    #(encjsonstr, iv) = encryptstring(jsonstr)
    testdict['secret_key'] = base64.b64encode(mysettings.DES3_SECRET_KEY)
    #testdict['ivec'] = iv
    testdict['ivec'] = ''
    testdict['answering_languages'] = json.dumps(mysettings.ANSWER_LANG_DICT)
    testdict['rules'] = json.dumps(mysettings.RULES_DICT)
    testdict['skill_quality'] = mysettings.SKILL_QUALITY
    tmpl = get_template("tests/test_wrapper.html")
    testdict.update(csrf(request))
    cxt = Context(testdict)
    testhtml = tmpl.render(cxt)
    #testhtml = testhtml.replace("####ENCJSON####", base64.b64encode(encjsonstr))
    testhtml = testhtml.replace("####ENCJSON####", base64.b64encode(jsonstr))
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        testhtml = testhtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(testhtml)

#mobile_showtestcandidatemode = showtestcandidatemode


def encryptstring(mystr):
    iv = Random.get_random_bytes(8)
    des = DES3.new(mysettings.DES3_SECRET_KEY, DES3.MODE_CBC, iv)
    if mystr.__len__() < 16:
        remlen = 16 % mystr.__len__()
        mystr = mystr + ' '*(remlen - 16) # padding with whitespace
    else:
        remlen = mystr.__len__() % 16
        mystr = mystr + ' '*(16 - remlen) # padding with whitespace
    encryptedstr = des.encrypt(mystr)
    return (encryptedstr, iv)


"""
An alternate implementation of encryptstring()
"""
"""
def encryptstring(mystr):
    cipher = AES.new(mysettings.DES3_SECRET_KEY, AES.MODE_ECB)
    # mystr.__len__() has to be a multiple of 16
    mystrlen = mystr.__len__()
    remlen = mystrlen % 16
    mystr = mystr + ' '*(16 - remlen)
    encodedstr = base64.b64encode(cipher.encrypt(mystr))
    # decoded = cipher.decrypt(base64.b64decode(msg_text)) # This operation has to be implemented in javascript.
    return encodedstr
"""

"""
Function to generate a link using which an evaluator (identified by the parameter
'evalemailid') would be able to assess the responses provided by the candidate
(identified by the parameter 'useremail') for the test specified using the 
parameter 'testobj'. The specific test invitation against which the candidate 
supplied the responses is  specified using the parameters 'tabref' and 'tabid'.
"""
@csrf_exempt
def getevaluationlink(request, testid, evalemailid, useremail, tabref, tabid):
    evaluation_api_url = skillutils.gethosturl(request) + "/" + mysettings.TEST_EVALUATION_URL
    queryparams = "testid=" + testid.__str__() + "&evalemail=" + evalemailid + "&candidateemail=" + useremail + "&tabref=" + tabref + "&tabid=" + tabid
    encquerystr = base64.b64encode(queryparams)
    evaluation_api_url += "?" + encquerystr
    # Modify the URL to use 'https'
    protocol, other = evaluation_api_url.split("://")
    protocol = 'https'
    evaluation_api_url = "://".join([protocol, other])
    return evaluation_api_url


"""
Outdated Note: Method to communicate with server while user is taking test.
The type of response is indicated by the value of mode. 
0 - Response to a challenge, 1 - User ends the test by closing
the challenge window, 2 - Test is over, 3 - User applies
for a break (if break is allowed), 4 - User returns from a
break (if break is allowed), 5 - Evaluator requests the
candidate's response, 6 - Evaluator submits the evaluation,
7 - User starts taking the test.

Current Functionality: This method gets invoked when the user  submits the test
or the test ends due to the time limit running out. Now, in the case of a number
of people taking a test, there is a fat chance that this method will be called
at the same time by several takers. Hence, this method simply dumps the testpages
content in a flat json file. Along with that testpages variable, we also store the
starttime, testid, useremail, endtime, tabref and tabid in the same file. Later,
a batch script will read the files and send the data in them to the database.
The name of the text file in which these vars will be dumped will be along the 
following lines: answerscripts/<testid>/<tabref>_<tabid>.json.
Note: The testpages variable will contain base64 encoded content. Decoding will 
be done by the batch process when it is transferred to DB.
"""

@csrf_exempt
def gettestdata(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1071')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    starttest, starttime, testid, useremail, testlink, endtime, status, testpagesenc, testpages = 0, '', -1, '', '', '', 1, '', ''
    mode, challengeid, challengetype, challengestatement, oneormore, resptext = -1, -1, "", "", False, ""
    chkboxselectedoptions, radioselection, filbtext, useremail, tabref, tabid = [], "", "", "", "", -1
    if request.POST.has_key('starttest'):
        starttest = request.POST['starttest']
    if request.POST.has_key('starttime'):
        starttime = request.POST['starttime']
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    if request.POST.has_key('useremail'):
        useremail = base64.b64decode(request.POST['useremail'])
    if request.POST.has_key('endtime'):
        endtime = request.POST['endtime']
    if request.POST.has_key('status'):
        status = request.POST['status']
    if request.POST.has_key('tabref'):
        tabref = request.POST['tabref']
    if request.POST.has_key('tabid'):
        tabid = request.POST['tabid']
    if request.POST.has_key('mode'):
        mode = request.POST['mode']
    if request.POST.has_key('testpages'):
        testpagesenc = request.POST['testpages']
    ts = int(time.time())
    if tabref == "usertest":
        utobj = UserTest.objects.get(id=tabid)
        utobj.starttime = starttime
        utobj.status = starttest
    elif tabref == "wouldbeusers":
        utobj = WouldbeUsers.objects.get(id=tabid)
        utobj.starttime = starttime
        utobj.status = starttest
    utobj.save()
    ts_str = str(ts)
    answerscriptpath = mysettings.MEDIA_ROOT + os.path.sep + mysettings.ANSWER_SCRIPT_DUMP_PATH + os.path.sep + testid + "_" + ts_str
    if not os.path.exists(answerscriptpath) and testpagesenc != "":
        os.makedirs(answerscriptpath)
    answerscriptfile = answerscriptpath  + os.path.sep +  tabref + "_" + tabid + ".json"
    clientIP = skillutils.get_client_ip(request)
    clientua = request.META['HTTP_USER_AGENT']
    answerscript = { 'testid' : testid, 'starttime' : starttime, 'endtime' : endtime, 'useremail' : useremail, 'status' : status, 'tabref' : tabref, 'tabid' : tabid, 'mode' : mode, 'clientIP' : clientIP, 'useragent' : clientua, 'testpages' : testpagesenc }
    message = ""
    try:
        answerscriptdumped = json.dumps(answerscript)
        if not testpagesenc or testpagesenc == "":
            message = "Could not find any answer script with this request. We are not accepting requests without answer scripts"
            response = HttpResponse(message)
            return response
        fp = open(answerscriptfile, "w")
        fp.write(answerscriptdumped)
        fp.close()
        message = "Successfully stored answer script for evaluation later. Once the evaluator evaluates your answer script, you will be informed of the score and the outcome of the test through email at '%s'. Thank you for choosing TestYard as your test partner."%useremail
    except:
        message = "Failed to store the answer script due to the following reason: %s\n\nPlease contact the administrator or the support personnel at 'support@testyard.com'."%sys.exc_info()[1].__str__()
    response = HttpResponse(message)
    return response


"""
@csrf_protect
def gettestdata(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1071')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    starttest, starttime, testid, useremail, testlink, endtime, status, testpagesenc, testpages = 0, '', -1, '', '', '', 1, '', ''
    mode, challengeid, challengetype, challengestatement, oneormore, resptext = -1, -1, "", "", False, ""
    chkboxselectedoptions, radioselection, filbtext, useremail, tabref, tabid = [], "", "", "", "", -1
    if request.POST.has_key('starttest'):
        starttest = request.POST['starttest']
    if request.POST.has_key('starttime'):
        starttime = request.POST['starttime']
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    if request.POST.has_key('useremail'):
        useremail = base64.b64decode(request.POST['useremail'])
    if request.POST.has_key('endtime'):
        endtime = request.POST['endtime']
    if request.POST.has_key('status'):
        status = request.POST['status']
    if request.POST.has_key('tabref'):
        tabref = request.POST['tabref']
    if request.POST.has_key('tabid'):
        tabid = request.POST['tabid']
    if request.POST.has_key('mode'):
        mode = request.POST['mode']
    if request.POST.has_key('testpages'):
        testpagesenc = request.POST['testpages']
    testpages = base64.b64decode(testpagesenc)
    if int(mode) == 0: # This is a response to a challenge by the test taker.
        challengeid = request.POST['challengeid']
        challengetype = request.POST['challengetype']
        oneormore = request.POST['oneormore']
        challengestatement = request.POST['challengestatement']
        useremail = base64.b64decode(request.POST['useremail'])
        respkeypattern = re.compile(r"^id_\d+$")
        if challengetype  == 'CODN' or challengetype == 'SUBJ' or challengetype == 'ALGO':
            for k in request.POST.keys():
                if respkeypattern.search(k):
                    resptext = request.POST[k]
                    break
        elif challengetype == 'MULT':
            if oneormore == 'true': # checkboxes
                checkoptionpattern = re.compile(r"^option_\d+$")
                for k in request.POST.keys():
                    if checkoptionpattern.search(k):
                        chkboxselectedoptions.append(request.POST[k])
            else: # radio buttons
                if request.POST.has_key('rdoption'):
                    radioselection = request.POST['rdoption']
        elif challengetype == 'FILB':
            filbpattern = re.compile(r"^filb_\d+$")
            for k in request.POST.keys():
                if filbpattern.search(k):
                    filbtext = request.POST[k]
                    break
    
    clientsware = request.META['HTTP_USER_AGENT']
    ipaddress = request.META['REMOTE_ADDR']
    testobj = None
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "Error: %s"%error_msg('1072')
        response = HttpResponse(message)
        return response
    usertestqset = None
    if tabref == 'usertest' and tabid > 0:
        usertestqset = UserTest.objects.filter(id=int(tabid)).filter(cancelled=False, active=True)
    elif tabref == 'wouldbeusers' and tabid > 0:
        usertestqset = WouldbeUsers.objects.filter(id=int(tabid)).filter(cancelled=False, active=True)
    else: # Unrecognized table ref and/or improper table Id.
        message = "Error: %s"%error_msg('1079')
        response = HttpResponse(message)
        return response
    if not usertestqset or len(usertestqset) == 0: # No matching record found.
        message = "Error: %s"%error_msg('1073')
        response = HttpResponse(message)
        return response
    usertestobj = usertestqset[0]
    curdatetime = datetime.datetime.now()
    if curdatetime > skillutils.mysqltopythondatetime(usertestobj.validtill.__str__()):
        message = "Error: %s\n"%error_msg('1074')
        response = HttpResponse(message)
        return response
    if int(mode) == 1 or int(mode) == 2: # User has ended test by closing the window using the 'End Test' button, or the test duration has run out.
        usertestobj.status = status
        usertestobj.endtime = urllib.unquote(endtime)
        usertestobj.save()
    if int(mode) == 7:
        usertestobj.status = status
        usertestobj.clientsware = clientsware
        usertestobj.ipaddress = ipaddress
        usertestobj.starttime = urllib.unquote(starttime)
        usertestobj.save()
    if int(mode) == 0:
        userresponseobj = UserResponse()
        userresponseobj.test = testobj
        userresponseobj.challenge = Challenge.objects.filter(id=challengeid)[0]
        userresponseobj.emailaddr = useremail
        if challengetype == 'CODN' or challengetype == 'ALGO' or challengetype == 'SUBJ':
            userresponseobj.answer = smart_text(resptext, encoding='utf-8')
        elif challengetype == 'MULT':
            if oneormore == 'true': # checkboxes
                selectedoptionstring = "#||#".join(chkboxselectedoptions)
                userresponseobj.answer = smart_text(selectedoptionstring, encoding='utf-8')
            else:
                userresponseobj.answer = smart_text(urllib.unquote(radioselection), encoding='utf-8')
        elif challengetype == 'FILB':
            userresponseobj.answer = smart_text(filbtext, encoding='utf-8')
        userresponseobj.candidate_comment = ""
        userresponseobj.tabref = tabref
        userresponseobj.tabid = tabid
        responsedatetime = str(datetime.datetime.now())
        responsedatetimeparts = responsedatetime.split(' ')
        message = ""
        if responsedatetimeparts.__len__() == 1:
            responsedatetimeparts.append('00:00:00')
        if responsedatetimeparts.__len__() == 2:
            responsedate, responsetime = responsedatetimeparts[0], responsedatetimeparts[1]
            responsedateparts = responsedate.split("-")
            if responsedateparts.__len__() == 3:
                responsedate = responsedateparts[0] + "-" + responsedateparts[1] + "-" + responsedateparts[2]
            responsetime = str(responsetime)[:8]
            responsedate = responsedate + ' ' + responsetime
        userresponseobj.responsedatetime = responsedate
        userresponseobj.attachment = ''
        userresponseobj.save()
    if int(status) == 2: # Test completed. Send email to evaluator(s) with test name, test Id, email address, table reference and table Id.
        fromaddr = "testyardteam@testyard.com"
        retval = 0
        evalobj = testobj.evaluator
        evalemails = [] # Evaluator's and test creator's email address(es)
        evalemailsdict = {}
        if testobj.creatorisevaluator:
            evalemails.append(testobj.creator.emailid)
            evalemailsdict[testobj.creator.emailid] = 1
        groupmemberpattern = re.compile("groupmember\d+_id")
        for dk in evalobj.__dict__.keys():
            groupmemberpatmatch = groupmemberpattern.search(dk)
            if not groupmemberpatmatch:
                continue
            if evalobj.__dict__[dk] and int(evalobj.__dict__[dk]) > 0:
                try:
                    if not evalemailsdict.has_key(User.objects.get(id=evalobj.__dict__[dk]).emailid):
                        evalemailsdict[User.objects.get(id=evalobj.__dict__[dk]).emailid] = 1
                        evalemails.append(User.objects.get(id=evalobj.__dict__[dk]).emailid)
                    else:
                        pass
                except:
                    pass
        message = ""
        log = None
        try:
            log = Logger(mysettings.LOG_PATH)
        except:
            print "Could not create 'Logger' object - logging will not be possible.\n"
        #import pdb;pdb.set_trace()
        for evalemailid in evalemails:
            emailsent = False
            testevallink = getevaluationlink(request, testobj.id, evalemailid, useremail, tabref, tabid)
            emailsubject = "Test taken by user with email Id '%s'"%useremail
            emailmessage = "Dear Sir, \
             The  test named '%s' has been completed by user with email Id '%s'. You can start evaluating the \
             responses by clicking on the following link: '%s'. Do please let us know in case of any issues or\
             irregularities. \
             Thanks, \
             %s "%(testobj.testname, useremail, testevallink, fromaddr)
        try:
            retval = send_mail(emailsubject, emailmessage, fromaddr, [ evalemailid, ], False)
            message = "Successfully sent email to evaluator identified by '%s' for test identified by name '%s' and Id '%s' and candidate identified by email '%s'"%(evalemailid, testobj.testname, testid, useremail)
                if log is not None:
                log.logmessage(message)
                else:
                    print message
                emailsent = True # Just for maintaining uniformity
        except: # Log this error with all the variables used above in try block. This will be used by admins to manually send the email.
            message = "Could not send email to evaluator '%s' regarding completion of test identified by name '%s' and Id '%s' for user identified by emailaddress '%s' with record in table '%s' and table Id '%s'"%(evalemailid, testobj.testname, testid, useremail, tabref, tabid)
                if log is not None:
                log.logmessage(message)
                else:
                    print message
                emailsent = False
        # Also send an email to test creator informing about the exception.
        creatoremail = ""
            if not emailsent or emailsent == False:
            try:
                retval = send_mail("Could not send email to evaluator", message, fromaddr, [creatoremail,], False)
            except:
                pass # worthless email service!!! (cursing...).
    #return HttpResponse(message)
    message = "Success: Test in progress."
    return HttpResponse(message)
"""

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def sendtestinvitations(request):
    message = ''
    if request.method != "POST":
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    if not request.POST.has_key('testid') or not request.POST.has_key('txtemailslist'):
        message = "Error: " +error_msg('1069')
        response = HttpResponse(message)
        return response
    testid = request.POST['testid']
    baseurl = request.POST['baseurl']
    emailsliststr = request.POST['txtemailslist']
    forcefreshurl = 0
    if request.POST.has_key('forcefreshurl'):
        forcefreshurl = request.POST['forcefreshurl']
    if forcefreshurl == '':
        forcefreshurl = 0
    joingroupflag = request.POST.get('joingroupflag', None)
    emailsliststr = re.sub(re.compile(r"%20", re.MULTILINE|re.DOTALL), mysettings.HEXCODE_CHAR_MAP['%20'], emailsliststr) # replace for whitespace
    emailsliststr = re.sub(re.compile(r"%2C", re.MULTILINE|re.DOTALL), ",", emailsliststr) # replace for comma
    emailslist = emailsliststr.split(",")
    # We might have a few email Ids repeated, even though there is a
    # javascript check to avoid this. So we create a dict with the 
    # email Ids as keys so that we can access an unique list.
    emailsdict = {}
    for email in emailslist:
        emailsdict[email] = 1
    emailslist = emailsdict.keys() # So now emailslist contains unique emails
    if emailslist.__len__() > mysettings.MAX_INVITES_PER_SESSION:
        resp = HttpResponse("Error: More than %s distinct emails are not permitted per request."%mysettings.MAX_INVITES_PER_SESSION)
        return resp
    testobj = None
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "Error: " + error_msg('1056')
        response = HttpResponse(message)
        return response
    # Check if the user is the creator of this test. Only creators of a test 
    # are allowed to send invitations to candidates (except when an invitation 
    # needs to be sent automatically due to a user's need to join a group).
    if testobj.creator.id != userobj.id and not joingroupflag:
        message = "Error: " + error_msg('1070')
        response = HttpResponse(message)
        return response
    # Find out when this test was created. Then find the subscription plan that the user was subscribed to during that period, and also the number of candidates permitted by that plan.
    testcreateddate = testobj.createdate
    dbconn, dbcursor = connectdb()
    #dbconn, dbcursor = connectdb_p()
    candidatescount = 5 # Default candidates count for Free Plan.
    plnname = 'Free Plan'
    freeplansql = "select candidates from Subscription_plan where planname='%s'"
    dbcursor.execute(freeplansql, (plnname,))
    allplanrecs = dbcursor.fetchall()
    if allplanrecs.__len__() > 0:
        candidatescount = allplanrecs[0][0]
    else:
        pass
    userplanid = -1
    subsuserplansql = "select pl.planname, pl.candidates, up.planstartdate, up.planenddate, up.id from Subscription_plan pl, Subscription_userplan up where up.user_id=%s and pl.id=up.plan_id and up.planstartdate < %s and up.planenddate > %s"
    # Note that the plan in question may have expired, but the user still has the right to send invitations to candidates in order to use the test created during the period of the plan.
    dbcursor.execute(subsuserplansql, (userobj.id, testcreateddate, testcreateddate))
    allusrplanrecs = dbcursor.fetchall()
    planstartdate = None
    planenddate = None
    if allusrplanrecs.__len__() == 0: # If no userplan is found, we should consider it to be a free plan.
        pass
    else:
        planname = allusrplanrecs[0][0]
        candidatescount = int(allusrplanrecs[0][1])
        planstartdate = allusrplanrecs[0][2]
        planenddate = allusrplanrecs[0][3]
        userplanid = allusrplanrecs[0][4]
    rightnow = datetime.datetime.now()
    if userplanid > 0: # Check to see if the user has extended his plan at present.
        planextsql = "select allowedinvites, periodstart, periodend from Subscription_planextensions where userplan_id=%s and user_id=%s and periodstart <= %s and periodend > %s"
        dbcursor.execute(planextsql, (userplanid, userobj.id, rightnow, rightnow))
        allextensionrecs = dbcursor.fetchall()
        if allextensionrecs.__len__() > 0:
            candidatescount = allextensionrecs[0][0]
            planstartdate = allextensionrecs[0][1]
            planenddate = allextensionrecs[0][2]
    # Now, find how many (distinct) candidates have already been invited to this test. If that count is less than the allowed
    # number of candidates permitted by the subscription plan, then go ahead. Otherwise, return a response with an error message.
    reguserinvitations_sql = "select count(distinct emailaddr) from Tests_usertest where test_id=%s and dateadded > %s and dateadded < %s"
    dbcursor.execute(reguserinvitations_sql, (testobj.id, planstartdate, planenddate))
    allreguserinvitations = dbcursor.fetchall()
    unreguserinvitations_sql = "select count(distinct emailaddr) from Tests_wouldbeusers where test_id=%s and dateadded > %s and dateadded < %s"
    dbcursor.execute(reguserinvitations_sql, (testobj.id, planstartdate, planenddate))
    allunreguserinvitations = dbcursor.fetchall()
    totalcandidates = 0
    if allreguserinvitations.__len__() > 0:
        totalcandidates += allreguserinvitations[0][0]
    if allunreguserinvitations.__len__() > 0:
        totalcandidates += allunreguserinvitations[0][0]
    if totalcandidates >= candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
        message = "Error: You have exhausted the allocated quota of test invitations for this test."
        return HttpResponse(message)
    # Now, check if the number of invitations being sent would surpass the quota of allowed invitations for the selected plan.
    postoperationuserinvitations = totalcandidates + emailslist.__len__()
    if postoperationuserinvitations > candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
        decrementfactor = postoperationuserinvitations - candidatescount
        message = "Error: The number of users being invited will surpass the allowed number of invitations that can be sent under your subscription plan. Please remove %s email addresses and try again."%decrementfactor
        return HttpResponse(message)
    disconnectdb(dbconn, dbcursor)
    # Done checking invitations quota.
    validfrom = ""
    if request.POST.has_key('validfrom') and request.POST['validfrom'] != "":
        validfromdt = request.POST['validfrom']
    else:
        validfromdt = str(datetime.datetime.now())
    validfromdtprts = validfromdt.split(".")
    if validfromdtprts.__len__() >= 1:
        validfromdt = validfromdtprts[0]
    validfromdtobj = datetime.datetime.strptime(validfromdt, '%d-%m-%Y %H:%M:%S')
    validfromdatetimeparts = validfromdt.split(' ')
    message = ""
    if validfromdatetimeparts.__len__() == 1:
        validfromdatetimeparts.append('00:00:00')
    if validfromdatetimeparts.__len__() == 2:
        validfromdate, validfromtime = validfromdatetimeparts[0], validfromdatetimeparts[1]
        validfromdateparts = validfromdate.split("-")
        if validfromdateparts.__len__() == 3:
            validfromdate = validfromdateparts[2] + "-" + validfromdateparts[1] + "-" + validfromdateparts[0]
        else:
            message = "Error: It seems there is some problem with your valid from date. The format of the date should be 'dd/mm/yyyy hh:mm'. Please rectify it and try again."
            response = HttpResponse(message)
            return response
        validfrom = validfromdate + " " + validfromtime
    else:
        message = "Error: It seems there is some problem with your valid from date. The format of the date should be 'dd/mm/yyyy hh:mm'. Please rectify it and try again."
        response = HttpResponse(message)
        return response
    validtill = ""
    if request.POST.has_key('validtill') and request.POST['validtill'] != "":
        validtilldt = request.POST['validtill']
        validtilldtprts = validtilldt.split(".")
        if validtilldtprts.__len__() >= 1:
            validtilldt = validtilldtprts[0]
        validtilldtobj = datetime.datetime.strptime(validtilldt, '%d-%m-%Y %H:%M:%S')
        validtilldatetimeparts = validtilldt.split(' ')
        message = ""
        if validtilldatetimeparts.__len__() == 1:
            validtilldatetimeparts.append('00:00:00')
        if validtilldatetimeparts.__len__() == 2:
            validtilldate, validtilltime = validtilldatetimeparts[0], validtilldatetimeparts[1]
            validtilldateparts = validtilldate.split("-")
            if validtilldateparts.__len__() == 3:
                validtilldate = validtilldateparts[2] + "-" + validtilldateparts[1] + "-" + validtilldateparts[0]
            else:
                message = "Error: It seems there is some problem with your valid till date. The format of the date should be 'dd/mm/yyyy hh:mm'. Please rectify it and try again."
                response = HttpResponse(message)
                return response
            validtill = validtilldate + " " + validtilltime
        else:
            message = "Error: It seems there is some problem with your valid till date. The format of the date should be 'dd/mm/yyyy hh:mm'. Please rectify it and try again."
            response = HttpResponse(message)
            return response
    else:
        validtill = "onwards"
    currentdateobj = datetime.datetime.now()
    try:
        if currentdateobj > validfromdtobj:
            message = "Your 'Valid From' date is prior to the current date/time. Please set it to the current date/time or some time in future"
            response = HttpResponse(message)
            return response
        if currentdateobj > validtilldtobj:
            message = "Your 'Valid Till' date is prior to the current date/time. Please set it to the current date/time or some time in future"
            response = HttpResponse(message)
            return response
        if validfromdtobj > validtilldtobj:
            message = "Your 'Valid From' date is later than the 'Valid Till' date/time. 'Valid Till' date/time should be later than 'Valid From' date/time. Please rectify this mistake and try again."
            response = HttpResponse(message)
            return response
    except:
        message = sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    # Find the list of all evaluator's emails.
    testevaluator = testobj.evaluator
    testevalemailidlist = []
    if testevaluator.groupmember1:
        testevalemailidlist.append(testevaluator.groupmember1.emailid)
    if testevaluator.groupmember2:
        testevalemailidlist.append(testevaluator.groupmember2.emailid)
    if testevaluator.groupmember3:
        testevalemailidlist.append(testevaluator.groupmember3.emailid)
    if testevaluator.groupmember4:
        testevalemailidlist.append(testevaluator.groupmember4.emailid)
    if testevaluator.groupmember5:
        testevalemailidlist.append(testevaluator.groupmember5.emailid)
    if testevaluator.groupmember6:
        testevalemailidlist.append(testevaluator.groupmember6.emailid)
    if testevaluator.groupmember7:
        testevalemailidlist.append(testevaluator.groupmember7.emailid)
    if testevaluator.groupmember8:
        testevalemailidlist.append(testevaluator.groupmember8.emailid)
    if testevaluator.groupmember9:
        testevalemailidlist.append(testevaluator.groupmember9.emailid)
    if testevaluator.groupmember10:
        testevalemailidlist.append(testevaluator.groupmember10.emailid)
    # Now we need to start sending out the invitations. In the process, we
    # will be creating a 'bit.ly' link that will be unique for each 
    # candidate. In order to take the test, the candidate will need to click
    # on that link. Also, since there may be cases where a email Id of
    # a candidate is not in our records, we will store the bitly link, the
    # email Id, the test Id, date of test, etc in a table 'Tests_wouldbeuser'.
    # Once such a user accesses the test link, the system will request the
    # user to sign up in order to take the test. Once the user signs up, a
    # record pertaining to the user and test will be entered in Tests_usertest
    # and the record in 'wouldbeuser' will be deleted.
    error_emails_list = [] # This will contain all those email ids to which the test invitation email could not be sent.
    for email in emailslist:
        """ This email should not be the test creator's or one of the evaluator's emails."""
        if email == testobj.creator.emailid or email in testevalemailidlist:
            continue
        uobjqset = User.objects.filter(emailid=email)
        uobj = None
        candidatename = ""
        testlink = ""
        usertestobj = None
        wouldbeuserobj = None
        if uobjqset and uobjqset.__len__() > 0: # User exists
            uobj = uobjqset[0]
            # Check if we had sent a URL to this user for this test already.
            # If so, we will use the URL created that time. A new record
            # will be inserted in the appropriate table.
            checkexistsusrtestobj = UserTest.objects.filter(test = testobj).filter(emailaddr = uobj.emailid)
            if checkexistsusrtestobj and checkexistsusrtestobj.__len__() > 0 and forcefreshurl == 0:
                usertestobj = UserTest()
                testlink = checkexistsusrtestobj[0].testurl
                usertestobj.testurl = testlink
                usertestobj.user = uobj
                usertestobj.emailaddr = uobj.emailid # or = email
            else:
                usertestobj = UserTest()
                usertestobj.user = uobj
                usertestobj.emailaddr = uobj.emailid # or = email
                t = gettesturlforuser(usertestobj.emailaddr, testid, baseurl)
                #print(t)
                usertestobj.testurl, usertestobj.stringid = t[0], t[1]
                testlink = usertestobj.testurl
            usertestobj.test = testobj
            usertestobj.status = 0 # The test hasn't been taken as yet.
            usertestobj.validfrom = validfrom
            usertestobj.validtill = validtill
            if validtill == 'onwards':
                usertestobj.validtill = '2025-12-31 00:00:00'
            candidatename = uobj.displayname
        else: # user doesn't exist. So populate the wouldbeusers table.
            checkexistswldbeobj = WouldbeUsers.objects.filter(test = testobj).filter(emailaddr = email)
            # Check if we had sent a URL to this user for this test already.
            # If so, we will use the URL created that time.
            wouldbeuserobj = WouldbeUsers()
            if checkexistswldbeobj and checkexistswldbeobj.__len__() > 0 and forcefreshurl == 0:
                testlink = checkexistswldbeobj[0].testurl
                wouldbeuserobj.testurl = testlink
                wouldbeuserobj.emailaddr = email
            else:
                wouldbeuserobj.emailaddr = email
                (wouldbeuserobj.testurl, wouldbeuserobj.stringid) = gettesturlforuser(wouldbeuserobj.emailaddr, testid, baseurl)
                testlink = wouldbeuserobj.testurl
            wouldbeuserobj.test = testobj
            wouldbeuserobj.status = 0
            wouldbeuserobj.validfrom = validfrom
            wouldbeuserobj.validtill = validtill
            wouldbeuserobj.active = True
            wouldbeuserobj.cancelled = False
            if wouldbeuserobj.validtill == 'onwards':
                wouldbeuserobj.validtill = '2025-12-31 00:00:00'
            candidatename = "candidate"
        # Now send the email... save the above records only when email has been sent successfully
        emailsubject = "A test has been scheduled for you on testyard"
        emailmessage = """Dear %s,
        <br/><br/>
    A test with the name '%s' has been scheduled for you by %s. <br/><br/>
        """%(candidatename, testobj.testname, userobj.displayname)
        if validtill == 'onwards':
            emailmessage += """ This test will remain valid from %s %s, """%(validfrom, validtill)
        else:
            emailmessage += """ This test will remain valid from %s till %s, """%(validfrom, validtill)
        testyardhosturl = skillutils.gethosturl(request)
        emailmessage += """and hence you are kindly requested to take the test<br/>
        within that interval. You would be able to access the test by clicking
        on the following link: <a href='%s'>%s</a>. 
        <br/><br/>
        You may add this test to your <a href='%s/skillstest/test/addtocalendar/?turl=%s'>google calendar</a>.
	<br/><br/>
        If clicking on the above link doesn't work for you, please copy it and 
        paste it in your browser's address bar and hit enter. Do please feel <br/>
        free to let us know in case of any issues. We would do our best to
        resolve it at the earliest.
        <br/><br/>
        We wish you all the best for the test.
        <br/><br/>
        Regards,<br/>
        The TestYard Team.<br/><br/>
        """%(testlink, testlink, testyardhosturl, urllib.quote_plus(testlink))
        fromaddr = "testyardteam@testyard.in"
        retval = 0
        try:
            send_emails.delay(emailsubject, emailmessage, fromaddr, email, False)
            #retval = send_mail(emailsubject, emailmessage, fromaddr, [email,], False)
            if usertestobj:
                usertestobj.save()
            else:
                wouldbeuserobj.save()
        except:
            if mysettings.DEBUG:
                print "Error: sendemail failed for %s - %s\n"%(email, sys.exc_info()[1].__str__())
            message = "Error: sendemail failed for %s - %s\n"%(email, sys.exc_info()[1].__str__())
            error_emails_list.append(email)
            #response = HttpResponse(message)
            #return response
            continue # Continue processing the rest of the emails in the list.
    message = "Success! All candidates are being queued for email with the link."
    # Dump all emails Ids to which email could not be sent
    failure = False
    for error_email in error_emails_list:
        print error_email
        emailfail = EmailFailure()
        emailfail.user = userobj
        emailfail.sessionid = sesscode
        emailfail.failedemailid = error_email
        emailfail.script = 'Tests.views.sendtestinvitations'
        emailfail.failurereason = sys.exc_info()[1].__str__()
        emailfail.tryagain = 1
        try:
            emailfail.save()
            failure = True
        except:
            message = sys.exc_info()[1].__str__()
            print message
    if failure:
        message = "The emails could not be sent. Please contact " + mysettings.MAILSENDER + " with the test details and email addresses"
    response = HttpResponse(message)
    return(response)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def manageinvitations(request):
    message = ''
    if request.method != "POST":
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = None
    pageno = 1
    if not request.POST.has_key('testid'):
        message = "Error: %s."%(error_msg('1059'))
        response = HttpResponse(message)
        return response
    testid = request.POST['testid'] # Got the test. Now extract data and populate the manageinvitation.html template
    if request.POST.has_key('pageno'):
        try:
            pageno = int(request.POST['pageno'])
        except:
            pass
    csrftoken = ""
    if request.POST.has_key('csrfmiddlewaretoken'):
        csrftoken = request.POST['csrfmiddlewaretoken']
    testobj = None
    try:
        testobj = Test.objects.filter(id=testid)[0]
    except:
        message = "Error: %s"%(error_msg('1056'))
        response = HttpResponse(message)
        return responses
    startctr = pageno * mysettings.TEST_INVITATION_BY_TYPE_CHUNK_SIZE - mysettings.TEST_INVITATION_BY_TYPE_CHUNK_SIZE
    endctr = pageno * mysettings.TEST_INVITATION_BY_TYPE_CHUNK_SIZE
    usertestqset = UserTest.objects.filter(test=testobj).order_by("user", "status", "-validfrom", "-active")[startctr:endctr]
    wouldbeuserqset = WouldbeUsers.objects.filter(test=testobj).order_by("-validfrom", "-active")[startctr:endctr]
    if usertestqset.__len__() == 0 and wouldbeuserqset.__len__() == 0 and pageno == 1:
        message = "<p style='color:#aa0000;font-weight:bold;font-size:medium;'>Info: You do not have any invitations sent by you to any other user.<a href='#/' onClick=\"thediv=document.getElementById('existinginvitationdiv%s');thediv.style='display:none';thediv.innerHTML='';\" class='btn btn-testyard1'>Close</a><input type='hidden' name='testsinvitecounter' id='testsinvitecounter' value='0'></p>"%testid
        response = HttpResponse(message)
        return response
    elif pageno > 1 and usertestqset.__len__() == 0 and wouldbeuserqset.__len__() == 0:
        message = "<p style='color:#aa0000;font-weight:bold;font-size:medium;'>Info: You do not have any more invitations sent by you to other users.</p>"
        message += "<div id='pagination' style='text-align:center;padding-left:10px;'><nav aria-label='Test Records Page Navigation'><ul class='pagination'><li class='page-item'><a href='#/' onclick=\"javascript:existinginvitations(%s, '%s', %s);\" class='page-link' style='font-size:medium;'>< Prev</a></li><li class='page-item'><a href='#/' onClick=\"thediv=document.getElementById('existinginvitationdiv%s');thediv.style='display:none';thediv.innerHTML='';\" class='btn btn-testyard1'>Close</a><input type='hidden' name='testsinvitecounter' id='testsinvitecounter' value='0'></li></ul></nav></div>"%(testid, csrftoken, pageno-1, testid)
        response = HttpResponse(message)
        return response
    invitations_dict = {'usertest' : {}, 'wouldbeusers' : {}, 'testname' : testobj.testname, 'testid' : testid, 'testsinvitecounter' : 0 }
    usertest_dict = invitations_dict['usertest']
    wouldbeusers_dict = invitations_dict['wouldbeusers']
    usertest_dict['LENGTH'] = usertestqset.__len__()
    wouldbeusers_dict['LENGTH'] = wouldbeuserqset.__len__()
    ctr = 100
    testsinvitecounter = 0
    for usertest in usertestqset:
        status = usertest.status
        if usertest.status == 0:
            status = "Test not yet taken"
        elif usertest.status == 1:
            status = "Test is being taken"
        elif usertest.status == 2:
            status = "Test  has been taken"
        else:
            status = "Unknown status"
        outcome = usertest.outcome
        if not usertest.outcome:
            outcome = "NA"
        score = usertest.score
        if not usertest.score:
            score = "NA"
        starttime, endtime = skillutils.readabledatetime(usertest.starttime), skillutils.readabledatetime(usertest.endtime)
        if not usertest.starttime and  not usertest.endtime:
            starttime, endtime = "NA", "NA"
        ipaddress = usertest.ipaddress
        if not usertest.ipaddress:
            ipaddress = "NA"
        clientsware = usertest.clientsware
        if not usertest.clientsware:
            clientsware = "NA"
        try:
            #usertest_dict[str(ctr)] = [ skillutils.readabledatetime(str(usertest.validfrom)), skillutils.readabledatetime(str(usertest.validtill)), usertest.testurl, usertest.emailaddr, usertest.user.displayname, status, outcome, score, starttime, endtime, ipaddress, clientsware, usertest.sessid, usertest.active, usertest.cancelled, usertest.id ]
            usertest_dict[str(ctr)] = [ str(usertest.validfrom), str(usertest.validtill), usertest.testurl, usertest.emailaddr, usertest.user.displayname, status, outcome, score, starttime, endtime, ipaddress, clientsware, usertest.sessid, usertest.active, usertest.cancelled, usertest.id ]
        except:
            pass
            #fs = open("/home/supriyo/work/dddd7.txt", "w")
            #fs.write(sys.exc_info()[1].__str__());
            #fs.close()
        ctr += 1
        testsinvitecounter += 1
    invitations_dict['usertest'] = usertest_dict
    ctr = 100
    for wouldbeuser in wouldbeuserqset:
        wouldbeuserid = ctr
        status = wouldbeuser.status
        if wouldbeuser.status == 0:
            status = "Test not yet taken"
        elif wouldbeuser.status == 1:
            status = "Test is being taken"
        elif wouldbeuser.status == 2:
            status = "Test  has been taken"
        else:
            status = "Unknown status"
        outcome = wouldbeuser.outcome
        if not wouldbeuser.outcome:
            outcome = "NA"
        score = wouldbeuser.score
        if not wouldbeuser.score:
            score = "NA"
        starttime, endtime = skillutils.readabledatetime(wouldbeuser.starttime), skillutils.readabledatetime(wouldbeuser.endtime)
        if not wouldbeuser.starttime and  not wouldbeuser.endtime:
            starttime, endtime = "NA", "NA"
        ipaddress = wouldbeuser.ipaddress
        if not wouldbeuser.ipaddress:
            ipaddress = "NA"
        clientsware = wouldbeuser.clientsware
        if not wouldbeuser.clientsware:
            clientsware = "NA"
        #wouldbeusers_dict[str(wouldbeuserid)] = [ skillutils.readabledatetime(str(wouldbeuser.validfrom)), skillutils.readabledatetime(str(wouldbeuser.validtill)), wouldbeuser.testurl, wouldbeuser.emailaddr, status, outcome, score, starttime, endtime, ipaddress, clientsware, wouldbeuser.active, wouldbeuser.cancelled, wouldbeuser.id ]
        wouldbeusers_dict[str(wouldbeuserid)] = [ str(wouldbeuser.validfrom), str(wouldbeuser.validtill), wouldbeuser.testurl, wouldbeuser.emailaddr, status, outcome, score, starttime, endtime, ipaddress, clientsware, wouldbeuser.active, wouldbeuser.cancelled, wouldbeuser.id ]
        ctr += 1
        testsinvitecounter += 1
    invitations_dict['wouldbeusers'] = wouldbeusers_dict
    invitations_dict['testsinvitecounter'] = testsinvitecounter
    nextpage = pageno + 1
    prevpage = pageno - 1
    invitations_dict['nextpage'] = nextpage
    invitations_dict['prevpage'] = prevpage
    invitations_dict['testid'] = testid
    tmpl = get_template("tests/manageinvitations.html")
    invitations_dict.update(csrf(request))
    cxt = Context(invitations_dict)
    invitationshtml = tmpl.render(cxt)
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        invitationshtml = invitationshtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(invitationshtml)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def invitationactivation(request):
    message = ''
    if request.method != "POST":
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    invitationurl, tablename, action = "", "", 1
    if request.POST.has_key('inviteid'):
        invitationurl = request.POST['inviteid']
    if request.POST.has_key('table'):
        tablename = request.POST['table']
    if request.POST.has_key('action'):
        action = request.POST['action']
    tabid = None
    if request.POST.has_key('tabid'):
        tabid = request.POST['tabid']
    inviteobj = None
    message = ''
    try:
        if tablename == 'usertest':
            if not tabid:
                inviteobj =UserTest.objects.filter(testurl=invitationurl)[0]
            else:
                inviteqset =UserTest.objects.filter(testurl=invitationurl).filter(id=tabid)
                if inviteqset.__len__() > 0:
                    inviteobj = inviteqset[0]
                else:
                    inviteobj = None
                    message = 'Could not create any invitation object (possibly due to Id mismatch - %s)'%invitationid
                    response = HttpResponse(message)
                    return response
        elif tablename == 'wouldbeusers':
            if not tabid:
                inviteobj = WouldbeUsers.objects.filter(testurl=invitationurl)[0]
            else:
                inviteqset = WouldbeUsers.objects.filter(testurl=invitationurl).filter(id=tabid)
                if inviteqset.__len__() > 0:
                    inviteobj = inviteqset[0]
                else:
                    inviteobj = None
                    message = 'Could not create any invitation object (possibly due to Id mismatch - %s)'%invitationid
                    response = HttpResponse(message)
                    return response
        else:
            pass
    except:
        message = 'Could not create any invitation object (possibly due to Id mismatch - %s)'%invitationid
        response = HttpResponse(message)
        return response
    if int(action) == 1: # Set active to True
        inviteobj.active = True
        inviteobj.save()
        message = "Invitation URL '%s' has been activated as per your request. Please close and reopen the invitations view to observe the changes."%inviteobj.testurl
    else:
        inviteobj.active = False
        inviteobj.save()
        message = "Invitation URL '%s' has been deactivated as per your request. Please close and reopen the invitations view to observe the changes."%inviteobj.testurl
    response = HttpResponse(message)
    return response
    

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def invitationcancellation(request):
    message = ''
    if request.method != "POST":
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    invitationurl, tablename, tabid = "", "",""
    if request.POST.has_key('inviteurl'):
        invitationurl = request.POST['inviteurl']
    if request.POST.has_key('table'):
        tablename = request.POST['table']
    if request.POST.has_key('tabid'):
        tabid = request.POST['tabid']
    inviteobj = None
    message = ''
    try:
        if tablename == 'usertest':
            inviteobj = UserTest.objects.filter(testurl=invitationurl, id=tabid)[0]
        elif tablename == 'wouldbeusers':
            inviteobj =WouldbeUsers.objects.filter(testurl=invitationurl, id=tabid)[0]
        else:
            pass
    except:
        message = 'Could not create any invitation object (possibly due to Id mismatch - %s)'%invitationid
        response = HttpResponse(message)
        return response
    inviteobj.cancelled = True
    inviteobj.save()
    message = "Invitation URL '%s' has been cancelled as per your request. Please close and reopen the invitations view to observe the changes."%inviteobj.testurl
    response = HttpResponse(message)
    return response


"""
Function to enable a evaluator of the test to evaluate the candidate's responses.
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def evaluate(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    pageno = 1
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = None
    candidateresponses = {}
    if not request.POST.has_key('testid'):
        candidateresponses['Error'] = error_msg('1059')
        response = HttpResponse(json.dumps(candidateresponses))
        return response
    if 'pageno' in request.POST.keys():
        try:
            pageno = int(request.POST['pageno'])
        except:
            pass
    startctr = mysettings.PAGE_CHUNK_SIZE * pageno - mysettings.PAGE_CHUNK_SIZE
    endctr = mysettings.PAGE_CHUNK_SIZE * pageno
    testid = request.POST['testid']
    # First ensure that the user is a valid evaluator for the given test
    testobj = None
    evalobj = None
    try:
        testobj = Test.objects.filter(id=testid)[0]
        evalobj = testobj.evaluator
    except:
        print "Could not retrieve test object or evaluator object for given test Id %s\n"%testid
        candidateresponses['Error'] = error_msg('1080')
        response = HttpResponse(json.dumps(candidateresponses))
        return response
    userisvalidevaluator = False
    useremail = userobj.emailid
    evaluatoremails = []
    if evalobj.groupmember1:
        evaluatoremails.append(evalobj.groupmember1.emailid)
    if evalobj.groupmember2:
        evaluatoremails.append(evalobj.groupmember2.emailid)
    if evalobj.groupmember3:
        evaluatoremails.append(evalobj.groupmember3.emailid)
    if evalobj.groupmember4:
        evaluatoremails.append(evalobj.groupmember4.emailid)
    if evalobj.groupmember5:
        evaluatoremails.append(evalobj.groupmember5.emailid)
    if evalobj.groupmember6:
        evaluatoremails.append(evalobj.groupmember6.emailid)
    if evalobj.groupmember7:
        evaluatoremails.append(evalobj.groupmember7.emailid)
    if evalobj.groupmember8:
        evaluatoremails.append(evalobj.groupmember8.emailid)
    if evalobj.groupmember9:
        evaluatoremails.append(evalobj.groupmember9.emailid)
    if evalobj.groupmember10:
        evaluatoremails.append(evalobj.groupmember10.emailid)
    if testobj.creatorisevaluator:
        evaluatoremails.append(testobj.creator.emailid)
    if useremail not in evaluatoremails:
        candidateresponses['Error'] = error_msg('1081')
        response = HttpResponse(json.dumps(candidateresponses))
        return response
    else:
        userisvalidevaluator = True
    if userisvalidevaluator: # Get a list of all candidates who have taken this test.
        utqset = UserTest.objects.filter(test=testobj)[startctr:endctr]
        wbuqset = WouldbeUsers.objects.filter(test=testobj)[startctr:endctr]
        #fp = open("/home/supriyo/work/testyard/tmpfiles/answes.txt","w+")
        for ut in utqset:
            if ut.active and not ut.cancelled and ut.status == 2:
                candidaterec = {'emailaddr' : ut.emailaddr, 'starttime' : str(ut.starttime), 'endtime' : str(ut.endtime), 'outcome' : ut.outcome, 'status' : ut.status, 'score' : ut.score, 'stringid' : ut.stringid, 'testurl' : ut.testurl, 'testid' : testid, 'testname' : testobj.testname, 'tabref' : 'usertest', 'tabid' : ut.id, 'candidateresponse' : {}, 'evaltestcomment' : ut.evaluator_comment, 'evalcommitstate' : ut.evalcommitstate, 'disqualified' : ut.disqualified, 'windowchangeattempts' : ut.windowchangeattempts}
                userresputqueryset = UserResponse.objects.filter(test=testobj, tabref='usertest', tabid=ut.id, emailaddr=ut.emailaddr)
                for userrespobj in userresputqueryset:
                    # Change for display of english statements in mathjax code starts here.
                    slashbracketpattern = re.compile("^.*\\]?([^\]\[]+)\\[.*$", re.DOTALL)
                    bracketpattern = re.compile("^.*\]([^\[\]]+)\[.*$", re.DOTALL)
                    
                    slashbracketpatternmatchobj = re.search(slashbracketpattern, userrespobj.answer)
                    bracketpatternmatchobj = re.search(bracketpattern, userrespobj.answer)
                    userrespobj.answer = userrespobj.answer.replace("'", "\'")
                    if slashbracketpatternmatchobj:
                        slashbracketgroups = slashbracketpatternmatchobj.groups()
                        ans = userrespobj.answer
                        for grp in slashbracketgroups:
                            userrespobj.answer = re.sub(slashbracketpattern, "\\]textrm{%s}\\["%str(grp), userrespobj.answer)
                            #fp.write(str(grp) + " ##################### " + userrespobj.answer)
                            textrmpattern = re.compile("^.*[t]?[extrm]{1}.*$", re.DOTALL)
                            textrmpatternobj = re.search(textrmpattern,userrespobj.answer)
                            if textrmpatternobj:
                                userrespobj.answer = ans # Nothing changes.
                    elif bracketpatternmatchobj:
                        bracketgroups = bracketpatternmatchobj.groups()
                        for grp in bracketgroups:
                            userrespobj.answer = re.sub(bracketpattern, "\]\textrm{%s}\["%str(grp), userrespobj.answer)
                    else:
                        userrespobj.answer = userrespobj.answer # Nothing changes. # We have not implemented with '$' delimiter.
                    
                    # Change for display of english statements in mathjax code ends here.
                    candidaterec['candidateresponse'][userrespobj.challenge.statement] = {'answer' : userrespobj.answer, 'responsedatetime' : skillutils.pythontomysqldatetime2(str(userrespobj.responsedatetime)), 'maxscore' : userrespobj.challenge.challengescore, 'negativescore' : userrespobj.challenge.negativescore, 'correctanswer' : userrespobj.challenge.responsekey, 'challengeid' : userrespobj.challenge.id, 'evaluation' : userrespobj.evaluation, 'evaluatorremarks' : userrespobj.evaluator_remarks, 'mathenv' :  userrespobj.challenge.mathenv}
                candidateresponses[ut.emailaddr + "####" + str(ut.id)] = candidaterec
        #fp.close()
        for wbu in wbuqset:
            if wbu.active and not wbu.cancelled and wbu.status == 2:
                candidaterec = {'emailaddr' : wbu.emailaddr, 'starttime' : str(wbu.starttime), 'endtime' : str(wbu.endtime), 'outcome' : wbu.outcome, 'status' : wbu.status, 'score' : wbu.score, 'stringid' : wbu.stringid, 'testurl' : wbu.testurl, 'testid' : testid, 'testname' : testobj.testname, 'tabref' : 'wouldbeusers', 'tabid' : wbu.id, 'candidateresponse' : {}, 'evaltestcomment' : wbu.evaluator_comment, 'evalcommitstate' : wbu.evalcommitstate, 'disqualified' : wbu.disqualified, 'windowchangeattempts' : wbu.windowchangeattempts}
                userrespwbequeryset = UserResponse.objects.filter(test=testobj, tabref='wouldbeusers', tabid=wbu.id, emailaddr=wbu.emailaddr)
                for userrespobj in userrespwbequeryset:
                    candidaterec['candidateresponse'][userrespobj.challenge.statement] = {'answer' : userrespobj.answer, 'responsedatetime' : skillutils.pythontomysqldatetime2(str(userrespobj.responsedatetime)), 'maxscore' : userrespobj.challenge.challengescore, 'negativescore' : userrespobj.challenge.negativescore, 'correctanswer' : userrespobj.challenge.responsekey, 'challengeid' : userrespobj.challenge.id, 'evaluation' : userrespobj.evaluation, 'evaluatorremarks' : userrespobj.evaluator_remarks, 'windowchangeattempts' : wbu.windowchangeattempts, 'mathenv' : userrespobj.challenge.mathenv }
                candidateresponses[wbu.emailaddr + "####" + str(wbu.id)] = candidaterec
    return HttpResponse(base64.b64encode(json.dumps(candidateresponses)))


##################################################################
## New function to display evaluation screen. Get the values 
## needed for "evaluate_responses.html". Serve that template
## with the required values. Also, rename "evaluateresponses_new"
## to "evaluateresponses" in the tests.html file. Rename the
## current "evaluateresponses" to "evaluateresponses_old".
##################################################################
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def showevaluationscr(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = None
    emailid = None
    utid = None
    candidateresponse  = {}
    if not request.POST.has_key('testid') or not request.POST.has_key('emailid') or not request.POST.has_key('utid'):
        evalscreenresponse['Error'] = error_msg('1059')
        response = HttpResponse(json.dumps(evalscreenresponse))
        return response
    testid = request.POST['testid']
    emailid = request.POST['emailid']
    utid = request.POST['utid'] # This would be the id from the 'usertest' table. The individual who is evaluating the test has to be a member of TestYard(tm).
    evalcommitstate = None
    candidateresponse['cctr']  = 1
    candidateresponse['candidateresp']  = ''
    candidateresponse['testid']  = testid
    candidateresponse['emailid'] = emailid
    candidateresponse['utid'] = utid
    candidateresponse['challengecounter'] = 0 # This is basically the number of challenges contained in the test.
    candidateresponse['tabref'] = None
    candidateresponse['tabid'] = utid
    candidateresponse['evalcommitstate'] = None
    candidateresponse['evaltestcomment'] = None
    testobj = Test.objects.get(id=testid)
    usrrespqset = UserResponse.objects.filter(test=testobj, emailaddr=emailid, tabid=utid)
    #candidateresponse['usrrespqset'] = list(usrrespqset).__len__() # Just to test...
    tabref = usrrespqset[0].tabref
    tabid = utid
    candidateresponse['tabid'] = tabid
    candidateresponse['tabref'] = tabref
    utobj = None
    if tabref == 'usertest':
        utobj = UserTest.objects.get(id=utid)
    elif tabref == 'wouldbeusers':
        utobj = WouldbeUsers.objects.get(id=utid)
    if utobj:
        candidateresponse['evaltestcomment'] = utobj.evaluator_comment.replace("%20", " ")
    usrtestobj = None
    if tabref == "usertest":
        usrtestobj = UserTest.objects.get(emailaddr=emailid, test=testobj, id=utid)
    elif tabref == "wouldbeusers":
        usrtestobj = WouldbeUsers.objects.get(emailaddr=emailid, test=testobj, id=utid)
    else: # Unsupported table, so something is fishy. Return a response with an appropriate message
        return HttpResponse("Invalid table name selected.")
    evalcommitstate = usrtestobj.evalcommitstate
    candidateresponse['evalcommitstate'] = evalcommitstate # Should be 0 or 1
    # The above variables would be the same for all challenges related to this test.
    #fp = open("/home/supriyo/work/testyard/extralogs/dumplog.log", "w")
    candidateresponse['challenge_statement'] = {}
    for usrrespobj in usrrespqset:
        candidateresponse['challengecounter'] += 1
        try:
            # Populate the dictionary 'candidateresponse'
            challenge = usrrespobj.challenge
            challenge_statement = challenge.statement
            candidateresponse['challenge_statement'][challenge_statement] = {}
            candidateresponse['challenge_statement'][challenge_statement]['maxscore'] = challenge.challengescore
            candidateresponse['challenge_statement'][challenge_statement]['challengeid'] = challenge.id
            candidateresponse['challenge_statement'][challenge_statement]['evaluatorremarks'] = usrrespobj.evaluator_remarks
            candidateresponse['challenge_statement'][challenge_statement]['evaluation'] = usrrespobj.evaluation
            candidateresponse['challenge_statement'][challenge_statement]['negativescore'] = challenge.negativescore
            candidateresponse['challenge_statement'][challenge_statement]['mathenv'] = challenge.mathenv
            correctanswer = challenge.responsekey
            #fp.write(str(challenge_statement) + " #### " + str(challenge.challengescore) + " #### " + str(challenge.id) + " #### " + str(usrrespobj.evaluator_remarks) + " #### " + str(usrrespobj.evaluation) + " #### " + str(challenge.negativescore) + " #### " + candidateresponse['challengecounter'] + "\n\n")
            correctanswerlist = []
            try:
                correctanswerlist = correctanswer.split("#||#") # This is the pattern separating the answers in multiple choice type questions.
            except:
                pass               
            if correctanswerlist.__len__() > 0:
                pass # Need to implement this
            candidateresponse['challenge_statement'][challenge_statement]['correctanswer'] = challenge.responsekey
            """
            # Change for display of english statements in mathjax code starts here.
            slashbracketpattern = re.compile("^.*\\]?([^\]\[]+)\\[.*$", re.DOTALL)
            bracketpattern = re.compile("^.*\]([^\[\]]+)\[.*$", re.DOTALL)
            slashbracketpatternmatchobj = re.search(slashbracketpattern)
            bracketpatternmatchobj = re.search(bracketpattern)
            if slashbracketpatternmatchobj:
                slashbracketgroups = slashbracketpatternmatchobj.groups()
                for grp in slashbracketgroups:
                    usrrespobj.answer = re.sub(slashbracketpattern, "\textrm{%s}"%str(grp))
            elif bracketpatternmatchobj:
                bracketgroups = bracketpatternmatchobj.groups()
                for grp in bracketgroups:
                    usrrespobj.answer = re.sub(bracketpattern, "\textrm{%s}"%str(grp))
            # Change for display of english statements in mathjax code ends here.
            """
            candidateresponse['challenge_statement'][challenge_statement]['answer'] = usrrespobj.answer
        except:
            continue
    results_dict = {}
    results_dict['candidateresponse'] = candidateresponse
    tmpl = get_template("tests/evaluate_responses.html")
    results_dict.update(csrf(request))
    try:
        cxt = Context(results_dict)
        candidateresponsehtml = tmpl.render(cxt)
    except:
        return HttpResponse(sys.exc_info()[1].__str__())
        #fp.write(sys.exc_info()[1].__str__())
    #fp.close()
    for htmlkey in mysettings.HTML_ENTITIES_CHAR_MAP.keys():
        candidateresponsehtml = candidateresponsehtml.replace(htmlkey, mysettings.HTML_ENTITIES_CHAR_MAP[htmlkey])
    return HttpResponse(candidateresponsehtml)
    


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def evaluateresponses(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = None
    emailid, tabid, tabref, evaltestcomment = "", -1, "", ""
    maxcctr = -1
    postdataencpluscsrf = request.body
    postdataenc, csrfparam = postdataencpluscsrf.split("&")
    postdata = base64.b64decode(postdataenc)
    postdict = {}
    postkeyvals = postdata.split("&")
    for keyval in postkeyvals:
        key, val = keyval.split("=")
        postdict[key] = val
    if not postdict.has_key('testid'):
        message = error_msg('1059')
        response = HttpResponse(message)
        return response
    testid = postdict['testid']
    # First ensure that the user is a valid evaluator for the given test
    testobj = None
    evalobj = None
    passscore = None
    evalcommitstate = '0'
    if postdict.has_key('maxcctr'):
        maxcctr = postdict['maxcctr']
    if postdict.has_key('testid'):
        testid = postdict['testid']
    if postdict.has_key('emailid'):
        emailid = postdict['emailid']
    if postdict.has_key('tabid'):
        tabid = postdict['tabid']
    if postdict.has_key('tabref'):
        tabref = postdict['tabref']
    if postdict.has_key('evaltestcomment'):
        evaltestcomment = postdict['evaltestcomment']
        if evaltestcomment.__len__() > 200:
            message = "The comment for the entire evaluation cannot be stored as it is more than 200 characters long."
            response = HttpResponse(message)
            return response
    if postdict.has_key('evalcommitstate'):
        evalcommitstate = postdict['evalcommitstate']
    if testid:
        testobj = Test.objects.get(id=testid)
        passscore = testobj.passscore
    # Check if this test may be evaluated again - if 10 days have passed since it was first evaluated, then it may not be evaluated anymore.
    utobj = None
    evallimit = 86400 * int(mysettings.NUM_DAYS_EVALUATION_COMMIT)
    if tabref == 'wouldbeusers':
        utobj = WouldbeUsers.objects.get(id=tabid)
        if not utobj.first_eval_timestamp:
            utobj.first_eval_timestamp = int(time.time())
        else:
            firsttimestamp = utobj.first_eval_timestamp
            currenttimestamp = int(time.time())
            if currenttimestamp - firsttimestamp > evallimit:
                message = error_msg('1083')
                response = HttpResponse(message)
                utobj.evalcommitstate = True
                utobj.save()
                return response
    elif tabref == 'usertest':
        utobj = UserTest.objects.get(id=tabid)
        if not utobj.first_eval_timestamp:
            utobj.first_eval_timestamp = int(time.time())
        else:
            firsttimestamp = utobj.first_eval_timestamp
            currenttimestamp = int(time.time())
            if currenttimestamp - firsttimestamp > evallimit:
                message = error_msg('1083')
                utobj.evalcommitstate = True
                utobj.save()
                response = HttpResponse(message)
                return response
    else:
        message = "Error: " + error_msg('1079')
        response = HttpResponse(message)
        return response
    totalscore = 0
    commitable = True
    for cctr in range(1, int(maxcctr) + 1):
        challengeid, assessment, comments, maxscore = -1, 0.0, "", 0
        if postdict.has_key('challengeid_' + cctr.__str__()):
            challengeid = postdict['challengeid_' + cctr.__str__()]
        else:
            continue # no use retrieving other elements with same id as we won't be able to store those values against a challenge.
        if postdict.has_key('assessment_' + cctr.__str__()):
            assessment = postdict['assessment_' + cctr.__str__()]
            try:
                assessment = float(assessment)
            except:
                if assessment != '': # In case of '' value, we will not commit the evaluation later.
                    message = "One of the assessment values is not numeric. Please rectify it and try again."
                    response = HttpResponse(message)
                    return response
            if not testobj.negativescoreallowed and assessment == -1.0:
                commitable = False
            elif testobj.negativescoreallowed == True and assessment == '':
                commitable = False
            totalscore = totalscore + float(assessment)
        if postdict.has_key('maxscore_' + cctr.__str__()):
            maxscore = postdict['maxscore_' + cctr.__str__()]
        if postdict.has_key('comments_' + cctr.__str__()):
            comments = postdict['comments_' + cctr.__str__()]
            if comments.__len__() > 200:
                message += "Warning: Comment # %s is greater than 200 characters long. Only part of the comment will be saved."%cctr
                comments = comments[:200]
        if float(assessment) > float(maxscore):
            message += "Error: assessment greater than maxscore in challenge with Id %s<br />"%challengeid
            continue
        challengeobj = Challenge.objects.filter(id=challengeid)
        userrespqset = UserResponse.objects.filter(test=testobj, challenge=challengeobj, emailaddr=emailid, tabref=tabref, tabid=tabid)
        if userrespqset.__len__() == 0:
            message += "Error: Could not find the challenge object to update <br />"
            continue
        userrespobj = userrespqset[0]
        userrespobj.evaluation = assessment
        userrespobj.evaluator_remarks = comments
        userrespobj.save() # User response updated with evaluation and evaluator remarks and saved.
    utobj.score = totalscore
    utobj.evaluator_comment = evaltestcomment
    if passscore and passscore <= totalscore:
        utobj.outcome = True
    elif not passscore and totalscore > 0:
        utobj.outcome = None
    elif passscore and passscore > totalscore:
        utobj.outcome = False
    else:
        pass
    if evalcommitstate == '1' and commitable:
        utobj.evalcommitstate = True
    else:
        utobj.evalcommitstate = False
        message += "The evaluation could not be commited as not all responses have been assessed. To commit the evaluation, you need to assess all the responses. "
    utobj.save()
    if utobj.evalcommitstate is True: # Send email to the candidate with the obtained score for the test.
        subject = "Your test '%s' has been evaluated"%utobj.test.testname
        fromaddr = utobj.test.creator.emailid
        email = utobj.emailaddr
        passscore = utobj.test.passscore
        #outcome = "<font color='#0000AA'>Pass</font>"
        outcome = "Pass"
        if utobj.score < passscore:
            #outcome = "<font color='#AA0000'>Fail</font>"
            outcome = "Fail"
        message = """
            Dear Candidate,
            

            Your answer script for the test '%s' has been evaluated and the results are as follows:

            Score: %s/%s (%s out of %s)
            Outcome: %s

            Pass Score: %s

            
            To be able to refer to the score permanently, we suggest you create an account on testyard. If
            you already have one, then you would be able to refer to the test's outcome in the 'Tests' tab.
            


            Thank you for using TestYard as a test partner.
            

            Regards,

            TestYard Test Facilitation Team.
        """%(utobj.test.testname, utobj.score, utobj.test.maxscore, utobj.score, utobj.test.maxscore, outcome, passscore)
        try:
            #retval = send_mail(subject, message, fromaddr, [email,], False)
            retval = send_emails.delay(subject, message, fromaddr, email, False)
        except:
            message = error_msg('1160')
            #print message

    message += "Handled %s answers for user with email address '%s'"%(maxcctr.__str__(), emailid)
    return HttpResponse(message)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def getevaluationdetails(request):
    return evaluate(request)

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def createtestbulkupload(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testlinkid, exist_test_id, upload1, upload2, upload3, upload4, upload5, upload6, upload7, upload8, upload9, upload10 = [ "" for i in range(12)]
    uploaddict = {}
    if request.POST.has_key('testlinkid'):
        testlinkid = request.POST['testlinkid']
    if request.POST.has_key('exist_test_id'):
        exist_test_id = request.POST['exist_test_id']
    username = userobj.displayname
    uploadPattern = re.compile(r"upload\d+")
    for postparam in request.FILES.keys():
        uploadfilename = None
        #print request.FILES[postparam].name
        fsz = request.FILES[postparam].size
        if fsz > mysettings.MAX_FILE_SIZE_ALLOWED:
            message += "Error: Couldn't upload file %s as it is larger than %s"%(request.FILES[postparam].name, str(mysettings.MAX_FILE_SIZE_ALLOWED/1000000) + " MB")
            continue
        if request.FILES.has_key(postparam) and request.FILES[postparam].name != "" and uploadPattern.search(postparam):
            uploadfilename = request.FILES[postparam].name.split(".")[0]
        else:
            continue
        uploadedext = skillutils.get_extension2(request.FILES[postparam].name)
        if uploadedext.lower() != 'csv' and uploadedext.lower() != 'xls' and uploadedext.lower() != 'xlsx' and uploadedext.lower() != 'xml':
            message += "Error: Invalid test file format for POST parameter '%s'. Please upload 'csv' or 'xls(x)' or 'xml' files only."%postparam
            continue
        fpath, message, testmedia = skillutils.handleuploadedfile2(request.FILES[postparam], mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "bulkupload", uploadfilename)
        uploadedfile = request.FILES[postparam].name
        uploaddict[uploadedfile] = [fpath, uploadedext]
    # At this point, all files have been uploaded. So we are ready to start creating the tests and their challenges.
    # Note: We will delete the entire 'bulkupload' folder (along with its contents) once all tests from those files have been created.
    testcount = 0
    for filename in uploaddict.keys():
        filepath = uploaddict[filename][0]
        fileext = uploaddict[filename][1] # This can be bogus if extension is greater than 3 chars long, e.g xlsx.
        if fileext == "xlsx":
            testcount += create_test_from_xlsx(filename, filepath, testlinkid, userobj)
        elif fileext == "xls":
            testcount += create_test_from_xls(filename, filepath, testlinkid, userobj)
        elif fileext == "csv":
            testcount += create_test_from_csv(filename, filepath, testlinkid, userobj)
        elif fileext == "xml":
            testcount += create_test_from_xml(filename, filepath, testlinkid, userobj)
        else:
            print "Unsupported format %s\n", fileext
            continue
    try:
        allfiles = os.listdir(mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "bulkupload")
        for filename in allfiles:
            os.unlink(mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "bulkupload" + os.path.sep + filename)
        os.rmdir(mysettings.MEDIA_ROOT + os.path.sep + username + os.path.sep + "bulkupload")
    except:
        print "Error removing bulkupload dir: %s\n"%sys.exc_info()[1].__str__()
    return HttpResponse("%s tests created"%testcount.__str__())


def _check_subscription_plan_limits(userobj, filepath, filetype="csv"):
    todaynow = datetime.datetime.now()
    # Check the allowed number of tests for this user - only in case of new tests
    dbconn, dbcursor = connectdb()
    #dbconn, dbcursor = connectdb_p()
    userplansql = "select pl.testsninterviews, up.plan_id, pl.planname, up.amountdue, up.planstartdate, up.planenddate from Subscription_userplan up, Subscription_plan pl where up.planstartdate < %s and up.planenddate > %s and up.planstatus=TRUE and up.plan_id=pl.id and up.user_id=%s"
    dbcursor.execute(userplansql, (todaynow, todaynow, userobj.id))
    allrecords = dbcursor.fetchall()
    if allrecords.__len__() > 0:
        testsandinterviewsquota = allrecords[0][0]
        planid = allrecords[0][1]
        planname = allrecords[0][2]
        amtdue = allrecords[0][3]
        planstartdate = allrecords[0][4]
        if planstartdate is None or planstartdate == "":
            message = "Error: Plan start date is missing for user '%s'. Please contact support with this error message."%userobj.displayname
            return (False, message)
        planenddate = allrecords[0][5]
        if planenddate is None or planenddate == "":
            message = "Error: Plan end date is missing for user '%s'. Please contact support with this error message."%userobj.displayname
            return (False, message)
        if amtdue > 0.00: # User has not paid the total amount for this plan.
            message = "Error: You have an amount due for the '%s' subscription. Please clear the amount and try again."%planname
            return (False, message)
        # Check how many tests and interviews this user has created using the subscribed plan.
        testscount, interviewscount, testsninterviewscount = 0, 0, 0
        testsql = "select count(*) from Tests_test where creator_id=%s and createdate > %s and createdate < %s"
        dbcursor.execute(testsql, (userobj.id, planstartdate, planenddate))
        alltestrecs = dbcursor.fetchall()
        if alltestrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
            testscount = 0
        else:
            testscount = int(alltestrecs[0][0])
        interviewsql = "select count(*) from Tests_interview where interviewer_id=%s and createdate > %s and createdate < %s"
        dbcursor.execute(interviewsql, (userobj.id, planstartdate, planenddate))
        allinterviewrecs = dbcursor.fetchall()
        if allinterviewrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
            interviewscount = 0
        else:
            interviewscount = int(allinterviewrecs[0][0])
        testsninterviewscount = testscount + interviewscount
        # Add the number of tests that would be created if the operation succeeds
        if filetype == "xlsx":
            try:
                wb = load_workbook(filename=filepath)
                ws = wb.worksheets[0]
                numtests = ws.rows.__len__() - 1
                testsninterviewscount = testsninterviewscount + numtests
            except:
                pass
        elif filetype == "xls":
            wb = open_workbook(filepath)
            numtests = 0
            for sheet in wb.sheets():
                numtests = sheet.nrows.__len__() - 1
                break
            testsninterviewscount = testsninterviewscount + numtests
        elif filetype == "csv":
            try:
                with open(filepath, 'rb') as csvfile:
                    csvreader = csv.reader(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
                    rowctr = 0
                    for row in csvreader:
                        rowctr += 1
                    testsninterviewscount = testsninterviewscount + rowctr - 1
            except:
                pass
        elif filetype == "xml":
            pass
        #print("%s ### %s =============== %s"%(planstartdate.strftime("%Y-%m-%d %H:%M:%S"), planenddate.strftime("%Y-%m-%d %H:%M:%S"), testsninterviewscount))
        if testsninterviewscount >= testsandinterviewsquota and planname != "Free Plan": # User has already consumed the allocated quota.
            message = "Error: You have already consumed the allocated count of tests and interviews in your subscription plan. Please extend your subscription to create more tests."
            return (False, message)
        else: # Allow user to go ahead and create test (unless user has consumed the number of tests and interviews allowed by Free Plan.
            if planname == "Free Plan" and int(mysettings.NEW_USER_FREE_TESTS_COUNT) != -1 and mysettings.NEW_USER_FREE_TESTS_COUNT <= testsninterviewscount:
                message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                return (False, message)
            elif planname == "Free Plan" and int(mysettings.NEW_USER_FREE_TESTS_COUNT) == -1 and testsninterviewscount >= testsandinterviewsquota:
                message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                return (False, message)
            else: 
                return (True, "success") # User's free quota of tests and interviews is not exhausted yet. So let the user continue.
    else: # So this is a free user
        testscount, interviewscount, testsninterviewscount = 0, 0, 0
        testsql = "select count(*) from Tests_test where creator_id=%s"
        dbcursor.execute(testsql, (userobj.id,))
        alltestrecs = dbcursor.fetchall()
        if alltestrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
            testscount = 0
        else:
            testscount = int(alltestrecs[0][0])
        interviewsql = "select count(*) from Tests_interview where interviewer_id=%s"
        dbcursor.execute(interviewsql, (userobj.id,))
        allinterviewrecs = dbcursor.fetchall()
        if allinterviewrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
            interviewscount = 0
        else:
            interviewscount = int(allinterviewrecs[0][0])
        testsninterviewscount = testscount + interviewscount
        freeplansql = "select testsninterviews from Subscription_plan where planname='Free Plan'"
        dbcursor.execute(freeplansql)
        planrecords = dbcursor.fetchall()
        if planrecords.__len__() >= 0:
            testsninterviewsquota = planrecords[0][0]
        else:
            testsninterviewsquota = 5 # Default free quota is this value.
        if int(mysettings.NEW_USER_FREE_TESTS_COUNT) != -1 and mysettings.NEW_USER_FREE_TESTS_COUNT <= testsninterviewscount:
            message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
            return (False, message)
        elif int(mysettings.NEW_USER_FREE_TESTS_COUNT) == -1 and testsninterviewscount >= testsninterviewsquota:
            message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
            return (False, message)
        else:
            pass # Allow user to continue creating tests and interviews.
    disconnectdb(dbconn, dbcursor)
    return (True, "success")

"""
Function to create a test from data in a xlsx file.
TO DO: Improper values or absence of values of mandatory parameters should raise appropriate exceptions.
"""
def create_test_from_xlsx(uploadedfile, filepath, testlinkid, userobj):
    (outcome, message) = _check_subscription_plan_limits(userobj, filepath, "xlsx")
    if outcome is False:
        return message
    try:
        wb = load_workbook(filename=filepath)
        ws = wb.worksheets[0]
        rowctr = 0
        testobj = None
        for row in ws.rows:
            if rowctr == 0: # This will be test metadata row
                # Create a test object
                testobj = Test()
                testobj.createdate = datetime.datetime.now()
                testobj.testname = row[0].value.decode('utf-8', "replace")
                if testobj.testname.__len__() > 100:
                    message = "Error: Couldn't create test as the testname exceeds 100 characters."
                    return message
                testobj.testtype = row[1].value.decode('utf-8', "replace")
                testobj.ruleset = row[2].value.decode('utf-8', "replace")
                testobj.topicname = ""
                if row[3].value != "" and row[3].value is not None:
                    testobj.topicname = row[3].value
                    testobj.topic = Topic.objects.filter(id=-1)[0]
                    testobj.topicname = testobj.topicname.replace("__", " ")
                else:
                    topicname = ""
                    if row[4].value != "" and row[4].value is not None:
                        topicname = row[4].value
                    if topicname.__len__() > 0 and testobj.topicname.__len__() == 0:
                        topic = Topic()
                        topic.topicname = topicname
                        topic.user = userobj
                        topic.createdate = datetime.datetime.now()
                        topic.isactive = True
                        topic.save()
                        testobj.topic = topic
                        testobj.topicname = topic.topicname
                if row[5].value != "":
                    testobj.maxscore = int(row[5].value)
                else:
                    testobj.maxscore = 0
                if row[6].value != "":
                    testobj.challengecount = int(row[6].value)
                else:
                    testobj.challengecount = 0
                if row[7].value != "":
                    samescore = int(row[7].value)
                else:
                    samescore = False
                if row[8].value != "":
                    testobj.negativescoreallowed = int(row[8].value)
                else:
                    testobj.negativescoreallowed = False
                if row[9].value != "":
                    testobj.passscore = int(row[9].value)
                else:
                    testobj.passscore = 0
                if row[10].value != "":
                    testduration = int(row[10].value)
                else:
                    testduration = 0
                testdurationunit = row[11].value.decode('utf-8', "replace")
                if testdurationunit == 'h':
                    testduration = testduration * 60 * 60
                elif testdurationunit == 'm':
                    testduration = testduration * 60
                elif testdurationunit == 's':
                    testduration = testduration
                testobj.duration = testduration
                testobj.creator = userobj
                if row[12].value != "":
                    maxchallengeduration = int(row[12].value)
                else:
                    maxchallengeduration = testduration
                if row[13].value != "":
                    maxchallengedurationunit = row[13].value
                else:
                    maxchallengedurationunit = testdurationunit
                evaluatorids = row[14].value.decode('utf-8', "replace")
                evaluatorgrpname = row[15].value.decode('utf-8', "replace")
                evaluatorgrpname = evaluatorgrpname.replace(" ", "__")
                evaluatorslist = evaluatorids.split(",")
                evaluatorobj = Evaluator()
                evaluatorobj.evalgroupname = evaluatorgrpname
                evalctr = 0
                for evalid in evaluatorslist:
                    if evalctr == 0:
                        evaluatorobj.groupmember1 = User.objects.get(emailid=evalid)
                    elif evalctr == 1:
                        evaluatorobj.groupmember2 = User.objects.get(emailid=evalid)
                    elif evalctr == 2:
                        evaluatorobj.groupmember3 = User.objects.get(emailid=evalid)
                    elif evalctr == 3:
                        evaluatorobj.groupmember4 = User.objects.get(emailid=evalid)
                    elif evalctr == 4:
                        evaluatorobj.groupmember5 = User.objects.get(emailid=evalid)
                    elif evalctr == 5:
                        evaluatorobj.groupmember6 = User.objects.get(emailid=evalid)
                    elif evalctr == 6:
                        evaluatorobj.groupmember7 = User.objects.get(emailid=evalid)
                    elif evalctr == 7:
                        evaluatorobj.groupmember8 = User.objects.get(emailid=evalid)
                    elif evalctr == 8:
                        evaluatorobj.groupmember9 = User.objects.get(emailid=evalid)
                    elif evalctr == 9:
                        evaluatorobj.groupmember10 = User.objects.get(emailid=evalid)
                    evalctr += 1
                evaluatorobj.save()
                testobj.evaluator = evaluatorobj
                if testobj.creator.emailid == testobj.evaluator.groupmember1.emailid or testobj.creator.emailid == testobj.evaluator.groupmember2.emailid or testobj.creator.emailid == testobj.evaluator.groupmember3.emailid or testobj.creator.emailid == testobj.evaluator.groupmember4.emailid or testobj.creator.emailid == testobj.evaluator.groupmember5.emailid or testobj.creator.emailid == testobj.evaluator.groupmember6.emailid or testobj.creator.emailid == testobj.evaluator.groupmember7.emailid or testobj.creator.emailid == testobj.evaluator.groupmember8.emailid or testobj.creator.emailid == testobj.evaluator.groupmember9.emailid or testobj.creator.emailid == testobj.evaluator.groupmember10.emailid:
                    testobj.creatorisevaluator = True
                else:
                    testobj.creatorisevaluator = False
                if row[16].value != "":
                    testobj.publishdate = row[16].value
                else:
                    testobj.publishdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                if row[17].value != "":
                    testobj.activationdate = row[17].value
                else:
                    testobj.activationdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                testobj.quality = row[18].value
                if row[19].value and row[19].value != "":
                    testobj.scope = row[19].value
                else:
                    testobj.scope = "private"
                if row[20].value == "" or not row[20].value:
                    testobj.allowedlanguages = 'enus'
                else:
                    testobj.allowedlanguages = row[20].value
                testobj.progenv = row[21].value
                if row[22].value != "":
                    testobj.multimediareqd = row[22].value
                else:
                    testobj.multimediareqd = False
                if row[23].value != "":
                    testobj.randomsequencing = row[23].value
                else:
                    testobj.randomsequencing = False
                testobj.allowmultiattempts = row[24].value
                testobj.maxattemptscount = row[25].value
                testobj.attemptsinterval = row[26].value
                testobj.attemptsintervalunit = row[27].value
                testobj.status = False
                print "Creating test with name '%s'"%testobj.testname
                testobj.save()
            elif rowctr == 1:
                pass
            elif rowctr > 1: # Challenge row.
                challengeobj = Challenge()
                challengeobj.test = testobj
                challengeobj.statement = row[0].value.decode('utf-8')
                challengeobj.challengescore = row[1].value
                if row[2].value and row[2].value != "":
                    challengeobj.timeframe = row[2].value
                else:
                    challengeobj.timeframe = testobj.duration
                if row[3].value and row[3].value != "":
                    challengeobj.negativescore = row[3].value
                else:
                    challengeobj.negativescore = 0
                challengeobj.mediafile = row[4].value
                challengeobj.additionalurl = row[5].value
                challengeobj.challengetype = row[6].value
                if row[7].value and row[7].value != "":
                    challengeobj.mustrespond = row[7].value
                else:
                    challengeobj.mustrespond = True
                challengeobj.oneormore = row[8].value
                if row[9].value and row[9].value != "":
                    challengeobj.challengequality = row[9].value
                else:
                    challengeobj.challengequality = testobj.quality
                challengeobj.responsekey = row[10].value
                if challengeobj.challengetype == 'MULT':
                    if row[11].value != "":
                        challengeobj.option1 = row[11].value
                    if row[12].value != "":
                        challengeobj.option2 = row[12].value
                    if row[13].value != "":
                        challengeobj.option3 = row[13].value
                    if row[14].value != "":
                        challengeobj.option4 = row[14].value
                    if row[15].value != "":
                        challengeobj.option5 = row[15].value
                    if row[16].value != "":
                        challengeobj.option6 = row[16].value
                    if row[17].value != "":
                        challengeobj.option7 = row[17].value
                    if row[18].value != "":
                        challengeobj.option8 = row[18].value
                challengeobj.save()
            rowctr += 1
    except:
        message = "Exception occurred while creating test from file '%s': %s\n"%(os.path.basename(filepath), sys.exc_info()[1].__str__())
        print message
        return message
    return 1


"""
TO DO: Improper values or absence of values of mandatory parameters should raise appropriate exceptions.
"""
def create_test_from_xls(uploadedfile, filepath, testlinkid, userobj):
    (outcome, message) = _check_subscription_plan_limits(userobj, filepath, "xls")
    if outcome is False:
        return message
    try:
        wb = open_workbook(filepath)
        for sheet in wb.sheets():
            rowctr = 0
            for row in range(sheet.nrows):
                if rowctr == 0:
                    testobj = Test()
                    testobj.createdate = datetime.datetime.now()
                    testobj.testname = sheet.cell(row,0).value.decode('utf-8', "replace")
                    if testobj.testname.__len__() > 100:
                        message = "Error: Couldn't create test as the testname exceeds 100 characters."
                        break
                    testobj.testtype = sheet.cell(row,1).value.decode('utf-8', "replace")
                    testobj.ruleset = sheet.cell(row,2).value.decode('utf-8', "replace")
                    if sheet.cell(row,3).value != "" and sheet.cell(row,3).value is not None:
                        testobj.topicname = sheet.cell(row,3).value
                        testobj.topic = Topic.objects.filter(id=-1)[0]
                        testobj.topicname = testobj.topicname.replace("__", " ")
                    else:
                        topicname = ""
                        if sheet.cell(row,4).value != "" and sheet.cell(row,4).value is not None:
                            topicname = sheet.cell(row,4).value
                        if topicname.__len__() > 0 and testobj.topicname.__len__() == 0:
                            topic = Topic()
                            topic.topicname = topicname
                            topic.user = userobj
                            topic.createdate = datetime.datetime.now()
                            topic.isactive = True
                            topic.save()
                            testobj.topic = topic
                            testobj.topicname = topic.topicname
                    if sheet.cell(row,5).value and sheet.cell(row,5).value != "":
                        testobj.maxscore = sheet.cell(row,5).value
                    else:
                        testobj.maxscore = 0
                    if sheet.cell(row,6).value and sheet.cell(row,6).value != "":
                        testobj.challengecount = sheet.cell(row,6).value
                    else:
                        testobj.challengecount = 0
                    if sheet.cell(row,7).value and sheet.cell(row,7).value != "":
                        samescore = sheet.cell(row,7).value
                    else:
                        samescore = 0
                    if sheet.cell(row,8).value and sheet.cell(row,8).value != "":
                        testobj.negativescoreallowed = sheet.cell(row,8).value
                    else:
                        testobj.negativescoreallowed = 0
                    if sheet.cell(row,9).value and sheet.cell(row,9).value != "":
                        testobj.passscore = sheet.cell(row,9).value
                    else:
                        testobj.passscore = 0
                    if sheet.cell(row,10).value and sheet.cell(row,10).value != "":
                        testduration = sheet.cell(row,10).value
                    else:
                        testduration = 0
                    testdurationunit = sheet.cell(row,11).value
                    if testdurationunit == 'h':
                        testduration = testduration * 60 * 60
                    elif testdurationunit == 'm':
                        testduration = testduration * 60
                    elif testdurationunit == 's':
                        testduration = testduration
                    testobj.duration = testduration
                    testobj.creator = userobj
                    if sheet.cell(row,12).value and sheet.cell(row,12).value != "":
                        maxchallengeduration = sheet.cell(row,12).value
                    else:
                        maxchallengeduration = testduration
                    if sheet.cell(row,13).value and sheet.cell(row,13).value != "":
                        maxchallengedurationunit = sheet.cell(row,13).value
                    else:
                        maxchallengedurationunit = testdurationunit
                    evaluatorids = sheet.cell(row,14).value.decode('utf-8', "replace")
                    evaluatorgrpname = sheet.cell(row,15).value.decode('utf-8', "replace")
                    evaluatorgrpname = evaluatorgrpname.replace(" ", "__")
                    evaluatorslist = evaluatorids.split(",")
                    evaluatorobj = Evaluator()
                    evaluatorobj.evalgroupname = evaluatorgrpname
                    evalctr = 0
                    for evalid in evaluatorslist:
                        if evalctr == 0:
                            evaluatorobj.groupmember1 = User.objects.get(emailid=evalid)
                        elif evalctr == 1:
                            evaluatorobj.groupmember2 = User.objects.get(emailid=evalid)
                        elif evalctr == 2:
                            evaluatorobj.groupmember3 = User.objects.get(emailid=evalid)
                        elif evalctr == 3:
                            evaluatorobj.groupmember4 = User.objects.get(emailid=evalid)
                        elif evalctr == 4:
                            evaluatorobj.groupmember5 = User.objects.get(emailid=evalid)
                        elif evalctr == 5:
                            evaluatorobj.groupmember6 = User.objects.get(emailid=evalid)
                        elif evalctr == 6:
                            evaluatorobj.groupmember7 = User.objects.get(emailid=evalid)
                        elif evalctr == 7:
                            evaluatorobj.groupmember8 = User.objects.get(emailid=evalid)
                        elif evalctr == 8:
                            evaluatorobj.groupmember9 = User.objects.get(emailid=evalid)
                        elif evalctr == 9:
                            evaluatorobj.groupmember10 = User.objects.get(emailid=evalid)
                        evalctr += 1
                    evaluatorobj.creationdate = datetime.datetime.now()
                    evaluatorobj.save()
                    testobj.evaluator = evaluatorobj
                    if (hasattr(testobj.evaluator, 'groupmember1') and testobj.creator.emailid == testobj.evaluator.groupmember1.emailid) or  (hasattr(testobj.evaluator, 'groupmember2') and testobj.creator.emailid == testobj.evaluator.groupmember2.emailid) or (hasattr(testobj.evaluator, 'groupmember3') and testobj.creator.emailid == testobj.evaluator.groupmember3.emailid) or (hasattr(testobj.evaluator, 'groupmember4') and  testobj.creator.emailid == testobj.evaluator.groupmember4.emailid) or (hasattr(testobj.evaluator, 'groupmember5') and testobj.creator.emailid == testobj.evaluator.groupmember5.emailid) or (hasattr(testobj.evaluator, 'groupmember6') and testobj.creator.emailid == testobj.evaluator.groupmember6.emailid) or (hasattr(testobj.evaluator, 'groupmember7') and testobj.creator.emailid == testobj.evaluator.groupmember7.emailid) or (hasattr(testobj.evaluator, 'groupmember8') and testobj.creator.emailid == testobj.evaluator.groupmember8.emailid) or (hasattr(testobj.evaluator, 'groupmember9') and testobj.creator.emailid == testobj.evaluator.groupmember9.emailid) or (hasattr(testobj.evaluator, 'groupmember10') and testobj.creator.emailid == testobj.evaluator.groupmember10.emailid):
                        testobj.creatorisevaluator = True
                    else:
                        testobj.creatorisevaluator = False
                    if sheet.cell(row,16).value and sheet.cell(row,16).value != "":
                        testobj.publishdate = sheet.cell(row,16).value
                    else:
                        testobj.publishdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                    if sheet.cell(row,17).value and sheet.cell(row,17).value != "":
                        testobj.activationdate = sheet.cell(row,17).value
                    else:
                        testobj.activationdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                    testobj.quality = sheet.cell(row,18).value
                    if sheet.cell(row,19).value and sheet.cell(row,19).value != "":
                        testobj.scope = sheet.cell(row,19).value
                    else:
                        testobj.scope = "private"
                    if sheet.cell(row,20).value and sheet.cell(row,20).value != "":
                        testobj.allowedlanguages = sheet.cell(row,20).value
                    else:
                        testobj.allowedlanguages = "enus"
                    testobj.progenv = sheet.cell(row,21).value
                    if sheet.cell(row,22).value != "":
                        testobj.multimediareqd = sheet.cell(row,22).value
                    else:
                        testobj.multimediareqd = False
                    if sheet.cell(row,23).value != "":
                        testobj.randomsequencing = sheet.cell(row,23).value
                    else:
                        testobj.randomsequencing = False
                    testobj.allowmultiattempts = sheet.cell(row,24).value
                    testobj.maxattemptscount = sheet.cell(row,25).value
                    testobj.attemptsinterval = sheet.cell(row,26).value
                    testobj.attemptsintervalunit = sheet.cell(row,27).value
                    testobj.status = False
                    print "Creating test with name '%s'"%testobj.testname
                    testobj.save()
                elif rowctr == 1:
                    pass
                elif rowctr > 1: # Challenge row.
                    challengeobj = Challenge()
                    challengeobj.test = testobj
                    challengeobj.statement = sheet.cell(row,0).value
                    challengeobj.challengescore = sheet.cell(row,1).value
                    if sheet.cell(row,2).value and sheet.cell(row,2).value != "":
                        challengeobj.timeframe = sheet.cell(row,2).value
                    else:
                        challengeobj.timeframe = testobj.duration
                    if sheet.cell(row,3).value and sheet.cell(row, 3).value != "":
                        challengeobj.negativescore = sheet.cell(row, 3).value
                    else:
                        challengeobj.negativescore = 0
                    challengeobj.mediafile = sheet.cell(row,4).value
                    challengeobj.additionalurl = sheet.cell(row,5).value
                    challengeobj.challengetype = sheet.cell(row,6).value
                    if sheet.cell(row,7).value and sheet.cell(row,7).value != "":
                        challengeobj.mustrespond = sheet.cell(row,7).value
                    else:
                        challengeobj.mustrespond = True
                    if sheet.cell(row,8).value and sheet.cell(row,8).value != "":
                        challengeobj.oneormore = sheet.cell(row,8).value
                    else:
                        challengeobj.oneormore = False
                    if sheet.cell(row,9).value and sheet.cell(row,9).value != "":
                        challengeobj.challengequality = sheet.cell(row,9).value
                    else:
                        challengeobj.challengequality = testobj.quality
                    challengeobj.responsekey = sheet.cell(row,10).value
                    if challengeobj.challengetype == 'MULT':
                        if sheet.cell(row,11).value != "":
                            challengeobj.option1 = sheet.cell(row,11).value
                        if sheet.cell(row,12).value != "":
                            challengeobj.option2 = sheet.cell(row,12).value
                        if sheet.cell(row,13).value != "":
                            challengeobj.option3 = sheet.cell(row,13).value
                        if sheet.cell(row,14).value != "":
                            challengeobj.option4 = sheet.cell(row,14).value
                        if sheet.cell(row,15).value != "":
                            challengeobj.option5 = sheet.cell(row,15).value
                        if sheet.cell(row,16).value != "":
                            challengeobj.option6 = sheet.cell(row,16).value
                        if sheet.cell(row,17).value != "":
                            challengeobj.option7 = sheet.cell(row,17).value
                        if sheet.cell(row,18).value != "":
                            challengeobj.option8 = sheet.cell(row,18).value
                    print "Creating challenge object '%s'...\n"%challengeobj.statement
                    challengeobj.save()
                rowctr += 1
    except:
        message = "Exception occurred while creating test from file '%s': %s\n"%(os.path.basename(filepath), sys.exc_info()[1].__str__())
        print message
        return message
    return 1


"""
TO DO: Improper values or absence of values of mandatory parameters should raise appropriate exceptions.
"""
def create_test_from_csv(uploadedfile, filepath, testlinkid, userobj):
    (outcome, message) = _check_subscription_plan_limits(userobj, filepath, "csv")
    if outcome is False:
        return message
    try:
        with open(filepath, 'rb') as csvfile:
            csvreader = csv.reader(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
            rowctr = 0
            for row in csvreader:
                if rowctr == 0:
                    testobj = Test()
                    testobj.createdate = datetime.datetime.now()
                    testobj.testname = row[0].decode('utf-8', 'replace')
                    if testobj.testname.__len__() > 100:
                        message = "Error: Couldn't create test as the testname exceeds 100 characters."
                        return message
                    testobj.testtype = row[1]
                    testobj.testtype = testobj.testtype.replace('"', '')
                    testobj.testtype = testobj.testtype.strip()
                    row[2] = row[2].replace('"', '')
                    row[2] = row[2].strip()
                    testobj.ruleset = row[2]
                    row[3] = row[3].replace('"', '')
                    row[3] = row[3].strip()
                    if row[3] != "" and row[3] is not None:
                        testobj.topicname = row[3]
                        testobj.topic = Topic.objects.filter(id=-1)[0]
                        testobj.topicname = testobj.topicname.replace("__", " ")
                    else:
                        topicname = ""
                        row[4] = row[4].replace('"', '')
                        row[4] = row[4].strip()
                        if row[4] != "" and row[4] is not None:
                            topicname = row[4]
                        if topicname.__len__() > 0 and testobj.topicname.__len__() == 0:
                            topic = Topic()
                            topic.topicname = topicname
                            topic.user = userobj
                            topic.createdate = datetime.datetime.now()
                            topic.isactive = True
                            topic.save()
                            testobj.topic = topic
                            testobj.topicname = topic.topicname
                    for j in range(5, 9):
                        row[j] = row[j].replace('"', '')
                        row[j] = row[j].strip()
                    if row[5] and row[5] != "":
                        testobj.maxscore = row[5]
                    else:
                        testobj.maxscore = 0
                    if row[6] and row[6] != "":
                        testobj.challengecount = row[6]
                    else:
                        testobj.challengecount = 0
                    if row[7] and row[7] != "":
                        samescore = row[7]
                    else:
                        samescore = 0
                    if row[8] and row[8] != "":
                        testobj.negativescoreallowed = row[8]
                    else:
                        testobj.negativescoreallowed = 0
                    if row[9] and row[9] != "":
                        testobj.passscore = row[9]
                    else:
                        testobj.passscore = 0
                    if row[10] and row[10] != "":
                        testduration = row[10]
                    else:
                        testduration = 0
                    testdurationunit = row[11]
                    if testdurationunit == 'h':
                        testduration = testduration * 60 * 60
                    elif testdurationunit == 'm':
                        testduration = testduration * 60
                    elif testdurationunit == 's':
                        testduration = testduration
                    else:
                        testduration = testduration # Any other unit will be ignored and considered as seconds. Do we want it this way?
                    testobj.duration = testduration
                    testobj.creator = userobj
                    testobj.status = False
                    if row[12] and row[12] != "":
                        maxchallengeduration = row[12]
                    else:
                        maxchallengeduration = testduration
                    if row[13] and row[13] != "":
                        maxchallengedurationunit = row[13]
                    else:
                        maxchallengedurationunit = testdurationunit
                    evaluatorids = row[14].decode('utf-8', "replace")
                    evaluatorgrpname = row[14].decode('utf-8', "replace")
                    evaluatorgrpname = evaluatorgrpname.replace('"', '')
                    evaluatorids = evaluatorids.replace('"', '')
                    evaluatorgrpname = evaluatorgrpname.replace(" ", "__")
                    evaluatorslist = evaluatorids.split(",")
                    evaluatorobj = Evaluator()
                    evaluatorobj.evalgroupname = evaluatorgrpname
                    evalctr = 0
                    quotePattern = re.compile('"')
                    for evalid in evaluatorslist:
                        evalid = re.sub(quotePattern, '', evalid) # Remove quote characters, if any.
                        evalid = ''.join(s for s in evalid if s in string.printable) # Remove non-printable characters, including some unicode chars. Should be changed if foreign languages are to be supported.
                        evalid = re.sub(mysettings.MULTIPLE_WS_PATTERN, '', evalid) # Remove inadvertant whitespaces, if any.
                        evalid = evalid.__str__() # Finally, we want ASCII string, not unicode.
                        evalid = evalid.strip()
                        #print evalid
                        if evalctr == 0:
                            evaluatorobj.groupmember1 = User.objects.get(emailid=evalid)
                        elif evalctr == 1:
                            evaluatorobj.groupmember2 = User.objects.get(emailid=evalid)
                        elif evalctr == 2:
                            evaluatorobj.groupmember3 = User.objects.get(emailid=evalid)
                        elif evalctr == 3:
                            evaluatorobj.groupmember4 = User.objects.get(emailid=evalid)
                        elif evalctr == 4:
                            evaluatorobj.groupmember5 = User.objects.get(emailid=evalid)
                        elif evalctr == 5:
                            evaluatorobj.groupmember6 = User.objects.get(emailid=evalid)
                        elif evalctr == 6:
                            evaluatorobj.groupmember7 = User.objects.get(emailid=evalid)
                        elif evalctr == 7:
                            evaluatorobj.groupmember8 = User.objects.get(emailid=evalid)
                        elif evalctr == 8:
                            evaluatorobj.groupmember9 = User.objects.get(emailid=evalid)
                        elif evalctr == 9:
                            evaluatorobj.groupmember10 = User.objects.get(emailid=evalid)
                        else:
                            pass
                        evalctr += 1
                    evaluatorobj.creationdate = datetime.datetime.now()
                    evaluatorobj.save()
                    testobj.evaluator = evaluatorobj
                    if (hasattr(testobj.evaluator, 'groupmember1') and testobj.creator.emailid == testobj.evaluator.groupmember1.emailid) or  (hasattr(testobj.evaluator, 'groupmember2') and testobj.creator.emailid == testobj.evaluator.groupmember2.emailid) or (hasattr(testobj.evaluator, 'groupmember3') and testobj.creator.emailid == testobj.evaluator.groupmember3.emailid) or (hasattr(testobj.evaluator, 'groupmember4') and  testobj.creator.emailid == testobj.evaluator.groupmember4.emailid) or (hasattr(testobj.evaluator, 'groupmember5') and testobj.creator.emailid == testobj.evaluator.groupmember5.emailid) or (hasattr(testobj.evaluator, 'groupmember6') and testobj.creator.emailid == testobj.evaluator.groupmember6.emailid) or (hasattr(testobj.evaluator, 'groupmember7') and testobj.creator.emailid == testobj.evaluator.groupmember7.emailid) or (hasattr(testobj.evaluator, 'groupmember8') and testobj.creator.emailid == testobj.evaluator.groupmember8.emailid) or (hasattr(testobj.evaluator, 'groupmember9') and testobj.creator.emailid == testobj.evaluator.groupmember9.emailid) or (hasattr(testobj.evaluator, 'groupmember10') and testobj.creator.emailid == testobj.evaluator.groupmember10.emailid):
                        testobj.creatorisevaluator = True
                    else:
                        testobj.creatorisevaluator = False
                    row[16] = row[16].strip()
                    row[17] = row[17].strip()
                    row[18] = row[18].replace('"', '')
                    row[18] = row[18].strip()
                    if row[16] and row[16] != "" and row[16] != '""':
                        testobj.publishdate = row[16]
                    else:
                        testobj.publishdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                    if row[17] and row[17] != "" and row[17] != '""':
                        testobj.activationdate = row[17]
                    else:
                        testobj.activationdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                    testobj.quality = row[18]
                    for j in range(19, 26):
                        row[j] = row[j].replace('"', '')
                        row[j] = row[j].strip()
                    if row[19] and row[19] != "":
                        testobj.scope = row[19]
                    else:
                        testobj.scope = "private"
                    if row[20] and row[20] != "":
                        testobj.allowedlanguages = row[20]
                    else:
                        testobj.allowedlanguages = "enus"
                    testobj.progenv = row[21]
                    if row[22] != "":
                        testobj.multimediareqd = row[22]
                    else:
                        testobj.multimediareqd = False
                    if row[23] != "":
                        testobj.randomsequencing = row[23]
                    else:
                        testobj.randomsequencing = False
                    testobj.allowmultiattempts = row[24]
                    testobj.maxattemptscount = row[25]
                    testobj.attemptsinterval = row[26]
                    testobj.attemptsintervalunit = row[27]
                    testobj.attemptsintervalunit = testobj.attemptsintervalunit.replace('"', '')
                    testobj.attemptsintervalunit = testobj.attemptsintervalunit.strip()
                    testobj.status = False
                    testobj.testlinkid = testlinkid
                    print "Creating test with name '%s'"%testobj.testname
                    testobj.save()
                elif rowctr == 1:
                    pass
                elif rowctr > 1:
                    challengeobj = Challenge()
                    challengeobj.test = testobj
                    challengeobj.statement = row[0]
                    row[1] = row[1].replace('"', '')
                    row[1] = row[1].strip()
                    row[2] = row[2].replace('"', '')
                    row[2] = row[2].strip()
                    challengeobj.challengescore = row[1]
                    if row[2] and row[2] != "":
                        challengeobj.timeframe = row[2]
                    else:
                        challengeobj.timeframe = testobj.duration
                    row[3] = row[3].replace('"', '')
                    row[3] = row[3].strip()
                    if row[3] and row[3] != "":
                        challengeobj.negativescore = row[3]
                    else:
                        challengeobj.negativescore = 0
                    row[4] = row[4].replace('"', '')
                    row[4] = row[4].strip()
                    row[5] = row[5].replace('"', '')
                    row[5] = row[5].strip()
                    row[7] = row[7].replace('"', '')
                    row[7] = row[7].strip()
                    row[8] = row[8].replace('"', '')
                    row[8] = row[8].strip()
                    challengeobj.mediafile = row[4]
                    challengeobj.additionalurl = row[5]
                    row[6] = row[6].replace('"', '')
                    row[6] = row[6].strip()
                    challengeobj.challengetype = row[6]
                    if row[7] and row[7] != "":
                        challengeobj.mustrespond = row[7]
                    else:
                        challengeobj.mustrespond = True
                    if row[8] and row[8] != "":
                        challengeobj.oneormore = row[8]
                    else:
                        challengeobj.oneormore = False
                    row[9] = row[9].replace('"', '')
                    row[9] = row[9].strip()
                    row[10] = row[10].replace('"', '')
                    row[10] = row[10].strip()
                    for j in range(11, row.__len__()):
                        row[j] = row[j].replace('"', '')
                        row[j] = row[j].strip()
                    if row[9] and row[9] != "":
                        challengeobj.challengequality = row[9]
                    else:
                        challengeobj.challengequality = testobj.quality
                    challengeobj.responsekey = row[10]
                    if challengeobj.challengetype == 'MULT':
                        if row.__len__() >= 12 and row[11] != "":
                            challengeobj.option1 = row[11]
                        if row.__len__() >= 13 and row[12] != "":
                            challengeobj.option2 = row[12]
                        if row.__len__() >= 14 and row[13] != "":
                            challengeobj.option3 = row[13]
                        if row.__len__() >= 15 and row[14] != "":
                            challengeobj.option4 = row[14]
                        if row.__len__() >= 16 and row[15] != "":
                            challengeobj.option5 = row[15]
                        if row.__len__() >= 17 and row[16] != "":
                            challengeobj.option6 = row[16]
                        if row.__len__() >= 18 and row[17] != "":
                            challengeobj.option7 = row[17]
                        if row.__len__() >= 19 and row[18] != "":
                            challengeobj.option8 = row[18]
                    challengeobj.testlinkid = testlinkid
                    print "Creating challenge object '%s'...\n"%challengeobj.statement
                    challengeobj.save()
                rowctr += 1
    except:
        message = "Exception occurred while creating test from file '%s': %s\n"%(os.path.basename(filepath), sys.exc_info()[1].__str__())
        #print message
        return message
    return 1


"""
TO DO: Improper values or absence of values of mandatory parameters should raise appropriate exceptions.
"""
def create_test_from_xml(uploadedfile, filepath, testlinkid, userobj):
    outcome, message = _check_subscription_plan_limits(userobj, filepath, "xml")
    if outcome is False:
        return message
    try:
        tree = et.parse(filepath)
        root = tree.getroot()
        childnodes = root.getchildren()
        challengeslist = []
        testobj = Test()
        for child in childnodes:
            if child.tag == 'metadata': # Test object should be created using data from this node.
                infonodes = child.getchildren()
                evaluatoremailslist = []
                evaluatorgroupname = ""
                for infond in infonodes:
                    if infond.tag == 'testname':
                        testobj.testname = infond.text
                        if testobj.testname.__len__() > 100:
                            message = "Error: Couldn't create test as the testname exceeds 100 characters."
                            return message
                    elif infond.tag == 'testtype':
                        testobj.testtype = infond.text
                    elif infond.tag == 'testrules':
                        testobj.ruleset = infond.text
                    elif infond.tag == 'testtopic' and infond.text != "":
                        testobj.topicname = infond.text
                        testobj.topicname = testobj.topicname.replace('__', ' ')
                        testobj.topic = Topic.objects.filter(id=-1)[0]
                    elif infond.tag == 'othertopicname' and infond.text == "":
                        topic = Topic()
                        topic.topicname = infond.text
                        topic.user = userobj
                        topic.createdate = datetime.datetime.now()
                        topic.isactive = True
                        topic.save()
                        testobj.topic = topic
                        testobj.topicname = topic.topicname
                    elif infond.tag == 'totalscore':
                        testobj.maxscore = infond.text
                    elif infond.tag == 'numberofchallenges':
                        testobj.challengecount = infond.text
                    elif infond.tag == 'eachchallengesamescore':
                        samescore = infond.text
                    elif infond.tag == 'incorrectresponsenegativescore':
                        testobj.negativescoreallowed = infond.text
                        if not testobj.negativescoreallowed or testobj.negativescoreallowed == '':
                            testobj.negativescoreallowed = 0
                    elif infond.tag == 'passscore':
                        testobj.passscore = infond.text
                    elif infond.tag == 'testduration':
                        testobj.duration = int(infond.text)
                        if not testobj.duration or testobj.duration == "":
                            raise ValueError('Test duration cannot be null')
                    elif infond.tag == 'testdurationunit':
                        testdurationunit = infond.text
                        if testdurationunit == 'h':
                            testobj.duration = testobj.duration * 60 * 60
                        elif testdurationunit == 'm':
                            testobj.duration = testobj.duration * 60
                        elif testdurationunit == 's':
                            testobj.duration = testobj.duration
                        else:
                            testobj.duration = testobj.duration
                    elif infond.tag == 'maxchallengeduration': # Not used at present
                        maxchallengeduration = infond.text
                    elif infond.tag == 'maxchallengedurationunit': # Not used at present
                        maxchallengedurationunit = infond.text
                    elif infond.tag == 'evaluatoremailids':
                        evaluatoremailids = infond.text
                        evaluatoremailslist = evaluatoremailids.split(",")
                    elif infond.tag == 'evaluatorgroupname':
                        evaluatorgroupname = infond.text
                    elif infond.tag == 'testpublishdate':
                        testobj.publishdate = infond.text
                        if not testobj.publishdate or testobj.publishdate == "":
                            testobj.publishdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                    elif infond.tag == 'testactivationdate':
                        testobj.activationdate = infond.text
                        if not testobj.activationdate or testobj.activationdate == "":
                            testobj.activationdate = skillutils.pythontomysqldatetime(datetime.datetime.now().__str__())
                    elif infond.tag == 'testtargetskilllevel':
                        testobj.quality = infond.text
                    elif infond.tag == 'testscope':
                        testobj.scope = infond.text
                        if not testobj.scope or testobj.scope == "":
                            testobj.scope = 'private'
                    elif infond.tag == 'answeringlanguage':
                        testobj.allowedlanguages = infond.text
                        if not testobj.allowedlanguages or testobj.allowedlanguages == "":
                            testobj.allowedlanguages = "enus"
                    elif infond.tag == 'programmingenvironment':
                        testobj.progenv = infond.text
                    elif infond.tag == 'needmultimedia':
                        testobj.multimediareqd = infond.text
                    elif infond.tag == 'randomsequencing':
                        testobj.randomsequencing = infond.text
                    elif infond.tag == 'allowmultipleattempts':
                        testobj.allowmultiattempts = infond.text
                    elif infond.tag == 'maxattemptsallowed':
                        testobj.maxattemptscount = infond.text
                    elif infond.tag == 'intervalbetweenattempts':
                        testobj.attemptsinterval = infond.text
                    elif infond.tag == 'intervalunits':
                        testobj.attemptsintervalunit = infond.text
                evaluatorobj = Evaluator()
                evaluatorobj.evalgroupname = evaluatorgroupname
                evalctr = 0
                for emailid in evaluatoremailslist:
                    grpmemberset = User.objects.filter(emailid=emailid)
                    if grpmemberset.__len__() == 0:
                        raise ValueError("No evaluators specified for the test")
                    if evalctr == 0:
                        evaluatorobj.groupmember1 = grpmemberset[0]
                    elif evalctr == 1:
                        evaluatorobj.groupmember2 = grpmemberset[0]
                    elif evalctr == 2:
                        evaluatorobj.groupmember3 = grpmemberset[0]
                    elif evalctr == 3:
                        evaluatorobj.groupmember4 = grpmemberset[0]
                    elif evalctr == 4:
                        evaluatorobj.groupmember5 = grpmemberset[0]
                    elif evalctr == 5:
                        evaluatorobj.groupmember6 = grpmemberset[0]
                    elif evalctr == 6:
                        evaluatorobj.groupmember7 = grpmemberset[0]
                    elif evalctr == 7:
                        evaluatorobj.groupmember8 = grpmemberset[0]
                    elif evalctr == 8:
                        evaluatorobj.groupmember9 = grpmemberset[0]
                    elif evalctr == 9:
                        evaluatorobj.groupmember10 = grpmemberset[0]
                    evalctr += 1
                evaluatorobj.save()
                testobj.evaluator = evaluatorobj
            elif child.tag == 'challenge': # Challenge object
                infonodes = child.getchildren()
                challengeobj = Challenge()
                for infond in infonodes:
                    if infond.tag == 'challengestatement':
                        challengeobj.statement = infond.text
                    elif infond.tag == 'challengescore':
                        challengeobj.challengescore = infond.text
                    elif infond.tag == 'challengeallocatedtime':
                        challengeobj.timeframe = infond.text
                    elif infond.tag == 'negativescore':
                        challengeobj.negativescore = infond.text
                        if not challengeobj.negativescore or challengeobj.negativescore == '':
                            challengeobj.negativescore = 0
                    elif infond.tag == 'challengeimagepath':
                        challengeobj.mediafile = infond.text
                    elif infond.tag == 'externalresourcepath':
                        challengeobj.additionalurl = infond.text
                    elif infond.tag == 'challengetype':
                        challengeobj.challengetype = infond.text
                    elif infond.tag == 'compulsoryforuser':
                        challengeobj.mustrespond = infond.text
                        if not challengeobj.mustrespond or challengeobj.mustrespond == "":
                            challengeobj.mustrespond = True
                    elif infond.tag == 'morethanoneoptioncorrect':
                        challengeobj.oneormore = infond.text
                    elif infond.tag == 'challengequality':
                        challengeobj.challengequality = infond.text
                        if not challengeobj.challengequality or challengeobj.challengequality == "":
                            challengeobj.challengequality = testobj.quality
                    elif infond.tag == 'correctresponse':
                        challengeobj.responsekey = infond.text
                    elif infond.tag == 'option_1':
                        challengeobj.option1 = infond.text
                    elif infond.tag == 'option_2':
                        challengeobj.option2 = infond.text
                    elif infond.tag == 'option_3':
                        challengeobj.option3 = infond.text
                    elif infond.tag == 'option_4':
                        challengeobj.option4 = infond.text
                    elif infond.tag == 'option_5':
                        challengeobj.option5 = infond.text
                    elif infond.tag == 'option_6':
                        challengeobj.option6 = infond.text
                    elif infond.tag == 'option_7':
                        challengeobj.option7 = infond.text
                    elif infond.tag == 'option_8':
                        challengeobj.option8 = infond.text
                    else:
                        pass # Unrecognized parameter
                challengeslist.append(challengeobj)
        testobj.creator = userobj
        testobj.status = False
        testobj.save()
        for challengeobj in challengeslist:
            challengeobj.test = testobj
            challengeobj.save()
        print "Created test named '%s' with %s challenges."%(testobj.testname, challengeslist.__len__())
    except:
        message = "Exception occurred while creating test from file '%s': %s\n"%(os.path.basename(filepath), sys.exc_info()[1].__str__())
        print message
        return message
    return 1


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def setvisibility(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid, visibility, vtype = -1, 0, ""
    if request.POST.has_key('testid'):
        testid = request.POST['testid'];
    if request.POST.has_key('vtype'):
        vtype = request.POST['vtype'];
    if request.POST.has_key('visibility'):
        visibility = request.POST['visibility'];
    testobj = None
    if testid != -1:
        testobj = Test.objects.get(id=testid)
    if vtype == 'test':
        if visibility == '0':
            testobj.scope = 'private'
            message = "Visibility value set to private"
        elif visibility == '1':
            testobj.scope = 'protected'
            message = "Visibility value set to protected"
        elif visibility == '2':
            testobj.scope = 'public'
            message = "Visibility value set to public"
        else:
            pass
        testobj.save()
    elif vtype == 'usertest':
        emailid = userobj.emailid
        utqset = UserTest.objects.filter(emailaddr=emailid, test=testobj)
        if utqset.__len__() == 0:
            utqset = WouldbeUsers.objects.filter(emailaddr=emailid, test=testobj)
        for utobj in utqset: # there may be multiple invitation records
            utobj.visibility = int(visibility)
            utobj.save()
        if visibility == '0':
            message = "Visibility value set to private"
        elif visibility == '1':
            message = "Visibility value set to protected"
        elif visibility == '2':
            message = "Visibility value set to public"
        else:
            message = "Unrecognized visibility value."
    else:
        message = "Unrecognized vtype value."
    response = HttpResponse(message)
    return response


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def showtestinfo(request):
    message = ""
    if request.method != 'POST':
        message = "Error: " + error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    testqset = Test.objects.filter(id=testid, scope='public')
    testobj = None
    if testqset.__len__() > 0:
        testobj = testqset[0]
    jsondata = {}
    if not testobj:
        jsondata = { 'warning' : "<font color='#AA0000'>This owner of this test has not marked it as publicly viewable test.</font>" }
    else:
        try:
            jsondata = { 'testname' : testobj.testname, 'topicname' : testobj.topicname, 'creatorname' : testobj.creator.displayname, 'testtype' : testobj.testtype, 'maxscore' : testobj.maxscore, 'passscore' : testobj.passscore, 'numchallenges' : testobj.challengecount, 'maxattemptscount' : testobj.maxattemptscount, 'attemptsinterval' : testobj.attemptsinterval, 'attemptsintervalunit' : testobj.attemptsintervalunit, 'quality' : testobj.quality, 'negativescoreallowed' : testobj.negativescoreallowed, 'warning' : '' }
        except:
            jsondata = { 'warning' : "<font color='#AA0000'>" + sys.exc_info()[1].__str__() + "</font>" }
    serialized_data = json.dumps(jsondata)
    return HttpResponse(serialized_data)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_exempt
def getcanvas(request):
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: " + error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid, challengeid = -1, -1
    if request.GET.has_key('testid') and request.GET['testid'] != "":
        testid = request.GET['testid']
    if request.GET.has_key('challengeid') and request.GET['challengeid'] != "":
        challengeid = request.GET['challengeid']
    contextdict = { 'savedrawingurl' : mysettings.SAVE_DRAWING_URL, 'testid' : str(testid), 'challengeid' : str(challengeid) }
    if request.method == 'GET':
        tmpl = get_template("tests/canvas.html")
        contextdict.update(csrf(request))
        cxt = Context(contextdict)
        canvashtml = tmpl.render(cxt)
    return HttpResponse(canvashtml)


"""
This method returns back the image filename if the challenge is a new one that is being created. If it 
is an existing challenge, then it returns back an empty string if successful. In all other cases of 
errors, it returns back a string that starts with "Error: ", and the rest of the message describes the
type of error encountered while processing the submitted image. Note: The mediafile created has a name
composed of the Id of the test and the timestamp of the moment at which the image is saved.
"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def savedrawing(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid, challengeid = -1, -1
    if request.POST.has_key('testid') and request.POST['testid'] != '':
        testid = request.POST['testid']
    if testid == -1:
        message = "Error: %s"%error_msg('1142')
        response = HttpResponse(message)
        return response
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = "Error: %s"%error_msg('1056')
        response = HttpResponse(message)
        return response
    canvasdata = request.POST['canvasdata']
    canvasdata = base64.b64decode(canvasdata)
    imageheader = "data:image/png;base64,"
    canvasdata = canvasdata.replace(imageheader, "")
    # Base64 decode it one more time to retrieve the binary data
    canvasdata = base64.b64decode(canvasdata)
    mediafilename = "t" + testid + "_" + str(int(time.time())) + ".png"
    mediapath = mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "tests" + os.path.sep + testid
    if not os.path.exists(mediapath):
        os.makedirs(mediapath)
    mediapath += os.path.sep + mediafilename
    returnval = ""
    try:
        fp = open(mediapath, "wb+")
        fp.write(canvasdata)
        fp.close()
    except:
        returnval = "Error: %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(returnval)
        return response
    if request.POST.has_key('challengeid') and request.POST['challengeid'] != '' and request.POST['challengeid'] != '-1':
        challengeid = request.POST['challengeid']
        challengeobj = Challenge.objects.get(id=challengeid)
        challengeobj.mediafile = mediafilename
        challengeobj.save()
        response = HttpResponse(returnval)
        return response
    else: # The challenge is not yet created, but is in the process of being created. So return back the filename so that the user may submit it while creating the challenge.
        response = HttpResponse(mediafilename)
        return response

"""
 This function is deprecated. Please refer to the function named "executecode".
"""
def executecodepad(request):
    """
    Method to execute code written by a test taker.
    """
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    codecontent_enc = request.POST['code']
    codecontent = base64.b64decode(codecontent_enc)
    #codecontent = codecontent.replace("__GT__", ">")
    #codecontent = codecontent.replace("__LT__", "<")
    #fp = open("/home/supriyo/work/stuff/codepadview.html", "w")
    #fp.write(codecontent)
    #fp.close()
    progenv = request.POST['progenv']
    # Now, enter codepad.org. Then navigate to the designated page where the run takes place.
    # Execute the code there and collect the output. Send the output back to the page from where the request originated.
    pagecontent = skillutils.getCodePadEditorPage()
    langradlist = []
    if pagecontent is None:
        response = HttpResponse("Could not get response from codepad.org. The site may be temporarily down or our programs may be having some difficulty. Please skip this operation or try again later.")
        #return response
    else:
        soup = BeautifulSoup(pagecontent)
        langradlist = soup.findAll("input", {'name' : 'lang'})
    reqdenv = progenv
    client_ip = skillutils.get_client_ip(request)
    client_port = skillutils.get_client_port(request)
    langfoundflag = 0;
    for langelem in langradlist:
        langname = langelem["value"]
        if langname is None or langname == "":
            continue
        if reqdenv.lower() == langname.lower():
            langfoundflag = 1
            break
    if not langfoundflag or langfoundflag == 0: # Requested environment not found in codepad. So run this on local environment.
        coderesult = skillutils.runCodeOnLocalResources(reqdenv, codecontent_enc, client_ip, client_port)
        return HttpResponse(coderesult)
    if not codecontent:
        message = "There is no code to run. Please enter some code to execute."
        response = HttpResponse(message)
        return response
    postdata = "lang=" + progenv + "&code=" + codecontent + "&run=True&submit=Submit"
    requestUrl = "http://codepad.org/"
    pageRequest = urllib2.Request(requestUrl, postdata, skillutils.gHttpHeaders)
    opener = urllib2.build_opener(urllib2.HTTPHandler(), urllib2.HTTPSHandler(), skillutils.NoRedirectHandler())
    pageResponse = None
    codepadsuccess = 0
    try:
        pageResponse = opener.open(pageRequest)
        respHeaders = pageResponse.info()
        if not respHeaders.has_key('Location') and not respHeaders.has_key('location'):
            coderesult = skillutils.runCodeOnLocalResources(reqdenv, codecontent_enc, client_ip, client_port)
            return HttpResponse(coderesult)
        requestUrl = respHeaders['Location']
        codepadsuccess = 1
    except:
        print "Could not make HTTP request to codepad: %s"%sys.exc_info()[1].__str__()
        coderesult = skillutils.runCodeOnLocalResources(reqdenv, codecontent_enc, client_ip, client_port)
        return HttpResponse(coderesult)
    pageRequest = urllib2.Request(requestUrl, None, skillutils.gHttpHeaders)
    try:
        pageResponse = opener.open(pageRequest)
    except:
        print "Redirect to the answer page failed: %s"%sys.exc_info()[1].__str__()
        coderesult = skillutils.runCodeOnLocalResources(reqdenv, codecontent_enc, client_ip, client_port)
        return HttpResponse(coderesult)
    currentPageContent = skillutils._decodeGzippedContent(skillutils.getPageContent(pageResponse))
    anchorPattern = re.compile('<a\s+name=[\"\']output[\"\']>', re.IGNORECASE)
    pageParts = re.split(anchorPattern, currentPageContent)
    morePageParts = pageParts[1].split("</table>")
    #soup = BeautifulSoup(currentPageContent)
    #anchortag = soup.find("a", {'name' : 'output'})
    #divtag = anchortag.findNext("div", {'class' : 'code'})
    #response = HttpResponse(divtag.renderContents())
    response = HttpResponse(morePageParts[0] + "</table>")
    return response


"""
@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
"""
def executecode(request):
    """
    Method to execute code written by a test taker.
    """
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    codecontent_enc, progenv, csrfmiddlewaretoken, testid, challengeid, emailid = "", "", "", "", "", ""
    if request.POST.has_key('code'):
        codecontent_enc = request.POST['code']
    if request.POST.has_key('progenv'):
        progenv = request.POST['progenv']
    if request.POST.has_key('csrfmiddlewaretoken'):
        csrfmiddlewaretoken = request.POST['csrfmiddlewaretoken']
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    if request.POST.has_key('challengeid'):
        challengeid = request.POST['challengeid']
    if request.POST.has_key('emailid'):
        emailid = request.POST['emailid']
    if not emailid or not challengeid or not testid or not progenv:
        message = "Error: Code cannot be executed as one or more required params are missing.\n"
        response = HttpResponse(message)
        return response 
    # Create an XML formatted string from the above parameters.
    xml = """<?xml version = "1.0"?>
	<challenge>
	    <email_id>%s</email_id>
	    <test_id>%s</test_id>
	    <challenge_id>%s</challenge_id>
	    <code>%s</code>
	    <proglang>%s</proglang>
	</challenge>"""%(emailid, testid, challengeid, codecontent_enc, progenv)
    # Open a socket connection to the appropriate port on the code execution service host.
    code_exec_host = mysettings.CODE_EXEC_HOST_IP
    code_exec_port = mysettings.CODE_EXEC_HOST_PORT
    max_code_size = mysettings.CODE_MAX_SIZE
    if xml.__len__() > max_code_size:
        message = "Error: Code cannot be executed as length of the XML request exceeds the permissible size acceptable at the executive host.\n"
        response = HttpResponse(message)
        return response 
    max_conn_backlogs = mysettings.MAX_CONN_BACKLOGS_LEN
    retdata = ""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect((code_exec_host, int(code_exec_port)))
        s.sendall(xml)
        retdata = s.recv(1024)
    except:
        print "Error: %s\n\n"%sys.exc_info()[1].__str__()
        retdata = "Had some problem with the code execution service."
    return HttpResponse(retdata)
    


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def disqualifycandidate(request):
    """
    Method to disqualify a candidate. Basically it sets the 'disqualified' 
    flag in 'Tests_usertest' or 'Tests_wouldbeusers' tables. Note that
    a candidate may be disqualified only by the owner/creator of the 
    concerned test.
    """
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    if request.POST.has_key('testid') and request.POST['testid'] != '':
        testid = request.POST['testid']
    if testid == -1:
        message = "Error: %s"%error_msg('1142')
        response = HttpResponse(message)
        return response
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = "Error: %s"%error_msg('1056')
        response = HttpResponse(message)
        return response
    emailid = ""
    if request.POST.has_key('emailid') and request.POST['emailid'] != '':
        emailid = request.POST['emailid']
    if emailid == '':
        message = error_msg('1157')
        response = HttpResponse(message)
        return response
    # Check if the user is the owner of the test.
    userisallowed = False
    if userobj == testobj.creator:
        userisallowed = True
    if not userisallowed:
        message = error_msg('1158')
        response = HttpResponse(message)
        return response
    usertestqset = UserTest.objects.filter(emailaddr=emailid, test=testobj)
    if usertestqset.__len__() == 0:
        usertestqset = WouldbeUsers.objects.filter(emailaddr=emailid, test=testobj)
    if usertestqset.__len__() == 0:
        message = error_msg('1159')
        response = HttpResponse(message)
        return response
    # There may be multiple instances of the same test scheduled for the same candidate. Need to disqualify all.
    for utobj in usertestqset:
        utobj.disqualified = True
        utobj.save()
    message = "Successfully disqualified the user for the given test."
    emailsubject = "Disqualification of your candidature from the test"
    emailmessage = """
    Dear candidate,
    
    Your candidature for the test named '%s' has been disqualified. If you
    consider to be unjust, please feel free to talk to our support team. They
    will help you to sort out the matter.

    Thanks,
    TestYard Support
    """%testobj.testname
    fromaddr = 'support@testyard.com'
    #str(emailmessage).content_subtype = 'html'
    """
    try:
        #retval = send_mail(emailsubject, emailmessage, fromaddr, [ emailid, ], False)
        retval = send_emails.delay(emailsubject, emailmessage, fromaddr, emailid, False)
    except:
        message += "Could not send email to the candidate."
    """
    response = HttpResponse(message)
    return response


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def copytest(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    else:
        message = error_msg('1055')
        response = HttpResponse(message)
        return response
    initiate = 1
    if request.POST.has_key('initiate'):
        initiate = request.POST['initiate']
    existingtestobj = Test.objects.get(id=testid)
    newtestobj = skillutils.copy_test(existingtestobj, userobj)
    message = """New test has been created successfully. Please note the following: 
    1. The evaluator for this test is you. You may change it later. 
    2. The publish date and activation date are scheduled 10 days ahead. You may change them as you wish. 
    3. To view the copied test, please refresh the screen."""
    return HttpResponse(message)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def gettestschedule(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    pageno = 1
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    else:
        message = error_msg('1055')
        response = HttpResponse(message)
        return response
    if 'pageno' in request.POST.keys():
        try:
            pageno = int(request.POST['pageno'])
        except:
            pass
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = error_msg('1058') + ": %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    startctr = mysettings.PAGE_CHUNK_SIZE * pageno - mysettings.PAGE_CHUNK_SIZE
    endctr = mysettings.PAGE_CHUNK_SIZE * pageno
    # Note: Don't allow creator to schedule test if the test hasn't been published and activated. Also 'status' should be True
    try:
        curdatetime = utilstimezone.now()
        publishdate = testobj.publishdate
        activationdate = testobj.activationdate
        status = testobj.status
    except:
        pass
    if not status:
        message = "<font color='#AA0000'>" + error_msg('1162') + "</font>"
        message += "<br /><font color='#0000AA'> If you think you are done with editing and ready to start scheduling this test, then you must activate it first. You may do so by clicking on the \"<i>Activate Test Now</i>\" link for this test at the far right hand side of the main screen."
        response = HttpResponse(message)
        return response
    if curdatetime < publishdate or curdatetime < activationdate:
        message = "<font color='#AA0000'>" + error_msg('1163') + "</font>"
        response = HttpResponse(message)
        return response
    # Get all schedules for this test.
    allschedulesqset = Schedule.objects.filter(test=testobj).order_by('-createdon')[startctr:endctr]
    # Now look for all schedules for this test in 'Tests_usertest' and 'Tests_wouldbeusers' tables.
    schedule_dict = {}
    for schedobj in allschedulesqset:
        scheduledtestsqset_ut = UserTest.objects.filter(test=testobj, schedule=schedobj)
        scheduledtestsqset_wbu = WouldbeUsers.objects.filter(test=testobj, schedule=schedobj)
        emailslist = []
        for utobj in scheduledtestsqset_ut:
            emailslist.append(utobj.emailaddr)
        for wbuobj in scheduledtestsqset_wbu:
            emailslist.append(wbuobj.emailaddr)
        past = 0
        try:
            if scheduledtestsqset_wbu.__len__() > 0 and curdatetime > scheduledtestsqset_wbu[0].validfrom:
                #print publishdate, " ##### ",activationdate, "#####", curdatetime
                past = 1
            elif scheduledtestsqset_ut.__len__() > 0 and curdatetime > scheduledtestsqset_ut[0].validfrom:
                #print publishdate, " ##### ",activationdate, "#####", curdatetime
                past = 1
            validfrom, validto = schedobj.slot.split('#||#')
            emailsliststr = ", ".join(emailslist[:2]) # Show only 2 email Ids. If the user wants to see more, she will click on the listing link.
            schedule_dict[str(schedobj.id)] = (validfrom, validto, past, emailsliststr)
        except:
            print sys.exc_info()[1].__str__()
    tmpl = get_template("tests/getscheduleinfo.html")
    nextpageno = pageno + 1
    prevpageno = pageno - 1
    contextdict = {'scheduleinfo' : schedule_dict, 'settestscheduleurl' : mysettings.SET_TEST_SCHEDULE_URL, 'testid' : testid, 'testname' : testobj.testname, 'showallemailidsurl' : mysettings.SHOW_ALL_EMAILS_URL, 'updatescheduleurl' : mysettings.UPDATE_SCHEDULE_URL, 'prevpageno' : prevpageno, 'nextpageno' : nextpageno, 'pageno' : pageno, 'gettestscheduleurl' : skillutils.gethosturl(request) + "/" + mysettings.GET_TEST_SCHEDULE_URL}
    contextdict.update(csrf(request))
    cxt = Context(contextdict)
    schedulehtml = tmpl.render(cxt)
    return HttpResponse(schedulehtml)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def updateschedule(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    scheduleid = -1
    scheduleobj = None
    starttime, endtime = "", ""
    if request.POST.has_key('schedid'):
        scheduleid = request.POST['schedid']
    else:
        message = error_msg('1185')
        response = HttpResponse(message)
        return response
    if request.POST.has_key('start'):
        starttime = request.POST['start']
    else:
        message = error_msg('1187')
        response = HttpResponse(message)
        return response
    if request.POST.has_key('end'):
        endtime = request.POST['end']
    else:
        message = error_msg('1187')
        response = HttpResponse(message)
        return response
    try:
        scheduleobj = Schedule.objects.get(id=scheduleid)
    except:
        message = error_msg('1186')
        response = HttpResponse(message)
        return response
    scheduleobj.slot = starttime + "#||#" + endtime
    scheduleobj.save()
    usertestqset = UserTest.objects.filter(schedule=scheduleobj)
    wouldbeusersqset = WouldbeUsers.objects.filter(schedule=scheduleobj)
    for utobj in usertestqset:
        try:
            starttime_date, starttime_time = starttime.split(" ")
            starttime_year, starttime_month, starttime_day = starttime_date.split("-")
            starttime_components = starttime_time.split(":")
            if starttime_components.__len__() > 2:
                starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], starttime_components[2]
            else:
                starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], "00"
            utobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second), 0, pytz.UTC)
            endtime_date, endtime_time = endtime.split(" ")
            endtime_year, endtime_month, endtime_day = endtime_date.split("-")
            endtime_components = endtime_time.split(":")
            if endtime_components.__len__() > 2:
                endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], endtime_components[2]
            else:
                endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], "00"
            utobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second), 0, pytz.UTC)
        except:
            message = sys.exc_info()[1].__str__()
            response = HttpResponse(message)
            return response
        try:
            utobj.save()
        except:
            utobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second))
            utobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second))
            utobj.save()
    for wbuobj in wouldbeusersqset:
        try:
            starttime_date, starttime_time = starttime.split(" ")
            starttime_year, starttime_month, starttime_day = starttime_date.split("-")
            starttime_components = starttime_time.split(":")
            if starttime_components.__len__() > 2:
                starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], starttime_components[2]
            else:
                starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], "00"
            wbuobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second), 0, pytz.UTC)
            endtime_date, endtime_time = endtime.split(" ")
            endtime_year, endtime_month, endtime_day = endtime_date.split("-")
            endtime_components = endtime_time.split(":")
            if endtime_components.__len__() > 2:
                endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], endtime_components[2]
            else:
                endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], "00"
            wbuobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second), 0, pytz.UTC)                  
        except:
            message = " Could not update record: %s"%sys.exc_info()[1].__str__()
            response = HttpResponse(message)
            return response
        try:
            wbuobj.save()
        except:
            message = "Error: %s Making ammends to rectify the situation... All will be well.\n"%sys.exc_info()[1].__str__()
            wbuobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second))
            wbuobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second))
            wbuobj.save()
    # TODO: Send email to all users intimating the change in the schedule
    message = "Updated existing schedule."
    response = HttpResponse(message)
    return response


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def showallemailids(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    scheduleid = -1
    pageno = 1
    if request.POST.has_key('schedid'):
        scheduleid = request.POST['schedid']
    else:
        message = error_msg('1185')
        response = HttpResponse(message)
        return response
    if request.POST.has_key('page'):
        try:
            pageno = int(request.POST['page'])
        except:
            pageno = 1
    scheduleobj = None
    try:
        scheduleobj = Schedule.objects.get(id=scheduleid)
    except:
        message = error_msg('1186')
        response = HttpResponse(message)
        return response
    testobj = scheduleobj.test
    scheduledtestsqset_ut = UserTest.objects.filter(test=testobj, schedule=scheduleobj)
    scheduledtestsqset_wbu = WouldbeUsers.objects.filter(test=testobj, schedule=scheduleobj)
    emailslist = []
    for utobj in scheduledtestsqset_ut:
        emailslist.append(utobj.emailaddr)
    for wbuobj in scheduledtestsqset_wbu:
        emailslist.append(wbuobj.emailaddr)
    segsize = 100
    segmentstart = segsize * pageno - segsize
    segmentend = segmentstart + segsize
    if emailslist.__len__() < segmentstart:
        message = json.dumps(['',])
        return HttpResponse(message)
    else:
        if emailslist.__len__() < segmentend:
            segmentend = emailslist.__len__()
        message = json.dumps(emailslist[segmentstart:segmentend])
        return HttpResponse(message)



@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def setschedule(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    else:
        message = error_msg('1055')
        response = HttpResponse(message)
        return response
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = error_msg('1058') + ": %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    
    joingroupflag = request.POST.get('joingroupflag', None)
    # Check if the user is the creator of this test. Only creators of a test 
    # are allowed to send invitations to candidates (except when an invitation 
    # needs to be sent automatically due to a user's need to join a group).
    if testobj.creator.id != userobj.id and not joingroupflag:
        message = "Error: " + error_msg('1070')
        response = HttpResponse(message)
        return response
    # Find the list of all evaluator's emails.
    testevaluator = testobj.evaluator
    testevalemailidlist = []
    if testevaluator.groupmember1:
        testevalemailidlist.append(testevaluator.groupmember1.emailid)
    if testevaluator.groupmember2:
        testevalemailidlist.append(testevaluator.groupmember2.emailid)
    if testevaluator.groupmember3:
        testevalemailidlist.append(testevaluator.groupmember3.emailid)
    if testevaluator.groupmember4:
        testevalemailidlist.append(testevaluator.groupmember4.emailid)
    if testevaluator.groupmember5:
        testevalemailidlist.append(testevaluator.groupmember5.emailid)
    if testevaluator.groupmember6:
        testevalemailidlist.append(testevaluator.groupmember6.emailid)
    if testevaluator.groupmember7:
        testevalemailidlist.append(testevaluator.groupmember7.emailid)
    if testevaluator.groupmember8:
        testevalemailidlist.append(testevaluator.groupmember8.emailid)
    if testevaluator.groupmember9:
        testevalemailidlist.append(testevaluator.groupmember9.emailid)
    if testevaluator.groupmember10:
        testevalemailidlist.append(testevaluator.groupmember10.emailid)
    start_new = request.POST['start_new']
    end_new = request.POST['end_new']
    start_new_dtobj = datetime.datetime.strptime(start_new, "%Y-%m-%d %H:%M:%S")
    end_new_dtobj = datetime.datetime.strptime(end_new, "%Y-%m-%d %H:%M:%S")
    if start_new_dtobj > end_new_dtobj:
        message = "Your schedule start date/time is later than your schedule end date/time. Kindly rectify this error and try again"
        return HttpResponse(message)
    if start_new and end_new:
        #duration = testobj.duration
        timeslot = start_new + "#||#" + end_new
        schedule = Schedule()
        schedule.test = testobj
        schedule.slot = timeslot
        schedule.save()
        # Now fetch all emails from emails_new
        try:
            emails_new = request.POST['emails_new']
        except:
            emails_new = ""
        validfrom, validtill = start_new, end_new
        new_emails_list = emails_new.split(",")
        # Find out when this test was created. Then find the subscription plan that the user was subscribed to during that period, and also the number of candidates permitted by that plan.
        testcreateddate = testobj.createdate
        dbconn, dbcursor = connectdb()
        #dbconn, dbcursor = connectdb_p()
        candidatescount = 5 # Default candidates count for Free Plan.
        plnname = 'Free Plan'
        freeplansql = "select candidates from Subscription_plan where planname='%s'"
        dbcursor.execute(freeplansql, (plnname,))
        allplanrecs = dbcursor.fetchall()
        if allplanrecs.__len__() > 0:
            candidatescount = allplanrecs[0][0]
        else:
            pass
        userplanid = -1
        subsuserplansql = "select pl.planname, pl.candidates, up.planstartdate, up.planenddate, up.id from Subscription_plan pl, Subscription_userplan up where up.user_id=%s and pl.id=up.plan_id and up.planstartdate < %s and up.planenddate > %s"
        # Note that the plan in question may have expired, but the user still has the right to send invitations to candidates in order to use the test created during the period of the plan.
        dbcursor.execute(subsuserplansql, (userobj.id, testcreateddate, testcreateddate))
        allusrplanrecs = dbcursor.fetchall()
        planstartdate = None
        planenddate = None
        if allusrplanrecs.__len__() == 0: # If no userplan is found, we should consider it to be a free plan.
            pass
        else:
            planname = allusrplanrecs[0][0]
            candidatescount = int(allusrplanrecs[0][1])
            planstartdate = allusrplanrecs[0][2]
            planenddate = allusrplanrecs[0][3]
            userplanid = allusrplanrecs[0][4]
        rightnow = datetime.datetime.now()
        if userplanid > 0: # Check to see if the user has extended his plan at present.
            planextsql = "select allowedinvites, periodstart, periodend from Subscription_planextensions where userplan_id=%s and user_id=%s and periodstart <= %s and periodend > %s"
            dbcursor.execute(planextsql, (userplanid, userobj.id, rightnow, rightnow))
            allextensionrecs = dbcursor.fetchall()
            if allextensionrecs.__len__() > 0:
                candidatescount = allextensionrecs[0][0]
                planstartdate = allextensionrecs[0][1]
                planenddate = allextensionrecs[0][2]
        # Now, find how many (distinct) candidates have already been invited to this test. If that count is less than the allowed
        # number of candidates permitted by the subscription plan, then go ahead. Otherwise, return a response with an error message.
        reguserinvitations_sql = "select count(distinct emailaddr) from Tests_usertest where test_id=%s and dateadded > %s and dateadded < %s"
        dbcursor.execute(reguserinvitations_sql, (testobj.id, planstartdate, planenddate))
        allreguserinvitations = dbcursor.fetchall()
        unreguserinvitations_sql = "select count(distinct emailaddr) from Tests_wouldbeusers where test_id=%s and dateadded > %s and dateadded < %s"
        dbcursor.execute(reguserinvitations_sql, (testobj.id, planstartdate, planenddate))
        allunreguserinvitations = dbcursor.fetchall()
        totalcandidates = 0
        if allreguserinvitations.__len__() > 0:
            totalcandidates += allreguserinvitations[0][0]
        if allunreguserinvitations.__len__() > 0:
            totalcandidates += allunreguserinvitations[0][0]
        if totalcandidates >= candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
            message = "Error: You have exhausted the allocated quota of test invitations for this test."
            return HttpResponse(message)
        # Now, check if the number of invitations being sent would surpass the quota of allowed invitations for the selected plan.
        postoperationuserinvitations = totalcandidates + emailslist.__len__()
        if postoperationuserinvitations > candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
            decrementfactor = postoperationuserinvitations - candidatescount
            message = "Error: The number of users being invited will surpass the allowed number of invitations that can be sent under your subscription plan. Please remove %s email addresses and try again."%decrementfactor
            return HttpResponse(message)
        disconnectdb(dbconn, dbcursor)
        # Done checking invitations quota.
        for new_email in new_emails_list:
            new_email = new_email.strip()
            # If this email belongs to the creator or one of the evaluators, skip it.
            if new_email == testobj.creator.emailid or new_email in testevalemailidlist:
                continue
            # Is the user registered with testyard?
            uobj = None
            try:
                uobj = User.objects.get(emailid=new_email)
            except:
                pass
            utobj = None
            if uobj is not None: # user is registered
                utobj = UserTest()
                utobj.user = uobj
            else:
                utobj = WouldbeUsers()
            utobj.emailaddr = new_email
            utobj.test = testobj
            try:
                start_new_date, start_new_time = start_new.split(" ")
                start_new_year, start_new_month, start_new_day = start_new_date.split("-")
                start_new_hour, start_new_minute, start_new_second = start_new_time.split(":")
                utobj.validfrom = datetime.datetime(int(start_new_year), int(start_new_month), int(start_new_day), int(start_new_hour), int(start_new_minute), int(start_new_second), 0, pytz.UTC)
                
                end_new_date, end_new_time = end_new.split(" ")
                end_new_year, end_new_month, end_new_day = end_new_date.split("-")
                end_new_hour, end_new_minute, end_new_second = end_new_time.split(":")
                utobj.validtill = datetime.datetime(int(end_new_year), int(end_new_month), int(end_new_day), int(end_new_hour), int(end_new_minute), int(end_new_second), 0, pytz.UTC)
            except:
                print sys.exc_info()[1].__str__()
            utobj.status = 0
            utobj.schedule = schedule
            baseurl = skillutils.gethosturl(request)
            (utobj.testurl, utobj.stringid) = gettesturlforuser(utobj.emailaddr, testid, baseurl)
            error_emails_list = []
            testyardhosturl = skillutils.gethosturl(request)
            candidatename = "candidate"
            emailsubject = "A test has been scheduled for you on testyard"
            emailmessage = """Dear %s,<br/><br/>

         A test with the name '%s' has been scheduled for you by <i>%s</i>.<br/><br/>

             """%(candidatename, testobj.testname, userobj.displayname)
            emailmessage += """
	    The test will start from %s and end at %s."""%(validfrom, validtill)
            emailmessage += """ and hence you are kindly requested to take the test
            within that interval. You would be able to access the test by clicking<br/>
            on the following link: <a href='%s'>%s</a>.<br/>
	    You may add this test to your <a href='%s/skillstest/test/addtocalendar/?turl=%s'>google calendar</a>.
	<br/><br/>
            If clicking on the above link doesn't work for you, please copy it and 
            paste it in your browser's address bar and hit enter. Do please feel<br/>
            free to let us know in case of any issues. We would do our best to
            resolve it at the earliest.<br/><br/>

            We wish you all the best for the test.<br/><br/>

            Regards,<br/>
            The TestYard Team.
            """%(utobj.testurl, utobj.testurl, testyardhosturl, urllib.quote_plus(utobj.testurl))
            fromaddr = "testyardteam@testyard.com"
            #str(emailmessage).content_subtype = 'html'
            retval = 0
            try:
                #retval = send_mail(emailsubject, emailmessage, fromaddr, [new_email,], False)
                retval = send_emails.delay(emailsubject, emailmessage, fromaddr, new_email, False)
                utobj.save()
                #print utobj.validfrom, utobj.validtill
            except:
                if mysettings.DEBUG:
                    print "Error: sendemail failed for %s - %s\n"%(new_email, sys.exc_info()[1].__str__())
                message = "Error: sendemail failed for %s - %s\n"%(new_email, sys.exc_info()[1].__str__())
                message += "Making ammends to rectify the situation... all will be well.\n"
                #utobj.validfrom = datetime.datetime(int(start_new_year), int(start_new_month), int(start_new_day), int(start_new_hour), int(start_new_minute), int(start_new_second))
                #utobj.validtill = datetime.datetime(int(end_new_year), int(end_new_month), int(end_new_day), int(end_new_hour), int(end_new_minute), int(end_new_second))
                #utobj.save()
                error_emails_list.append(new_email)
                continue # Continue processing the rest of the emails in the list.
        message = "Success! All candidates have been queued for email with the link."
        # Dump all emails Ids to which email could not be sent
        for error_email in error_emails_list:
            emailfail = EmailFailure()
            emailfail.user = userobj
            emailfail.sessionid = sesscode
            emailfail.failedemailid = error_email
            emailfail.script = 'Tests.views.setschedule'
            emailfail.failurereason = sys.exc_info()[1].__str__()
            emailfail.tryagain = 1
            try:
                emailfail.save()
            except:
                message = sys.exc_info()[1].__str__()
                print message
    # Check out other schedule variables
    start_pattern = re.compile("start_(\d+)$")
    end_pattern = re.compile("end_(\d+)$")
    existing_schedules = {}
    for fieldname in request.POST.keys():
        start_match = start_pattern.search(fieldname)
        end_match = end_pattern.search(fieldname)
        if start_match:
            scheduleid = start_match.groups()[0]
            if not existing_schedules.has_key(str(scheduleid)):
                existing_schedules[str(scheduleid)] = [ request.POST[fieldname], '']
            else:
                existing_schedules[str(scheduleid)][0] = request.POST[fieldname]
        elif end_match:
            scheduleid = end_match.groups()[0]
            if not existing_schedules.has_key(str(scheduleid)):
                existing_schedules[str(scheduleid)] = [ '', request.POST[fieldname] ]
            else:
                existing_schedules[str(scheduleid)][1] = request.POST[fieldname]
        else:
            pass
    for scheduleid in existing_schedules.keys():
        scheduleobj = Schedule.objects.get(id=scheduleid)
        starttime, endtime = existing_schedules[scheduleid]
        if starttime == "" or endtime == "":
            continue
        scheduleobj.slot = starttime + "#||#" + endtime
        scheduleobj.save()
        usertestqset = UserTest.objects.filter(schedule=scheduleobj)
        wouldbeusersqset = WouldbeUsers.objects.filter(schedule=scheduleobj)
        for utobj in usertestqset:
            try:
                starttime_date, starttime_time = starttime.split(" ")
                starttime_year, starttime_month, starttime_day = starttime_date.split("-")
                starttime_components = starttime_time.split(":")
                if starttime_components.__len__() > 2:
                    starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], starttime_components[2]
                else:
                    starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], "00"
                utobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second), 0, pytz.UTC)
                endtime_date, endtime_time = endtime.split(" ")
                endtime_year, endtime_month, endtime_day = endtime_date.split("-")
                endtime_components = endtime_time.split(":")
                if endtime_components.__len__() > 2:
                    endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], endtime_components[2]
                else:
                    endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], "00"
                utobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second), 0, pytz.UTC)
            except:
                print sys.exc_info()[1].__str__()
            try:
                utobj.save()
            except:
                utobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second))
                utobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second))
                utobj.save()
        for wbuobj in wouldbeusersqset:
            try:
                starttime_date, starttime_time = starttime.split(" ")
                starttime_year, starttime_month, starttime_day = starttime_date.split("-")
                starttime_components = starttime_time.split(":")
                if starttime_components.__len__() > 2:
                    starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], starttime_components[2]
                else:
                    starttime_hour, starttime_minute, starttime_second = starttime_components[0], starttime_components[1], "00"
                wbuobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second), 0, pytz.UTC)
                endtime_date, endtime_time = endtime.split(" ")
                endtime_year, endtime_month, endtime_day = endtime_date.split("-")
                endtime_components = endtime_time.split(":")
                if endtime_components.__len__() > 2:
                    endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], endtime_components[2]
                else:
                    endtime_hour, endtime_minute, endtime_second = endtime_components[0], endtime_components[1], "00"
                wbuobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second), 0, pytz.UTC)                  
            except:
                print sys.exc_info()[1].__str__()
                message += " Could not update record."
                response = HttpResponse(message)
                return response
            try:
                wbuobj.save()
            except:
                message += "Error: %s Making ammends to rectify the situation... All will be well.\n"%sys.exc_info()[1].__str__()
                wbuobj.validfrom = datetime.datetime(int(starttime_year), int(starttime_month), int(starttime_day), int(starttime_hour), int(starttime_minute), int(starttime_second))
                wbuobj.validtill = datetime.datetime(int(endtime_year), int(endtime_month), int(endtime_day), int(endtime_hour), int(endtime_minute), int(endtime_second))
                wbuobj.save() 
    message += " Updated existing schedules. However, the following email Ids couldn't be handled: " + ", ".join(error_emails_list) 
    response = HttpResponse(message)
    return response


@csrf_protect
def savetestresponses(request):
    pass


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def activatetestbycreator(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    else:
        message = error_msg('1055')
        response = HttpResponse(message)
        return response
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = error_msg('1058') + ": %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    # Check if the user is the creator of the test. If not, back out with an appropriate message.
    if testobj.creator != userobj:
        message = "User is not the creator of this test. Hence she/he may not activate the test."
        response = HttpResponse(message)
        return response
    testobj.status = True
    testobj.save()
    message = "The test has been activated. You may now schedule this test for test takers and the test takers will now be able to take the test. Please refresh the page to view the updated status of the test."
    response = HttpResponse(message)
    return response


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def deactivatetestbycreator(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.DASHBOARD_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.MANAGE_TEST_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    testid = -1
    if request.POST.has_key('testid'):
        testid = request.POST['testid']
    else:
        message = error_msg('1055')
        response = HttpResponse(message)
        return response
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = error_msg('1058') + ": %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    # Check if the user is the creator of the test. If not, back out with an appropriate message.
    if testobj.creator != userobj:
        message = "User is not the creator of this test. Hence she/he may not activate the test."
        response = HttpResponse(message)
        return response
    testobj.status = False
    testobj.save()
    message = "The test has been deactivated. You may not schedule this test for test takers now. Test takers will not be able to take the test now. Please refresh the page to view the updated status of the test."
    response = HttpResponse(message)
    return response


def captureaudiovisual(request):
    #tmpl = get_template("tests/interview_candidate_screen.html")
    tmpl = get_template("tests/audiovisual.html")
    tests_user_dict = {}
    tests_user_dict['blob_upload_url'] = mysettings.BLOB_UPLOAD_URL
    tests_user_dict['interviewlinkid'] = request.POST['interviewlinkid']
    tests_user_dict['realtime'] = '1'
    tests_user_dict['updateinterviewmetaurl'] = mysettings.UPDATE_INTERVIEW_META_URL
    tests_user_dict['signal_server'] = mysettings.SIGNAL_SERVER
    tests_user_dict.update(csrf(request))
    cxt = Context(tests_user_dict)
    audiovisualhtml = tmpl.render(cxt)
    return HttpResponse(audiovisualhtml)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def askquestion(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    interviewlinkid = None
    int_questions_dict = {}
    int_questions_dict['askquestionurl'] = mysettings.ASK_QUESTION_URL
    int_questions_dict['challengestoreurl'] = mysettings.CHALLENGE_STORE_URL
    int_questions_dict['pagetitle'] = "Record Question"
    if request.POST.has_key('interviewlinkid'):
        interviewlinkid = request.POST['interviewlinkid']
    else:
        int_questions_dict['errmsg'] = error_msg('1171')
    question_num = 0
    if request.POST.has_key('question_num'):
        question_num = request.POST['question_num']
        if question_num == '':
            current_question_num = '0'
        else:
            current_question_num = int(question_num) + 1
    medium = "audiovisual"
    if request.POST.has_key('medium'):
        medium = request.POST['medium']
    int_questions_dict['interviewlinkid'] = interviewlinkid
    int_questions_dict['question_num'] = current_question_num
    int_questions_dict['updateinterviewmetaurl'] = mysettings.UPDATE_INTERVIEW_META_URL
    int_questions_dict['answerfilename'] = ""
    int_questions_dict['signal_server'] = mysettings.SIGNAL_SERVER
    if medium == "audiovisual":
        #tmpl = get_template("tests/interview_candidate_screen.html")
        tmpl = get_template("tests/audiovisual.html")
    elif medium == "audio":
        tmpl = get_template("tests/audio.html")
    else:
        print "Unsupported medium requested."
        tmpl = "Unsupported medium requested."
    int_questions_dict.update(csrf(request))
    cxt = Context(int_questions_dict)
    mediumhtml = tmpl.render(cxt)
    return HttpResponse(mediumhtml)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def uploadblobdata(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    interviewlinkid, blobdata, filename, question_num = "", "", "intro.wav", ""
    if request.POST.has_key('question_num'):
        question_num = request.POST['question_num']
    else:
        question_num = '1'
    if question_num != '' and int(question_num) > 0:
        filename = "q" + question_num + ".wav"
    if request.POST.has_key('interviewlinkid'):
        interviewlinkid = request.POST['interviewlinkid']
    else: # This question/challenge doesn't seem to be associated with any interview. Hence we may drop it.
        response = HttpResponse(error_msg('1167'))
        return response
    if int(question_num) >= 1 and interviewlinkid:
        try:
            interviewquestionobj = InterviewQuestions()
            interviewobj = Interview.objects.get(interviewlinkid = interviewlinkid)
            interviewquestionobj.interview = interviewobj
            interviewquestionobj.questionfilename = filename
            interviewquestionobj.questionnumber = question_num
            interviewquestionobj.interviewlinkid = interviewlinkid
            interviewquestionobj.status = True
            interviewquestionobj.maxscore = interviewobj.maxscore
            interviewquestionobj.timelimit = interviewobj.maxduration
            interviewquestionobj.save()
        except:
            print sys.exc_info()[1].__str__()
            resp = HttpResponse(error_msg('1172'))
            return resp
    if request.FILES.has_key('file'):
        interviewqset = Interview.objects.filter(interviewlinkid=interviewlinkid)
        if interviewqset.__len__() == 0:
            response = HttpResponse(error_msg('1169'))
            return response
        if interviewqset.__len__() > 1:
            return HttpResponse(error_msg('1170'))
        interviewtitle = interviewqset[0].title
        directoryname = re.sub(re.compile("\s+"), "_", interviewtitle)
        # Check if there are any files in 'filepath = mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "interviews" + os.path.sep + directoryname'. If not, the filename should be 'intro.wav'.
        fileslist = os.listdir(mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "interviews" + os.path.sep + directoryname)
        if fileslist.__len__() == 0:
            filename = interviewqset[0].introfilepath
        filepath = mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "interviews" + os.path.sep + directoryname + os.path.sep + filename
        os.chmod(mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "interviews" + os.path.sep + directoryname, stat.S_IRWXG|stat.S_IRWXO|stat.S_IRWXU)
        blobdata = request.FILES['file']
        with open(filepath, 'wb+') as destination:
            for chunk in blobdata.chunks():
                destination.write(chunk)
        destination.close()
        #if interviewqset[0].realtime:
        #    print "EEEEEEEEEEEEEEEEEEEEEEEE"
        #    streaminterview(filepath)
    else:
        response = HttpResponse(error_msg('1168'))
        return response
    return HttpResponse("The content has been uploaded successfully")


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def uploadrecording(request):
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    # Get the data from the request - recording, interviewtitle and csrfmiddlewaretoken
    recordingblob = request.POST['recording']
    interviewtitle = request.POST['interviewtitle']
    return HttpResponse("")


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def createinterview(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    interviewtitle, interviewtopic, totalscore, maxresponsestarttime, numchallenges, interviewduration, medium, publishdate, language, realtime, skilltarget, interviewscope, randomsequencing, interviewlinkid, introbtntext, introfilename, emailinvitationtarget, scheduledatetime, chkrightnow, interviewdatetime, maxinterviewerscount, intervieweremails = "", "", '100', '300', '20', '3600', "audiovisual", "", "English-US", 1, "","", 0, "", "Add Intro", "intro.wav", "", "", 0, None, 1, ""
    if request.POST.has_key('interviewtitle'):
        interviewtitle = request.POST['interviewtitle']
    # Check to see if an interview with the same title exists in the current user's list.
    interviewqset = Interview.objects.filter(title=interviewtitle)
    if interviewqset.__len__() > 0: # Return an error message
        msg = "Error: %s"%error_msg('1166')
        return HttpResponse(msg)
    if request.POST.has_key('interviewtopic'):
        interviewtopic = request.POST['interviewtopic']
    if request.POST.has_key('totalscore') and request.POST['totalscore'] != "":
        totalscore = request.POST['totalscore']
    totalscore = totalscore.strip()
    if request.POST.has_key('maxresponsestarttime') and request.POST['maxresponsestarttime'] != "":
        maxresponsestarttime = request.POST['maxresponsestarttime']
    if request.POST.has_key('numchallenges') and request.POST['numchallenges'] != "":
        numchallenges = request.POST['numchallenges']
    if request.POST.has_key('interviewduration') and request.POST['interviewduration'] != "":
        interviewduration = request.POST['interviewduration']
    if request.POST.has_key('medium'):
        medium = request.POST['medium']
    if request.POST.has_key('publishdate'):
        publishdate = request.POST['publishdate']
    #if request.POST.has_key('scheduledatetime'):
    #    scheduledatetime = request.POST['scheduledatetime']
    if request.POST.has_key('language') and request.POST['language'] != "":
        language = request.POST['language']
    if request.POST.has_key('max_interviewers_count') and request.POST['max_interviewers_count'] != "":
        maxinterviewerscount = request.POST['max_interviewers_count']
    if request.POST.has_key('interviewer_emails') and request.POST['interviewer_emails'] != "":
        intervieweremails = request.POST['interviewer_emails']
    if int(maxinterviewerscount) == 1:
        intervieweremails = userobj.emailid
    else:
        intervieweremails += ",%s"%userobj.emailid
    intervieweremailslist = intervieweremails.split(",")
    if request.POST.has_key('realtime') and request.POST.has_key('scheduledatetime'):
        realtime = request.POST['realtime']
        scheduledatetime = request.POST['scheduledatetime']
    else:  
        realtime = 0
        scheduledatetime = ''
    if not scheduledatetime:
        scheduledatetime = request.POST['interviewdatetime']
    if request.POST.has_key('chkrightnow') and request.POST['chkrightnow'] == '1':
        scheduledatetime = datetime.datetime.now()
    if realtime: # Email invitation to all interviewees should be sent.
      emailinvitationtarget = request.POST['invitationemailaddr']
      #scheduledatetime = request.POST['scheduledatetime']
    if request.POST.has_key('skilltarget'):
        skilltarget = request.POST['skilltarget']
    if request.POST.has_key('interviewscope'):
        interviewscope = request.POST['interviewscope']
    randomsequencing = 0
    if request.POST.has_key('interviewlinkid'):
        interviewlinkid = request.POST['interviewlinkid']
    if request.POST.has_key('chkrightnow'):
        chkrightnow = 1
    if request.POST.has_key('introfilename'):
        introfilename = request.POST['introfilename']
    else:
        introfilename = "intro.wav"
    if request.POST.has_key('btncreateinterview'):
        introbtntext = request.POST['btncreateinterview']
    if request.POST.has_key('interviewdatetime'):
        interviewdatetime = request.POST['interviewdatetime']
    
    if interviewdatetime == "":
        interviewdatetime = None
    # Check if we already have an interview with the same interviewlinkid. If so, we do not create the interview.
    intqset = Interview.objects.filter(interviewlinkid=interviewlinkid)
    if intqset.__len__() > 0:
        resp = HttpResponse(error_msg('1170'))
        return resp
    else: # User is trying to create a new interview. Check quotas as per subscription plan
        dbconn, dbcursor = connectdb()
        #dbconn, dbcursor = connectdb_p()
        todaynow = datetime.datetime.now()
        userplansql = "select pl.testsninterviews, up.plan_id, pl.planname, up.amountdue, up.planstartdate, up.planenddate from Subscription_userplan up, Subscription_plan pl where up.planstartdate < %s and up.planenddate > %s and up.planstatus=TRUE and up.plan_id=pl.id and up.user_id=%s"
        dbcursor.execute(userplansql, (todaynow, todaynow, userobj.id))
        allrecords = dbcursor.fetchall()
        if allrecords.__len__() > 0:
            testsandinterviewsquota = allrecords[0][0]
            planid = allrecords[0][1]
            planname = allrecords[0][2]
            amtdue = float(allrecords[0][3])
            planstartdate = allrecords[0][4]
            if planstartdate is None or planstartdate == "":
                message = "Error: Plan start date is missing for user '%s'. Please contact support with this error message."%userobj.displayname
                return HttpResponse(message)
            planenddate = allrecords[0][5]
            if planenddate is None or planenddate == "":
                message = "Error: Plan end date is missing for user '%s'. Please contact support with this error message."%userobj.displayname
                return HttpResponse(message)
            if amtdue > 0.00: # User has not paid the total amount for this plan.
                message = "Error: You have an amount due for the '%s' subscription. Please clear the amount and try again."%planname
                return HttpResponse(message)
            # Check how many tests and interviews this user has created using the subscribed plan.
            testscount, interviewscount, testsninterviewscount = 0, 0, 0
            testsql = "select count(*) from Tests_test where creator_id=%s and createdate > %s and createdate < %s"
            dbcursor.execute(testsql, (userobj.id, planstartdate, planenddate))
            alltestrecs = dbcursor.fetchall()
            if alltestrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                testscount = 0
            else:
                testscount = int(alltestrecs[0][0])
            interviewsql = "select count(*) from Tests_interview where interviewer_id=%s and createdate > %s and createdate < %s"
            dbcursor.execute(interviewsql, (userobj.id, planstartdate, planenddate))
            allinterviewrecs = dbcursor.fetchall()
            if allinterviewrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                interviewscount = 0
            else:
                interviewscount = int(allinterviewrecs[0][0])
            testsninterviewscount = testscount + interviewscount
            if testsninterviewscount >= testsandinterviewsquota and planname != "Free Plan": # User has already consumed the allocated quota.
                message = "Error: You have already consumed the allocated count of tests and interviews in your subscription plan. Please extend your subscription to create more tests and interviews."
                return HttpResponse(message)
            else: # Allow user to go ahead and create test (unless user has consumed the number of tests and interviews allowed by Free Plan.
                if planname == "Free Plan" and int(mysettings.NEW_USER_FREE_TESTS_COUNT) != -1 and mysettings.NEW_USER_FREE_TESTS_COUNT <= testsninterviewscount:
                    message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                    return HttpResponse(message)
                elif planname == "Free Plan" and int(mysettings.NEW_USER_FREE_TESTS_COUNT) == -1 and testsninterviewscount >= testsandinterviewsquota:
                     message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                     return HttpResponse(message)
                else:
                    pass # User's free quota of tests and interviews is not exhausted yet. So let the user continue.
        else: # So this is a free user
            testscount, interviewscount, testsninterviewscount = 0, 0, 0
            testsql = "select count(*) from Tests_test where creator_id=%s"
            dbcursor.execute(testsql, (userobj.id,))
            alltestrecs = dbcursor.fetchall()
            if alltestrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                testscount = 0
            else:
                testscount = int(alltestrecs[0][0])
            interviewsql = "select count(*) from Tests_interview where interviewer_id=%s"
            dbcursor.execute(interviewsql, (userobj.id,))
            allinterviewrecs = dbcursor.fetchall()
            if allinterviewrecs.__len__() == 0: #This case should not happen. If it happens, then there is an error somewhere.
                interviewscount = 0
            else:
                interviewscount = int(allinterviewrecs[0][0])
            testsninterviewscount = testscount + interviewscount
            freeplansql = "select testsninterviews from Subscription_plan where planname='Free Plan'"
            dbcursor.execute(freeplansql)
            planrecords = dbcursor.fetchall()
            if planrecords.__len__() >= 0:
                testsninterviewsquota = planrecords[0][0]
            else:
                testsninterviewsquota = 5 # Default free quota is this value.
            if int(mysettings.NEW_USER_FREE_TESTS_COUNT) != -1 and mysettings.NEW_USER_FREE_TESTS_COUNT <= testsninterviewscount:
                message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                return HttpResponse(message)
            elif int(mysettings.NEW_USER_FREE_TESTS_COUNT) == -1 and testsninterviewscount >= testsninterviewsquota:
                message = "Error: You have already consumed the allocated count of tests and interviews for a free account. Please buy a subscription plan to create more tests and interviews."
                return HttpResponse(message)
            else:
                pass # Allow user to continue creating tests and interviews.
        disconnectdb(dbconn, dbcursor)
    interviewobj = Interview()
    if introbtntext == 'Add Intro'  or introbtntext == 'Start Video Call': # We are here for the first time
        interviewobj.title = interviewtitle
    elif introbtntext == 'Open Intro':
        try:
            interviewobj = Interview.objects.filter(title=interviewtitle, interviewer=userobj)[0]
        except:
            pass
        interviewobj.title = interviewtitle
    interviewobj.challengescount = numchallenges
    interviewobj.maxresponsestarttime = maxresponsestarttime
    interviewobj.topicname = interviewtopic
    interviewobj.interviewer = userobj
    interviewobj.medium = medium
    interviewobj.language = language
    interviewobj.createdate = ""
    publishdateparts = publishdate.split("-")
    pubmon2digit = mysettings.MONTHS_DICT[publishdateparts[1]]
    publishdate_mysqlcompliant = publishdateparts[2] + "-" + pubmon2digit + "-" + publishdateparts[0]
    interviewobj.publishdate = publishdate_mysqlcompliant + " 00:00:00"
    #datetimePattern = re.compile("\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}")
    #if datetimePattern.search(scheduledatetime):
    #    interviewobj.publishdate = scheduledatetime
    interviewobj.status = False # The interview is being created... so this ought to be false.
    interviewobj.maxscore = totalscore
    interviewobj.maxduration = interviewduration
    #interviewobj.randomsequencing = randomsequencing
    interviewobj.randomsequencing = 0
    interviewobj.interviewlinkid = interviewlinkid
    interviewobj.scope = interviewscope
    #interviewobj.quality = skilltarget
    interviewobj.interviewers_count = maxinterviewerscount
    interviewobj.interviewer_ids = intervieweremails
    interviewobj.challengesfilepath = ""
    interviewobj.introfilepath = introfilename
    interviewobj.scheduledtime = interviewdatetime
    hashtoken = binascii.hexlify(os.urandom(16))
    testyardhosturl = skillutils.gethosturl(request)
    interviewurlparams = mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  interviewlinkid + "&hash=" + hashtoken + "&attend=" + emailinvitationtarget
    if emailinvitationtarget: # Send an email invitation link to the email address.
        message = """<pre>Dear Candidate,
                     <br/><br/>
                     This is an invitation to attend an interview with %s on %s hours. Please click on the link below to load the interview interface. If it doesn't work, then copy the link and paste it in your browser's address bar and hit <enter>.<br/>
                     
                     <a href="%s/%s" target='_blank'>%s/%s</a> <br/>
                     
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.
                     <br/>
                     Important Note: Please use Chrome, Firefox or Opera to attend the interview. Browsers other than these 3 may not support every feature used by the interview application.<br/><br/>

                     Good Luck!<br/>
                     The TestYard Interview Team.</pre>
    """%(userobj.displayname, scheduledatetime, skillutils.gethosturl(request), interviewurlparams, skillutils.gethosturl(request), interviewurlparams, testyardhosturl, urllib.quote_plus(str(interviewlinkid)))
        subject = "TestYard Interview Invitation"
        fromaddr = userobj.emailid
        #str(message).content_subtype = 'html'
        # Send email
        try:
            #retval = send_mail(subject, message, fromaddr, [emailinvitationtarget,], False)
            retval = send_emails.delay(subject, message, fromaddr, emailinvitationtarget, False)
        except:
            if mysettings.DEBUG:
                print "sendemail failed for %s - %s\n"%(emailinvitationtarget, sys.exc_info()[1].__str__())
    uniqueintervieweremailsdict = {}
    for interviewerem in intervieweremailslist:
        if interviewerem.strip() == "":
            continue
        if interviewerem not in uniqueintervieweremailsdict.keys():
            uniqueintervieweremailsdict[interviewerem] = 1
        else:
            pass
    if intervieweremailslist.__len__() >= 1: # Send an email with the interviewer's link to the email addresses
        message = """<pre>Dear Interviewer,
                     <br/><br/>
                     You have invited a candidate identified by email '%s' to interview at %s. Please click on the link below to load the interview interface. If it doesn't work, then copy the link and paste it in your browser's address bar and hit <enter>.<br/>
                     
                     %s/%s <br/>
                     
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.
                     <br/>
                     Important Note: Please use Chrome, Firefox or Opera to attend the interview. Browsers other than these 3 may not support every feature used by the interview application.<br/><br/>

                     Good Luck!<br/>
                     The TestYard Interview Team.</pre>
         """%(emailinvitationtarget, scheduledatetime, skillutils.gethosturl(request), mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  interviewlinkid + "&hash=" + hashtoken, testyardhosturl, urllib.quote_plus(str(interviewlinkid)))
        subject = "TestYard Interview Scheduled"
        fromaddr = userobj.emailid
        for interviewerem in uniqueintervieweremailsdict.keys():
            if interviewerem != "":
                try:
                    retval = send_emails.delay(subject, message, fromaddr, interviewerem, False)
                except:
                    if mysettings.DEBUG:
                        print "sendemail failed for %s - %s\n"%(interviewerem, sys.exc_info()[1].__str__())
    interviewobj.filetype = "wav"
    int_user_dict = {}
    try:
        interviewobj.save()
        # Create the directory for the interview
        directoryname = re.sub(re.compile("\s+"), "_", interviewtitle)
        dirpath = mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "interviews" + os.path.sep + directoryname
        if not os.path.exists(dirpath):
            os.makedirs(dirpath)
    except:
        int_user_dict['errmsg'] = "Error: " + sys.exc_info()[1].__str__()
        return HttpResponse(int_user_dict['errmsg'])
    # Inform both users that interview has been scheduled. If scheduled datetime > current datetime, 
    # then the audio.html/audiovisual.html screen should not be shown. Instead, show the URL for both
    # users and request the users to appear at the linked page at the scheduled time. The respective 
    # URLs will be sent to both the users through email. Additionally, the interviewer will have the 
    # URL listed in her/his list of interviews. If the interviewee is also a member of TestYard, then
    # she too will have the URL listed  in her/his list of interviews.
    intcandidateobj = None
    if realtime and medium == 'audiovisual':
        intcandidateobj = InterviewCandidates()
        intcandidateobj.interview = interviewobj
        intcandidateobj.emailaddr = emailinvitationtarget
        if scheduledatetime == 'YYYY-MM-DD hh:mm:ss' or scheduledatetime == "":
            scheduledatetime = datetime.datetime.now()
        intcandidateobj.scheduledtime = scheduledatetime
        intcandidateobj.interviewlinkid = interviewlinkid
        #caur = skillutils.gethosturl(request) + "/" + mysettings.ATTEND_INTERVIEW_URL + interviewlinkid + "/" + "?lid=" + interviewlinkid
        try:
            intcandidateobj.interviewurl = skillutils.gethosturl(request) + "/" + mysettings.ATTEND_INTERVIEW_URL + interviewlinkid + "/" + "?lid=" + interviewlinkid + "&hash=" + hashtoken
            intcandidateobj.save()
        except:
            print "Error: %s"%sys.exc_info()[1].__str__()
        currentdatetime = datetime.datetime.now()
        if not chkrightnow:
            scheduledatetime_dt = datetime.datetime.strptime(scheduledatetime, "%Y-%m-%d %H:%M:%S")
        else:
            scheduledatetime_dt = scheduledatetime
        #if datetime.datetime.strptime(scheduledatetime_dt, "%Y-%m-%d %H:%M:%S") > currentdatetime:
        if scheduledatetime_dt > currentdatetime:
            message = """<pre>Dear Candidate,<br/><br/>
                     

                     This is an invitation to attend an interview named '%s' with %s on %s.  Please click on the link below to load the interview interface. If it doesn't work, then copy the link and paste it in your browser's address bar and hit <enter>.<br/>
                     
                     %s<br/>
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.<br/>
                     
                     Important Note: Please use a recent version of Chrome, Firefox or Opera to attend the interview. Browsers other than these 3 may not support every feature used by the interview application.<br/><br/>
                     
                     Good Luck!
                     The TestYard Interview Team.</pre>
    """%(interviewobj.title, userobj.displayname, scheduledatetime, intcandidateobj.interviewurl + "&attend=" + emailinvitationtarget, skillutils.gethosturl(request), urllib.quote_plus(str(interviewlinkid)))
            subject = "TestYard Interview Invitation"
            fromaddr = userobj.emailid
            #str(message).content_subtype = 'html'
            # Send email
            try:
                #retval = send_mail(subject, message, fromaddr, [emailinvitationtarget,], False)
                retval = send_emails.delay(subject, message, fromaddr, emailinvitationtarget, False)
            except:
                if mysettings.DEBUG:
                    retmsg = "sendemail failed for %s - %s\n"%(emailinvitationtarget, sys.exc_info()[1].__str__())
                    return HttpResponse(retmsg)
            # Now, send a similar email to the creator/conductor of the interview with the appropriate interview URL.
            message = """<pre>Dear Interviewer,<br/><br/>
        You have successfully set up an interview titled '%s' for %s at %s hours. You may click on the following link to access the interview application that will assist you in conducting the interview at the aforementioned time.<br/>

        %s <br/>
        
        You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.<br/><br/>

        Good Luck!

        The TestYard Team.</pre>
        """%(interviewobj.title, emailinvitationtarget, scheduledatetime, intcandidateobj.interviewurl, skillutils.gethosturl(request), urllib.quote_plus(str(interviewlinkid)))
            #str(message).content_subtype = 'html'
            subject = "Interview Scheduled"
            fromaddr = userobj.emailid
            toaddr = userobj.emailid
            try:
                #retval = send_mail(subject, message, fromaddr, [toaddr,], False)
                retval = send_emails.delay(subject, message, fromaddr, toaddr, False)
            except:
                if mysettings.DEBUG:
                    retmsg = "sendmail failed for %s - %s\n"%(toaddr, sys.exc_info()[1].__str__())
                    return HttpResponse(retmsg)
            html = "The interview has been scheduled at %s hours, and the candidate is being informed about it by email."%scheduledatetime
            html += "You may conduct the interview at the mentioned hour by clicking on the following link: %s"%intcandidateobj.interviewurl
            html += "<br />The interview link (above) is also being sent to your email address. It might take a little while based of how busy our email servers are."
            #str(html).content_subtype = 'html'
            return HttpResponse(html)
        else:
            pass
    elif realtime and medium == 'audio':
        pass
    elif not realtime and medium == 'audio':
        pass
    else: # This will handle the case where medium is audiovisual but the realtime flag is off. You can have an audiovisual interview only at realtime.
        pass
        
    int_user_dict['challengestoreurl'] = mysettings.CHALLENGE_STORE_URL
    int_user_dict['interviewlinkid'] = interviewlinkid
    int_user_dict['askquestionurl'] = mysettings.ASK_QUESTION_URL
    int_user_dict['updateinterviewmetaurl'] = mysettings.UPDATE_INTERVIEW_META_URL
    int_user_dict['pagetitle'] = "^ Please click on 'Allow' button above to allow the computer to use webcam and microphone. ^ Add an Introductory Comment"
    int_user_dict['question_num'] = "0"
    int_user_dict['medium'] = medium
    if intcandidateobj:
        int_user_dict['interview_url'] = intcandidateobj.interviewurl
    else:
        int_user_dict['interview_url'] = ""
    int_user_dict['interview_data_upload_url'] = mysettings.INTERVIEW_DATA_UPLOAD_URL
    int_user_dict['hashtoken'] = hashtoken
    int_user_dict['interviewtitle'] = interviewtitle
    int_user_dict['signal_server'] = mysettings.SIGNAL_SERVER
    interviewfilename = interviewtitle
    interviewfilename = interviewfilename.replace("-", "_")
    interviewfilename = interviewfilename.replace(" ", "_")
    targetpart = emailinvitationtarget
    targetpart = targetpart.replace("-", "_")
    targetpart = targetpart.replace(".", "_")
    targetpart = targetpart.replace("@", "_AT_")
    interviewfilename += "_" + targetpart
    interviewfilename += "_" + int(time.time()).__str__() + ".mp4"
    int_user_dict['interviewfilename'] = interviewfilename
    int_user_dict['scheduledatetime'] = scheduledatetime
    int_user_dict['email'] = userobj.emailid
    int_user_dict['emailinvitationtarget'] = emailinvitationtarget
    if medium == 'audio':
        tmpl = get_template("tests/audio.html")
    else:
        currentdatetime = datetime.datetime.now()
        if not chkrightnow:
            scheduledatetime_dt = datetime.datetime.strptime(scheduledatetime, "%Y-%m-%d %H:%M:%S")
        else:
            scheduledatetime_dt = scheduledatetime
        if scheduledatetime_dt > currentdatetime:
            tmpl = get_template("tests/waitscreen.html")
        else:
            #tmpl = get_template("tests/interview_candidate_screen.html")
            tmpl = get_template("tests/audiovisual.html")
    int_user_dict.update(csrf(request))
    cxt = Context(int_user_dict)
    audiovisualhtml = tmpl.render(cxt)
    return HttpResponse(audiovisualhtml)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def uploadinterviewdata(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message) # TODO: Need to handle this case in a more user-friendly way.
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    imagefilename, audiofilename, imagecontent, audiocontent, interviewlinkid, hashtoken, filename = "", "", None, None, "", "", ""
    if request.POST.has_key('image'):
        # handle image content
        imagecontent = base64.b64decode(request.POST['image'])
    elif request.POST.has_key('audio'):
        # handle audio content
        audiocontent = base64.b64decode(request.POST['audio'])
    if request.POST.has_key('interviewlinkid'):
        interviewlinkid = request.POST['interviewlinkid']
    else:
        message = "No interview link Id found. Cannot save data for this interview. Please save the data locally for future reference."
        response = HttpResponse(message) 
        return response
    if request.POST.has_key('hashtoken'):
        hashtoken = request.POST['hashtoken']
    if request.POST.has_key('filename'):
        filename = request.POST['filename']
    else:
        message = "Could not upload interview data as filename is missing. Please save the data locally for future reference."
        response = HttpResponse(message)
        return response
    intobj = None
    intobjqset = Interview.objects.filter(interviewlinkid=interviewlinkid)
    if intobjqset.__len__() > 0:
        intobj = intobjqset[0]
    else:
        message = "Could not find the interview record identified by the interview link Id '%s'"%interviewlinkid
        response = HttpResponse(message)
        return response
    intdirname = intobj.title
    intdirname = intdirname.replace(" ", "_")
    mediapath = mysettings.MEDIA_ROOT + os.path.sep + userobj.displayname + os.path.sep + "interviews" + os.path.sep + intdirname
    if not os.path.exists(mediapath):
        os.makedirs(mediapath)
    filepath = mediapath + os.path.sep + filename
    fp = open(filepath, "wb")
    if imagecontent is not None:
        fp.write(imagecontent)
    elif audiocontent is not None:
        fp.write(audiocontent)
    else:
        message = "Unhandled content type. Interview data could not be saved."
        fp.close()
        response = HttpResponse(message)
        return response
    fp.close()
    message = "Successfully uploaded interview data file."
    resp = HttpResponse(message)
    return resp


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def checknameavailability(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    inttitle = request.POST['inttitle']
    intqset = Interview.objects.filter(title=inttitle, interviewer=userobj)
    if intqset.__len__() > 0:
        return HttpResponse('0') # Not available
    else:
        return HttpResponse('1') # Available


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def displayinterviewschedulescreen(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    intlinkid = ""
    if not request.POST.has_key('intlinkid'):
        message = "Error: Required parameter interviewlink Id missing."
        response = HttpResponse(message)
        return response
    intlinkid = request.POST['intlinkid']
    if intlinkid.strip() == "":
        message = "Error: Required parameter interviewlinkId is empty."
        response = HttpResponse(message)
        return response
    interviewobj = None
    try:
        interviewobj = Interview.objects.get(interviewlinkid=intlinkid)
    except:
        message = "Error: Could not fetch interview record for the given interviewlinkId - %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    intdatadict = {}
    intdatadict['interviewtitle'] = interviewobj.title
    intdatadict['interviewer'] = interviewobj.interviewer.displayname
    intdatadict['interviewid'] = interviewobj.id
    interviewcandidatesqset = InterviewCandidates.objects.filter(interview=interviewobj).order_by('-scheduledtime')
    intdatadict['existing_interviews'] = []
    for intcandidate in interviewcandidatesqset:
        starttime = intcandidate.actualstarttime
        schedulestatus = False
        if not starttime or starttime == "NULL" or starttime is None:
            schedulestatus = True
        candidateemailslist = intcandidate.emailaddr.split(",")
        candidateemailscount = candidateemailslist.__len__()
        d = {'email' : intcandidate.emailaddr, 'scheduled' : intcandidate.scheduledtime.strftime("%Y-%m-%d %H:%M:%S"), 'schedulestatus' : schedulestatus, 'intcandidateid' : intcandidate.id, 'interviewid' : interviewobj.id, 'emailscount' : candidateemailscount}
        intdatadict['existing_interviews'].append(d)
    tmpl = get_template("tests/schedule_interview.html")
    intdatadict.update(csrf(request))
    cxt = Context(intdatadict)
    schedulescreenhtml = tmpl.render(cxt)
    contentdict = {'html' : schedulescreenhtml}
    return HttpResponse(json.dumps(contentdict))


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def savenewinterviewschedule(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    interviewid = ""
    if not request.POST.has_key('interviewid'):
        message = "Error: Required parameter interview Id missing."
        response = HttpResponse(message)
        return response
    interviewid = request.POST['interviewid']
    if interviewid.strip() == "":
        message = "Error: Required parameter interview Id is empty."
        response = HttpResponse(message)
        return response
    interviewobj = None
    try:
        interviewobj = Interview.objects.get(id=interviewid)
    except:
        message = "Error retrieving interview object"
        response = HttpResponse(message)
        return response
    # Now we do have the requisite interview object, let's check out if all other values are valid.
    newsched_email, newsched_datetime = "", ""
    if not request.POST.has_key("newsched_email"):
        message = "Error: Required parameter email Ids missing."
        response = HttpResponse(message)
        return response
    if not request.POST.has_key("newsched_datetime"):
        message = "Error: Required parameter Schedule date and time missing."
        response = HttpResponse(message)
        return response
    newsched_email = request.POST['newsched_email']
    newsched_datetime = request.POST['newsched_datetime']
    if newsched_email.strip() == "" or newsched_datetime.strip() == "":
        message = "One or more of the required field values are empty"
        response = HttpResponse(message)
        return response
    emailpattern = re.compile("@.*\.")
    datetimepattern = re.compile("\d{4}\-\d{2}\-\d{2}T\d{2}\:\d{2}")
    if not re.search(datetimepattern, newsched_datetime):
        message = "Schedule datetime is not in appropriate format."
        response = HttpResponse(message)
        return response
    emailslist = newsched_email.split(",")
    emailidslist = []
    uniqemails = {}
    for email in emailslist:
        if not re.search(emailpattern, email): # If email Id is invalid, we simply skip it.
            continue
        if email not in uniqemails.keys():
            emailidslist.append(email)
            uniqemails[email] = 1
        else:
            pass
    # Find out when this interview was created. Then find the subscription plan that the user was subscribed to during that period, and also the number of candidates permitted by that plan.
    interviewcreateddate = interviewobj.createdate
    dbconn, dbcursor = connectdb()
    #dbconn, dbcursor = connectdb_p()
    candidatescount = 5 # Default candidates count for Free Plan.
    plnname = 'Free Plan'
    freeplansql = "select candidates from Subscription_plan where planname=%s"
    dbcursor.execute(freeplansql, (plnname,))
    allplanrecs = dbcursor.fetchall()
    if allplanrecs.__len__() > 0:
        candidatescount = allplanrecs[0][0]
    else:
        pass
    userplanid = -1
    subsuserplansql = "select pl.planname, pl.candidates, up.planstartdate, up.planenddate, up.id from Subscription_plan pl, Subscription_userplan up where up.user_id=%s and pl.id=up.plan_id and up.planstartdate < %s and up.planenddate > %s"
    # Note that the plan in question may have expired, but the user still has the right to send invitations to candidates in order to use the test created during the period of the plan.
    dbcursor.execute(subsuserplansql, (userobj.id, interviewcreateddate, interviewcreateddate))
    allusrplanrecs = dbcursor.fetchall()
    planstartdate = None
    planenddate = None
    if allusrplanrecs.__len__() == 0: # If no userplan is found, we should consider it to be a free plan.
        pass
    else:
        planname = allusrplanrecs[0][0]
        candidatescount = int(allusrplanrecs[0][1])
        planstartdate = allusrplanrecs[0][2]
        planenddate = allusrplanrecs[0][3]
        userplanid = allusrplanrecs[0][4]
    rightnow = datetime.datetime.now()
    if userplanid > 0: # Check to see if the user has extended his plan at present.
        planextsql = "select allowedinvites, periodstart, periodend from Subscription_planextensions where userplan_id=%s and user_id=%s and periodstart <= %s and periodend > %s"
        dbcursor.execute(planextsql, (userplanid, userobj.id, rightnow, rightnow))
        allextensionrecs = dbcursor.fetchall()
        if allextensionrecs.__len__() > 0:
            candidatescount = allextensionrecs[0][0]
            planstartdate = allextensionrecs[0][1]
            planenddate = allextensionrecs[0][2]
    # Now, find how many (distinct) candidates have already been invited to this test. If that count is less than the allowed
    # number of candidates permitted by the subscription plan, then go ahead. Otherwise, return a response with an error message.
    reguserinvitations_sql = "select count(distinct emailaddr) from Tests_interviewcandidates where interview_id=%s and dateadded > %s and dateadded < %s"
    dbcursor.execute(reguserinvitations_sql, (interviewobj.id, planstartdate, planenddate))
    allreguserinvitations = dbcursor.fetchall()
    totalcandidates = 0
    if allreguserinvitations.__len__() > 0:
        totalcandidates += allreguserinvitations[0][0]
    if totalcandidates >= candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
        message = "Error: You have exhausted the allocated quota of invitations for this interview."
        return HttpResponse(message)
    # Now, check if the number of invitations being sent would surpass the quota of allowed invitations for the selected plan.
    postoperationuserinvitations = totalcandidates + emailidslist.__len__()
    if postoperationuserinvitations > candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
        decrementfactor = postoperationuserinvitations - candidatescount
        message = "Error: The number of users being invited will surpass the allowed number of invitations that can be sent under your subscription plan. Please remove %s email addresses and try again."%decrementfactor
        return HttpResponse(message)
    disconnectdb(dbconn, dbcursor)
    # Done checking invitations quota.
    # So emailidslist is not a list of unique emails sent to us for scheduling
    intcandidateobj = InterviewCandidates()
    intcandidateobj.interview = interviewobj
    intcandidateobj.emailaddr = ",".join(emailidslist)
    newsched_datetime = newsched_datetime.replace("T", " ")
    newsched_datetime_parts = newsched_datetime.split(" ")
    if newsched_datetime_parts.__len__() > 1:
        newschedtimeparts = newsched_datetime_parts[1].split(":")
        if newschedtimeparts.__len__() == 2:
            newsched_datetime = newsched_datetime_parts[0] + " " + newsched_datetime_parts[1] + ":00"
    intcandidateobj.scheduledtime = datetime.datetime.strptime(newsched_datetime, "%Y-%m-%d %H:%M:%S")
    interviewlinkid = skillutils.generate_random_string()
    intcandidateobj.interviewlinkid = interviewlinkid
    hashtoken = binascii.hexlify(os.urandom(16))
    intcandidateobj.interviewurl = skillutils.gethosturl(request) + "/" + mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  intcandidateobj.interviewlinkid + "&hash=" + hashtoken + "&attend=" + ",".join(emailidslist)
    try:
        intcandidateobj.save()
    except:
        message = "The interview schedule could not be saved: %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    # Send out the requisite emails.
    if emailidslist.__len__() > 0: # Send an email invitation link to the email addresses.
        interviewlinkurlwithparams = skillutils.gethosturl(request) + "/" + mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  interviewlinkid + "&hash=" + hashtoken + "&attend=" + ",".join(emailidslist)
        message = """<html><head></head><body>Dear Candidate,
                     <br/><br/>
                     This is an invitation to attend an interview titled '%s' with %s on %s hours. Please click on the<br/> 
                     link below to load the interview interface. If it doesn't work, then copy <br/>
                     the link and paste it in your browser's address bar and hit <enter>.<br/><br/>
                     
                     <a href="%s" target='_blank'>%s/%s</a> <br/><br/>
                     
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.
                     <br/><br/>
                     Important Note: Please use Chrome, Firefox or Opera to attend the interview.<br/>
                     Browsers other than these 3 may not support every feature used by the inter-<br/>
                     view application.<br/><br/>

                     Good Luck!<br/>
                     The TestYard Interview Team.</body></html>
    """%(interviewobj.title, userobj.displayname, newsched_datetime, interviewlinkurlwithparams, skillutils.gethosturl(request), mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  interviewlinkid + "&hash=" + hashtoken + "&attend=" + ",".join(emailidslist), skillutils.gethosturl(request), urllib.quote_plus(str(interviewlinkid)))
        subject = "TestYard Interview Invitation"
        fromaddr = userobj.emailid
        # Send email
        for targetemail in emailidslist:
            try:
                #retval = send_mail(subject, message, fromaddr, [targetemail,], False)
                retval = send_emails.delay(subject, message, fromaddr, targetemail, False)
            except:
                if mysettings.DEBUG:
                    print "sendemail failed for %s - %s\n"%(targetemail, sys.exc_info()[1].__str__())
    # Send an email to the interviewers
    interviewlinkurlwithparams = skillutils.gethosturl(request) + "/" + mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  interviewlinkid + "&hash=" + hashtoken
    message = """<html><head></head><body>Dear Interviewer,
                     <br/><br/>
                     You have invited candidates identified by emails %s to interview titled '%s' at %s. Please click on the<br/> 
                     link below to load the interview interface. If it doesn't work, then copy <br/>
                     the link and paste it in your browser's address bar and hit <enter>.<br/><br/>
                     
                     <a href="%s" target='_blank'>%s/%s</a> <br/><br/>
                     
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.
                     <br/><br/>
                     Important Note: Please use Chrome, Firefox or Opera to attend the interview.<br/>
                     Browsers other than these 3 may not support every feature used by the inter-<br/>
                     view application.<br/><br/>

                     Good Luck!<br/>
                     The TestYard Interview Team.</body></html>
         """%(",".join(emailidslist), interviewobj.title, newsched_datetime, interviewlinkurlwithparams, skillutils.gethosturl(request), mysettings.ATTEND_INTERVIEW_URL + "?lid=" +  interviewlinkid + "&hash=" + hashtoken, skillutils.gethosturl(request), urllib.quote_plus(str(interviewlinkid)))
    subject = "TestYard Interview Scheduled"
    fromaddr = userobj.emailid
    try:
        retval = send_emails.delay(subject, message, mysettings.MAILSENDER, fromaddr, False)
    except:
        if mysettings.DEBUG:
            print "sendemail failed for %s - %s\n"%(fromaddr, sys.exc_info()[1].__str__())
    message = "The interview schedule has been successfully saved and the requisite emails are being sent out."
    response = HttpResponse(message)
    return response
        

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def rescheduleinterview(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    interviewid, intcandidateid = -1, -1
    if not request.POST.has_key('interviewid'):
        message = "Error: Required parameter interview Id missing."
        response = HttpResponse(message)
        return response
    interviewid = request.POST['interviewid']
    if interviewid.strip() == "":
        message = "Error: Required parameter interview Id is empty."
        response = HttpResponse(message)
        return response
    interviewobj = None
    try:
        interviewobj = Interview.objects.get(id=int(interviewid))
    except:
        message = "Error retrieving interview object"
        response = HttpResponse(message)
        return response
    # Now we do have the requisite interview object, let's check out if all other values are valid.
    if not request.POST.has_key('intcandidateid'):
        message = "Error: Required parameter interview candidate Id missing."
        response = HttpResponse(message)
        return response
    intcandidateid = request.POST['intcandidateid']
    if intcandidateid.strip() == "":
        message = "Error: Required parameter interview candidate Id is empty."
        response = HttpResponse(message)
        return response
    intcandidateobj = None
    try:
        intcandidateobj = InterviewCandidates.objects.get(id=int(intcandidateid))
        if intcandidateobj.interview.id != int(interviewid):
            message = "The interview object doesn't apply to this given interview candidate"
            response = HttpResponse(message)
            return response
    except:
        message = "Error retrieving interview candidate object"
        response = HttpResponse(message)
        return response
    resched_datetime = ""
    if not request.POST.has_key('resched_datetime'):
        message = "Required parameter schedule missing"
        response = HttpResponse(message)
        return response
    resched_datetime = request.POST['resched_datetime']
    if resched_datetime.strip() == "":
        message = "Required parameter schedule is empty"
        response = HttpResponse(message)
        return response
    datetimepattern = re.compile("\d{4}\-\d{2}\-\d{2}T\d{2}\:\d{2}")
    if not re.search(datetimepattern, resched_datetime):
        message = "Schedule datetime is not in appropriate format."
        response = HttpResponse(message)
        return response
    resched_datetime = resched_datetime.replace("T", " ")
    resched_datetime_parts = resched_datetime.split(" ")
    if resched_datetime_parts.__len__() > 1:
        reschedtimeparts = resched_datetime_parts[1].split(":")
        if reschedtimeparts.__len__() == 2:
            resched_datetime = resched_datetime_parts[0] + " " + resched_datetime_parts[1] + ":00"
        else:
            pass
    intcandidateobj.scheduledtime = datetime.datetime.strptime(resched_datetime, "%Y-%m-%d %H:%M:%S")
    candidateemails = intcandidateobj.emailaddr
    candidateemailslist = candidateemails.split(",")
    interviewlinkid = intcandidateobj.interviewlinkid
    interviewurl = intcandidateobj.interviewurl
    try:
        intcandidateobj.save()
    except:
        message = "Could not save the given schedule: %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    # Send requisite emails to candidates
    if candidateemailslist.__len__() > 0:
        message = """<html><head></head><body><pre>Dear Candidate,
                     <br/><br/>
                     This is an invitation to attend an interview titled '%s' with %s on %s hours. Please click on the link below to load the interview interface. If it doesn't work, then copy the link and paste it in your browser's address bar and hit "Enter".<br/>
                     
                     <a href="%s" target='_blank'>%s</a> <br/>
                     
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.
                     <br/><br/>
                     Important Note: Please use Chrome, Firefox or Opera to attend the interview. Browsers other than these 3 may not support every feature used by the interview application.<br/><br/>

                     Good Luck!<br/>
                     The TestYard Interview Team.</pre></body></html>
    """%(interviewobj.title, userobj.displayname, resched_datetime, interviewurl, interviewurl, skillutils.gethosturl(request), urllib.quote_plus(str(interviewlinkid)))
        subject = "TestYard Interview Invitation"
        fromaddr = userobj.emailid
        # Send email
        for targetemail in candidateemailslist:
            try:
                #retval = send_mail(subject, message, fromaddr, [targetemail,], False)
                retval = send_emails.delay(subject, message, fromaddr, targetemail, False)
            except:
                if mysettings.DEBUG:
                    print "sendemail failed for %s - %s\n"%(targetemail, sys.exc_info()[1].__str__())
    # Send an email to the interviewer too.
    interviewerurl = interviewurl
    try:
        interviewerurl = interviewurl.split("&attend")[0]
    except:
        pass
    message = """<html><head></head><body><pre>Dear Interviewer,
                     <br/><br/>
                     You have invited candidates identified by emails %s to interview titled '%s' at %s. Please click on the link below to load the interview interface. If it doesn't work, then copy the link and paste it in your browser's address bar and hit "Enter".<br/>
                     
                     <a href="%s" target='_blank'>%s</a> <br/>
                     
                     You may add the schedule to your <a href='%s/skillstest/interview/addtocalendar/?inturl=%s'>google calendar</a>.
                     <br/>
                     Important Note: Please use Chrome, Firefox or Opera to attend the interview. Browsers other than these 3 may not support every feature used by the interview application.<br/><br/>

                     Good Luck!<br/>
                     The TestYard Interview Team.</pre></body></html>
         """%(",".join(candidateemailslist), interviewobj.title, resched_datetime, interviewerurl, interviewerurl, skillutils.gethosturl(request), urllib.quote_plus(str(interviewlinkid)))
    subject = "TestYard Interview Scheduled"
    fromaddr = userobj.emailid
    try:
        retval = send_emails.delay(subject, message, mysettings.MAILSENDER, fromaddr, False)
    except:
        if mysettings.DEBUG:
            print "sendemail failed for %s - %s\n"%(fromaddr, sys.exc_info()[1].__str__())
    message = "The interview schedule has been successfully saved and the requisite emails are being sent out."
    response = HttpResponse(message)
    return response
    

@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def interviewchallengestore(request):
    pass


def streaminterview(filename):
    CHUNK = 1024
    wf = wave.open(filename, 'rb')
    p = pyaudio.PyAudio()
    stream = p.open(format=p.get_format_from_width(wf.getsampwidth()), channels=wf.getnchannels(), rate=wf.getframerate(), output=True)
    data = wf.readframes(CHUNK)
    while len(data) > 0:
        stream.write(data)
        data = wf.readframes(CHUNK)
    stream.stop_stream()
    stream.close()
    p.terminate()


def attendinterview(request):
    if request.method != 'GET':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    interviewlinkid = ""
    if not request.GET.has_key('lid'):
        message = error_msg('1173')
        response = HttpResponse(message)
        return response
    interviewlinkid = request.GET['lid']
    hashtoken = request.GET['hash']
    intobjqset = Interview.objects.filter(interviewlinkid=interviewlinkid)    
    if intobjqset.__len__() > 0:
        intobj = intobjqset[0]
    else:
        intobj = None
    try:
        intcandobj = InterviewCandidates.objects.get(interview=intobj)
    except:
        intcandobj = None
    if intcandobj is None:
        try:
            intcandobj = InterviewCandidates.objects.get(interviewlinkid=interviewlinkid)
            intobj = intcandobj.interview
        except:
            intcandobj = None
    curdatetime = datetime.datetime.now()
    scheduledatetime = None
    if intcandobj is not None:
        intcandobj.actualstarttime = datetime.datetime.now()
        scheduledatetime = intcandobj.scheduledtime
        year, month, day, hour, minute, second = scheduledatetime.year, scheduledatetime.month, scheduledatetime.day, scheduledatetime.hour, scheduledatetime.minute, scheduledatetime.second
        scheduledatetime = datetime.datetime(year,month,day, hour, minute, second, 0, pytz.UTC)
        curdatetime = pytz.utc.localize(curdatetime)
        #scheduledatetime_dt = datetime.datetime.strptime(scheduledatetime, "%Y-%m-%d %H:%M:%S")
        #intcandobj.save()
    int_user_dict = {}
    int_user_dict['signal_server'] = mysettings.SIGNAL_SERVER
    if intobj is not None:
        interviewschedulestart = intobj.scheduledtime
        #fp.write(str(time.strptime(interviewschedulestart, "%Y-%m-%d %H:%M:%S")))
        #fp.close()
        if interviewschedulestart is not None:
            yeari, monthi, dayi, houri, minutei, secondi = interviewschedulestart.year, interviewschedulestart.month, interviewschedulestart.day, interviewschedulestart.hour, interviewschedulestart.minute, interviewschedulestart.second
            interviewschedulestart = datetime.datetime(yeari,monthi,dayi, houri, minutei, secondi, 0, pytz.UTC)
            interviewscheduleend = interviewschedulestart
        if interviewschedulestart is not None: # if this is not None
            interviewscheduleend  = interviewschedulestart + datetime.timedelta(0, intobj.maxduration)
            #if curdatetime > interviewschedulestart and curdatetime < interviewscheduleend:
            if curdatetime > interviewschedulestart:
                #tmpl = get_template("tests/interview_candidate_screen.html")
                tmpl = get_template("tests/audiovisual.html")
                int_user_dict['interviewtitle'] = intobj.title
                interviewfilename = intobj.title
                interviewfilename = interviewfilename.replace("-", "_")
                interviewfilename = interviewfilename.replace(" ", "_")
                interviewfilename += "_" + int(time.time()).__str__() + ".mp4"
                int_user_dict['interviewfilename'] = interviewfilename
                try:
                    intcandobj.save()
                except:
                    pass
            else:
                tmpl = get_template("tests/waitscreen.html")
                int_user_dict['interviewtitle'] = intobj.title
                interviewfilename = intobj.title
                interviewfilename = interviewfilename.replace("-", "_")
                interviewfilename = interviewfilename.replace(" ", "_")
                interviewfilename += "_" + int(time.time()).__str__() + ".mp4"
                int_user_dict['interviewfilename'] = interviewfilename
                int_user_dict['scheduledatetime'] = interviewschedulestart
                int_user_dict['email'] = intcandobj.emailaddr
                int_user_dict['hashtoken'] = hashtoken
                int_user_dict['curdatetime'] = curdatetime
                int_user_dict.update(csrf(request))
                cxt = Context(int_user_dict)
                intscreen = tmpl.render(cxt)
                return HttpResponse(intscreen)
    if scheduledatetime is not None and scheduledatetime <= curdatetime:
        #tmpl = get_template("tests/interview_candidate_screen.html")
        tmpl = get_template("tests/audiovisual.html")
        if intobj:
            int_user_dict['interviewtitle'] = intobj.title
            interviewfilename = intobj.title
            interviewfilename = interviewfilename.replace("-", "_")
            interviewfilename = interviewfilename.replace(" ", "_")
            interviewfilename += "_" + int(time.time()).__str__() + ".mp4"
            int_user_dict['interviewfilename'] = interviewfilename
            try:
                intcandobj.save()
            except:
                pass
    else:
        tmpl = get_template("tests/waitscreen.html")
        if intobj is not None:
            int_user_dict['interviewtitle'] = intobj.title
            interviewfilename = intobj.title
            interviewfilename = interviewfilename.replace("-", "_")
            interviewfilename = interviewfilename.replace(" ", "_")
            interviewfilename += "_" + int(time.time()).__str__() + ".mp4"
            int_user_dict['interviewfilename'] = interviewfilename
            int_user_dict['scheduledatetime'] = scheduledatetime
            int_user_dict['email'] = intcandobj.emailaddr
        else:
            int_user_dict['interviewtitle'] = ""
            int_user_dict['scheduledatetime'] = ""
            int_user_dict['email'] = ""
    int_user_dict['hashtoken'] = hashtoken
    int_user_dict['curdatetime'] = curdatetime
    int_user_dict.update(csrf(request))
    cxt = Context(int_user_dict)
    intcandscreen = tmpl.render(cxt)
    return HttpResponse(intcandscreen)



@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def updateinterviewmeta(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    intlinkid = ""
    if request.POST.has_key('interviewlinkid'):
        intlinkid = request.POST['interviewlinkid']
        intobj = Interview.objects.get(interviewlinkid=intlinkid)
        intquestionsqset = InterviewQuestions.objects.filter(interviewlinkid=intlinkid)
        questionscount = list(intquestionsqset).__len__()
        intobj.challengescount = questionscount
        try:
            intobj.save()
        except:
            response = HttpResponse("Could not save interview property - %s"%sys.exc_info()[1].__str__())
            return response
        response = HttpResponse("Updated Interview object successfully")
        return response
    else:
        response = HttpResponse("The request is corrupt. Please try again. If it fails again, please contact the TestYard administrator at '%s'"%mysettings.MAILSENDER)
        return response


@skillutils.is_session_valid
@csrf_protect
def mobile_listtestsandinterviews(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sessid, username = "", ""
    if request.POST.has_key('sessionid'):
        sessid = request.POST['sessionid']
    if request.POST.has_key('username'):
        username = request.POST['username']
    # Check if the sessionid provided actually relates to the given username. if not, send an empty response.
    sessionqset = Session.objects.filter(sessioncode=sessid)
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    if userobj.displayname != username:
        print "Username is not the same as displayname. username: %s, displayname: %s"%(username, userobj.displayname)
        return HttpResponse(message) # Empty message sent.
    jsonResponse = {}
    utestcandidateqset = UserTest.objects.filter(user=userobj).order_by('starttime')
    testcreatorqset = Test.objects.filter(creator=userobj)
    evaluatorqset1 = Evaluator.objects.filter(groupmember1=userobj)
    evaluatorqset2 = Evaluator.objects.filter(groupmember2=userobj)
    evaluatorqset3 = Evaluator.objects.filter(groupmember3=userobj)
    evaluatorqset4 = Evaluator.objects.filter(groupmember4=userobj)
    evaluatorqset5 = Evaluator.objects.filter(groupmember5=userobj)
    evaluatorqset6 = Evaluator.objects.filter(groupmember6=userobj)
    evaluatorqset7 = Evaluator.objects.filter(groupmember7=userobj)
    evaluatorqset8 = Evaluator.objects.filter(groupmember8=userobj)
    evaluatorqset9 = Evaluator.objects.filter(groupmember9=userobj)
    evaluatorqset10 = Evaluator.objects.filter(groupmember10=userobj)
    evaluatorlist = list(chain(evaluatorqset1, evaluatorqset2, evaluatorqset3, evaluatorqset4, evaluatorqset5, evaluatorqset6, evaluatorqset7, evaluatorqset8, evaluatorqset9, evaluatorqset10))
    testevaluatorqlist = []
    for evalobj in evaluatorlist:
        testsqset = Test.objects.filter(evaluator=evalobj)
        for testobj in testsqset:
            testevaluatorqlist.append(testobj)
    interviewqset = InterviewCandidates.objects.filter(emailaddr=userobj.emailid)
    interviewerqset = Interview.objects.filter(interviewer=userobj)
    # Form the json object
    jsonResponse['asCandidate'] = {}
    candidatesdata = {}
    ctr = 0
    currentdatetime = datetime.datetime.now()
    for ctr in range(0, utestcandidateqset.__len__()):
        usertestobj = utestcandidateqset[ctr]
        testname = usertestobj.test.testname
        testscore = usertestobj.score
        outcome = usertestobj.outcome
        testdate = usertestobj.starttime
        testurl = ""
        if not usertestobj.status and usertestobj.active and not usertestobj.cancelled and not usertestobj.disqualified: # and (usertestobj.starttime is None or skillutils.mysqltopythondatetime(usertestobj.starttime) < currentdatetime) and (usertestobj.endtime is None or skillutils.mysqltopythondatetime(usertestobj.endtime) > currentdatetime):
            testurl = usertestobj.testurl
        testdate_str = str(testdate)
        testdate_str_parts = testdate_str.split("+")
        testdate_str = testdate_str_parts[0]
        testdate_serializable = testdate_str
        testtopic = usertestobj.test.topic.topicname
        if not testtopic:
            testtopic = usertestobj.test.topicname
        candidatesdata[testname] = (testscore, outcome, testdate_serializable, testtopic, testurl)
    jsonResponse['asCandidate'] = candidatesdata
    jsonResponse['asCreator'] = {}
    creatordata = {}
    ctr = 0
    for ctr in range(0, testcreatorqset.__len__() - 1):
        testobj = testcreatorqset[ctr]
        testname = testobj.testname
        testtopic = testobj.topic.topicname
        if not testtopic:
            testtopic = testobj.topicname
        takerscount = list(UserTest.objects.filter(test=testobj)).__len__()
        creatordata[testname] = (takerscount, testtopic)
    jsonResponse['asCreator'] = creatordata
    jsonResponse['asEvaluator'] = {}
    evaluatordata = {}
    ctr = 0
    for ctr in range(0, testevaluatorqlist.__len__() - 1):
        testobj = testevaluatorqlist[ctr]
        testname = testobj.testname
        testtopic = testobj.topic.topicname
        if not testtopic:
            testtopic = testobj.topicname
        evaluatordata[testname] = (testtopic,)
    jsonResponse['asEvaluator'] = evaluatordata
    jsonResponse['asInterviewCandidates'] = {}
    interviewcandidates = {}
    for interviewcandiateobj in interviewqset:
        interviewname = interviewcandiateobj.interview.title
        interviewtopic = ""
        interviewconductdate = ""
        try:
            interviewtopic = interviewcandiateobj.interview.topic.topicname
            if not interviewtopic:
                interviewtopic = interviewcandiateobj.interview.topicname
            interviewconductdate = interviewcandiateobj.actualstarttime
        except:
            print "INTERVIEW CANDIDATES: %s"%sys.exc_info()[1].__str__()
        try:
            interviewcandidates[interviewname] = (interviewtopic, interviewconductdate )
        except:
            interviewcandidates[interviewname] = [interviewtopic, interviewconductdate ]
    jsonResponse['asInterviewCandidates'] = interviewcandidates
    jsonResponse['asInterviewer'] = {}
    interviewerdata = {}
    for interviewobj in interviewerqset:
        interviewname = interviewobj.title
        interviewtopic = ""
        interviewer = ""
        try:
            interviewtopic = interviewobj.topic.topicname
            if not interviewtopic:
                interviewtopic = interviewobj.topicname
            interviewer = interviewobj.interviewer.displayname
        except:
            interviewtopic = ""
            print "INTERVIEWERS: %s"%sys.exc_info()[1].__str__()
        interviewerdata[interviewname] = (interviewtopic, interviewer )
    jsonResponse['asInterviewer'] = interviewerdata
    try:
        jsonstr = json.dumps(jsonResponse)
    except:
        print "Could not 'dumps' jsonResponse - Error: %s"%sys.exc_info()[1].__str__()
        jsonstr = "{}"
    respobj = HttpResponse(jsonstr)
    return respobj
 

@skillutils.is_session_valid
@csrf_protect
def mobile_createtest(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        jsondict = {'message' : message, 'testlinkid' : ''}
        jsonstr = json.dumps(jsondict)
        response = HttpResponse(jsonstr)
        return response
    sessid, username = "", ""
    if not request.POST.has_key('data'):
        message = "Error: %s"%error_msg('1174')
        jsondict = {'message' : message, 'testlinkid' : ''}
        jsonstr = json.dumps(jsondict)
        response = HttpResponse(jsonstr)
        return response
    postdata = request.POST['data']
    keystring, ivstring = "test", "test"
    decryptedPostdata = skillutils.des3Decrypt(postdata, keystring, ivstring)
    decodedPostdata = base64.b64decode(postdata)
    keyValuePairs = decodedPostdata.split("&")
    testobj = Test()
    datepattern = re.compile("^\d{4}\-\d{2}\-\d{2}$")
    maxchallengeduration = 0
    evendistributionofscores = False
    sessionid = ""
    username = ""
    userobj = None
    topicobj = None
    topicname = ""
    testobj.status = 0
    for keyval in keyValuePairs:
        attr,value = keyval.split("=")
        if attr != "username":
            continue
        username = value
        userobj = User.objects.get(displayname=username)
        break
    for keyval in keyValuePairs:
        attr, value = keyval.split("=")
        if attr == "testname":
            value = value.replace("+", " ")
            emptyPattern = re.compile("^\s*$")
            if emptyPattern.search(value):
                message = "Error: %s"%error_msg('1177')
                response = HttpResponse(message)
                return response
            # Check if a test with  the same name already exists. If so, return a message saying the same.
            testsqset = Test.objects.filter(testname=value)
            if testsqset.__len__() > 0:
                message = "A test with the same name already exists. Please try with a different test name"
                response = HttpResponse(message)
                return response
            testobj.testname = value
        elif attr == "progenvselected":
            value = value.replace("+", " ")
            testobj.progenv = value
        elif attr == "testscore":
            testobj.maxscore = value
        elif attr == "multimediaAllowedFlag":
            if value == '1':
                testobj.multimediareqd = True
            elif value == '0':
                testobj.multimediareqd = False
        elif attr == "randomSequencedFlag":
            if value == '1':
                testobj.randomsequencing = True
            elif value == '0':
                testobj.randomsequencing = False
        elif attr == "multipleAttemptsFlag":
            if value == '1':
                testobj.allowmultiattempts = True
            elif value == '0':
                testobj.allowmultiattempts = False
        elif attr == "passscore":
            testobj.passscore = value
        elif attr == "testscopeselected":
            testobj.scope = value
        elif attr == "challengescount":
            testobj.challengecount = value
        elif attr == "testruleselected":
            testobj.ruleset = value
        elif attr == "evalgroupname":
            evaluatorobj = None
            if value != "":
                value = value.replace("%40", "@")
                try:
                    evaluatorobj = Evaluator.objects.get(evalgroupname=value)
                except:
                    evaluatorobj = None
                if evaluatorobj:
                    testobj.evaluator = evaluatorobj
                else:
                    evaluatorobj = Evaluator()
                    evaluatorobj.evalgroupname = value
                    testobj.evaluator = None
            else:
                evaluatorobj = Evaluator()
                evaluatorobj.evalgroupname = None
                testobj.evaluator = None
        elif attr == "evalemailids":
            value = value.replace("%40", "@")
            emailidsstr = value
            emailidslist = emailidsstr.split(",")
            emailidctr = 1
            if userobj.emailid not in emailidslist:
                emailidslist.append(userobj.emailid)
            if not evaluatorobj.evalgroupname:
                if emailidslist[0] == "":
                    emailidparts = emailidslist[1].split("@") # The first element is '', so we ignore it.
                else:
                    emailidparts = emailidslist[0].split("@")
                evaluatorobj.evalgroupname = emailidparts[0]
            for emailid in emailidslist:
                if emailid == "":
                    continue
                emailid = emailid.replace("%40", "@")
                if emailidctr == 1:
                    try:
                        evaluatorobj.groupmember1 = User.objects.get(emailid=emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 2:
                    try:
                        evaluatorobj.groupmember2 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 3:
                    try:
                        evaluatorobj.groupmember3 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 4:
                    try:
                        evaluatorobj.groupmember4 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 5:
                    try:
                        evaluatorobj.groupmember5 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 6:
                    try:
                        evaluatorobj.groupmember6 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 7:
                    try:
                        evaluatorobj.groupmember7 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 8:
                    try:
                        evaluatorobj.groupmember8 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 9:
                    try:
                        evaluatorobj.groupmember9 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                elif emailidctr == 10:
                    try:
                        evaluatorobj.groupmember10 = User.objects.get(emailid = emailid)
                    except:
                        message = "Error: %s"%error_msg('1176')
                        jsondict = {'message' : message, 'testlinkid' : ''}
                        jsonstr = json.dumps(jsondict)
                        response = HttpResponse(jsonstr)
                        return response
                emailidctr += 1
            evaluatorobj.save()
            if not testobj.evaluator:
                testobj.evaluator = evaluatorobj
        elif attr == "testactivationdate": # Sanitize the date value first. Should be in yyyy-mm-dd format
            matchobj = datepattern.search(value)
            if not matchobj:
                message = "Error: " + error_msg('1175')
                jsondict = {'message' : message, 'testlinkid' : ''}
                jsonstr = json.dumps(jsondict)
                response = HttpResponse(jsonstr)
                return response
            testobj.activationdate = value
        elif attr == "maxchallengeduration_secs":
            maxchallengeduration = value
        elif attr == "answeringlanguageselected":
            value = value.replace("+", " ")
            if value == "English - US":
                testobj.allowedlanguages = "enus"
            elif value == "English - UK":
                testobj.allowedlanguages = "enuk"
            elif value == "Latin":
                testobj.allowedlanguages = "lat"
            elif value == "French":
                testobj.allowedlanguages = "fr"
            elif value == "Hindi":
                testobj.allowedlanguages = "hndi"
            elif value == "Bengali - WB":
                testobj.allowedlanguages = "bngw"
            elif value == "Bengali - Bangladesh":
                testobj.allowedlanguages = "bnge"
        elif attr == "samescorevalue":
            evendistributionofscores = True
        elif attr == "targetskilllevelselected":
            val = "BEG"
            if value == "Intermediate":
                val = "INT"
            elif value == "Proficient":
                val = "PRO"
            testobj.quality = val
        elif attr == "testduration":
            testobj.duration = int(value) * 60 # The value is in minutes, so we convert it to seconds.
        elif attr == "testtypeselected":
            val = "COMP"
            if value == "Coding": 
                val = "COD"
            elif value == "Fill up the Blanks":
                val = "FILB"
            elif value == "Algorithm":
                val = "ALGO"
            elif value == "Subjective":
                val = "SUBJ"
            elif value == "Multiple Choice":
                val = "MULT"
            testobj.testtype = val
        elif attr == "negativescorevalue":
            if value == "Yes":
                testobj.negativescoreallowed = True
            else:
                testobj.negativescoreallowed = False
        elif attr == "testtopicselected":
            value = value.replace("+", " ")
            testobj.topicname = value
            topicname = value
            topicobj = None
        elif attr == "testpublishdate":
            matchobj = datepattern.search(value)
            if not matchobj:
                message = "Error: " + error_msg('1175')
                jsondict = {'message' : message, 'testlinkid' : ''}
                jsonstr = json.dumps(jsondict)
                response = HttpResponse(jsonstr)
                return response
            testobj.publishdate = value
    topicobj = Topic()
    topicobj.topicname = topicname
    topicobj.user = userobj
    topicobj.isactive = True
    topicobj.save()
    testobj.topic = topicobj
    testobj.creator = userobj
    testobj.testlinkid = skillutils.generate_random_string()
    try:
        testobj.save()
    except:
        message = "Error: Could not create the test - %s\n"%sys.exc_info()[1].__str__()
        jsondict = {'message' : message, 'testlinkid' : ''}
        jsonstr = json.dumps(jsondict)
        response = HttpResponse(jsonstr)
        return response
    message = "Test object created successfully"
    jsondict = {'message' : message, 'testlinkid' : testobj.testlinkid}
    jsonstr = json.dumps(jsondict)
    response = HttpResponse(jsonstr)
    return response


@skillutils.is_session_valid
@csrf_protect
def mobile_addchallenge(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sessid, username = "", ""
    if not request.POST.has_key('data'):
        message = "Error: %s"%error_msg('1174')
        response = HttpResponse(message)
        return response
    postdata = request.POST['data']
    keystring, ivstring = "test", "test"
    decryptedPostdata = skillutils.des3Decrypt(postdata, keystring, ivstring)
    decodedPostdata = base64.b64decode(postdata)
    keyValuePairs = decodedPostdata.split("&")
    challengeStatement, externalResourceUrl, responseLinesCount, challengeScore, negativeScore, maxTimeLimit, challengeQuality, compulsoryChallenge, testname, challengeType, testType, testLinkId, oneOrMoreValues, responsekey = "", "", "", "", "", "", "", "", "", "", "", "", "", ""
    mcOptions = [""] * 8
    optionKeyPattern = re.compile(r"option(\d+)Value")
    print keyValuePairs
    for keyval in keyValuePairs:
        key, val = keyval.split("=")
        val = val.replace("+", " ")
        if key == "challengeStatement":
            val = val.replace("%3F", "?")
            val = val.replace("%2C", ",")
            val = val.replace("%27", "'")
            val = val.replace("%22", '"')
            val = val.replace("%2F", "/")
            challengeStatement = val
        elif key == "externalResourceUrl":
            externalResourceUrl = val
        elif key == "maxResponseLinesCount":
            responseLinesCount = val
        elif key == "challengeScore":
            challengeScore = val
        elif key == "negativeScore":
            negativeScore = val
        elif key == "maxTimeLimit":
            maxTimeLimit = val
        elif key == "challengeQuality":
            challengeQuality = val
        elif key == "compulsoryChallenge":
            compulsoryChallenge = val
            if compulsoryChallenge == "":
                compulsoryChallenge = False
        elif key == "testname":
            testname = val
        elif key == "challengeType":
            challengeType = val
        elif key == "testtypeselected":
            testType = val
        elif key == "testlinkid":
            testLinkId = val
        elif key == "filbResponseStr":
             responsekey = val
        elif key == "oneOrMoreValues":
            if val == "No":
                oneOrMoreValues = False
            else:
                oneOrMoreValues = True
        else:
            if challengeType == "Multiple Choice" or testType == "Multiple Choice":
                optionKeyMatch = optionKeyPattern.search(key)
                if optionKeyMatch:
                    optionKeyId = optionKeyMatch.groups()[0]
                    if val == "":
                        continue
                    mcOptions[int(optionKeyId)] = val
    testobj = None
    if testname != "":
        try:
            testqset = Test.objects.filter(testname=testname)
            testobj = testqset[0]
            if maxTimeLimit == "":
                maxTimeLimit = testobj.duration
        except:
            message = "Could not find the test named '%s' - Error: %s"%(testname, sys.exc_info()[1].__str__())
            response = HttpResponse(message)
            return response
    else:
        message = "The testname parameter was empty in the request."
        response = HttpResponse(message)
        return response
    challengeobj = Challenge()
    challengeobj.test = testobj
    challengeobj.statement = challengeStatement
    if challengeType == "Subjective":
        challengeobj.challengetype = "SUBJ"
    elif challengeType == "Multiple Choice":
        challengeobj.challengetype = "MULT"
    elif challengeType == "Fill up the Blanks":
        challengeobj.challengetype = "FILB"
    elif challengeType == "Algorithm":
        challengeobj.challengetype = "ALGO"
    elif challengeType == "Coding":
        challengeobj.challengetype = "CODN"
    else:
        challengeobj.challengetype = ""
    challengeobj.challengescore = challengeScore
    challengeobj.negativescore = negativeScore
    if compulsoryChallenge == "1":
        challengeobj.mustrespond = True
    else:
        challengeobj.mustrespond = False
    challengeobj.responsekey = ""
    challengeobj.mediafile = ""
    challengeobj.additionalurl = externalResourceUrl
    challengeobj.timeframe = maxTimeLimit
    if challengeQuality == "Intermediate":
        challengeQuality = "INT"
    elif challengeQuality == "Proficient":
        challengeQuality = "PRO"
    elif challengeQuality == "Beginner":
        challengeQuality = "BEG"
    challengeobj.challengequality = challengeQuality
    challengeobj.testlinkid = testobj.testlinkid
    challengeobj.responsekey = responsekey
    challengeobj.oneormore = False
    try:
        challengeobj.save()
    except:
        message = "Could not save the challenge. Please try again. Error: %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    response = HttpResponse("The challenge was successfully saved.")
    return response


@skillutils.is_session_valid
@csrf_protect
def mobile_listcreatortests(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        msgdict = {'message' : message}
        jsonstr = json.dumps(msgdict)
        response = HttpResponse(jsonstr)
        return response
    sessid, username = "", ""
    if request.POST.has_key('sessionid'):
        sessid = request.POST['sessionid']
    if sessid == "":
        message = "Error: %s"%error_msg('1178')
        msgdict = {'message' : message}
        jsonstr = json.dumps(msgdict)
        response = HttpResponse(jsonstr)
        return response
    if request.POST.has_key('username'):
        username = request.POST['username']
    if username == "":
        message = "Error: %s\n"%error_msg('1179')
        msgdict = {'message' : message}
        jsonstr = json.dumps(msgdict)
        response = HttpResponse(jsonstr)
        return response
    userobj = None
    try:
        userobj = User.objects.get(displayname=username)
    except:
        message = error_msg('1180')
        msgdict = {'message' : message}
        jsonstr = json.dumps(msgdict)
        response = HttpResponse(jsonstr)
        return response
    testinfo = {}
    testrecords = Test.objects.filter(creator=userobj)
    for testobj in testrecords:
        testid = testobj.id
        testname = testobj.testname
        testtopic = testobj.topicname
        if testtopic == "":
            testtopic = testobj.topic.topicname
        maxscore = testobj.maxscore
        duration = testobj.duration
        testinfo[testname] = [testtopic, maxscore, duration, testname, testid.__str__()]
    jsonstr = json.dumps(testinfo)
    #print jsonstr
    response = HttpResponse(jsonstr)
    return response


@skillutils.is_session_valid
@csrf_protect
def mobile_setschedule(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    sessid, username = "", ""
    sessid = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sessid)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    if not request.POST.has_key('data'):
        message = "Error: %s"%error_msg('1174')
        response = HttpResponse(message)
        return response
    postdata = request.POST['data']
    keystring, ivstring = "schedule", "schedule"
    decryptedPostdata = skillutils.des3Decrypt(postdata, keystring, ivstring)
    decodedPostdata = base64.b64decode(postdata)
    keyValuePairs = decodedPostdata.split("&")
    emailIds, scheduleDate, scheduleTime, testid, username = "", "", "", "", ""
    for keyval in keyValuePairs:
        key, val = keyval.split("=")
        if key == "emailtext":
            emailIds = val
        elif key == "sched_date":
            scheduleDate = val
        elif key == "sched_time":
            scheduleTime = val
        elif key == "testid":
            testid = val
        elif key == "username":
            username = val
    emailIds = emailIds.replace("%2C", ",")
    emailIds = emailIds.replace("%40", "@")
    emailsList = emailIds.split(",")
    if userobj.displayname != username:
        message = "Error: %s"%error_msg('1181')
        response = HttpResponse(message)
        return response
    testobj = None
    try:
        testobj = Test.objects.get(id=testid)
    except:
        message = error_msg('1058') + ": %s"%sys.exc_info()[1].__str__()
        response = HttpResponse(message)
        return response
    # Check if the user is the creator of this test. Only creators of a test 
    # are allowed to send invitations to candidates (except when an invitation 
    # needs to be sent automatically due to a user's need to join a group).
    if testobj.creator.id != userobj.id:
        message = "Error: " + error_msg('1070')
        response = HttpResponse(message)
        return response
    # Find out when this test was created. Then find the subscription plan that the user was subscribed to during that period, and also the number of candidates permitted by that plan.
    testcreateddate = testobj.createdate
    dbconn, dbcursor = connectdb()
    #dbconn, dbcursor = connectdb_p()
    candidatescount = 5 # Default candidates count for Free Plan.
    plnname = 'Free Plan'
    freeplansql = "select candidates from Subscription_plan where planname='%s'"
    dbcursor.execute(freeplansql, (plnname,))
    allplanrecs = dbcursor.fetchall()
    if allplanrecs.__len__() > 0:
        candidatescount = allplanrecs[0][0]
    else:
        pass
    userplanid = -1
    subsuserplansql = "select pl.planname, pl.candidates, up.planstartdate, up.planenddate, up.id from Subscription_plan pl, Subscription_userplan up where up.user_id=%s and pl.id=up.plan_id and up.planstartdate < %s and up.planenddate > %s"
    # Note that the plan in question may have expired, but the user still has the right to send invitations to candidates in order to use the test created during the period of the plan.
    dbcursor.execute(subsuserplansql, (userobj.id, testcreateddate, testcreateddate))
    allusrplanrecs = dbcursor.fetchall()
    planstartdate = None
    planenddate = None
    if allusrplanrecs.__len__() == 0: # If no userplan is found, we should consider it to be a free plan.
        pass
    else:
        planname = allusrplanrecs[0][0]
        candidatescount = int(allusrplanrecs[0][1])
        planstartdate = allusrplanrecs[0][2]
        planenddate = allusrplanrecs[0][3]
        userplanid = allusrplanrecs[0][4]
    rightnow = datetime.datetime.now()
    if userplanid > 0: # Check to see if the user has extended his plan at present.
        planextsql = "select allowedinvites, periodstart, periodend from Subscription_planextensions where userplan_id=%s and user_id=%s and periodstart <= %s and periodend > %s"
        dbcursor.execute(planextsql, (userplanid, userobj.id, rightnow, rightnow))
        allextensionrecs = dbcursor.fetchall()
        if allextensionrecs.__len__() > 0:
            candidatescount = allextensionrecs[0][0]
            planstartdate = allextensionrecs[0][1]
            planenddate = allextensionrecs[0][2]
    # Now, find how many (distinct) candidates have already been invited to this test. If that count is less than the allowed
    # number of candidates permitted by the subscription plan, then go ahead. Otherwise, return a response with an error message.
    reguserinvitations_sql = "select count(distinct emailaddr) from Tests_usertest where test_id=%s and dateadded > %s and dateadded < %s"
    dbcursor.execute(reguserinvitations_sql, (testobj.id, planstartdate, planenddate))
    allreguserinvitations = dbcursor.fetchall()
    unreguserinvitations_sql = "select count(distinct emailaddr) from Tests_wouldbeusers where test_id=%s and dateadded > %s and dateadded < %s"
    dbcursor.execute(reguserinvitations_sql, (testobj.id, planstartdate, planenddate))
    allunreguserinvitations = dbcursor.fetchall()
    totalcandidates = 0
    if allreguserinvitations.__len__() > 0:
        totalcandidates += allreguserinvitations[0][0]
    if allunreguserinvitations.__len__() > 0:
        totalcandidates += allunreguserinvitations[0][0]
    if totalcandidates >= candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
        message = "Error: You have exhausted the allocated quota of test invitations for this test."
        return HttpResponse(message)
    # Now, check if the number of invitations being sent would surpass the quota of allowed invitations for the selected plan.
    postoperationuserinvitations = totalcandidates + emailslist.__len__()
    if postoperationuserinvitations > candidatescount and mysettings.UNLIMITED_INVITATIONS == False:
        decrementfactor = postoperationuserinvitations - candidatescount
        message = "Error: The number of users being invited will surpass the allowed number of invitations that can be sent under your subscription plan. Please remove %s email addresses and try again."%decrementfactor
        return HttpResponse(message)
    disconnectdb(dbconn, dbcursor)
    # Done checking invitations quota.
    # Now, find the list of all evaluator's emails.
    testevaluator = testobj.evaluator
    testevalemailidlist = []
    try:
        if testevaluator.groupmember1:
            testevalemailidlist.append(testevaluator.groupmember1.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember2:
            testevalemailidlist.append(testevaluator.groupmember2.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember3:
            testevalemailidlist.append(testevaluator.groupmember3.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember4:
            testevalemailidlist.append(testevaluator.groupmember4.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember5:
            testevalemailidlist.append(testevaluator.groupmember5.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember6:
            testevalemailidlist.append(testevaluator.groupmember6.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember7:
            testevalemailidlist.append(testevaluator.groupmember7.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember8:  
            testevalemailidlist.append(testevaluator.groupmember8.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember9:
            testevalemailidlist.append(testevaluator.groupmember9.emailid)
    except:
        pass
    try:
        if testevaluator.groupmember10:
            testevalemailidlist.append(testevaluator.groupmember10.emailid)
    except:
        pass
    testduration = testobj.duration
    # Provide a 3 day validity - this has to be worked on to allow the user to select a period for which this invitation will be valid.
    try:
        scheduleTime = scheduleTime.replace("%3A", ":")
        fromdatetimestr = scheduleDate + " " + scheduleTime + ":00"
        fromdatetimeobj = datetime.datetime.strptime(fromdatetimestr, '%Y-%m-%d %H:%M:%S')
        todatetimeobj = fromdatetimeobj + datetime.timedelta(days=3)
        toyearstr = str(todatetimeobj.year)
    except:
        print "#############################" + sys.exc_info()[1].__str__()
    tomonthstr = str(todatetimeobj.month)
    if tomonthstr.__len__() < 2:
        tomonthstr = "0" + tomonthstr
    todaystr = str(todatetimeobj.day)
    if todaystr.__len__() < 2:
        todaystr = "0" + todaystr
    tohourstr = str(todatetimeobj.hour)
    if tohourstr.__len__() < 2:
        tohourstr = "0" + tohourstr 
    tominutestr = str(todatetimeobj.minute)
    if tominutestr.__len__() < 2:
        tominutestr = "0" + tominutestr
    todatetimestr = toyearstr + "-" + tomonthstr + "-" + todaystr + " " + tohourstr + ":" + tominutestr + ":00"
    timeslot = fromdatetimestr + "#||#" + todatetimestr
    schedule = Schedule()
    schedule.test = testobj
    schedule.slot = timeslot
    schedule.save()
    validfrom, validtill = fromdatetimestr, todatetimestr
    for new_email in emailsList:
        new_email = new_email.strip()
        new_email = new_email.replace("+", "") # The user sometimes places whitespace characters around email Id. 
        # That gets encoded as '+'. We need to remove those.
        # If this email belongs to the creator or one of the evaluators, skip it.
        if new_email == testobj.creator.emailid or new_email in testevalemailidlist:
            continue
        # Is the user registered with testyard?
        uobj = None
        try:
            uobj = User.objects.get(emailid=new_email)
        except:
            pass
        utobj = None
        if uobj is not None: # user is registered
            utobj = UserTest()
            utobj.user = uobj
        else:
            utobj = WouldbeUsers()
        utobj.emailaddr = new_email
        utobj.test = testobj
        utobj.clientsware = "Android Browser App"
        try:
            utobj.validfrom = fromdatetimeobj
            utobj.validtill = todatetimeobj
        except:
            print sys.exc_info()[1].__str__()
        utobj.status = 0
        utobj.schedule = schedule
        baseurl = skillutils.gethosturl(request)
        (utobj.testurl, utobj.stringid) = gettesturlforuser(utobj.emailaddr, testid, baseurl)
        error_emails_list = []
        candidatename = "candidate"
        emailsubject = "A test has been scheduled for you on testyard"
        emailmessage = """Dear %s,<br/><br/>

         A test with the name '%s' has been scheduled for you by <i>%s</i>. <br/><br/>

             """%(candidatename, testobj.testname, userobj.displayname)
        emailmessage += """The test will start from %s IST and end at %s IST, """%(validfrom, validtill)
        emailmessage += """and hence you are kindly requested to take the test<br/>
            within that interval. You would be able to access the test by clicking
            on the following link: <a href='%s' target=_blank>%s</a>.<br/><br/>
            You may add this test to your <a href='%s/skillstest/test/addtocalendar/?turl=%s'>google calendar</a>.
	<br/><br/>
            If clicking on the above link doesn't work for you, please copy it and 
            paste it in your browser's address bar and hit enter. Do please feel<br/>
            free to let us know in case of any issues. We would do our best to
            resolve it at the earliest.<br/><br/>
            
            We wish you all the best for the test.<br/><br/>
            
            Regards,<br/>
            The TestYard Team.
            """%(utobj.testurl, utobj.testurl, baseurl, urllib.quote_plus(utobj.testurl))
        fromaddr = "testyardteam@testyard.com"
        retval = 0
        try:
            #retval = send_mail(emailsubject, emailmessage, fromaddr, [new_email,], False)
            retval = send_emails.delay(emailsubject, emailmessage, fromaddr, new_email, False)
            utobj.save()
        except:
            if mysettings.DEBUG:
                print "Error: sendemail failed for %s - %s\n"%(new_email, sys.exc_info()[1].__str__())
            message = "Error: sendemail failed for %s - %s\n"%(new_email, sys.exc_info()[1].__str__())
            error_emails_list.append(new_email)
            continue # Continue processing the rest of the emails in the list.
    if retval == 0:
        message = "Success! All candidates except the following ones have been queued for email with the link: %s"%(", ".join(error_emails_list))
    else:
        message = "Success! All candidates have been queued for email with the link."
    # Dump all emails Ids to which email could not be sent
    for error_email in error_emails_list:
        emailfail = EmailFailure()
        emailfail.user = userobj
        emailfail.sessionid = sessid
        emailfail.failedemailid = error_email
        emailfail.script = 'Tests.views.mobile_setschedule'
        emailfail.failurereason = sys.exc_info()[1].__str__()
        emailfail.tryagain = 1
        try:
            emailfail.save()
        except:
            message = "Error: %s"%sys.exc_info()[1].__str__()
            print message
    # In mobile we do not provide for the user to edit existing schedules. That can only be done from the web interface.
    response = HttpResponse(message)
    return response


def linkedinredirect(request):
    code = request.GET.get('code', '')
    state = request.GET.get('state', '')
    sesscode = request.GET.get('sesscode', '')
    sid = sesscode.replace('"', '')
    rolename = request.GET.get('role', '')
    postlinkedinobj = None
    #First check if the CSRF token received matches our csrftoken in the DB for this session (identified by the sesscode variable)
    try:
        postlinkedinobj = PostLinkedin.objects.filter(sessionid=sid).order_by("-current_ts")[0]
    except:
        message = error_msg('1183') + " - Error: %s\n"%sys.exc_info()[1].__str__()
        return HttpResponse(message)
    if postlinkedinobj.sessionid != sid:
        message = error_msg('1184')
        return HttpResponse(message)
    """
    # Check if the state value equals our stored access token
    if postlinkedinobj.csrftoken != state:
        message = "The response could not be authenticated due to the difference in CSRF token values. Please contact support@testyard.in for resolving this."
        return HttpResponse(message, status=401)
    """
    if not rolename:
        rolename = postlinkedinobj.role
    # Now, we will try to retrieve our authorization code.
    httpHeaders = {}
    for k in skillutils.gHttpHeaders.keys():
        httpHeaders[k] = skillutils.gHttpHeaders[k]
    getdatadict = {"response_type" : "code", "client_id" : mysettings.OAUTH_API_KEY, "redirect_uri" : mysettings.REDIRECT_URI, "scope" : "r_liteprofile w_member_social", "state" : ""} 
    getdata = urllib.urlencode(getdatadict)
    httpHeaders['Host'] = "www.linkedin.com"
    httpHeaders['Content-Type'] = "application/x-www-form-urlencoded"
    authrequest = urllib2.Request("https://www.linkedin.com/oauth/v2/authorization?" + getdata, None, httpHeaders)
    opener = urllib2.build_opener(urllib2.HTTPHandler(), urllib2.HTTPSHandler(), skillutils.NoRedirectHandler())
    authresponse = None
    try:
        authresponse = opener.open(authrequest)
        respHeaders = authresponse.info()
    except:
        print "Could not make HTTP authorization request: %s"%sys.exc_info()[1].__str__()
        return HttpResponse("Could not make HTTP authorization request: %s"%sys.exc_info()[1].__str__())
    state,code = "", ""
    ff0 = open("/home/supriyo/work/testyard/tmpfiles/linkedinresponse_0.txt", "w")
    while respHeaders.has_key('Location') or respHeaders.has_key('location'): 
        if respHeaders.has_key('Location'):   
            redirectUri = respHeaders['Location']
        else:
            redirectUri = respHeaders['location']
        ff0.write("REDIRECTURI = " + redirectUri + "\n\n")
        httpHeaders['Referer'] = "https://www.linkedin.com/oauth/v2/authorization"
        authrequest = urllib2.Request(redirectUri, None, httpHeaders)
        try:
            authresponse = opener.open(authrequest)
            respHeaders = authresponse.info()
        except:
            print "Could not make HTTP authorization request: %s"%sys.exc_info()[1].__str__()
            return HttpResponse("Could not make HTTP authorization request: %s"%sys.exc_info()[1].__str__())
    requrl = authresponse.url
    uriparts = requrl.split("?")
    if uriparts.__len__() > 1:
        datachunks = uriparts[1].split("&")
        for chunk in datachunks:
            chunkparts = chunk.split("=")
            if chunkparts[0] == "code":
                code = chunkparts[1]
            elif chunkparts[0] == "state":
                state = chunkparts[1]
            else:
                pass
    
    ff0.write(getdata + "\n\n code = %s\n================= \n\nredirectUri = %s\n\n"%(code, requrl))
    ff0.close()
    postdatadict = {"grant_type" : "client_credentials", "code" : code, "redirect_uri" : mysettings.REDIRECT_URI, "client_id" : mysettings.OAUTH_API_KEY, "client_secret" : mysettings.OAUTH_SECRET_KEY} 
    postdata = urllib.urlencode(postdatadict)
    httpHeaders['Host'] = "www.linkedin.com"
    httpHeaders['Content-Type'] = "application/x-www-form-urlencoded"
    tokenrequest = urllib2.Request("https://www.linkedin.com/oauth/v2/accessToken", postdata, httpHeaders)
    ff1 = open("/home/supriyo/work/testyard/tmpfiles/linkedinresponse_1.txt", "w")
    ff1.write(postdata.__str__())
    ff1.write("\n\n#################################\n\n")
    ff1.write(httpHeaders.__str__())
    ff1.close()
    tokenResponse = None
    try:
        tokenResponse = opener.open(tokenrequest)
        respHeaders = tokenResponse.info()
    except:
        print "Could not make HTTP request to acquire authorization code: %s"%sys.exc_info()[1].__str__()
        return HttpResponse("Could not make HTTP request to acquire authorization code: %s"%sys.exc_info()[1].__str__())
    tokenResponseContent = skillutils._decodeGzippedContent(tokenResponse.read())
    urn = ""
    responsejson = json.loads(tokenResponseContent)
    accesstoken = responsejson['access_token']
    
    # Make a GET request to /v2/me to get the URN value.
    httpHeaders = {}
    httpHeaders['Content-Type'] = "application/json"
    httpHeaders['x-li-format'] = "json"
    httpHeaders['X-Restli-Protocol-Version'] = "2.0.0"
    httpHeaders['Authorization'] = "Bearer " + accesstoken
    httpHeaders['X-Target-URI'] = "https://api.linkedin.com"
    httpHeaders['Host'] = "api.linkedin.com"
    httpHeaders['Connection'] = "Keep-Alive"
    urnurl = "https://api.linkedin.com/v2/me"
    urnrequest = urllib2.Request(urnurl, None, httpHeaders)
    try:
        urnresponse = opener.open(urnrequest)
    except:
        print "Could not successfully make the request to /v2/me -Error: %s"%sys.exc_info()[1].__str__()
        return HttpResponse("Could not successfully make the request to /v2/me -Error: %s"%sys.exc_info()[1].__str__())
    urncontent = skillutils._decodeGzippedContent(urnresponse.read())
    urnjson = json.loads(urncontent)
    
    # Otherwise, now we are all poised to post the message on linkedin.
    #shareurl = "https://api.linkedin.com/v1/people/~/shares?format=json"
    shareurl = "https://api.linkedin.com/v2/ugcPosts"
    #sharedata = { "content" : {"title" : "", "description" : "", 'submitted-url' : "https://testyard.in/", 'submitted-image-url' : "https://testyard.in/"}, "comment" : "", "visibility" : { "code" : "anyone" }}
    sharedata = { "author": "", "lifecycleState": "PUBLISHED", "specificContent": { "com.linkedin.ugc.ShareContent": { "shareCommentary": { "text": "" }, "shareMediaCategory": "NONE" } }, "visibility": { "com.linkedin.ugc.MemberNetworkVisibility": "CONNECTIONS" } }
    sharedata["author"] = "urn:li:person:" + urn
    if rolename == "creator":
        objectname = ""
        if postlinkedinobj.test:
            objectname = postlinkedinobj.test.testname
            topicname = ""
            if postlinkedinobj.test.topic:
                topicname = postlinkedinobj.test.topic.topicname
            else:
                topicname = postlinkedinobj.test.topicname
            objecttopic = postlinkedinobj.test.topic
        """
        sharedata["comment"] = "Created a test named '%s' under '%s' topic"%(objectname, topicname)
        sharedata["content"]["title"] = "Test Created"
        sharedata["content"]["description"] = "Test Created"
        """
        sharedata["specificContent"]["com.linkedin.ugc.ShareContent"]["shareCommentary"]["text"] = "Created a test named '%s' under '%s' topic"%(objectname, topicname)
    elif rolename == "evaluator":
        if postlinkedinobj.test:
            objectname = postlinkedinobj.test.testname
            topicname = ""
            if postlinkedinobj.test.topic:
                topicname = postlinkedinobj.test.topic.topicname
            else:
                topicname = postlinkedinobj.test.topicname
        """
        sharedata["comment"] = "Chosen as one of the evaluators for '%s' under '%s' topic."%(objectname, topicname)
        sharedata["content"]["title"] = "Evaluator Selected"
        sharedata["content"]["description"] = "Evaluator Selected"
        """
        sharedata["specificContent"]["com.linkedin.ugc.ShareContent"]["shareCommentary"]["text"] = "Chosen as one of the evaluators for '%s' under '%s' topic."%(objectname, topicname)
    elif rolename == "candidate":
        if postlinkedinobj.test:
            objectname = postlinkedinobj.test.testname
            topicname = ""
            if postlinkedinobj.test.topic:
                topicname = postlinkedinobj.test.topic.topicname
            else:
                topicname = postlinkedinobj.test.topicname
        """
        sharedata["comment"] = "A request to take the test named '%s' under '%s' topic has been made to you."%(objectname, topicname)
        sharedata["content"]["title"] = "Test Request"
        sharedata["content"]["description"] = "Test Request"
        """
        sharedata["specificContent"]["com.linkedin.ugc.ShareContent"]["shareCommentary"]["text"] = "A request to take the test named '%s' under '%s' topic has been made to you."%(objectname, topicname)
    elif rolename == "interviewconductor":
        if postlinkedinobj.interview:
            objectname = postlinkedinobj.interview.title
            topicname = ""
            try:
                if postlinkedinobj.interview.topic:
                    topicname = postlinkedinobj.interview.topic.topicname
                else:
                    topicname = postlinkedinobj.interview.topicname
            except:
                pass
        """
        sharedata["comment"] = "You have been requested to conduct the interview titled '%s' under '%s' topic."%(objectname, topicname)
        sharedata["content"]["title"] = "Conduct Interview Request"
        sharedata["content"]["description"] = "Conduct Interview Request"
        """
        sharedata["specificContent"]["com.linkedin.ugc.ShareContent"]["shareCommentary"]["text"] = "You have been requested to conduct the interview titled '%s' under '%s' topic."%(objectname, topicname)
    elif rolename == "interviewattended":
        if postlinkedinobj.interview:
            objectname = postlinkedinobj.interview.title
            topicname = ""
            try:
                if postlinkedinobj.interview.topic:
                    topicname = postlinkedinobj.interview.topic.topicname
                else:
                    topicname = postlinkedinobj.interview.topicname
            except:
                pass
        """
        sharedata["comment"] = "You have been requested to attend the interview titled '%s' under '%s' topic."%(objectname, topicname)
        sharedata["content"]["title"] = "Attend Interview"
        sharedata["content"]["description"] = "Attend Interview"
        """
        sharedata["specificContent"]["com.linkedin.ugc.ShareContent"]["shareCommentary"]["text"] = "You have been requested to attend the interview titled '%s' under '%s' topic."%(objectname, topicname)
    else:
        sharedata["specificContent"]["com.linkedin.ugc.ShareContent"]["shareCommentary"]["text"] = "Unrecognized rolename"
        """
        sharedata["comment"] = "Unrecognized rolename."
        sharedata["content"]["title"] = "Unrecognized rolename"
        sharedata["content"]["description"] = "Unrecognized rolename"
        """
    httpHeaders = {}
    httpHeaders['Content-Type'] = "application/json"
    httpHeaders['x-li-format'] = "json"
    httpHeaders['X-Restli-Protocol-Version'] = "2.0.0"
    httpHeaders['Authorization'] = "Bearer " + accesstoken
    httpHeaders['X-Target-URI'] = "https://api.linkedin.com"
    httpHeaders['Host'] = "api.linkedin.com"
    httpHeaders['Connection'] = "Keep-Alive"
    postdata = json.dumps(sharedata)
    httpHeaders['Content-Length'] = str(postdata.__len__())
    params = {}
    kw = dict(data=postdata, params=params, headers=httpHeaders, timeout=60)
    response = requests.request("POST", shareurl, **kw)
    jsonresponse = response.json()
    if response.status_code == 201:
        message = "<table border='0' cellspacing='1' cellpadding='4' bgcolor='#E1F6FF'><tr><td><font color='#0000AA'>The post was successfully shared on linkedin. Please click on the following button to close this window.</font></td></tr><tr><td><input type='button' name='btnclose' value='Close Me' onClick='javascript:self.close();'></td></tr></table>"
    else:
        message = "<table border='0' cellspacing='1' cellpadding='4' bgcolor='#E1F6FF'><tr><td><font color='#0000AA'>The post could not be shared on linkedin. Reason: %s. Please click on the following button to close this window.</font></td></tr><tr><td><input type='button' name='btnclose' value='Close Me' onClick='javascript:self.close();'></td></tr></table>"%jsonresponse['message']
    return HttpResponse(message)


@skillutils.is_session_valid
@csrf_protect
def linkedinpostinform(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    sessid, username = "", ""
    sessid = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sessid)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    postlinkedinrec = PostLinkedin()
    #"objectid=" + objectid + "&type=" + type + "&role=" + role + "&csrftoken=" + state + "&currentts=" + new Date().getTime();
    objectid = request.POST.get('objectid', '')
    objecttype = request.POST.get('type', '')
    role = request.POST.get('role', '')
    csrftoken = request.POST.get('csrfmiddlewaretoken', '')
    currentts = request.POST.get('currentts', '')
    sesscode = request.POST.get('sessionid', '')
    try:
        if objecttype == "test":
            testobj = Test.objects.get(id=objectid)
            postlinkedinrec.test = testobj
        elif objecttype == "interview":
            objectidparts = objectid.split("/")
            intlinkid = objectidparts[len(objectidparts) - 2]
            interviewobj = Interview.objects.get(interviewlinkid=intlinkid)
            postlinkedinrec.interview = interviewobj
        postlinkedinrec.role = role
        postlinkedinrec.csrftoken = csrftoken
        postlinkedinrec.current_ts = currentts
        postlinkedinrec.user = userobj
        #sesscode = sesscode.replace('"', '')
        postlinkedinrec.sessionid = sessid
        postlinkedinrec.save()
        msg = error_msg('1182')
        response = HttpResponse(msg)
    except:
        response = HttpResponse(sys.exc_info()[1].__str__())
    return response
    

@skillutils.is_session_valid
@csrf_protect
def convertspeechtotext(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    sessid, username = "", ""
    sessid = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sessid)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    r = sr.Recognizer()
    

def reportwindowchange(request):
    """
    This method sets the window change attempts field in usertest or wouldbeusers tables to the
    number of times the test taker has tried to go to another window from the test window. Attempts
    to go to another tab/window might mean that the test taker was trying to employ unfair means
    of passing the test.
    """
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    tabref, tabid, reportincident, csrftoken = "", 0, 0, ""
    tabref = request.POST.get("tabref")
    tabid = request.POST.get("tabid")
    reportincident = request.POST.get("reportincident", 0)
    csrftoken = request.POST.get("csrfmiddlewaretoken", "");
    if not tabref or not tabid or not reportincident or not csrftoken:
        return HttpResponse("")
    utobjqset = UserTest.objects.filter(id=int(tabid))
    if tabref == "wouldbeusers":
        utobjqset = WouldbeUsers.objects.filter(id=int(tabid))
    if list(utobjqset).__len__() < 1:
        return HttpResponse("")
    utobj = utobjqset[0]
    changeattempts = utobj.windowchangeattempts
    changeattempts = int(changeattempts) + 1
    utobj.windowchangeattempts = changeattempts
    utobj.save()
    return HttpResponse(str(changeattempts))


def chkinternetconnection(request):
    if request.method == 'GET':
        return HttpResponse("1")
    else:
        return HttpResponse("0")


def gettimeremaining(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponse(message)
        return response
    tabref, tabid, csrftoken = "", 0, ""
    tabref = request.POST.get("tabref")
    tabid = request.POST.get("tabid")
    csrftoken = request.POST.get("csrfmiddlewaretoken", "");
    if not tabref or not tabid or not csrftoken:
        return HttpResponse("")
    utobjqset = UserTest.objects.filter(id=int(tabid))
    if tabref == "wouldbeusers":
        utobjqset = WouldbeUsers.objects.filter(id=int(tabid))
    if list(utobjqset).__len__() < 1:
        return HttpResponse("")
    utobj = utobjqset[0]
    # Get the test object and find the duration of the test.
    testobj = utobj.test
    duration = testobj.duration
    # Find the start time of the usertest/wouldbeuser object.
    starttime = utobj.starttime
    tz_info = starttime.tzinfo
    currenttime = datetime.datetime.now(tz_info)
    # Get time elapsed
    timeelapsed = currenttime - starttime
    timeelapsedseconds = timeelapsed.total_seconds()
    timeleft = duration - timeelapsedseconds
    return HttpResponse(str(timeleft))


@skillutils.is_session_valid
@csrf_protect
def mobile_searchtest(request):
    """
    This function searches for tests created by the logged in user to have them displayed so that
    she may schedule it for candidates.
    """
    message = ""
    if request.method != 'POST':
        msgdict = {"Error" : "%s"%error_msg('1004')}
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    sessid = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sessid)
    if not sessionqset or sessionqset.__len__() == 0:
        msgdict = {"Error" : "%s"%error_msg('1008')}
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    if not request.POST.has_key('testname'):
        msgdict = {"Error" : "Required parameter 'testname' is missing. Can't proceed further."}
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    testname = request.POST['testname']
    if testname == "":
        alltestsbyuser = Tests.objects.filter(creator=userobj)
    else:
        alltestsbyuser = Tests.objects.filter(creator=userobj, testname__icontains=testname)
    matchedtests = {}
    for testobj in alltestsbyuser:
        tname = testobj.testname
        tid = testobj.id
        matchedtests[tname] = tid
    jsonstr = json.dumps(matchedtests)
    return HttpResponse(jsonstr)


@skillutils.is_session_valid
@csrf_protect
def mobile_pendingtests(request):
    """
    List of tests pending to be taken by the user.
    Called from "TakeTestActivity.java".
    """
    pass


@skillutils.is_session_valid
@csrf_protect
def mobile_opentest(request):
    """
    Open the selected test pending to be taken by the user.
    Called from "TakeTestActivity.java".
    """
    pass


@skillutils.is_session_valid
@csrf_protect
def mobile_createinterview(request):
    """
    Create an interview using mobile handset.
    Called from "ConductInterviewActivity.java"
    """
    pass


@skillutils.is_session_valid
@csrf_protect
def mobile_listinterviews(request):
    """
    List out the scheduled interviews (in future) for the given user.
    Called from "AttendInterviewActivity.java"
    """
    message = ""
    if request.method != 'POST':
        msgdict = {"Error" : "%s"%error_msg('1004')}
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    sessid = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sessid)
    if not sessionqset or sessionqset.__len__() == 0:
        msgdict = {"Error" : "%s"%error_msg('1008')}
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    # List out the list of scheduled interviews for this user.
    curdatetime = datetime.datetime.now()
    interviewqset = InterviewCandidates.objects.filter(emailaddr=userobj.emailid, scheduledtime__gte=curdatetime).order_by('-scheduledtime')
    datadict = {}
    for interviewobj in interviewqset:
        interviewname = interviewobj.interview.title
        interviewurl = interviewobj.interviewurl
        scheduledtime = interviewobj.scheduledtime
        if not datadict.has_key(interviewname):
            datadict[interviewname] = [interviewurl, scheduledtime]
    jsonstr = json.dumps(datadict)
    return HttpResponse(jsonstr)


@skillutils.is_session_valid
@csrf_protect
def latexkeyboard(request):
    message = ""
    if request.method == 'GET':
        # Load the latex keyboard
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    sessid = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sessid)
    if not sessionqset or sessionqset.__len__() == 0:
        msgdict = {"Error" : "%s"%error_msg('1008')}
        message = json.dumps(msgdict)
        response = HttpResponse(message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    curdatetime = datetime.datetime.now()
    if request.method == 'POST':
        pass
    # Implement accessing the answer data and putting it in text files.


def addtogooglecalendar(request):
    message = ""
    if request.method == 'GET':
        testurl = ""
        context = {}
        if 'turl' in request.GET.keys():
            testurl = urllib.unquote_plus(request.GET['turl'])
        if testurl == "":
            msgdict = {"Error" : "Nothing to add to calendar"}
            tmpl = get_template("tests/gcalendar.html")
            msgdict.update(csrf(request))
            cxt = Context(msgdict)
            gcalendar = tmpl.render(cxt)
            return HttpResponse(gcalendar)
        utqset = None
        try:
            utqset = UserTest.objects.filter(testurl=testurl)
        except:
            pass
        if utqset.__len__() == 0:
            utqset = WouldbeUsers.objects.filter(testurl=testurl)
        if utqset is not None and utqset.__len__() > 0:
            utobj = utqset[0]
            testname = utobj.test.testname
            validfrom = utobj.validfrom
            validto = utobj.validfrom + datetime.timedelta(seconds=utobj.test.duration)
            context = {'testurl' : testurl, 'calendarurl' : skillutils.gethosturl(request) + '/skillstest/test/addtocalendar/', 'clientid' : mysettings.GOOGLE_CALENDAR_CLIENT_ID, 'apikey' : mysettings.GOOGLE_CALENDAR_API_KEY, 'testname' : testname, 'start' : validfrom.isoformat(), 'end' : validto.isoformat()}
        else:
            msgdict = {"Error" : "Nothing to add to calendar"}
            tmpl = get_template("tests/gcalendar.html")
            msgdict.update(csrf(request))
            cxt = Context(msgdict)
            gcalendar = tmpl.render(cxt)
            return HttpResponse(gcalendar)
        tmpl = get_template("tests/gcalendar.html")
        context.update(csrf(request))
        cxt = Context(context)
        gcalendar = tmpl.render(cxt)
        return HttpResponse(gcalendar)
    else: # Weed out all other call methods
        msgdict = {"Error" : "Invalid method of call"}
        tmpl = get_template("tests/gcalendar.html")
        msgdict.update(csrf(request))
        cxt = Context(msgdict)
        gcalendar = tmpl.render(cxt)
        return HttpResponse(gcalendar)



def addinterviewtogooglecalendar(request):
    message = ""
    if request.method == 'GET':
        interviewurl = ""
        context = {}
        if 'inturl' in request.GET.keys():
            interviewurl = urllib.unquote_plus(request.GET['inturl']) # Actually, this would be the interviewlinkid
        if interviewurl == "":
            msgdict = {"Error" : "Nothing to add to calendar"}
            tmpl = get_template("tests/gcalendar.html")
            msgdict.update(csrf(request))
            cxt = Context(msgdict)
            gcalendar = tmpl.render(cxt)
            return HttpResponse(gcalendar)
        intcandidateqset = InterviewCandidates.objects.filter(interviewlinkid=interviewurl)
        if not intcandidateqset or intcandidateqset.__len__() == 0:
            msgdict = {"Error" : "Nothing to add to calendar"}
            message = json.dumps(msgdict)
            response = HttpResponse(message)
            return response
        interviewname = intcandidateqset[0].interview.title
        validfrom = intcandidateqset[0].scheduledtime
        validto = validfrom + datetime.timedelta(hours=1)
        inturl = intcandidateqset[0].interviewurl
        context = {'testurl' : inturl, 'clientid' : mysettings.GOOGLE_CALENDAR_CLIENT_ID, 'apikey' : mysettings.GOOGLE_CALENDAR_API_KEY, 'testname' : interviewname, 'start' : validfrom.isoformat(), 'end' : validto.isoformat(), 'calendarurl' : skillutils.gethosturl(request) + '/skillstest/interview/addtocalendar/',}
        tmpl = get_template("tests/gcalendar.html")
        context.update(csrf(request))
        cxt = Context(context)
        gcalendar = tmpl.render(cxt)
        return HttpResponse(gcalendar)
    else:
        msgdict = {"Error" : "Invalid method of call"}
        tmpl = get_template("tests/gcalendar.html")
        msgdict.update(csrf(request))
        cxt = Context(msgdict)
        gcalendar = tmpl.render(cxt)
        return HttpResponse(gcalendar)


@skillutils.is_session_valid
@skillutils.session_location_match
@csrf_protect
def downloadmydata(request):
    message = ""
    if request.method != 'POST':
        message = "Error: %s"%error_msg('1004')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.PROFILE_URL + "?msg=%s"%message)
        return response
    sesscode = request.COOKIES['sessioncode']
    usertype = request.COOKIES['usertype']
    sessionqset = Session.objects.filter(sessioncode=sesscode)
    if not sessionqset or sessionqset.__len__() == 0:
        message = "Error: %s"%error_msg('1008')
        response = HttpResponseBadRequest(skillutils.gethosturl(request) + "/" + mysettings.CREATE_INTERVIEW_URL + "?msg=%s"%message)
        return response
    sessionobj = sessionqset[0]
    userobj = sessionobj.user
    fmt = "xml"
    if 'format' in request.POST.keys():
        fmt = request.POST['format']
    if fmt != "xml" and fmt != "csv":
        message = "Error: Received invalid format specification."
        response = HttpResponse(message)
        return response
    # First, check for user's subscription plan; If user is a "Business Plan" user, 
    # then we will need to provide a test and interview data download link.
    # NOTE: If the user is on a Business Plan extension, she/he/other should not
    # be able to download the data. This is only for active Business Plan users.
    busplanobj = None
    busplanflag = False
    try:
        busplanobj = Plan.objects.get(planname="Business Plan")
    except:
        pass
    if busplanobj is not None:
        userplanqset = UserPlan.objects.filter(plan=busplanobj, user=userobj, planstatus=True).order_by('-subscribedon')
        curdate = datetime.datetime.now()
        for userplanobj in userplanqset:
            if curdate >= userplanobj.planstartdate.replace(tzinfo=None) and curdate <  userplanobj.planenddate.replace(tzinfo=None):
                busplanflag = True
                break
    if busplanflag is False:
        message = "Error: User '%s' is not an 'Active' Business Plan user. This facility is available to 'Active' Business Plan users only."%userobj.displayname
        response = HttpResponse(message)
        return response
    # Get all tests created by this user, tests taken by this user, and Interviews created and attended by this user.
    dbconn, dbcursor = skillutils.connectdb()
    testrecords_ascreator = []
    testrecords_ascandidate = []
    interviewrecords_ascreator = []
    interviewrecords_ascandidate = []
    testsdict_ascreator, testsdict_ascandidate, interviewsdict_ascreator, interviewsdict_ascandidate = {}, {}, {}, {}
    # Get tests created by this user:
    mytestssql_creator = "select t.id, t.testname, t.topicname, tt.topicname, t.creatorisevaluator, e.evalgroupname, t.testtype, t.createdate, t.maxscore, t.passscore, t.ruleset, t.duration, t.allowedlanguages, t.challengecount, t.activationdate, t.publishdate, t.status, t.allowmultiattempts, t.maxattemptscount, t.attemptsinterval, t.attemptsintervalunit, t.randomsequencing, t.multimediareqd, t.testlinkid, t.progenv, t.scope, t.quality, t.negativescoreallowed, c.id, c.statement, c.challengetype, c.maxresponsesizeallowable, c.option1, c.option2, c.option3, c.option4, c.option5, c.option6, c.option7, c.option8, c.challengescore, c.negativescore, c.mustrespond, c.responsekey, c.mediafile, c.additionalurl, c.timeframe, c.subtopic_id, c.challengequality, c.testlinkid, c.oneormore, c.proglang, c.mathenv from Tests_test t, Tests_challenge c, Tests_evaluator e, Tests_topic tt where t.id=c.test_id and t.evaluator_id=e.id and t.topic_id=tt.id and t.creator_id=%s"
    try:
        dbcursor.execute(mytestssql_creator, (userobj.id,))
        testrecords_ascreator = dbcursor.fetchall()
    except:
        skillutils.disconnectdb(dbconn, dbcursor)
        message = "Error4: Something went wrong - %s"%sys.exc_info()[1].__str__() # TODO: Need to create a log of these messages and send them back with the final response.
        #response = HttpResponse(message)
        #return response
        testrecords_ascreator = []
    for testrecord in  testrecords_ascreator:
        testid = str(testrecord[0])
        testname = str(testrecord[1])
        test_topicname = str(testrecord[2])
        topic_topicname = str(testrecord[3])
        creatorisevaluator = str(testrecord[4])
        evalgroupname = str(testrecord[5])
        testtype = str(testrecord[6])
        if testrecord[7] is not None:
            test_createdate = testrecord[7].strftime("%Y-%m-%d %H:%M:%S")
        else:
            test_createdate = ""
        maxscore = str(testrecord[8])
        passscore = str(testrecord[9])
        ruleset = str(testrecord[10])
        test_duration = str(testrecord[11])
        allowedlanguages = str(testrecord[12])
        challengescount = str(testrecord[13])
        if testrecord[14] is not None:
            activationdate = testrecord[14].strftime("%Y-%m-%d %H:%M:%S")
        else:
            activationdate = ""
        if testrecord[15] is not None:
            publishdate = testrecord[15].strftime("%Y-%m-%d %H:%M:%S")
        else:
            publishdate = ""
        test_status = str(testrecord[16])
        allowmultipleattempts = str(testrecord[17])
        maxattemptscount = str(testrecord[18])
        attemptsinterval = str(testrecord[19])
        attemptsintervalunit = str(testrecord[20])
        randomsequencing = str(testrecord[21])
        multimediarequired = str(testrecord[22])
        testlinkid = str(testrecord[23])
        progenv = str(testrecord[24])
        test_scope = str(testrecord[25])
        test_quality = str(testrecord[26])
        negativescoreallowed = str(testrecord[27])
        # Note: the keys in the dicts are the same as the tags in the xml file (see Sample.xml) for uploading a test.
        # When we create the xml file, we will use these keys to create the nodes in the xml doc.
        challenge = {'challengeid' : str(testrecord[28]), 'challengestatement' : str(testrecord[29]), 'challengetype' : str(testrecord[30]), 'maxresponsesizeallowable' : str(testrecord[31]), 'option1' : str(testrecord[32]), 'option2' : str(testrecord[33]), 'option3' : str(testrecord[34]), 'option4' : str(testrecord[35]), 'option5' : str(testrecord[36]), 'option6' : str(testrecord[37]), 'option7' : str(testrecord[38]), 'option8' : str(testrecord[39]), 'challengescore' : str(testrecord[40]), 'negativescore' : str(testrecord[41]), 'compulsoryforuser' : str(testrecord[42]), 'correctresponse' : str(testrecord[43]), 'challengeimagepath' : str(testrecord[44]), 'externalresourcepath' : str(testrecord[45]), 'challengeallocatedtime' : str(testrecord[46]), 'challenge_subtopic_id' : str(testrecord[47]), 'challengequality' : str(testrecord[48]), 'challenge_testlinkid' : str(testrecord[49]), 'morethanoneoptioncorrect' : str(testrecord[50]), 'proglang' : str(testrecord[51]), 'mathenv' : str(testrecord[52])}
        if testname in testsdict_ascreator.keys():
            challengeslist = json.loads(testsdict_ascreator[testname]['challenges'])
            #print(challengeslist)
            challengeslist.append(challenge)
            testsdict_ascreator[testname]['challenges'] = json.dumps(challengeslist)
        else:
            challengeslist = []
            challengeslist.append(challenge)
            testsdict_ascreator[testname] = {'testid' : str(testid), 'testname' : str(testname), 'testtopic' : str(test_topicname), 'othertopicname' : str(topic_topicname), 'creatorisevaluator' : str(creatorisevaluator), 'evaluatorgroupname' : str(evalgroupname), 'testtype' : str(testtype), 'test_createdate' : str(test_createdate), 'totalscore' : str(maxscore), 'passscore' : str(passscore), 'testrules' : str(ruleset), 'testduration' : str(test_duration), 'testdurationunit' : 'seconds', 'answeringlanguage' : str(allowedlanguages), 'numberofchallenges' : str(challengescount), 'testactivationdate' : str(activationdate), 'testpublishdate' : str(publishdate), 'teststatus' : str(test_status), 'allowmultipleattempts' : str(allowmultipleattempts), 'eachchallengesamescore' : False, 'maxattemptsallowed' : str(maxattemptscount), 'intervalbetweenattempts' : str(attemptsinterval), 'intervalunits' : str(attemptsintervalunit), 'randomsequencing' : str(randomsequencing), 'needssoundimage' : str(multimediarequired), 'testlinkid' : str(testlinkid), 'programmingenvironment' : str(progenv), 'testscope' : str(test_scope), 'testtargetskilllevel' : str(test_quality), 'incorrectresponsenegativescore' : str(negativescoreallowed), 'challenges' : json.dumps(challengeslist)}
    # So, now testsdict_ascreator is a dict with test names as keys and a dict with the test attributes and challenges list as value.
    # Get tests taken by this user:
    mytestssql_candidate = "select ut.id, ut.test_id, t.testname, t.topicname, ut.emailaddr, ut.testurl, ut.validfrom, ut.validtill, ut.status, ut.outcome, ut.score, ut.starttime, ut.endtime, ut.ipaddress, ut.clientsware, ut.sessid, ut.active, ut.cancelled, ut.stringid, ut.evaluator_comment, ut.first_eval_timestamp, ut.visibility, ut.evalcommitstate, ut.disqualified, ut.schedule_id, ut.windowchangeattempts, ut.dateadded from Tests_usertest ut, Tests_test t where t.id=ut.test_id and ut.user_id=%s"
    try:
        dbcursor.execute(mytestssql_candidate, (userobj.id,))
        testrecords_ascandidate = dbcursor.fetchall()
    except:
        skillutils.disconnectdb(dbconn, dbcursor)
        message = "Error3: Something went wrong - %s"%sys.exc_info()[1].__str__() # TODO: Need to create a log of these messages and send them back with the final response.
        #response = HttpResponse(message)
        #return response
        testrecords_ascandidate = []
    for testrecord in testrecords_ascandidate:
        utid = testrecord[0]
        testid = testrecord[1]
        testname = testrecord[2]
        topicname = testrecord[3]
        emailaddress = testrecord[4]
        testurl = testrecord[5]
        if testrecord[6] is not None:
            validfrom = testrecord[6].strftime("%Y-%m-%d %H:%M:%S")
        else:
            validfrom = ""
        if testrecord[7] is not None:
            validtill = testrecord[7].strftime("%Y-%m-%d %H:%M:%S")
        else:
            validtill = ""
        status = testrecord[8]
        outcome = testrecord[9]
        score = testrecord[10]
        if testrecord[11] is not None:
            starttime = testrecord[11].strftime("%Y-%m-%d %H:%M:%S")
        else:
            starttime = ""
        if testrecord[12] is not None:
            endtime = testrecord[12].strftime("%Y-%m-%d %H:%M:%S")
        else:
            endtime = ""
        ipaddress = testrecord[13]
        clientsware = testrecord[14]
        sessionid = testrecord[15]
        active = testrecord[16]
        cancelled = testrecord[17]
        stringid = testrecord[18]
        evaluatorcomment = testrecord[19]
        firstevaltimestamp = testrecord[20]
        visibility = testrecord[21]
        evalcommitstate = testrecord[22]
        disqualified = testrecord[23]
        scheduleid = testrecord[24]
        windowchangeattempts = testrecord[25]
        dateadded = testrecord[26].strftime("%Y-%m-%d %H:%M:%S")
        testsdict_ascandidate[testname] = {'usertestid' : str(utid), 'testid' : str(testid), 'topicname' : str(topicname), 'emailaddress' : str(emailaddress), 'testurl' : str(testurl), 'validfrom' : str(validfrom), 'validtill' : str(validtill), 'status' : str(status), 'outcome' : str(outcome), 'score' : str(score), 'starttime' : str(starttime), 'endtime' : str(endtime), 'ipaddress' : str(ipaddress), 'clientsware' : str(clientsware), 'sessionid' : str(sessionid), 'active' : str(active), 'cancelled' : str(cancelled), 'stringid' : str(stringid), 'evaluatorcomment' : str(evaluatorcomment), 'firstevaltimestamp' : str(firstevaltimestamp), 'visibility' : str(visibility), 'evalcommitstate' : str(evalcommitstate), 'disqualified' : str(disqualified), 'scheduleid' : str(scheduleid), 'windowchangeattempts' : str(windowchangeattempts), 'dateadded' : str(dateadded)}
    # Get interviews created by this user:
    myinterviewssql_creator = "select i.id, i.title, i.challengescount, i.maxresponsestarttime, i.topicname, i.medium, i.language, i.createdate, i.publishdate, i.status, i.maxscore, i.maxduration, i.randomsequencing, i.interviewlinkid, i.scope, i.quality, i.challengesfilepath, i.introfilepath, i.filetype, i.realtime, i.scheduledtime, i.interviewers_count, i.interviewer_ids from Tests_interview i where i.interviewer_id=%s"
    try:
        dbcursor.execute(myinterviewssql_creator, (userobj.id,))
        interviewrecords_ascreator = dbcursor.fetchall()
    except:
        skillutils.disconnectdb(dbconn, dbcursor)
        message = "Error2: Something went wrong - %s"%sys.exc_info()[1].__str__() # TODO: Need to create a log of these messages and send them back with the final response.
        #response = HttpResponse(message)
        #return response
        interviewrecords_ascreator = []
    for interviewrecord in interviewrecords_ascreator:
        intid = interviewrecord[0]
        interviewtitle = interviewrecord[1]
        numchallenges = interviewrecord[2]
        maxresponsestarttime = interviewrecord[3]
        inttopicname = interviewrecord[4]
        medium = interviewrecord[5]
        intlanguage = interviewrecord[6]
        if interviewrecord[8] is not None:
            intcreatedate = interviewrecord[7].strftime("%Y-%m-%d %H:%M:%S")
        else:
            intcreatedate = ""
        if interviewrecord[8] is not None:
            intpublishdate = interviewrecord[8].strftime("%Y-%m-%d %H:%M:%S")
        else:
            intpublishdate = ""
        intstatus = interviewrecord[9]
        intmaxscore = interviewrecord[10]
        intmaxduration = interviewrecord[11]
        randomsequencing = interviewrecord[12]
        intlinkid = interviewrecord[13]
        intscope = interviewrecord[14]
        intquality = interviewrecord[15]
        intchallengesfilepath = interviewrecord[16]
        introfilepath = interviewrecord[17]
        intfiletype = interviewrecord[18]
        intrealtime = interviewrecord[19]
        if interviewrecord[20] is not None:
            intscheduletime = interviewrecord[20].strftime("%Y-%m-%d %H:%M:%S")
        else:
            intscheduletime = ""
        numofinterviewers = interviewrecord[21]
        interviewerids = interviewrecord[22]
        interviewsdict_ascreator[interviewtitle] = {'interviewtitle' : str(interviewtitle), 'interviewid' : str(intid), 'numchallenges' : str(numchallenges), 'maxresponsestarttime' : str(maxresponsestarttime), 'interviewtopicname' : str(inttopicname), 'medium' : str(medium), 'language' : str(intlanguage), 'createdate' : str(intcreatedate), 'publishdate' : str(intpublishdate), 'interviewstatus' : str(intstatus), 'maxscore' : str(intmaxscore), 'maxduration' : str(intmaxduration), 'randomsequencing' : str(randomsequencing), 'interviewlinkid' : str(intlinkid), 'interviewscope' : str(intscope), 'interviewquality' : str(intquality), 'challengesfilepath' : str(intchallengesfilepath), 'introfilepath' : str(introfilepath), 'filetype' : str(intfiletype), 'realtime' : str(intrealtime), 'scheduletime' : str(intscheduletime), 'numofinterviewers' : str(numofinterviewers), 'interviewerids' : str(interviewerids)} 
    # Get interviews attended by this user:
    myinterviewssql_candidate = "select ic.id, ic.interview_id, i.title, ic.emailaddr, ic.scheduledtime, ic.actualstarttime, ic.interviewlinkid, ic.totaltimetaken, ic.interviewurl, ic.dateadded from Tests_interviewcandidates ic, Tests_interview i where ic.interview_id=i.id and ic.emailaddr=%s"
    try:
        dbcursor.execute(myinterviewssql_candidate, (userobj.emailid,))
        interviewrecords_ascandidate = dbcursor.fetchall()
    except:
        skillutils.disconnectdb(dbconn, dbcursor)
        message = "Error1: Something went wrong - %s"%sys.exc_info()[1].__str__() # TODO: Need to create a log of these messages and send them back with the final response.
        interviewrecords_ascandidate = []
        #response = HttpResponse(message)
        #return response
    for interviewrecord in interviewrecords_ascandidate:
        interviewcandidateid = interviewrecord[0]
        interviewid = interviewrecord[1]
        interviewtitle = interviewrecord[2]
        candidateemailaddress = interviewrecord[3]
        if interviewrecord[4] is not None:
            scheduledtime = interviewrecord[4].strftime("%Y-%m-%d %H:%M:%S")
        else:
            scheduledtime = ""
        if interviewrecord[5] is not None:
            actualstarttime = interviewrecord[5].strftime("%Y-%m-%d %H:%M:%S")
        else:
            actualstarttime = ""
        interviewlinkid = interviewrecord[6]
        totaltime = interviewrecord[7]
        interviewurl = interviewrecord[8]
        if interviewrecord[9] is not None:
            interviewdate = interviewrecord[9].strftime("%Y-%m-%d %H:%M:%S")
        else:
            interviewdate = ""
        interviewsdict_ascandidate[interviewtitle] = {'interviewcandidateid' : str(interviewcandidateid), 'interviewid' : str(interviewid), 'interviewtitle' : str(interviewtitle), 'candidateemailaddress' : str(candidateemailaddress), 'scheduledtime' : str(scheduledtime), 'actualstarttime' : str(actualstarttime), 'interviewlinkid' : str(interviewlinkid), 'totaltime' : str(totaltime), 'interviewurl' : str(interviewurl), 'interviewdate' : str(interviewdate)}
    dumpfilepath = ""
    mimetype = "application/x-zip-compressed"
    if fmt == "xml":
        dumpfilepath = skillutils.dumpasxml(userobj.displayname, testsdict_ascreator, testsdict_ascandidate, interviewsdict_ascreator, interviewsdict_ascandidate)
    elif fmt == "csv":
        dumpfilepath = skillutils.dumpascsv(userobj.displayname, testsdict_ascreator, testsdict_ascandidate, interviewsdict_ascreator, interviewsdict_ascandidate)
        skillutils.disconnectdb(dbconn, dbcursor)
    else:
        skillutils.disconnectdb(dbconn, dbcursor)
        message = "Could not identify the format."
        response = HttpResponse(message)
        return response
    # Content-disposition of dumpfilepath
    dumpcontents = b''
    fp = open(dumpfilepath, "rb")
    dumpcontents = fp.read()
    fp.close()
    #fname = dumpfilepath.split(os.path.sep)[-1]
    fname = os.path.basename(dumpfilepath)
    response = HttpResponse(dumpcontents)
    response['Content-Type'] = mimetype
    response['Content-Length'] = dumpcontents.__len__()
    response['Content-Disposition'] = 'attachment; filename="%s"'%fname
    return response

"""
Endpoints for proctored tests: 
1) saveproctorinfo - This would be used by the browser extensions (4 in number). They will send this info to be stored in DB and filesystem.
HTTP request would include: test_id, user_email, usertest_id/wouldbeuser_id (would be determined by user_email), file, start_time, end_time.
2) identifyuser - This would be used by the browser extensions as above. HTTP request would include: user_photo (file), user_email, test_id.
Response from both endpoints: 204 for success, 403 for failure.
"""
